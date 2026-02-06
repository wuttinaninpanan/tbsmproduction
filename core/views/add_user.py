import csv
import secrets

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator

from core.decorators import admin_required
from core.models.main import UserProfile

try:
	import openpyxl  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None


def download_user_import_template(request):
	"""Download a template for importing users.

	Default format is CSV (Excel-compatible). If `?format=xlsx` is provided and
	`openpyxl` is installed, an .xlsx file will be returned.
	"""
	fmt = (request.GET.get("format") or "csv").strip().lower()

	headers = ["username", "email", "full_name", "role", "is_active", "password"]
	rows = [
		["jane.doe", "jane.doe@example.com", "Jane Doe", "user", "TRUE", ""],
		["production.staff", "staff@example.com", "สมชาย ใจดี", "staff", "TRUE", ""],
		["admin01", "admin@example.com", "ผู้ดูแล ระบบ", "admin", "TRUE", "ChangeMe123!"],
	]

	if fmt == "xlsx":
		if openpyxl is None:
			return HttpResponse(
				"XLSX format is not available (openpyxl is not installed).",
				status=400,
				content_type="text/plain; charset=utf-8",
			)
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "users"
		ws.append(headers)
		for r in rows:
			ws.append(r)
		for col in range(1, len(headers) + 1):
			ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

		response = HttpResponse(
			content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)
		response["Content-Disposition"] = 'attachment; filename="user_import_template.xlsx"'
		wb.save(response)
		return response

	# CSV (default)
	response = HttpResponse(content_type="text/csv; charset=utf-8")
	response["Content-Disposition"] = 'attachment; filename="user_import_template.csv"'
	# Add UTF-8 BOM so Excel on Windows opens Thai correctly.
	response.write("\ufeff")
	writer = csv.writer(response, lineterminator="\r\n")
	writer.writerow(headers)
	writer.writerows(rows)
	return response


def _bool_from_any(value, default=True):
	if value is None:
		return default
	if isinstance(value, bool):
		return value
	s = str(value).strip().lower()
	if s in {"1", "true", "t", "yes", "y", "on", "active"}:
		return True
	if s in {"0", "false", "f", "no", "n", "off", "disabled", "inactive"}:
		return False
	return default


def _split_full_name(full_name):
	name = (full_name or "").strip()
	if not name:
		return "", ""
	parts = name.split()
	if len(parts) == 1:
		return parts[0], ""
	return parts[0], " ".join(parts[1:])


def _apply_role(user, role):
	r = (role or "user").strip().lower()
	if r == "admin":
		user.is_staff = True
		user.is_superuser = True
	elif r == "staff":
		user.is_staff = True
		user.is_superuser = False
	else:
		user.is_staff = False
		user.is_superuser = False


def _normalized_key(key):
	return (key or "").strip().lower().replace(" ", "_")


def _parse_csv(uploaded_file):
	data = uploaded_file.read()
	try:
		text = data.decode("utf-8-sig")
	except UnicodeDecodeError:
		text = data.decode("cp874", errors="replace")
	reader = csv.DictReader(text.splitlines())
	for row in reader:
		yield {(_normalized_key(k)): v for k, v in (row or {}).items()}


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


@method_decorator(admin_required, name='dispatch')
class AddUserViews(TemplateView):
	template_name = "add_user.html"

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		# Safe defaults so template can render without a backend yet.
		ctx.setdefault("username", "")
		ctx.setdefault("email", "")
		ctx.setdefault("full_name", "")
		ctx.setdefault("department", "")
		ctx.setdefault("role", "user")
		ctx.setdefault("note", "")
		# Optional: allow passing a Django form later.
		ctx.setdefault("form", {})
		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "create").strip().lower()
		User = get_user_model()

		if action == "import":
			uploaded = request.FILES.get("excel_file")
			if not uploaded:
				messages.error(request, "กรุณาเลือกไฟล์ Excel/CSV ก่อนนำเข้า")
				return redirect("add_user")

			filename = (uploaded.name or "").lower()
			if filename.endswith(".csv"):
				rows = _parse_csv(uploaded)
			elif filename.endswith(".xlsx"):
				try:
					rows = _parse_xlsx(uploaded)
				except RuntimeError:
					messages.error(request, "ยังไม่รองรับไฟล์ .xlsx ในสภาพแวดล้อมนี้ (ต้องติดตั้ง openpyxl)")
					return redirect("add_user")
			else:
				messages.error(request, "รองรับเฉพาะไฟล์ .xlsx หรือ .csv")
				return redirect("add_user")

			created = 0
			updated = 0
			skipped = 0

			with transaction.atomic():
				for row in rows:
					username = (row.get("username") or row.get("user") or "").strip()
					if not username:
						skipped += 1
						continue

					email = (row.get("email") or "").strip()
					full_name = (row.get("full_name") or row.get("name") or row.get("fullname") or "").strip()
					role = (row.get("role") or "user").strip().lower()
					is_active = _bool_from_any(row.get("is_active"), default=True)
					password = row.get("password")
					password = (str(password).strip() if password is not None else "")
					if not password:
						password = secrets.token_urlsafe(10)

					first_name, last_name = _split_full_name(full_name)

					user = User.objects.filter(username=username).first()
					if user is None:
						user = User.objects.create_user(username=username, email=email, password=password)
						created += 1
					else:
						# Update basic fields; only reset password if explicitly provided.
						if email:
							user.email = email
						if password and (row.get("password") is not None) and str(row.get("password")).strip():
							user.set_password(password)
						updated += 1

					if first_name:
						user.first_name = first_name
					if last_name:
						user.last_name = last_name
					user.is_active = is_active
					_apply_role(user, role)
					user.save()

			messages.success(
				request,
				f"นำเข้าผู้ใช้งานสำเร็จ: เพิ่มใหม่ {created} | อัปเดต {updated} | ข้าม {skipped}",
			)
			return redirect("manage_user")

		# action == create (manual form)
		username = (request.POST.get("username") or "").strip()
		email = (request.POST.get("email") or "").strip()
		full_name = (request.POST.get("full_name") or "").strip()
		role = (request.POST.get("role") or "user").strip().lower()
		shift = (request.POST.get("shift") or "shift_day").strip()
		is_active = bool(request.POST.get("is_active"))
		password = request.POST.get("password") or ""
		password_confirm = request.POST.get("password_confirm") or ""

		if not username:
			messages.error(request, "กรุณากรอก Username")
			return render(request, self.template_name, self.get_context_data())
		if not password:
			messages.error(request, "กรุณากรอกรหัสผ่าน")
			return render(request, self.template_name, self.get_context_data(username=username, email=email, full_name=full_name, role=role))
		if password != password_confirm:
			messages.error(request, "รหัสผ่านและยืนยันรหัสผ่านไม่ตรงกัน")
			return render(request, self.template_name, self.get_context_data(username=username, email=email, full_name=full_name, role=role))
		if User.objects.filter(username=username).exists():
			messages.error(request, "Username นี้มีอยู่แล้ว")
			return render(request, self.template_name, self.get_context_data(username=username, email=email, full_name=full_name, role=role))

		first_name, last_name = _split_full_name(full_name)
		user = User.objects.create_user(username=username, email=email, password=password)
		user.first_name = first_name
		user.last_name = last_name
		user.is_active = is_active
		_apply_role(user, role)
		user.save()

		# สร้าง UserProfile สำหรับเก็บข้อมูล shift
		UserProfile.objects.get_or_create(user=user, defaults={'shift': shift})

		messages.success(request, "เพิ่มผู้ใช้งานสำเร็จ")
		return redirect("manage_user")

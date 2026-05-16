from __future__ import annotations

import re
import uuid

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.views import View
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator

from core.auth.decorators import staff_required
from core.models.item_line import ItemLine
from core.models.item_list import Item_list
from core.models.item_stage import ItemStage
from core.models.line import Line
from core.models.line_process import LineProcess
from core.services.auditlog import log_event


try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None


def _normalized_key(key: str) -> str:
    key = (key or "").strip().lower()
    key = re.sub(r"[^0-9a-z]+", "_", key)
    key = re.sub(r"_+", "_", key).strip("_")
    return key


def _excel_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _row_get_first(row: dict, *keys: str) -> str:
    for k in keys:
        if not k:
            continue
        v = row.get(k)
        s = _excel_to_str(v).strip()
        if s != "":
            return s
    return ""


def _parse_xlsx(uploaded_file):
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return
    keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
    for values in rows:
        row = {}
        for idx, value in enumerate(values):
            k = keys[idx] if idx < len(keys) else ""
            if not k:
                continue
            row[k] = value
        yield row


def download_manage_line_import_template(request):
    """Download a 2-sheet template covering both Line records and ItemLine mappings.

    Sheet `lines`     : line_name | process_type | description
    Sheet `item_lines`: sd_code   | line_name    | item_stage
    """
    if openpyxl is None:
        return HttpResponse(
            "XLSX format is not available (openpyxl is not installed).",
            status=400,
            content_type="text/plain; charset=utf-8",
        )

    # ----- Sample values pulled from existing master data -----
    process_samples: list[str] = []
    for p in LineProcess.objects.order_by("display_name", "name")[:2]:
        label = (getattr(p, "display_name", "") or getattr(p, "name", "") or "").strip()
        if label:
            process_samples.append(label)
    while len(process_samples) < 2:
        process_samples.append(f"PROCESS-{chr(ord('A') + len(process_samples))}")

    sample_sd = "SD-0001"
    item = Item_list.objects.order_by("sd_code").first()
    if item is not None and (item.sd_code or "").strip():
        sample_sd = item.sd_code.strip()
    sample_line = "LINE-01"
    real_line = Line.objects.order_by("line_name").first()
    if real_line is not None and (real_line.line_name or "").strip():
        sample_line = real_line.line_name.strip()
    sample_stage = "Semi finished goods"
    stage = ItemStage.objects.order_by("display_name", "name").first()
    if stage is not None:
        sample_stage = (stage.display_name or stage.name or sample_stage).strip()

    wb = openpyxl.Workbook()

    # Sheet 1: Production lines
    ws_lines = wb.active
    ws_lines.title = "lines"
    line_headers = ["line_name", "process_type", "description"]
    ws_lines.append(line_headers)
    for row in (
        ["LINE-01", process_samples[0], "Main line"],
        ["LINE-02", process_samples[1], ""],
    ):
        ws_lines.append(row)
    for col in range(1, len(line_headers) + 1):
        ws_lines.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    # Sheet 2: ItemLine mappings
    ws_il = wb.create_sheet(title="item_lines")
    il_headers = ["sd_code", "line_name", "item_stage"]
    ws_il.append(il_headers)
    ws_il.append([sample_sd, sample_line, sample_stage])
    for col in range(1, len(il_headers) + 1):
        ws_il.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="manage_line_import_template.xlsx"'
    wb.save(response)
    return response


def _iter_sheet_rows(wb, sheet_name: str):
    """Yield dict rows from a named sheet. Returns empty iter if sheet missing."""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return
    keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
    for values in rows:
        row: dict[str, object] = {}
        for idx, value in enumerate(values):
            k = keys[idx] if idx < len(keys) else ""
            if not k:
                continue
            row[k] = value
        yield row


def _is_uuid(value: str) -> bool:
    try:
        uuid.UUID(str(value))
    except Exception:
        return False
    return True


def _page_items(num_pages: int, current: int) -> list[int | None]:
    if num_pages <= 0:
        return []
    if num_pages <= 10:
        return list(range(1, num_pages + 1))
    items: list[int | None] = [1]
    if current > 4:
        items.append(None)
    start = max(2, current - 1)
    end = min(num_pages - 1, current + 1)
    if current <= 4:
        start, end = 2, 4
    if current >= num_pages - 3:
        start, end = num_pages - 3, num_pages - 1
    for n in range(start, end + 1):
        if 1 < n < num_pages:
            items.append(n)
    if current < num_pages - 3:
        items.append(None)
    items.append(num_pages)
    compressed: list[int | None] = []
    for it in items:
        if compressed and compressed[-1] == it:
            continue
        if it is None and compressed and compressed[-1] is None:
            continue
        compressed.append(it)
    return compressed


@method_decorator(staff_required, name="dispatch")
class ManageLineViews(TemplateView):
    template_name = "manage_line.html"

    def get(self, request, *args, **kwargs):
        action = (request.GET.get("action") or "").strip().lower()
        if action == "download_template":
            return download_manage_line_import_template(request)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        request = self.request

        q = (request.GET.get("q") or "").strip()
        per_page_raw = (request.GET.get("per_page") or "").strip()
        page = (request.GET.get("page") or "1").strip() or "1"

        allowed_per_page = {100, 200, 500, 1000}
        try:
            per_page = int(per_page_raw or 100)
        except Exception:
            per_page = 100
        if per_page not in allowed_per_page:
            per_page = 100

        qs = Line.objects.select_related("line_process").annotate(item_count=Count("item_lines"))
        if q:
            qs = qs.filter(
                Q(line_name__icontains=q)
                | Q(description__icontains=q)
                | Q(line_process__name__icontains=q)
                | Q(line_process__display_name__icontains=q)
            )

        qs = qs.order_by("line_name")
        paginator = Paginator(qs, per_page)
        page_obj = paginator.get_page(page)

        rows = []
        for line in page_obj.object_list:
            rows.append(
                {
                    "id": str(line.id),
                    "line_name": line.line_name,
                    "description": line.description or "",
                    "process_type_id": str(line.line_process_id) if line.line_process_id else "",
                    "process_type_name": getattr(line.line_process, "display_name", "")
                    or getattr(line.line_process, "name", ""),
                    "item_count": getattr(line, "item_count", 0) or 0,
                }
            )

        ctx["rows"] = rows
        ctx["processes"] = list(
            LineProcess.objects.order_by("display_name", "name").values(
                "id",
                "name",
                "display_name",
            )
        )

        ctx["q"] = q
        ctx["page_obj"] = page_obj
        ctx["paginator"] = paginator
        ctx["per_page"] = per_page
        ctx["rows_total"] = paginator.count
        ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
        ctx["total_count"] = paginator.count
        return ctx

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "").strip().lower()
        obj_id = (request.POST.get("id") or "").strip()

        if action == "import_master_data":
            if openpyxl is None:
                messages.error(request, "ไม่สามารถนำเข้า XLSX ได้: ยังไม่ได้ติดตั้ง openpyxl")
                return redirect(request.get_full_path())
            upload = request.FILES.get("excel_file")
            if upload is None:
                messages.error(request, "กรุณาเลือกไฟล์ Excel (.xlsx)")
                return redirect(request.get_full_path())
            name = (getattr(upload, "name", "") or "").lower()
            if not name.endswith(".xlsx"):
                messages.error(request, "รองรับเฉพาะไฟล์ .xlsx")
                return redirect(request.get_full_path())

            # Load workbook once; dispatch to whichever sheets exist.
            try:
                wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
            except Exception as e:
                messages.error(request, f"อ่านไฟล์ Excel ไม่สำเร็จ: {e}")
                return redirect(request.get_full_path())

            # Stats per sheet
            ln_created = ln_updated = ln_skipped = ln_proc_nf = 0
            il_created = il_updated = il_skipped = 0
            il_item_nf = il_line_nf = il_stage_nf = 0
            has_lines_sheet = "lines" in wb.sheetnames
            has_item_lines_sheet = "item_lines" in wb.sheetnames
            # Backward compat: if neither named sheet exists, treat the active sheet as `lines`.
            use_active_as_lines = not has_lines_sheet and not has_item_lines_sheet

            try:
                with transaction.atomic():
                    # ----- Sheet: lines -----
                    if has_lines_sheet or use_active_as_lines:
                        rows_iter = (
                            _iter_sheet_rows(wb, "lines")
                            if has_lines_sheet
                            else _parse_xlsx(upload)
                        )
                        # _parse_xlsx re-opens the file; rewind upload stream
                        if use_active_as_lines:
                            try:
                                upload.seek(0)
                            except Exception:
                                pass
                            rows_iter = _parse_xlsx(upload)
                        for row in rows_iter:
                            line_name = _row_get_first(row, "line_name", "line", "production_line")
                            description = _row_get_first(row, "description", "desc")
                            process_key = _row_get_first(
                                row,
                                "process_type",
                                "process_type_name",
                                "process",
                                "process_name",
                                "line_process",
                            )

                            if not line_name or not process_key:
                                ln_skipped += 1
                                continue

                            process = (
                                LineProcess.objects.filter(display_name__iexact=process_key).first()
                                or LineProcess.objects.filter(name__iexact=process_key).first()
                            )
                            if process is None:
                                ln_proc_nf += 1
                                continue

                            existing = Line.objects.filter(line_name__iexact=line_name).first()
                            if existing is None:
                                Line.objects.create(
                                    line_name=line_name,
                                    description=description,
                                    line_process=process,
                                    user=request.user,
                                )
                                ln_created += 1
                            else:
                                existing.line_name = line_name
                                existing.description = description
                                existing.line_process = process
                                existing.save(update_fields=["line_name", "description", "line_process", "updated_at"])
                                ln_updated += 1

                    # ----- Sheet: item_lines -----
                    if has_item_lines_sheet:
                        for row in _iter_sheet_rows(wb, "item_lines"):
                            sd_code = _row_get_first(row, "sd_code", "sd")
                            line_name = _row_get_first(row, "line_name", "line")
                            stage_key = _row_get_first(row, "item_stage", "stage")
                            if not sd_code or not line_name or not stage_key:
                                il_skipped += 1
                                continue

                            item = Item_list.objects.filter(sd_code__iexact=sd_code).first()
                            if item is None:
                                il_item_nf += 1
                                continue
                            line_obj = Line.objects.filter(line_name__iexact=line_name).first()
                            if line_obj is None:
                                il_line_nf += 1
                                continue
                            stage_obj = (
                                ItemStage.objects.filter(display_name__iexact=stage_key).first()
                                or ItemStage.objects.filter(name__iexact=stage_key).first()
                            )
                            if stage_obj is None:
                                il_stage_nf += 1
                                continue

                            existing_il = ItemLine.objects.filter(item=item, line=line_obj).first()
                            if existing_il is None:
                                ItemLine.objects.create(
                                    item=item,
                                    line=line_obj,
                                    item_stage=stage_obj,
                                    user=request.user,
                                )
                                il_created += 1
                            else:
                                existing_il.item_stage = stage_obj
                                existing_il.save(update_fields=["item_stage", "updated_at"])
                                il_updated += 1
            except Exception as e:
                log_event(
                    request,
                    action="line:import_master_data",
                    status="failure",
                    message="นำเข้า Line/ItemLine ไม่สำเร็จ",
                    metadata={"filename": getattr(upload, "name", ""), "error": str(e)},
                )
                messages.error(request, f"เกิดข้อผิดพลาดระหว่างนำเข้า: {e}")
                return redirect(request.get_full_path())

            # Build a combined summary message
            parts: list[str] = []
            if has_lines_sheet or use_active_as_lines:
                parts.append(
                    f"Line: +{ln_created}, อัปเดต {ln_updated}, ข้าม {ln_skipped}, ไม่พบ Process {ln_proc_nf}"
                )
            if has_item_lines_sheet:
                parts.append(
                    f"ItemLine: +{il_created}, อัปเดต {il_updated}, ข้าม {il_skipped}, "
                    f"ไม่พบ Item {il_item_nf}, ไม่พบ Line {il_line_nf}, ไม่พบ Stage {il_stage_nf}"
                )
            if not parts:
                messages.warning(request, "ไม่พบ sheet ที่ชื่อ `lines` หรือ `item_lines` ในไฟล์")
            else:
                messages.success(request, "นำเข้าสำเร็จ — " + " | ".join(parts))

            transaction.on_commit(
                lambda: log_event(
                    request,
                    action="line:import_master_data",
                    message="นำเข้า Line/ItemLine สำเร็จ",
                    metadata={
                        "filename": getattr(upload, "name", ""),
                        "line_created": ln_created,
                        "line_updated": ln_updated,
                        "line_skipped": ln_skipped,
                        "line_process_not_found": ln_proc_nf,
                        "il_created": il_created,
                        "il_updated": il_updated,
                        "il_skipped": il_skipped,
                        "il_item_not_found": il_item_nf,
                        "il_line_not_found": il_line_nf,
                        "il_stage_not_found": il_stage_nf,
                    },
                )
            )
            return redirect(request.get_full_path())

        if action == "bulk_delete_lines":
            bulk_ids = request.POST.getlist("bulk_id")
            ids = [pk for pk in bulk_ids if _is_uuid((pk or "").strip())]
            if not ids:
                messages.error(request, "กรุณาเลือกรายการที่ต้องการลบ")
                return redirect(request.get_full_path())

            deleted = 0
            blocked = 0
            not_found = 0
            try:
                with transaction.atomic():
                    for pk in ids:
                        obj = Line.objects.filter(pk=pk).first()
                        if obj is None:
                            not_found += 1
                            continue
                        try:
                            obj.delete()
                            deleted += 1
                        except ProtectedError:
                            blocked += 1
            except Exception as e:
                log_event(
                    request,
                    action="line:bulk_delete",
                    status="failure",
                    message="ลบ Line แบบ bulk ไม่สำเร็จ",
                    metadata={"selected": len(ids), "error": str(e)},
                )
                messages.error(request, f"เกิดข้อผิดพลาด: {e}")
                return redirect(request.get_full_path())

            transaction.on_commit(
                lambda: log_event(
                    request,
                    action="line:bulk_delete",
                    message="ลบ Line แบบ bulk",
                    metadata={
                        "selected": len(ids),
                        "deleted": deleted,
                        "blocked": blocked,
                        "not_found": not_found,
                    },
                )
            )

            if blocked:
                messages.warning(
                    request,
                    f"ลบสำเร็จ {deleted} รายการ (ติดใช้งาน/ลบไม่ได้ {blocked}, ไม่พบ {not_found})",
                )
            else:
                messages.success(request, f"ลบสำเร็จ {deleted} รายการ")
            return redirect(request.get_full_path())

        if action in {"update", "delete"}:
            if not _is_uuid(obj_id):
                messages.error(request, "ไม่พบรหัสรายการ")
                return redirect(request.get_full_path())

        if action == "delete":
            obj = Line.objects.filter(pk=obj_id).first()
            if obj is None:
                messages.error(request, "ไม่พบรายการ")
                return redirect(request.get_full_path())
            try:
                obj.delete()
            except ProtectedError:
                messages.error(request, "ไม่สามารถลบได้: รายการนี้ถูกใช้งานอยู่")
                return redirect(request.get_full_path())

            transaction.on_commit(
                lambda: log_event(
                    request,
                    action="line:delete",
                    message="ลบ Line",
                    metadata={"id": obj_id, "line_name": obj.line_name},
                )
            )
            messages.success(request, "ลบรายการสำเร็จ")
            return redirect(request.get_full_path())

        line_name = (request.POST.get("line_name") or "").strip()
        description = (request.POST.get("description") or "").strip()
        process_type_id = (request.POST.get("process_type_id") or "").strip()

        if not line_name:
            messages.error(request, "กรุณาระบุ Production line")
            return redirect(request.get_full_path())
        if not _is_uuid(process_type_id):
            messages.error(request, "กรุณาเลือก Process type")
            return redirect(request.get_full_path())

        process = LineProcess.objects.filter(pk=process_type_id).first()
        if process is None:
            messages.error(request, "ไม่พบ Process type")
            return redirect(request.get_full_path())

        if action == "update":
            obj = Line.objects.filter(pk=obj_id).first()
            if obj is None:
                messages.error(request, "ไม่พบรายการ")
                return redirect(request.get_full_path())
            obj.line_name = line_name
            obj.description = description
            obj.line_process = process
            obj.save(update_fields=["line_name", "description", "line_process", "updated_at"])

            transaction.on_commit(
                lambda: log_event(
                    request,
                    action="line:update",
                    message="แก้ไข Line",
                    metadata={"id": obj_id, "line_name": line_name, "process_type": str(process.id)},
                )
            )
            messages.success(request, "แก้ไขรายการสำเร็จ")
            return redirect(request.get_full_path())

        # Default: create
        try:
            obj = Line.objects.create(
                line_name=line_name,
                description=description,
                line_process=process,
                user=request.user,
            )
        except Exception as e:
            log_event(
                request,
                action="line:create",
                status="failure",
                message="เพิ่ม Line ไม่สำเร็จ",
                metadata={"error": str(e)},
            )
            messages.error(request, f"เพิ่มข้อมูลไม่สำเร็จ: {e}")
            return redirect(request.get_full_path())

        transaction.on_commit(
            lambda: log_event(
                request,
                action="line:create",
                message="เพิ่ม Line",
                metadata={"id": str(obj.id), "line_name": obj.line_name, "process_type": str(process.id)},
            )
        )
        messages.success(request, "เพิ่มข้อมูลสำเร็จ")
        return redirect(request.get_full_path())


@method_decorator(staff_required, name="dispatch")
class ManageLineEditViews(TemplateView):
    template_name = "manage_line_edit.html"

    def _get_line(self, line_id):
        return (
            Line.objects.select_related("line_process")
            .filter(pk=line_id)
            .first()
        )

    def get(self, request, *args, **kwargs):
        line = self._get_line(kwargs.get("id"))
        if line is None:
            messages.error(request, "ไม่พบ Line")
            return redirect("manage_line")
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        line = self._get_line(kwargs.get("id"))
        ctx["line"] = line
        if line is None:
            ctx["item_lines"] = []
            ctx["processes"] = []
            ctx["stages"] = []
            return ctx

        item_lines_qs = (
            ItemLine.objects.select_related("item", "item_stage")
            .filter(line=line)
            .order_by("item__sd_code", "item__part_number")
        )
        ctx["item_lines"] = [
            {
                "id": str(il.id),
                "item_id": str(il.item_id) if il.item_id else "",
                "sd_code": getattr(il.item, "sd_code", "") if il.item_id else "",
                "part_number": getattr(il.item, "part_number", "") if il.item_id else "",
                "part_name": getattr(il.item, "part_name", "") if il.item_id else "",
                "stage_id": str(il.item_stage_id) if il.item_stage_id else "",
                "stage_name": (
                    getattr(il.item_stage, "display_name", "")
                    or getattr(il.item_stage, "name", "")
                )
                if il.item_stage_id
                else "",
            }
            for il in item_lines_qs
        ]
        ctx["processes"] = list(
            LineProcess.objects.order_by("display_name", "name").values(
                "id", "name", "display_name"
            )
        )
        ctx["stages"] = list(
            ItemStage.objects.order_by("display_name", "name").values(
                "id", "name", "display_name"
            )
        )
        return ctx

    def post(self, request, *args, **kwargs):
        line = self._get_line(kwargs.get("id"))
        if line is None:
            messages.error(request, "ไม่พบ Line")
            return redirect("manage_line")

        action = (request.POST.get("action") or "").strip().lower()

        if action == "update_line":
            line_name = (request.POST.get("line_name") or "").strip()
            description = (request.POST.get("description") or "").strip()
            process_type_id = (request.POST.get("process_type_id") or "").strip()

            if not line_name:
                messages.error(request, "กรุณาระบุ Production line")
                return redirect("manage_line_edit", id=line.id)
            if not _is_uuid(process_type_id):
                messages.error(request, "กรุณาเลือก Process type")
                return redirect("manage_line_edit", id=line.id)
            process = LineProcess.objects.filter(pk=process_type_id).first()
            if process is None:
                messages.error(request, "ไม่พบ Process type")
                return redirect("manage_line_edit", id=line.id)

            line.line_name = line_name
            line.description = description
            line.line_process = process
            line.save(
                update_fields=["line_name", "description", "line_process", "updated_at"]
            )
            transaction.on_commit(
                lambda: log_event(
                    request,
                    action="line:update",
                    message="แก้ไข Line",
                    metadata={
                        "id": str(line.id),
                        "line_name": line_name,
                        "process_type": str(process.id),
                    },
                )
            )
            messages.success(request, "แก้ไขรายการสำเร็จ")
            return redirect("manage_line_edit", id=line.id)

        if action == "add_item":
            item_id = (request.POST.get("item_id") or "").strip()
            stage_id = (request.POST.get("stage_id") or "").strip()
            if not _is_uuid(item_id):
                messages.error(request, "กรุณาเลือก Item")
                return redirect("manage_line_edit", id=line.id)
            if not _is_uuid(stage_id):
                messages.error(request, "กรุณาเลือก Stage")
                return redirect("manage_line_edit", id=line.id)
            item = Item_list.objects.filter(pk=item_id).first()
            stage = ItemStage.objects.filter(pk=stage_id).first()
            if item is None:
                messages.error(request, "ไม่พบ Item")
                return redirect("manage_line_edit", id=line.id)
            if stage is None:
                messages.error(request, "ไม่พบ Stage")
                return redirect("manage_line_edit", id=line.id)
            try:
                obj = ItemLine.objects.create(
                    item=item, line=line, item_stage=stage, user=request.user
                )
            except IntegrityError:
                messages.error(request, "ข้อมูลซ้ำ: Item นี้ผูกกับ Line นี้แล้ว")
                return redirect("manage_line_edit", id=line.id)
            transaction.on_commit(
                lambda: log_event(
                    request,
                    action="item_line:create",
                    message="เพิ่ม Item เข้า Line",
                    metadata={
                        "id": str(obj.id),
                        "line": line.line_name,
                        "sd_code": item.sd_code,
                        "stage": str(stage.id),
                    },
                )
            )
            messages.success(request, "เพิ่ม Item สำเร็จ")
            return redirect("manage_line_edit", id=line.id)

        if action == "update_item_stage":
            item_line_id = (request.POST.get("item_line_id") or "").strip()
            stage_id = (request.POST.get("stage_id") or "").strip()
            if not _is_uuid(item_line_id) or not _is_uuid(stage_id):
                messages.error(request, "ข้อมูลไม่ถูกต้อง")
                return redirect("manage_line_edit", id=line.id)
            obj = ItemLine.objects.filter(pk=item_line_id, line=line).first()
            stage = ItemStage.objects.filter(pk=stage_id).first()
            if obj is None or stage is None:
                messages.error(request, "ไม่พบรายการ")
                return redirect("manage_line_edit", id=line.id)
            obj.item_stage = stage
            obj.save(update_fields=["item_stage", "updated_at"])
            transaction.on_commit(
                lambda: log_event(
                    request,
                    action="item_line:update",
                    message="แก้ Stage ของ Item ใน Line",
                    metadata={"id": str(obj.id), "stage": str(stage.id)},
                )
            )
            messages.success(request, "อัปเดต Stage สำเร็จ")
            return redirect("manage_line_edit", id=line.id)

        if action == "delete_item":
            item_line_id = (request.POST.get("item_line_id") or "").strip()
            if not _is_uuid(item_line_id):
                messages.error(request, "ข้อมูลไม่ถูกต้อง")
                return redirect("manage_line_edit", id=line.id)
            obj = ItemLine.objects.filter(pk=item_line_id, line=line).first()
            if obj is None:
                messages.error(request, "ไม่พบรายการ")
                return redirect("manage_line_edit", id=line.id)
            try:
                obj.delete()
            except ProtectedError:
                messages.error(request, "ลบไม่ได้: รายการนี้ถูกใช้งานอยู่")
                return redirect("manage_line_edit", id=line.id)
            transaction.on_commit(
                lambda: log_event(
                    request,
                    action="item_line:delete",
                    message="ลบ Item ออกจาก Line",
                    metadata={"id": item_line_id, "line": line.line_name},
                )
            )
            messages.success(request, "ลบ Item ออกจาก Line สำเร็จ")
            return redirect("manage_line_edit", id=line.id)

        messages.error(request, "ไม่รองรับการทำงานนี้")
        return redirect(request.get_full_path())


@method_decorator(staff_required, name="dispatch")
class LineItemSearchView(View):
    """AJAX endpoint สำหรับค้น Item เพื่อใช้ใน autocomplete ของหน้า manage-line edit."""

    def get(self, request, *args, **kwargs):
        q = (request.GET.get("q") or "").strip()
        line_id = (request.GET.get("line_id") or "").strip()

        qs = Item_list.objects.all()
        if q:
            qs = qs.filter(
                Q(sd_code__icontains=q)
                | Q(part_number__icontains=q)
                | Q(part_name__icontains=q)
                | Q(item_code__icontains=q)
            )

        if _is_uuid(line_id):
            already_linked = ItemLine.objects.filter(line_id=line_id).values_list(
                "item_id", flat=True
            )
            qs = qs.exclude(pk__in=list(already_linked))

        qs = qs.order_by("sd_code", "part_number")[:20]
        results = [
            {
                "id": str(it.id),
                "sd_code": it.sd_code or "",
                "part_number": it.part_number or "",
                "part_name": it.part_name or "",
                "item_code": it.item_code or "",
            }
            for it in qs
        ]
        return JsonResponse({"results": results})

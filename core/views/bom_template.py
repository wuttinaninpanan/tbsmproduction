from __future__ import annotations

import re
import uuid
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.shortcuts import redirect

from core.auth.decorators import staff_required
from core.models.bill_of_material import BillOfMaterial
from core.models.bill_of_material_item_master import BillOfMaterialItemMater
from core.models.item_category import ItemCategory
from core.models.item_line import ItemLine
from core.models.item_list import Item_list, extract_item_number, format_item_code, is_spreadsheet_error
from core.services.item_import import SimilarSdIndex, fill_only_update
from core.models.item_stage import ItemStage
from core.models.line import Line
from core.models.portion import Portion
from core.models.side import Side
from core.models.inout import InOut
from core.models.way import Way


try:
	import openpyxl  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None


# Single canonical sheet layout shared by export, the import template, and the
# importer — so a file exported here can be edited and re-imported unchanged.
# Columns "0".."10" mark the tree level; hierarchy on import is driven by M2M
# (the parent item's Item Code).
BOM_SHEET_HEADERS = [str(n) for n in range(11)] + [
	"Item Code",
	"M2M",
	"SD Code",
	"Part No.",
	"Part Name",
	"Category",
	"Stage",
	"Line",
	"Quantity",
	"Unit",
	"Sequence",
	"Revision",
	"Latest ECI",
	"Scrap %",
	"SKU",
	"Weight (kg)",
	"Purchased Price",
	"Cost",
	"Comment",
	"Portion",
	"Side",
	"InOut",
	"Way",
]


def _normalized_key(key: str) -> str:
	key = (key or "").strip().lower()
	key = re.sub(r"[^0-9a-z]+", "_", key)
	key = re.sub(r"_+", "_", key).strip("_")
	return key


def _excel_to_str(value) -> str:
	if value is None:
		return ""
	if isinstance(value, bool):
		return "TRUE" if value else "FALSE"
	if isinstance(value, int):
		return str(value)
	if isinstance(value, float):
		if value.is_integer():
			return str(int(value))
		return str(value)
	return str(value).strip()


def _excel_number(value):
	"""Return a float so Excel treats the cell as numeric; "" for empty/None."""
	if value is None:
		return ""
	try:
		return float(value)
	except (TypeError, ValueError):
		return ""


def _row_get_first(row: dict, *keys: str) -> str:
	for k in keys:
		if not k:
			continue
		v = row.get(k)
		s = _excel_to_str(v).strip()
		if s != "":
			return s
	return ""


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


def _safe_decimal(value, default: Decimal = Decimal("0")) -> Decimal:
	if value is None:
		return default
	if isinstance(value, Decimal):
		return value
	if isinstance(value, (int, float)):
		try:
			return Decimal(str(value))
		except (InvalidOperation, ValueError):
			return default
	value = str(value).strip()
	if value == "":
		return default
	try:
		return Decimal(value)
	except (InvalidOperation, ValueError):
		return default


def _sanitize_sku_base(value: str) -> str:
	from django.utils.text import slugify
	value = (value or "").strip()
	if not value:
		return ""
	base = slugify(value).replace("-", "_")
	base = (base or "").strip("_")
	return base.upper()


def _generate_unique_sku(*, part_number: str, sd_code: str) -> str:
	"""Build a SKU that fits Item_list.sku (max 100) and isn't already used."""
	base = _sanitize_sku_base(part_number) or _sanitize_sku_base(sd_code) or "ITEM"
	for _ in range(20):
		suffix = uuid.uuid4().hex[:8].upper()
		max_base_len = 100 - 1 - len(suffix)
		trimmed = base[:max_base_len]
		candidate = f"{trimmed}-{suffix}" if trimmed else suffix
		if not Item_list.objects.filter(sku__iexact=candidate).exists():
			return candidate
	return uuid.uuid4().hex[:12].upper()


def download_bom_data_excel(request):
	"""Export the current BOM data as XLSX.

	Rows are ordered by FG group, depth-first within each FG. Each row carries:
	  * the level columns (0..10) + component identity (Item Code, M2M, SD Code,
	    Part No., Part Name, Category, Stage, Line);
	  * BOM line data — Quantity / Unit / Sequence — i.e. how the item is used in
	    its immediate parent (blank on the top-level FG row);
	  * BOM header data — Revision / Latest ECI / Scrap % — taken from the BOM
	    owned by the item on that row (blank for leaf items with no sub-BOM);
	  * extra Item_list fields — SKU, Weight, Purchased Price, Cost, Comment,
	    Portion, Side, InOut, Way.

	The import-relevant headers (SD Code, M2M, Part No., Quantity, Unit, ...) keep
	their names so the file can still be edited and re-imported; the importer
	ignores the extra reference columns.
	"""
	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)

	all_rows = list(
		BillOfMaterialItemMater.objects
		.select_related(
			"bom__item",
			"bom__item__stage",
			"bom__item__category",
			"component",
			"component__stage",
			"component__category",
			"component__portion",
			"component__side",
			"component__inout",
			"component__way",
		)
		.all()
	)

	item_line_map: dict = {}
	for il in ItemLine.objects.select_related("line").all():
		item_line_map.setdefault(il.item_id, il.line)

	# item_id -> the BillOfMaterial header owned by that item (FG / sub-assembly).
	# Used to surface revision / latest_eci / scrap_percent on the item's row.
	bom_by_item_id: dict = {
		b.item_id: b for b in BillOfMaterial.objects.all()
	}

	children_by_parent: dict = {}
	parent_items: dict = {}
	component_ids: set = set()
	for r in all_rows:
		parent_item_id = r.bom.item_id if r.bom_id and r.bom and r.bom.item_id else None
		if parent_item_id:
			children_by_parent.setdefault(parent_item_id, []).append(r)
			parent_items[parent_item_id] = r.bom.item
		if r.component_id:
			component_ids.add(r.component_id)

	for lst in children_by_parent.values():
		lst.sort(key=lambda x: (
			x.sequence or 0,
			getattr(x.component, "sd_code", "") or "",
		))

	fg_ids = [pid for pid in parent_items.keys() if pid not in component_ids]
	fg_ids.sort(key=lambda pid: getattr(parent_items.get(pid), "sd_code", "") or "")

	def _stage_display(stage):
		if stage is None:
			return ""
		name = getattr(stage, "display_name", "") or getattr(stage, "name", "")
		prefix = getattr(stage, "code_prefix", "") or ""
		return f"{name} — {prefix}" if prefix else (name or "")

	def _row_for(item, level=0, m2m="", bom_row=None):
		category = getattr(item, "category", None)
		stage = getattr(item, "stage", None)
		line = item_line_map.get(getattr(item, "id", None))
		own_bom = bom_by_item_id.get(getattr(item, "id", None))
		portion = getattr(item, "portion", None)
		side = getattr(item, "side", None)
		inout = getattr(item, "inout", None)
		way = getattr(item, "way", None)

		level_cells = [""] * 11
		if isinstance(level, int) and 0 <= level <= 10:
			level_cells[level] = level

		# How this item is used inside its immediate parent (blank for FG rows).
		qty = _excel_number(getattr(bom_row, "quantity", None)) if bom_row else ""
		unit = (getattr(bom_row, "unit", "") or "") if bom_row else ""
		sequence = getattr(bom_row, "sequence", "") if bom_row else ""

		# BOM header owned by this item (FG / sub-assembly); blank for leaf items.
		revision = (getattr(own_bom, "revision", "") or "") if own_bom else ""
		latest_eci = (getattr(own_bom, "latest_eci", "") or "") if own_bom else ""
		scrap = _excel_number(getattr(own_bom, "scrap_percent", None)) if own_bom else ""

		return level_cells + [
			getattr(item, "item_code", "") or "",
			m2m or "",
			getattr(item, "sd_code", "") or "",
			getattr(item, "part_number", "") or "",
			getattr(item, "part_name", "") or "",
			getattr(category, "name", "") or "" if category else "",
			_stage_display(stage),
			getattr(line, "line_name", "") or "" if line else "",
			# --- BOM line (BillOfMaterialItemMater) ---
			qty,
			unit,
			sequence,
			# --- BOM header (BillOfMaterial) ---
			revision,
			latest_eci,
			scrap,
			# --- Item_list extras ---
			getattr(item, "sku", "") or "",
			_excel_number(getattr(item, "weight", None)),
			_excel_number(getattr(item, "purchased_price", None)),
			_excel_number(getattr(item, "cost", None)),
			getattr(item, "comment", "") or "",
			getattr(portion, "title", "") or "" if portion else "",
			getattr(side, "title", "") or "" if side else "",
			getattr(inout, "title", "") or "" if inout else "",
			getattr(way, "title", "") or "" if way else "",
		]

	output_rows: list = []
	visiting: set = set()
	emitted: set = set()

	def dfs(parent, level):
		parent_code = getattr(parent, "item_code", "") or ""
		for r in children_by_parent.get(parent.id, []):
			comp = r.component
			if comp is None:
				continue
			row_key = (parent.id, comp.id)
			if row_key in emitted:
				continue
			emitted.add(row_key)
			output_rows.append(_row_for(comp, level=level + 1, m2m=parent_code, bom_row=r))
			if comp.id not in visiting:
				visiting.add(comp.id)
				try:
					dfs(comp, level + 1)
				finally:
					visiting.discard(comp.id)

	for fg_id in fg_ids:
		fg_item = parent_items.get(fg_id)
		if fg_item is None:
			continue
		# FG row first — level 0, M2M shown as "-" since FG has no parent.
		output_rows.append(_row_for(fg_item, level=0, m2m="-"))
		dfs(fg_item, 0)

	headers = BOM_SHEET_HEADERS

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "bom_template"
	ws.append(headers)
	for r in output_rows:
		ws.append(r)
	for col in range(1, len(headers) + 1):
		# Narrow level columns (1..11), wider data columns afterwards.
		width = 5 if col <= 11 else 22
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="bom_template_data.xlsx"'
	wb.save(response)
	return response


def download_bom_import_template(request):
	"""Download an XLSX template that matches the export layout (BOM_SHEET_HEADERS).

	Hierarchy is driven by ``M2M`` = the parent item's **Item Code** (leave blank
	or "-" for a top-level FG). The level columns 0..10 are visual only. Because
	M2M references an Item Code, parents must already exist (e.g. when round-
	tripping an exported file); brand-new parents have no Item Code yet.
	"""
	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)

	headers = BOM_SHEET_HEADERS
	idx = {h: i for i, h in enumerate(headers)}

	def make_row(level, **vals):
		r = [""] * len(headers)
		if isinstance(level, int) and 0 <= level <= 10:
			r[idx[str(level)]] = level
		for key, value in vals.items():
			r[idx[key]] = value
		return r

	rows = [
		make_row(0, **{"Item Code": "G000001", "M2M": "-", "SD Code": "150-15", "Part No.": "71013-X1424", "Part Name": "FRAME SUB-ASSY, FR SEAT BACK, RH", "Revision": "A"}),
		make_row(1, **{"Item Code": "W000002", "M2M": "G000001", "SD Code": "GP2-01", "Part No.": "71151-X1401", "Part Name": "FRAME, FR SEAT BACK, RH", "Quantity": 1, "Unit": "PCS", "Sequence": 1, "Revision": "A"}),
		make_row(1, **{"M2M": "G000001", "SD Code": "FP2-01", "Part No.": "71162-X1411", "Part Name": "PLATE, FR SEAT BACK, LWR", "Quantity": 1, "Unit": "PCS", "Sequence": 2}),
		make_row(2, **{"M2M": "W000002", "SD Code": "STT001", "Part No.": "SOLVEST-114-00", "Part Name": "SOLVEST114 (20 kg./Pail)", "Quantity": 0.5, "Unit": "KG", "Sequence": 1}),
	]

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "bom_template"
	ws.append(headers)
	for r in rows:
		ws.append(r)
	for col in range(1, len(headers) + 1):
		width = 5 if col <= 11 else 22
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="bom_template_import.xlsx"'
	wb.save(response)
	return response


def _reclassify_bom_stages_and_codes():
	"""Re-run the FG/WIP/Raw classification + item_code generation across all
	items currently referenced in BOMs. Preserves each item's numeric portion
	when only the prefix changes (so external references stay matchable)."""
	from core.models.item_list import (
		extract_item_number as _extract_item_number,
		format_item_code as _format_item_code,
	)

	pairs = list(
		BillOfMaterialItemMater.objects.values_list("component_id", "bom__item_id")
	)
	item_to_parents: dict = {}
	parent_ids: set = set()
	component_ids: set = set()
	for component_id, parent_item_id in pairs:
		if component_id and parent_item_id:
			item_to_parents.setdefault(component_id, set()).add(parent_item_id)
		if parent_item_id:
			parent_ids.add(parent_item_id)
		if component_id:
			component_ids.add(component_id)

	all_items_in_boms = parent_ids | component_ids
	if not all_items_in_boms:
		return

	levels: dict = {}
	visiting: set = set()

	def compute_level(item_id):
		if item_id in levels:
			return levels[item_id]
		parents = item_to_parents.get(item_id)
		if not parents:
			levels[item_id] = 0
			return 0
		if item_id in visiting:
			return 0
		visiting.add(item_id)
		try:
			parent_levels = [compute_level(pid) for pid in parents]
		finally:
			visiting.discard(item_id)
		lvl = max(parent_levels) + 1 if parent_levels else 0
		levels[item_id] = lvl
		return lvl

	for item_id in all_items_in_boms:
		compute_level(item_id)
	if not levels:
		return

	max_level = max(levels.values())
	stages = {
		s.name: s for s in ItemStage.objects.filter(name__in=["fg", "wip", "raw_mat"])
	}
	fg_stage = stages.get("fg")
	wip_stage = stages.get("wip")
	raw_stage = stages.get("raw_mat")
	if not (fg_stage and wip_stage and raw_stage):
		return

	target_by_item: dict = {}
	for iid, lvl in levels.items():
		if lvl == 0:
			target_by_item[iid] = fg_stage
		elif lvl == max_level and max_level > 0:
			target_by_item[iid] = raw_stage
		else:
			target_by_item[iid] = wip_stage

	items = list(Item_list.objects.filter(id__in=list(all_items_in_boms)))
	for item in items:
		target = target_by_item.get(item.id)
		if not target:
			continue
		prefix = (target.code_prefix or "").strip()
		stage_changed = item.stage_id != target.id

		if stage_changed:
			item.stage = target
			existing_num = _extract_item_number(item.item_code)
			if prefix and existing_num is not None:
				item.item_code = _format_item_code(prefix, existing_num)
			else:
				item.item_code = None

		if stage_changed:
			if item.item_code:
				item.save(update_fields=["stage", "item_code", "updated_at"])
			else:
				item.save()  # triggers Item_list.save() auto-generation
		elif not item.item_code and item.stage_id:
			item.save()


def _is_uuid(value: str) -> bool:
	try:
		uuid.UUID(str(value))
	except Exception:
		return False
	return True


def _page_items(num_pages: int, current: int) -> list[int | None]:
	if num_pages <= 0:
		return []
	if num_pages <= 10:
		return list(range(1, num_pages + 1))
	items: list[int | None] = [1]
	if current > 4:
		items.append(None)
	start = max(2, current - 1)
	end = min(num_pages - 1, current + 1)
	if current <= 4:
		start, end = 2, 4
	if current >= num_pages - 3:
		start, end = num_pages - 3, num_pages - 1
	for n in range(start, end + 1):
		if 1 < n < num_pages:
			items.append(n)
	if current < num_pages - 3:
		items.append(None)
	items.append(num_pages)
	compressed: list[int | None] = []
	for it in items:
		if compressed and compressed[-1] == it:
			continue
		if it is None and compressed and compressed[-1] is None:
			continue
		compressed.append(it)
	return compressed


def _sd_prefix(sd_code: str) -> str:
	if not sd_code:
		return ""
	return sd_code.split("-", 1)[0]


def _sd_suffix(sd_code: str) -> str:
	"""Everything after the first '-' in sd_code (e.g. 'NOSD-72510-X7V45' -> '72510-X7V45')."""
	if not sd_code:
		return ""
	parts = sd_code.split("-", 1)
	return parts[1] if len(parts) > 1 else ""


@method_decorator(staff_required, name="dispatch")
class BomTemplateView(TemplateView):
	template_name = "core/bom_template.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			return download_bom_import_template(request)
		if action == "download_data":
			return download_bom_data_excel(request)
		return super().get(request, *args, **kwargs)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request

		q = (request.GET.get("q") or "").strip()
		per_page_raw = (request.GET.get("per_page") or "").strip()
		page = (request.GET.get("page") or "1").strip() or "1"

		allowed_per_page = {100, 200, 500, 1000}
		try:
			per_page = int(per_page_raw or 100)
		except Exception:
			per_page = 100
		if per_page not in allowed_per_page:
			per_page = 100

		all_rows = list(
			BillOfMaterialItemMater.objects
			.select_related(
				"bom__item",
				"bom__item__stage",
				"bom__item__category",
				"component",
				"component__stage",
				"component__category",
			)
			.all()
		)

		# Map item_id -> first ItemLine (used to display the row's Line).
		item_line_map: dict = {}
		for il in ItemLine.objects.values("item_id", "line_id"):
			item_line_map.setdefault(il["item_id"], il["line_id"])
		# line_id -> line_name for cheap lookup when rendering rows.
		line_name_by_id = {
			str(row["id"]): row["line_name"] or ""
			for row in Line.objects.values("id", "line_name")
		}

		children_by_parent: dict = {}
		parent_items: dict = {}
		component_ids: set = set()
		for r in all_rows:
			parent_item_id = r.bom.item_id if r.bom_id and r.bom and r.bom.item_id else None
			if parent_item_id:
				children_by_parent.setdefault(parent_item_id, []).append(r)
				parent_items[parent_item_id] = r.bom.item
			if r.component_id:
				component_ids.add(r.component_id)

		for lst in children_by_parent.values():
			lst.sort(key=lambda x: (
				x.sequence or 0,
				getattr(x.component, "sd_code", "") or "",
			))

		fg_ids = [pid for pid in parent_items.keys() if pid not in component_ids]
		fg_ids.sort(key=lambda pid: getattr(parent_items.get(pid), "sd_code", "") or "")

		flat: list = []
		emitted_row_ids: set = set()
		visiting: set = set()

		def dfs(item_id, level, fg_id):
			for row in children_by_parent.get(item_id, []):
				if row.id in emitted_row_ids:
					continue
				flat.append((level + 1, row, fg_id))
				emitted_row_ids.add(row.id)
				cid = row.component_id
				if cid and cid not in visiting:
					visiting.add(cid)
					try:
						dfs(cid, level + 1, fg_id)
					finally:
						visiting.discard(cid)

		for fg_id in fg_ids:
			dfs(fg_id, 0, fg_id)

		# Orphan rows (in case the BOM tree contains cycles excluding it from any FG)
		for r in all_rows:
			if r.id in emitted_row_ids:
				continue
			parent_item_id = r.bom.item_id if r.bom_id and r.bom and r.bom.item_id else None
			flat.append((1, r, parent_item_id))
			emitted_row_ids.add(r.id)

		rich_rows: list = []
		for level, obj, fg_id in flat:
			component = obj.component if obj.component_id else None
			parent_item = obj.bom.item if obj.bom_id and obj.bom and obj.bom.item_id else None
			fg_item = parent_items.get(fg_id) if fg_id else None
			component_sd_code = getattr(component, "sd_code", "") or ""
			parent_item_code = getattr(parent_item, "item_code", "") or ""
			component_stage = getattr(component, "stage", None) if component else None
			fg_sd_code = getattr(fg_item, "sd_code", "") or ""
			fg_stage = getattr(fg_item, "stage", None) if fg_item else None
			fg_category = getattr(fg_item, "category", None) if fg_item else None
			fg_part_number_raw = getattr(fg_item, "part_number", "") or ""
			component_category = getattr(component, "category", None) if component else None

			component_id = getattr(component, "id", None)
			fg_item_id_val = getattr(fg_item, "id", None)
			component_line_id = item_line_map.get(component_id) if component_id else None
			fg_line_id = item_line_map.get(fg_item_id_val) if fg_item_id_val else None
			rich_rows.append({
				"id": str(obj.id),
				"level": level,
				"item_id": str(component_id) if component_id else "",
				"category_id": str(component.category_id) if component and component.category_id else "",
				"stage_id": str(component.stage_id) if component and component.stage_id else "",
				"line_id": str(component_line_id) if component_line_id else "",
				"item_code": getattr(component, "item_code", "") or "",
				"m2m": parent_item_code,
				"sd_code": component_sd_code,
				"part_number": getattr(component, "part_number", "") or "",
				"part_name": getattr(component, "part_name", "") or "",
				"quantity": obj.quantity,
				"category_name": getattr(component_category, "name", "") or "",
				"stage_name": getattr(component_stage, "display_name", "") or "",
				"line_name": line_name_by_id.get(str(component_line_id), "") if component_line_id else "",
				"fg_part_name": getattr(fg_item, "part_name", "") or "",
				"fg_item_code": getattr(fg_item, "item_code", "") or "",
				"fg_part_number": fg_part_number_raw,
				"fg_sd_code": fg_sd_code,
				"fg_category_name": getattr(fg_category, "name", "") or "",
				"fg_stage_name": getattr(fg_stage, "display_name", "") or "",
				"fg_item_id": str(fg_item_id_val) if fg_item_id_val else "",
				"fg_category_id": str(fg_item.category_id) if fg_item and fg_item.category_id else "",
				"fg_stage_id": str(fg_item.stage_id) if fg_item and fg_item.stage_id else "",
				"fg_line_id": str(fg_line_id) if fg_line_id else "",
				"fg_line_name": line_name_by_id.get(str(fg_line_id), "") if fg_line_id else "",
			})

		if q:
			ql = q.lower()

			def match(r):
				return (
					ql in (r["sd_code"] or "").lower()
					or ql in (r["m2m"] or "").lower()
					or ql in (r["item_code"] or "").lower()
					or ql in (r["part_number"] or "").lower()
					or ql in (r["part_name"] or "").lower()
					or ql in (r["stage_name"] or "").lower()
					or ql in (r["fg_sd_code"] or "").lower()
					or ql in (r["fg_part_name"] or "").lower()
				)

			rich_rows = [r for r in rich_rows if match(r)]

		paginator = Paginator(rich_rows, per_page)
		page_obj = paginator.get_page(page)

		ctx["rows"] = list(page_obj.object_list)
		ctx["q"] = q
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
		ctx["total_count"] = paginator.count
		ctx["level_range"] = list(range(0, 11))
		# columns after level: Item Code, M2M, SD Code, Part No., Part Name, Quantity, Category, Stage, Line = 9
		ctx["group_colspan"] = len(ctx["level_range"]) + 9
		ctx["categories"] = list(
			ItemCategory.objects.order_by("name").values("id", "name")
		)
		ctx["stages"] = list(
			ItemStage.objects.order_by("display_name").values(
				"id", "display_name", "code_prefix"
			)
		)
		ctx["lines"] = list(
			Line.objects.order_by("line_name").values("id", "line_name")
		)
		# Source items for Add modal — same filter as /products/ listing:
		# Item_list rows that have at least one ItemLine assignment.
		ctx["available_products"] = list(
			Item_list.objects
			.filter(id__in=ItemLine.objects.values("item_id"))
			.order_by("sd_code", "item_code")
			.values("id", "item_code", "sd_code", "part_number", "part_name")
		)
		# Items that already have a BillOfMaterial (FG candidates for parent-of-component selection).
		fg_ids_qs = BillOfMaterial.objects.values_list("item_id", flat=True)
		ctx["fg_items"] = list(
			Item_list.objects
			.filter(id__in=fg_ids_qs)
			.order_by("sd_code", "item_code")
			.values("id", "item_code", "sd_code", "part_number", "part_name")
		)
		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		if action == "import_master_data":
			return self._handle_import(request, *args, **kwargs)
		if action == "add_to_bom":
			return self._handle_add_to_bom(request)
		if action == "delete_bom_row":
			return self._handle_delete_bom_row(request)
		if action == "delete_fg_bom":
			return self._handle_delete_fg_bom(request)
		if action != "save_row":
			return JsonResponse({"ok": False, "error": "ไม่รองรับการทำงานนี้"}, status=400)

		item_id = (request.POST.get("item_id") or "").strip()
		category_id = (request.POST.get("category_id") or "").strip()
		stage_id = (request.POST.get("stage_id") or "").strip()
		line_id = (request.POST.get("line_id") or "").strip()

		if not _is_uuid(item_id):
			return JsonResponse({"ok": False, "error": "ไม่พบรหัสรายการ"}, status=400)

		item = Item_list.objects.filter(pk=item_id).first()
		if item is None:
			return JsonResponse({"ok": False, "error": "ไม่พบ Item"}, status=404)

		category = None
		if category_id:
			if not _is_uuid(category_id):
				return JsonResponse({"ok": False, "error": "Category ไม่ถูกต้อง"}, status=400)
			category = ItemCategory.objects.filter(pk=category_id).first()
			if category is None:
				return JsonResponse({"ok": False, "error": "ไม่พบ Category"}, status=404)

		stage = None
		if stage_id:
			if not _is_uuid(stage_id):
				return JsonResponse({"ok": False, "error": "Stage ไม่ถูกต้อง"}, status=400)
			stage = ItemStage.objects.filter(pk=stage_id).first()
			if stage is None:
				return JsonResponse({"ok": False, "error": "ไม่พบ Stage"}, status=404)

		line = None
		if line_id:
			if not _is_uuid(line_id):
				return JsonResponse({"ok": False, "error": "Line ไม่ถูกต้อง"}, status=400)
			line = Line.objects.filter(pk=line_id).first()
			if line is None:
				return JsonResponse({"ok": False, "error": "ไม่พบ Line"}, status=404)

		try:
			with transaction.atomic():
				updated_fields: list = []
				if item.category_id != (category.id if category else None):
					item.category = category
					updated_fields.append("category")

				stage_changed = item.stage_id != (stage.id if stage else None)
				if stage_changed:
					item.stage = stage
					updated_fields.append("stage")

					new_prefix = (
						(getattr(stage, "code_prefix", "") or "").strip()
						if stage else ""
					)
					existing_num = extract_item_number(item.item_code)

					if new_prefix and existing_num is not None:
						# Reuse the same number; only swap the prefix so external
						# references to this item stay matchable.
						item.item_code = format_item_code(new_prefix, existing_num)
					else:
						# Either no prefix (stage cleared) or no existing number.
						# Clear so model.save() can generate a fresh one if applicable.
						item.item_code = None
					updated_fields.append("item_code")

				if stage_changed:
					# Full save() triggers item_code generation via Item_list.save
					# when item_code is empty and stage is set.
					item.save()
				elif updated_fields:
					updated_fields.append("updated_at")
					item.save(update_fields=updated_fields)

				# Manage ItemLine: if a line is selected, ensure (item, line) row exists.
				# We do not delete existing assignments to other lines.
				if line is not None:
					stage_for_link = item.stage or stage
					if stage_for_link is None:
						return JsonResponse(
							{"ok": False, "error": "ต้องเลือก Stage ก่อนผูก Line"},
							status=400,
						)
					ItemLine.objects.update_or_create(
						item=item,
						line=line,
						defaults={
							"item_stage": stage_for_link,
							"user": request.user,
						},
					)
		except Exception as e:
			return JsonResponse({"ok": False, "error": str(e)}, status=500)

		return JsonResponse({
			"ok": True,
			"item_code": item.item_code or "",
			"category_name": getattr(item.category, "name", "") if item.category_id else "",
			"stage_name": getattr(item.stage, "display_name", "") if item.stage_id else "",
		})

	def _resolve_fk(self, model, value):
		value = (value or "").strip()
		if not value or not _is_uuid(value):
			return None
		return model.objects.filter(pk=value).first()

	def _handle_add_to_bom(self, request):
		source_item_id = (request.POST.get("source_item_id") or "").strip()
		parent_item_id = (request.POST.get("parent_item_id") or "").strip()

		if not _is_uuid(source_item_id):
			messages.error(request, "กรุณาเลือก Item จาก Products")
			return redirect(request.get_full_path())
		source = Item_list.objects.filter(pk=source_item_id).first()
		if source is None:
			messages.error(request, "ไม่พบ Item ที่เลือก")
			return redirect(request.get_full_path())

		# No parent → register source as a FG (create BillOfMaterial header).
		if not parent_item_id:
			try:
				with transaction.atomic():
					_, created = BillOfMaterial.objects.get_or_create(
						item=source,
						defaults={
							"revision": "A",
							"latest_eci": "",
							"user": request.user,
						},
					)
			except IntegrityError as e:
				messages.error(request, f"เพิ่มไม่สำเร็จ (ข้อมูลซ้ำ): {e}")
				return redirect(request.get_full_path())
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return redirect(request.get_full_path())
			if created:
				messages.success(request, f"เพิ่ม FG '{source.sd_code or source.part_name}' สำเร็จ")
			else:
				messages.info(request, "FG นี้มี BOM อยู่แล้ว")
			return redirect(request.get_full_path())

		# Has parent → link source as a component of parent's BOM.
		if not _is_uuid(parent_item_id):
			messages.error(request, "Parent FG ไม่ถูกต้อง")
			return redirect(request.get_full_path())
		parent = Item_list.objects.filter(pk=parent_item_id).first()
		if parent is None:
			messages.error(request, "ไม่พบ Parent FG")
			return redirect(request.get_full_path())
		if parent.id == source.id:
			messages.error(request, "ไม่สามารถเลือกตัวเองเป็น Component ได้")
			return redirect(request.get_full_path())

		quantity = _safe_decimal(request.POST.get("quantity") or "1", default=Decimal("1"))
		unit = (request.POST.get("unit") or "PCS").strip() or "PCS"
		sequence_raw = (request.POST.get("sequence") or "").strip()
		try:
			sequence = int(sequence_raw) if sequence_raw else None
		except (ValueError, TypeError):
			sequence = None

		try:
			with transaction.atomic():
				bom, _ = BillOfMaterial.objects.get_or_create(
					item=parent,
					defaults={
						"revision": "A",
						"latest_eci": "",
						"user": request.user,
					},
				)
				if BillOfMaterialItemMater.objects.filter(bom=bom, component=source).exists():
					messages.warning(request, "Component นี้อยู่ใน BOM อยู่แล้ว")
					return redirect(request.get_full_path())
				if sequence is None:
					last = (
						BillOfMaterialItemMater.objects.filter(bom=bom)
						.order_by("-sequence")
						.first()
					)
					sequence = (last.sequence + 1) if last and last.sequence else 1
				BillOfMaterialItemMater.objects.create(
					bom=bom,
					component=source,
					quantity=quantity,
					unit=unit,
					sequence=sequence,
					user=request.user,
				)
		except IntegrityError as e:
			messages.error(request, f"เพิ่มไม่สำเร็จ (ข้อมูลซ้ำ): {e}")
			return redirect(request.get_full_path())
		except Exception as e:
			messages.error(request, f"เกิดข้อผิดพลาด: {e}")
			return redirect(request.get_full_path())

		messages.success(
			request,
			f"เพิ่ม Component '{source.sd_code or source.part_name}' เข้า BOM '{parent.sd_code or parent.part_name}' สำเร็จ",
		)
		return redirect(request.get_full_path())

	def _handle_delete_bom_row(self, request):
		row_id = (request.POST.get("id") or "").strip()
		if not _is_uuid(row_id):
			messages.error(request, "ไม่พบรหัสรายการ")
			return redirect(request.get_full_path())
		obj = BillOfMaterialItemMater.objects.filter(pk=row_id).first()
		if obj is None:
			messages.error(request, "ไม่พบรายการ BOM")
			return redirect(request.get_full_path())
		try:
			with transaction.atomic():
				obj.delete()
		except ProtectedError:
			messages.error(request, "ลบไม่ได้: รายการนี้ถูกใช้งานอยู่")
			return redirect(request.get_full_path())
		except Exception as e:
			messages.error(request, f"เกิดข้อผิดพลาด: {e}")
			return redirect(request.get_full_path())
		messages.success(request, "ลบรายการ BOM สำเร็จ")
		return redirect(request.get_full_path())

	def _handle_delete_fg_bom(self, request):
		fg_id = (request.POST.get("id") or "").strip()
		if not _is_uuid(fg_id):
			messages.error(request, "ไม่พบรหัสรายการ")
			return redirect(request.get_full_path())
		bom = BillOfMaterial.objects.filter(item_id=fg_id).first()
		if bom is None:
			messages.error(request, "ไม่พบ BOM ของ FG นี้")
			return redirect(request.get_full_path())
		try:
			with transaction.atomic():
				bom.delete()  # CASCADE → child BillOfMaterialItemMater rows
		except ProtectedError:
			messages.error(request, "ลบไม่ได้: BOM นี้ถูกใช้งานอยู่")
			return redirect(request.get_full_path())
		except Exception as e:
			messages.error(request, f"เกิดข้อผิดพลาด: {e}")
			return redirect(request.get_full_path())
		messages.success(request, "ลบ BOM ของ FG สำเร็จ")
		return redirect(request.get_full_path())

	def _handle_import(self, request, *args, **kwargs):
		if openpyxl is None:
			messages.error(request, "ไม่สามารถนำเข้า XLSX ได้: ยังไม่ได้ติดตั้ง openpyxl")
			return redirect(request.get_full_path())
		upload = request.FILES.get("excel_file")
		if upload is None:
			messages.error(request, "กรุณาเลือกไฟล์ Excel (.xlsx)")
			return redirect(request.get_full_path())
		name = (getattr(upload, "name", "") or "").lower()
		if not name.endswith(".xlsx"):
			messages.error(request, "รองรับเฉพาะไฟล์ .xlsx")
			return redirect(request.get_full_path())
		items_created = 0
		items_existing = 0
		fg_count = 0
		bom_links_created = 0
		bom_links_updated = 0
		bom_links_duplicate = 0
		parent_not_found = 0
		skipped = 0
		bad_value_skipped = 0
		bad_value_samples: set = set()
		items_updated = 0
		items_similar_skipped = 0
		similar_samples: set = set()
		bom_cache: dict = {}
		seq_by_parent: dict = {}
		line_assignments: list = []

		# Seed similar-sd detection from the DB; new rows are added as created.
		sim_index = SimilarSdIndex()
		sim_index.seed(Item_list.objects.exclude(sd_code="").values_list("sd_code", flat=True))

		# Master lookups keyed by the displayed value (lower-cased) — the same
		# values the export writes, so a round-tripped file resolves cleanly.
		cat_by_name = {(c.name or "").strip().lower(): c for c in ItemCategory.objects.all() if c.name}
		portion_by_title = {(p.title or "").strip().lower(): p for p in Portion.objects.all() if p.title}
		side_by_title = {(s.title or "").strip().lower(): s for s in Side.objects.all() if s.title}
		inout_by_title = {(i.title or "").strip().lower(): i for i in InOut.objects.all() if i.title}
		way_by_title = {(w.title or "").strip().lower(): w for w in Way.objects.all() if w.title}

		def _resolve_item(sd_code, part_number, part_name):
			"""Resolve the Item_list for a row, keyed on sd_code (rule 1).

			Returns the existing item, a newly-created one, or ``None`` when the
			sd_code is a near-duplicate of an existing one (caller skips the row).
			"""
			nonlocal items_created, items_existing, items_similar_skipped
			item = Item_list.objects.filter(sd_code__iexact=sd_code).first()
			if item is not None:
				items_existing += 1
				return item
			# Brand-new sd_code -> flag near-duplicates before creating.
			sim = sim_index.similar_to(sd_code)
			if sim is not None:
				items_similar_skipped += 1
				if len(similar_samples) < 5:
					similar_samples.add(f'{sd_code}~{sim}')
				return None
			sku = _generate_unique_sku(part_number=part_number, sd_code=sd_code)
			item = Item_list.objects.create(
				sd_code=sd_code,
				part_number=part_number,
				part_name=part_name or "",
				sku=sku,
				user=request.user,
			)
			items_created += 1
			sim_index.add(sd_code)
			return item

		def _apply_item_fields(item, row):
			"""Fill-only update of the editable Item_list columns (rules 4/5/6):
			only columns that are EMPTY on the item are filled from the sheet; a
			non-empty value is never overwritten and a blank cell never wipes it.
			Stage / item_code are owned by the reclassify step and left alone."""
			nonlocal items_updated
			cat_name = _row_get_first(row, "category", "category_name")
			category = None
			if cat_name and cat_name.lower() != "(ไม่ระบุ)":
				category = cat_by_name.get(cat_name.lower())

			def _num(key):
				raw = _excel_to_str(row.get(key)).strip()
				return _safe_decimal(raw) if raw != "" else None

			sku = _row_get_first(row, "sku")
			if sku and Item_list.objects.filter(sku__iexact=sku).exclude(pk=item.pk).exists():
				sku = ""  # can't fill with a sku owned by another item

			incoming = {
				"part_number": _row_get_first(row, "part_no", "part_number", "partnumber", "pn") or None,
				"part_name": _row_get_first(row, "part_name", "partname", "name") or None,
				"sku": sku or None,
				"comment": _row_get_first(row, "comment") or None,
				"weight": _num("weight_kg"),
				"purchased_price": _num("purchased_price"),
				"cost": _num("cost"),
				"category": category,
				"portion": portion_by_title.get((_row_get_first(row, "portion") or "").lower()),
				"side": side_by_title.get((_row_get_first(row, "side") or "").lower()),
				"inout": inout_by_title.get((_row_get_first(row, "inout") or "").lower()),
				"way": way_by_title.get((_row_get_first(row, "way") or "").lower()),
			}
			changed = fill_only_update(item, incoming)
			if changed:
				item.save(update_fields=changed + ["updated_at"])
				items_updated += 1

		def _apply_own_bom_header(item, row):
			"""Set Revision / Latest ECI / Scrap % on the BOM owned by this item
			(those columns describe the item's own header, not its parent's)."""
			rev = _row_get_first(row, "revision")
			eci = _row_get_first(row, "latest_eci")
			scrap_raw = _excel_to_str(row.get("scrap")).strip()
			if not (rev or eci or scrap_raw):
				return
			bom, _ = BillOfMaterial.objects.get_or_create(
				item=item,
				defaults={"revision": "A", "latest_eci": "", "user": request.user},
			)
			h_changed: list = []
			if rev and bom.revision != rev:
				bom.revision = rev
				h_changed.append("revision")
			if eci and bom.latest_eci != eci:
				bom.latest_eci = eci
				h_changed.append("latest_eci")
			if scrap_raw != "":
				sc = _safe_decimal(scrap_raw)
				if bom.scrap_percent != sc:
					bom.scrap_percent = sc
					h_changed.append("scrap_percent")
			if h_changed:
				bom.save(update_fields=h_changed + ["updated_at"])

		try:
			with transaction.atomic():
				for row in _parse_xlsx(upload):
					sd_code = _row_get_first(row, "sd_code", "sdcode", "sd")
					part_number = _row_get_first(row, "part_no", "part_number", "partnumber", "pn")
					part_name = _row_get_first(row, "part_name", "partname", "name")
					# Parent is identified by M2M = the parent item's Item Code.
					m2m = _row_get_first(row, "m2m", "parent_item_code")
					line_name = _row_get_first(row, "line", "line_name")

					# Reject spreadsheet error literals (e.g. "#REF!") in the key
					# fields so a broken export never lands as real master data.
					if is_spreadsheet_error(sd_code) or is_spreadsheet_error(part_number):
						if len(bad_value_samples) < 5:
							bad_value_samples.add(sd_code if is_spreadsheet_error(sd_code) else part_number)
						bad_value_skipped += 1
						skipped += 1
						continue
					if not sd_code or not part_number:
						skipped += 1
						continue

					item = _resolve_item(sd_code, part_number, part_name)
					if item is None:
						# Near-duplicate sd_code — skip the whole row.
						skipped += 1
						continue
					_apply_item_fields(item, row)
					# A sub-assembly / FG owns its own BOM header (Revision/Scrap%…).
					_apply_own_bom_header(item, row)

					is_fg = m2m in ("", "-")
					if is_fg:
						# Top-level FG → ensure it owns a BOM header.
						BillOfMaterial.objects.get_or_create(
							item=item,
							defaults={"revision": "A", "latest_eci": "", "user": request.user},
						)
						fg_count += 1
					else:
						parent_item = Item_list.objects.filter(item_code__iexact=m2m).first()
						if parent_item is None:
							parent_not_found += 1
							skipped += 1
							continue
						bom = bom_cache.get(parent_item.id)
						if bom is None:
							bom, _ = BillOfMaterial.objects.get_or_create(
								item=parent_item,
								defaults={"revision": "A", "latest_eci": "", "user": request.user},
							)
							bom_cache[parent_item.id] = bom

						quantity = _safe_decimal(
							_excel_to_str(row.get("quantity")).strip(), default=Decimal("1")
						)
						unit = _excel_to_str(row.get("unit")).strip() or "PCS"
						seq_raw = _excel_to_str(row.get("sequence")).strip()
						try:
							file_seq = int(float(seq_raw)) if seq_raw else None
						except (ValueError, TypeError):
							file_seq = None
						existing = BillOfMaterialItemMater.objects.filter(bom=bom, component=item).first()
						if existing is not None:
							# Re-import of an edited file → apply qty/unit/sequence changes.
							e_changed: list = []
							if existing.quantity != quantity:
								existing.quantity = quantity
								e_changed.append("quantity")
							if existing.unit != unit:
								existing.unit = unit
								e_changed.append("unit")
							if file_seq is not None and existing.sequence != file_seq:
								existing.sequence = file_seq
								e_changed.append("sequence")
							if e_changed:
								existing.save(update_fields=e_changed + ["updated_at"])
								bom_links_updated += 1
							else:
								bom_links_duplicate += 1
						else:
							sequence = file_seq
							if sequence is None:
								seq_by_parent[parent_item.id] = seq_by_parent.get(parent_item.id, 0) + 1
								sequence = seq_by_parent[parent_item.id]
							BillOfMaterialItemMater.objects.create(
								bom=bom,
								component=item,
								quantity=quantity,
								unit=unit,
								sequence=sequence,
								user=request.user,
							)
							bom_links_created += 1

					# Defer line assignment until reclassify has set item.stage.
					if line_name:
						line_obj = Line.objects.filter(line_name__iexact=line_name).first()
						if line_obj is not None:
							line_assignments.append((item.id, line_obj.id))
		except Exception as e:
			messages.error(request, f"นำเข้าไม่สำเร็จ: {e}")
			return redirect(request.get_full_path())
		# Recompute stage classification + item_code for all BOM items.
		try:
			_reclassify_bom_stages_and_codes()
		except Exception as e:
			messages.warning(request, f"นำเข้าเสร็จ แต่ reclassify ไม่สำเร็จ: {e}")

		# Now apply Line assignments — items have valid stage at this point.
		for iid, lid in line_assignments:
			item = Item_list.objects.filter(pk=iid).first()
			line_obj = Line.objects.filter(pk=lid).first()
			if item is None or line_obj is None or item.stage_id is None:
				continue
			ItemLine.objects.update_or_create(
				item=item,
				line=line_obj,
				defaults={"item_stage": item.stage, "user": request.user},
			)

		parts = []
		if items_created:
			parts.append(f"สร้าง Item ใหม่ {items_created}")
		if items_updated:
			parts.append(f"เติมข้อมูล Item {items_updated}")
		if items_existing:
			parts.append(f"ใช้ Item เดิม {items_existing}")
		if items_similar_skipped:
			parts.append(f"ข้าม SD คล้ายซ้ำ {items_similar_skipped}")
		if fg_count:
			parts.append(f"FG {fg_count}")
		if bom_links_created:
			parts.append(f"ผูก BOM {bom_links_created}")
		if bom_links_updated:
			parts.append(f"อัปเดต BOM {bom_links_updated}")
		if bom_links_duplicate:
			parts.append(f"BOM ซ้ำ {bom_links_duplicate}")
		if parent_not_found:
			parts.append(f"ไม่พบ parent จาก M2M {parent_not_found}")
		if skipped:
			parts.append(f"ข้าม {skipped}")
		messages.success(request, "นำเข้าสำเร็จ — " + ", ".join(parts))
		if bad_value_skipped:
			sample = ", ".join(sorted(bad_value_samples))
			messages.warning(
				request,
				f"ข้าม {bad_value_skipped} แถวที่มีค่า error จากสเปรดชีต (เช่น {sample}) "
				"ใน SD Code / Part No. — กรุณาแก้ไฟล์ต้นทางแล้วนำเข้าใหม่",
			)
		if items_similar_skipped:
			sample = ", ".join(sorted(similar_samples))
			messages.warning(
				request,
				f"ข้าม {items_similar_skipped} แถวที่ SD Code คล้ายของเดิม (เช่น {sample}) "
				"— ตรวจว่าซ้ำหรือไม่ก่อนนำเข้าใหม่",
			)
		return redirect(request.get_full_path())
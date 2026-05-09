from __future__ import annotations

import uuid

from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.core.paginator import Paginator

from core.services.auditlog import log_event
from core.auth.decorators import staff_required
from core.models import DefectMode
from core.models.inspection.inspection_model import InspectionModels

try:
	import openpyxl  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None


def _normalized_key(key: str) -> str:
	return (key or "").strip().lower().replace(" ", "_")


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


def download_manage_defectmode_import_template(request):
	"""Download a template for importing defect modes."""
	headers = [
		"name_th",
		"name_en",
		"name_jp",
		"description_th",
		"description_en",
		"description_jp",
	]
	rows = [
		[
			"รอยแตก",
			"Crack",
			"ひび割れ",
			"",
			"",
			"",
		],
		[
			"รอยขีดข่วน",
			"Scratch",
			"傷",
			"",
			"",
			"",
		],
	]
	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "defectmode"
	ws.append(headers)
	for r in rows:
		ws.append(r)
	for col in range(1, len(headers) + 1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 26

	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="manage_defectmode_import_template.xlsx"'
	wb.save(response)
	return response


@method_decorator(staff_required, name="dispatch")
class ManageDefectModeViews(TemplateView):
	template_name = "manage_defectmode.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			return download_manage_defectmode_import_template(request)
		response = super().get(request, *args, **kwargs)
		response['Cache-Control'] = 'no-store, no-cache, must-revalidate'
		response['Pragma'] = 'no-cache'
		response['Expires'] = '0'
		return response

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request

		q = (request.GET.get("q") or "").strip()
		per_page_raw = (request.GET.get("per_page") or "").strip()
		page = (request.GET.get("page") or "1").strip() or "1"

		qs = DefectMode.objects.all()
		if q:
			qs = qs.filter(
				Q(name_th__icontains=q)
				| Q(name_en__icontains=q)
				| Q(name_jp__icontains=q)
				| Q(description_th__icontains=q)
				| Q(description_en__icontains=q)
				| Q(description_jp__icontains=q)
			)

		allowed_per_page = {20, 50, 100, 200}
		try:
			per_page = int(per_page_raw or 20)
		except Exception:
			per_page = 20
		if per_page not in allowed_per_page:
			per_page = 20

		qs = qs.order_by("name_en", "name_th")
		paginator = Paginator(qs, per_page)
		page_obj = paginator.get_page(page)
		inspection_models_qs = InspectionModels.objects.order_by("class_name")

		rows = []
		for d in page_obj.object_list:
			rows.append(
				{
					"id": str(d.id),
					"name_th": d.name_th,
					"name_en": d.name_en,
					"name_jp": d.name_jp,
					"description_th": d.description_th or "",
					"description_en": d.description_en or "",
					"description_jp": d.description_jp or "",
					"inspection_model_id": str(d.inspection_model_id) if d.inspection_model_id else "",
					"inspection_model_name": d.inspection_model.class_name if d.inspection_model else "-",
					"class_name": d.class_name or "",
				}
			)

		def _page_items(num_pages: int, current: int) -> list[int | None]:
			if num_pages <= 0:
				return []
			if num_pages <= 10:
				return list(range(1, num_pages + 1))
			items: list[int | None] = [1]
			if current > 4:
				items.append(None)
			start = max(2, current - 1)
			end = min(num_pages - 1, current + 1)
			if current <= 4:
				start, end = 2, 4
			if current >= num_pages - 3:
				start, end = num_pages - 3, num_pages - 1
			for n in range(start, end + 1):
				if 1 < n < num_pages:
					items.append(n)
			if current < num_pages - 3:
				items.append(None)
			items.append(num_pages)
			compressed: list[int | None] = []
			for it in items:
				if compressed and compressed[-1] == it:
					continue
				if it is None and compressed and compressed[-1] is None:
					continue
				compressed.append(it)
			return compressed

		ctx["rows"] = rows
		ctx["inspection_models"] = list(inspection_models_qs.values("id", "class_name"))
		ctx["q"] = q
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
		ctx["total_count"] = paginator.count

		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		obj_id = (request.POST.get("id") or "").strip()
		uploaded = request.FILES.get("excel_file")

		name_th = (request.POST.get("name_th") or "").strip()
		name_en = (request.POST.get("name_en") or "").strip()
		name_jp = (request.POST.get("name_jp") or "").strip()
		description_th = (request.POST.get("description_th") or "").strip()
		description_en = (request.POST.get("description_en") or "").strip()
		description_jp = (request.POST.get("description_jp") or "").strip()
		inspection_model_id_raw = (request.POST.get("inspection_model_id") or "").strip()
		class_name = (request.POST.get("class_name") or "").strip()

		def _is_uuid(value: str) -> bool:
			try:
				uuid.UUID(str(value))
			except Exception:
				return False
			return True

		if action == "bulk_delete_defects":
			bulk_ids = request.POST.getlist("bulk_id")
			ids: list[str] = []
			for raw in bulk_ids:
				raw = (raw or "").strip()
				if _is_uuid(raw):
					ids.append(raw)
			if not ids:
				messages.error(request, "กรุณาเลือกรายการที่ต้องการลบ")
				return self.get(request, *args, **kwargs)

			deleted = 0
			blocked = 0
			not_found = 0
			try:
				with transaction.atomic():
					for pk in ids:
						obj = DefectMode.objects.filter(pk=pk).first()
						if obj is None:
							not_found += 1
							continue
						try:
							obj.delete()
							deleted += 1
						except ProtectedError:
							blocked += 1
			except Exception as e:
				log_event(
					request,
					action="defectmode:bulk_delete_defects",
					status="failure",
					message="ลบ defect mode แบบ bulk ไม่สำเร็จ",
					metadata={"selected": len(ids), "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

			transaction.on_commit(
				lambda: log_event(
					request,
					action="defectmode:bulk_delete_defects",
					message="ลบ defect mode แบบ bulk",
					metadata={
						"selected": len(ids),
						"deleted": deleted,
						"blocked": blocked,
						"not_found": not_found,
						"ids": ids[:50],
					},
				)
			)

			if blocked:
				messages.warning(
					request,
					f"ลบสำเร็จ {deleted} รายการ, ลบไม่ได้ {blocked} รายการ (มี Component Part Record อ้างอิง), ไม่พบ {not_found}",
				)
			else:
				messages.success(request, f"ลบสำเร็จ {deleted} รายการ" + (f" (ไม่พบ {not_found})" if not_found else ""))
			return self.get(request, *args, **kwargs)

		if action == "import_master_data":
			if not uploaded:
				messages.error(request, "กรุณาเลือกไฟล์ Excel/CSV ก่อนนำเข้า")
				return self.get(request, *args, **kwargs)

			filename = (uploaded.name or "").lower()
			try:
				if filename.endswith(".xlsx"):
					rows_iter = _parse_xlsx(uploaded)
				else:
					messages.error(request, "รองรับเฉพาะไฟล์ Excel (.xlsx) เท่านั้น")
					return self.get(request, *args, **kwargs)
			except RuntimeError:
				messages.error(request, "ยังไม่รองรับไฟล์ .xlsx ในสภาพแวดล้อมนี้ (ต้องติดตั้ง openpyxl)")
				return self.get(request, *args, **kwargs)

			created = 0
			skipped = 0
			dup = 0
			try:
				with transaction.atomic():
					for row in rows_iter:
						row_name_th = str(row.get("name_th") or "").strip()
						row_name_en = str(row.get("name_en") or "").strip()
						row_name_jp = str(row.get("name_jp") or "").strip()
						row_desc_th = str(row.get("description_th") or "").strip()
						row_desc_en = str(row.get("description_en") or "").strip()
						row_desc_jp = str(row.get("description_jp") or "").strip()

						if not row_name_th or not row_name_en or not row_name_jp:
							skipped += 1
							continue
						# Prevent obvious duplicates by EN name (case-insensitive)
						if DefectMode.objects.filter(name_en__iexact=row_name_en).exists():
							dup += 1
							continue
						DefectMode.objects.create(
							name_th=row_name_th,
							name_en=row_name_en,
							name_jp=row_name_jp,
							description_th=row_desc_th,
							description_en=row_desc_en,
							description_jp=row_desc_jp,
							user=request.user,
						)
						created += 1
			except Exception as e:
				log_event(
					request,
					action="defectmode:import_master_data",
					status="failure",
					message="นำเข้า defect mode ไม่สำเร็จ",
					metadata={"filename": getattr(uploaded, "name", ""), "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาดระหว่างนำเข้า: {e}")
				return self.get(request, *args, **kwargs)

			messages.success(request, f"นำเข้าสำเร็จ: +{created}, ซ้ำ {dup}, ข้าม {skipped}")
			transaction.on_commit(
				lambda: log_event(
					request,
					action="defectmode:import_master_data",
					message="นำเข้า defect mode สำเร็จ",
					metadata={
						"filename": getattr(uploaded, "name", ""),
						"created": created,
						"dup": dup,
						"skipped": skipped,
					},
				)
			)
			return self.get(request, *args, **kwargs)


		if action == "create_defect":
			if not name_th or not name_en or not name_jp:
				messages.error(request, "กรุณากรอกชื่อ Defect mode ให้ครบทั้ง TH/EN/JP")
				return self.get(request, *args, **kwargs)
			inspection_model_obj = None
			if inspection_model_id_raw and _is_uuid(inspection_model_id_raw):
				inspection_model_obj = InspectionModels.objects.filter(pk=inspection_model_id_raw).first()
			try:
				with transaction.atomic():
					if DefectMode.objects.filter(name_en__iexact=name_en).exists():
						raise IntegrityError("Defect mode ซ้ำ: name_en มีอยู่แล้ว")
					obj = DefectMode.objects.create(
						name_th=name_th,
						name_en=name_en,
						name_jp=name_jp,
						description_th=description_th,
						description_en=description_en,
						description_jp=description_jp,
						inspection_model=inspection_model_obj,
						class_name=class_name,
						user=request.user,
					)
					messages.success(request, "เพิ่ม Defect mode สำเร็จ")
					transaction.on_commit(
						lambda: log_event(
							request,
							action="defectmode:create_defect",
							message="เพิ่ม defect mode",
							metadata={
								"defect_id": str(obj.pk),
								"name_th": name_th,
								"name_en": name_en,
								"name_jp": name_jp,
							},
						)
					)
					return self.get(request, *args, **kwargs)
			except IntegrityError as e:
				log_event(
					request,
					action="defectmode:create_defect",
					status="failure",
					message="เพิ่ม defect mode ไม่สำเร็จ (IntegrityError)",
					metadata={"name_en": name_en, "error": str(e)},
				)
				messages.error(request, f"บันทึกไม่สำเร็จ (ข้อมูลซ้ำหรือผิดเงื่อนไข): {e}")
				return self.get(request, *args, **kwargs)
			except Exception as e:
				log_event(
					request,
					action="defectmode:create_defect",
					status="failure",
					message="เพิ่ม defect mode ไม่สำเร็จ",
					metadata={"name_en": name_en, "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

		if action == "update_defect":
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return self.get(request, *args, **kwargs)
			if not name_th or not name_en or not name_jp:
				messages.error(request, "กรุณากรอกชื่อ Defect mode ให้ครบทั้ง TH/EN/JP")
				return self.get(request, *args, **kwargs)
			inspection_model_obj = None
			if inspection_model_id_raw and _is_uuid(inspection_model_id_raw):
				inspection_model_obj = InspectionModels.objects.filter(pk=inspection_model_id_raw).first()
			try:
				with transaction.atomic():
					defect = DefectMode.objects.get(pk=obj_id)

					updated_fields = []
					old_name_th = defect.name_th
					old_name_en = defect.name_en
					old_name_jp = defect.name_jp
					old_desc_th = defect.description_th or ""
					old_desc_en = defect.description_en or ""
					old_desc_jp = defect.description_jp or ""

					if defect.name_en != name_en and DefectMode.objects.filter(name_en__iexact=name_en).exclude(pk=defect.pk).exists():
						raise IntegrityError("Defect mode ซ้ำ: name_en มีอยู่แล้ว")

					if defect.name_th != name_th:
						defect.name_th = name_th
						updated_fields.append("name_th")
					if defect.name_en != name_en:
						defect.name_en = name_en
						updated_fields.append("name_en")
					if defect.name_jp != name_jp:
						defect.name_jp = name_jp
						updated_fields.append("name_jp")

					if (defect.description_th or "") != description_th:
						defect.description_th = description_th
						updated_fields.append("description_th")
					if (defect.description_en or "") != description_en:
						defect.description_en = description_en
						updated_fields.append("description_en")
					if (defect.description_jp or "") != description_jp:
						defect.description_jp = description_jp
						updated_fields.append("description_jp")

					new_im_id = inspection_model_obj.pk if inspection_model_obj else None
					if defect.inspection_model_id != new_im_id:
						defect.inspection_model = inspection_model_obj
						updated_fields.append("inspection_model")
					if (defect.class_name or "") != class_name:
						defect.class_name = class_name
						updated_fields.append("class_name")

					if updated_fields:
						updated_fields.append("updated_at")
						defect.save(update_fields=updated_fields)
						messages.success(request, "บันทึกการแก้ไขสำเร็จ")
						transaction.on_commit(
							lambda: log_event(
								request,
								action="defectmode:update_defect",
								message="แก้ไข defect mode",
								metadata={
									"defect_id": str(defect.pk),
									"changed_fields": [f for f in updated_fields if f != "updated_at"],
									"name_th": {"from": old_name_th, "to": defect.name_th} if old_name_th != defect.name_th else None,
									"name_en": {"from": old_name_en, "to": defect.name_en} if old_name_en != defect.name_en else None,
									"name_jp": {"from": old_name_jp, "to": defect.name_jp} if old_name_jp != defect.name_jp else None,
									"description_th": {"from": old_desc_th, "to": defect.description_th} if old_desc_th != (defect.description_th or "") else None,
									"description_en": {"from": old_desc_en, "to": defect.description_en} if old_desc_en != (defect.description_en or "") else None,
									"description_jp": {"from": old_desc_jp, "to": defect.description_jp} if old_desc_jp != (defect.description_jp or "") else None,
								},
							)
						)
					else:
						messages.info(request, "ไม่มีการเปลี่ยนแปลง")
					return self.get(request, *args, **kwargs)
			except ProtectedError:
				log_event(
					request,
					action="defectmode:update_defect",
					status="failure",
					message="แก้ไข defect mode (global) ไม่สำเร็จ (ProtectedError)",
					metadata={"defect_id": obj_id},
				)
				messages.error(request, "บันทึกไม่สำเร็จ: มีข้อมูล Component Part Record อ้างอิงอยู่")
				return self.get(request, *args, **kwargs)
			except IntegrityError as e:
				log_event(
					request,
					action="defectmode:update_defect",
					status="failure",
					message="แก้ไข defect mode (global) ไม่สำเร็จ (IntegrityError)",
					metadata={"defect_id": obj_id, "error": str(e)},
				)
				messages.error(request, f"บันทึกไม่สำเร็จ (ข้อมูลซ้ำหรือผิดเงื่อนไข): {e}")
				return self.get(request, *args, **kwargs)
			except Exception as e:
				log_event(
					request,
					action="defectmode:update_defect",
					status="failure",
					message="แก้ไข defect mode (global) ไม่สำเร็จ",
					metadata={"defect_id": obj_id, "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

		if action == "delete_defect":
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return self.get(request, *args, **kwargs)
			try:
				with transaction.atomic():
					defect = DefectMode.objects.get(pk=obj_id)
					meta = {
						"defect_id": str(defect.pk),
						"name_th": defect.name_th,
						"name_en": defect.name_en,
						"name_jp": defect.name_jp,
					}
					defect.delete()
					messages.success(request, "ลบ Defect mode สำเร็จ")
					transaction.on_commit(
						lambda: log_event(
							request,
							action="defectmode:delete_defect",
							message="ลบ defect mode",
							metadata=meta,
						)
					)
					return self.get(request, *args, **kwargs)
			except ProtectedError:
				log_event(
					request,
					action="defectmode:delete_defect",
					status="failure",
					message="ลบ defect mode (global) ไม่สำเร็จ (ProtectedError)",
					metadata={"defect_id": obj_id},
				)
				messages.error(
					request,
					"ลบไม่ได้: มีข้อมูล Component Part Record อ้างอิงอยู่ (โปรดลบ/ย้ายรายการที่เกี่ยวข้องก่อน)",
				)
				return self.get(request, *args, **kwargs)
			except Exception as e:
				log_event(
					request,
					action="defectmode:delete_defect",
					status="failure",
					message="ลบ defect mode (global) ไม่สำเร็จ",
					metadata={"defect_id": obj_id, "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

		messages.error(request, "ไม่รองรับการทำงานนี้")
		return self.get(request, *args, **kwargs)

from __future__ import annotations

import re
import uuid

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView

from core.auth.decorators import staff_required
from core.models.item_line import ItemLine
from core.models.item_list import Item_list
from core.models.item_stage import ItemStage
from core.models.line import Line
from core.services.auditlog import log_event

try:
	import openpyxl  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None


def _is_uuid(value: str) -> bool:
	try:
		uuid.UUID(str(value))
	except Exception:
		return False
	return True


def _page_items(num_pages: int, current: int) -> list[int | None]:
	if num_pages <= 0:
		return []
	if num_pages <= 10:
		return list(range(1, num_pages + 1))
	items: list[int | None] = [1]
	if current > 4:
		items.append(None)
	start = max(2, current - 1)
	end = min(num_pages - 1, current + 1)
	if current <= 4:
		start, end = 2, 4
	if current >= num_pages - 3:
		start, end = num_pages - 3, num_pages - 1
	for n in range(start, end + 1):
		if 1 < n < num_pages:
			items.append(n)
	if current < num_pages - 3:
		items.append(None)
	items.append(num_pages)
	compressed: list[int | None] = []
	for it in items:
		if compressed and compressed[-1] == it:
			continue
		if it is None and compressed and compressed[-1] is None:
			continue
		compressed.append(it)
	return compressed


def _normalized_key(key: str) -> str:
	key = (key or "").strip().lower()
	key = re.sub(r"[^0-9a-z]+", "_", key)
	key = re.sub(r"_+", "_", key).strip("_")
	return key


def _excel_to_str(value) -> str:
	if value is None:
		return ""
	if isinstance(value, bool):
		return "TRUE" if value else "FALSE"
	if isinstance(value, int):
		return str(value)
	if isinstance(value, float):
		if value.is_integer():
			return str(int(value))
		return str(value)
	return str(value).strip()


def _row_get_first(row: dict, *keys: str) -> str:
	for k in keys:
		if not k:
			continue
		v = row.get(k)
		s = _excel_to_str(v).strip()
		if s != "":
			return s
	return ""


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


def download_manage_item_line_import_template(request):
	"""Download a template for importing item-line mapping."""
	# Prefer real samples (if available).
	sample_sku = "SKU-0001"
	sample_line = "LINE-01"
	sample_stage = "FG"
	item = Item_list.objects.order_by("sku").first()
	if item is not None:
		sample_sku = item.sku
	line = Line.objects.order_by("line_name").first()
	if line is not None:
		sample_line = line.line_name
	stage = ItemStage.objects.order_by("display_name", "name").first()
	if stage is not None:
		sample_stage = (stage.display_name or stage.name).strip() or sample_stage

	headers = ["sku", "line_name", "item_stage"]
	rows = [[sample_sku, sample_line, sample_stage]]
	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "item_line"
	ws.append(headers)
	for r in rows:
		ws.append(r)
	for col in range(1, len(headers) + 1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="manage_item_line_import_template.xlsx"'
	wb.save(response)
	return response


@method_decorator(staff_required, name="dispatch")
class ManageItemLineViews(TemplateView):
	template_name = "manage_item_line.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			return download_manage_item_line_import_template(request)
		return super().get(request, *args, **kwargs)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request

		q = (request.GET.get("q") or "").strip()
		per_page_raw = (request.GET.get("per_page") or "").strip()
		page = (request.GET.get("page") or "1").strip() or "1"

		allowed_per_page = {20, 50, 100, 200}
		try:
			per_page = int(per_page_raw or 20)
		except Exception:
			per_page = 20
		if per_page not in allowed_per_page:
			per_page = 20

		qs = ItemLine.objects.select_related("item", "line", "item_stage").all()
		if q:
			qs = qs.filter(
				Q(item__sku__icontains=q)
				| Q(item__part_name__icontains=q)
				| Q(line__line_name__icontains=q)
				| Q(item_stage__name__icontains=q)
				| Q(item_stage__display_name__icontains=q)
			)
		qs = qs.order_by("item__sku", "line__line_name")
		paginator = Paginator(qs, per_page)
		page_obj = paginator.get_page(page)

		rows = []
		for obj in page_obj.object_list:
			rows.append(
				{
					"id": str(obj.id),
					"item_id": str(obj.item_id) if obj.item_id else "",
					"sku": getattr(obj.item, "sku", "") if obj.item_id else "",
					"part_name": getattr(obj.item, "part_name", "") if obj.item_id else "",
					"line_id": str(obj.line_id) if obj.line_id else "",
					"line_name": getattr(obj.line, "line_name", "") if obj.line_id else "",
					"item_stage_id": str(obj.item_stage_id) if obj.item_stage_id else "",
					"item_stage_name": (
						getattr(obj.item_stage, "display_name", "") or getattr(obj.item_stage, "name", "")
					)
					if obj.item_stage_id
					else "",
				}
			)

		ctx["rows"] = rows
		ctx["q"] = q
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
		ctx["total_count"] = paginator.count

		ctx["items"] = list(Item_list.objects.order_by("sku").values("id", "sku", "part_name"))
		ctx["lines"] = list(Line.objects.order_by("line_name").values("id", "line_name"))
		ctx["stages"] = list(ItemStage.objects.order_by("display_name", "name").values("id", "name", "display_name"))
		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		obj_id = (request.POST.get("id") or "").strip()

		if action == "import_master_data":
			if openpyxl is None:
				messages.error(request, "ไม่สามารถนำเข้า XLSX ได้: ยังไม่ได้ติดตั้ง openpyxl")
				return self.get(request, *args, **kwargs)
			upload = request.FILES.get("excel_file")
			if upload is None:
				messages.error(request, "กรุณาเลือกไฟล์ Excel (.xlsx)")
				return self.get(request, *args, **kwargs)
			name = (getattr(upload, "name", "") or "").lower()
			if not name.endswith(".xlsx"):
				messages.error(request, "รองรับเฉพาะไฟล์ .xlsx")
				return self.get(request, *args, **kwargs)

			created = 0
			updated = 0
			skipped = 0
			item_not_found = 0
			line_not_found = 0
			stage_not_found = 0
			try:
				with transaction.atomic():
					for row in _parse_xlsx(upload):
						sku = _row_get_first(row, "sku", "item_sku")
						line_name = _row_get_first(row, "line_name", "line")
						stage_key = _row_get_first(row, "item_stage", "stage")
						if not sku or not line_name or not stage_key:
							skipped += 1
							continue

						item = Item_list.objects.filter(sku__iexact=sku).first()
						if item is None:
							item_not_found += 1
							continue
						line = Line.objects.filter(line_name__iexact=line_name).first()
						if line is None:
							line_not_found += 1
							continue
						stage = (
							ItemStage.objects.filter(display_name__iexact=stage_key).first()
							or ItemStage.objects.filter(name__iexact=stage_key).first()
						)
						if stage is None:
							stage_not_found += 1
							continue

						obj = ItemLine.objects.filter(item=item, line=line).first()
						if obj is None:
							ItemLine.objects.create(item=item, line=line, item_stage=stage, user=request.user)
							created += 1
						else:
							obj.item_stage = stage
							obj.save(update_fields=["item_stage", "updated_at"])
							updated += 1
			except Exception as e:
				log_event(
					request,
					action="item_line:import_master_data",
					status="failure",
					message="นำเข้า ItemLine ไม่สำเร็จ",
					metadata={"filename": getattr(upload, "name", ""), "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาดระหว่างนำเข้า: {e}")
				return self.get(request, *args, **kwargs)

			messages.success(
				request,
				f"นำเข้าสำเร็จ: +{created}, อัปเดต {updated}, ข้าม {skipped}, ไม่พบ Item {item_not_found}, ไม่พบ Line {line_not_found}, ไม่พบ Stage {stage_not_found}",
			)
			transaction.on_commit(
				lambda: log_event(
					request,
					action="item_line:import_master_data",
					message="นำเข้า ItemLine สำเร็จ",
					metadata={
						"filename": getattr(upload, "name", ""),
						"created": created,
						"updated": updated,
						"skipped": skipped,
						"item_not_found": item_not_found,
						"line_not_found": line_not_found,
						"stage_not_found": stage_not_found,
					},
				),
			)
			return self.get(request, *args, **kwargs)

		if action == "bulk_delete_item_lines":
			bulk_ids = request.POST.getlist("bulk_id")
			ids = [pk for pk in bulk_ids if _is_uuid((pk or "").strip())]
			if not ids:
				messages.error(request, "กรุณาเลือกรายการที่ต้องการลบ")
				return self.get(request, *args, **kwargs)

			deleted = 0
			blocked = 0
			not_found = 0
			try:
				with transaction.atomic():
					for pk in ids:
						obj = ItemLine.objects.filter(pk=pk).first()
						if obj is None:
							not_found += 1
							continue
						try:
							obj.delete()
							deleted += 1
						except ProtectedError:
							blocked += 1
			except Exception as e:
				log_event(
					request,
					action="item_line:bulk_delete",
					status="failure",
					message="ลบ ItemLine แบบ bulk ไม่สำเร็จ",
					metadata={"selected": len(ids), "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

			transaction.on_commit(
				lambda: log_event(
					request,
					action="item_line:bulk_delete",
					message="ลบ ItemLine แบบ bulk",
					metadata={"selected": len(ids), "deleted": deleted, "blocked": blocked, "not_found": not_found},
				),
			)
			if blocked:
				messages.warning(request, f"ลบสำเร็จ {deleted} รายการ (ลบไม่ได้ {blocked}, ไม่พบ {not_found})")
			else:
				messages.success(request, f"ลบสำเร็จ {deleted} รายการ")
			return self.get(request, *args, **kwargs)

		if action in {"update", "delete"}:
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return self.get(request, *args, **kwargs)

		if action == "delete":
			obj = ItemLine.objects.filter(pk=obj_id).first()
			if obj is None:
				messages.error(request, "ไม่พบรายการ")
				return self.get(request, *args, **kwargs)
			try:
				obj.delete()
			except ProtectedError:
				messages.error(request, "ไม่สามารถลบได้: รายการนี้ถูกใช้งานอยู่")
				return self.get(request, *args, **kwargs)
			transaction.on_commit(
				lambda: log_event(
					request,
					action="item_line:delete",
					message="ลบ ItemLine",
					metadata={"id": obj_id},
				),
			)
			messages.success(request, "ลบรายการสำเร็จ")
			return self.get(request, *args, **kwargs)

		item_id = (request.POST.get("item_id") or "").strip()
		line_id = (request.POST.get("line_id") or "").strip()
		stage_id = (request.POST.get("item_stage_id") or "").strip()
		if not _is_uuid(item_id):
			messages.error(request, "กรุณาเลือก Item")
			return self.get(request, *args, **kwargs)
		if not _is_uuid(line_id):
			messages.error(request, "กรุณาเลือก Line")
			return self.get(request, *args, **kwargs)
		if not _is_uuid(stage_id):
			messages.error(request, "กรุณาเลือก Stage")
			return self.get(request, *args, **kwargs)

		item = Item_list.objects.filter(pk=item_id).first()
		line = Line.objects.filter(pk=line_id).first()
		stage = ItemStage.objects.filter(pk=stage_id).first()
		if item is None:
			messages.error(request, "ไม่พบ Item")
			return self.get(request, *args, **kwargs)
		if line is None:
			messages.error(request, "ไม่พบ Line")
			return self.get(request, *args, **kwargs)
		if stage is None:
			messages.error(request, "ไม่พบ Stage")
			return self.get(request, *args, **kwargs)

		if action == "update":
			obj = ItemLine.objects.filter(pk=obj_id).first()
			if obj is None:
				messages.error(request, "ไม่พบรายการ")
				return self.get(request, *args, **kwargs)
			obj.item = item
			obj.line = line
			obj.item_stage = stage
			try:
				obj.save(update_fields=["item", "line", "item_stage", "updated_at"])
			except IntegrityError:
				messages.error(request, "ข้อมูลซ้ำ: Item และ Line นี้มีอยู่แล้ว")
				return self.get(request, *args, **kwargs)
			transaction.on_commit(
				lambda: log_event(
					request,
					action="item_line:update",
					message="แก้ไข ItemLine",
					metadata={"id": obj_id, "sku": item.sku, "line": line.line_name, "stage": str(stage.id)},
				),
			)
			messages.success(request, "แก้ไขรายการสำเร็จ")
			return self.get(request, *args, **kwargs)

		# Default: create
		try:
			obj = ItemLine.objects.create(item=item, line=line, item_stage=stage, user=request.user)
		except IntegrityError:
			messages.error(request, "ข้อมูลซ้ำ: Item และ Line นี้มีอยู่แล้ว")
			return self.get(request, *args, **kwargs)
		except Exception as e:
			log_event(
				request,
				action="item_line:create",
				status="failure",
				message="เพิ่ม ItemLine ไม่สำเร็จ",
				metadata={"error": str(e)},
			)
			messages.error(request, f"เพิ่มข้อมูลไม่สำเร็จ: {e}")
			return self.get(request, *args, **kwargs)

		transaction.on_commit(
			lambda: log_event(
				request,
				action="item_line:create",
				message="เพิ่ม ItemLine",
				metadata={"id": str(obj.id), "sku": item.sku, "line": line.line_name, "stage": str(stage.id)},
			),
		)
		messages.success(request, "เพิ่มข้อมูลสำเร็จ")
		return self.get(request, *args, **kwargs)

from __future__ import annotations

import csv
import secrets

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView

from core.auth.decorators import staff_required
from core.models import UserProfile

from core.services.auditlog import log_event

try:
	import openpyxl  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None


def download_user_import_template(request):
	"""Download a template for importing users.

	Default format is CSV (Excel-compatible). If `?format=xlsx` is provided and
	`openpyxl` is installed, an .xlsx file will be returned.
	"""
	fmt = (request.GET.get("format") or "xlsx").strip().lower()

	headers = [
		"username",
		"email",
		"full_name",
		"role",
		"is_active",
		"password",
		"shift",
		"group",
	]
	rows = [
		["jane.doe", "jane.doe@example.com", "Jane Doe", "user", "TRUE", "", "shift_day", "R&D"],
		["production.staff", "staff@example.com", "สมชาย ใจดี", "staff", "TRUE", "", "shift_a", "Production"],
		["admin01", "admin@example.com", "ผู้ดูแล ระบบ", "admin", "TRUE", "ChangeMe123!", "shift_day", "Accounting"],
	]

	if fmt == "xlsx":
		if openpyxl is None:
			return HttpResponse(
				"XLSX format is not available (openpyxl is not installed).",
				status=400,
				content_type="text/plain; charset=utf-8",
			)
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "users"
		ws.append(headers)
		for r in rows:
			ws.append(r)
		for col in range(1, len(headers) + 1):
			ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

		response = HttpResponse(
			content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)
		response["Content-Disposition"] = 'attachment; filename="user_import_template.xlsx"'
		wb.save(response)
		return response

	# CSV
	response = HttpResponse(content_type="text/csv; charset=utf-8")
	response["Content-Disposition"] = 'attachment; filename="user_import_template.csv"'
	response.write("\ufeff")
	writer = csv.writer(response, lineterminator="\r\n")
	writer.writerow(headers)
	writer.writerows(rows)
	return response


def _bool_from_any(value, default=True):
	if value is None:
		return default
	if isinstance(value, bool):
		return value
	s = str(value).strip().lower()
	if s in {"1", "true", "t", "yes", "y", "on", "active"}:
		return True
	if s in {"0", "false", "f", "no", "n", "off", "disabled", "inactive"}:
		return False
	return default


def _split_full_name(full_name):
	name = (full_name or "").strip()
	if not name:
		return "", ""
	parts = name.split()
	if len(parts) == 1:
		return parts[0], ""
	return parts[0], " ".join(parts[1:])


def _apply_role(user, role):
	r = (role or "user").strip().lower()
	if r == "admin":
		user.is_staff = True
		user.is_superuser = True
	elif r == "staff":
		user.is_staff = True
		user.is_superuser = False
	else:
		user.is_staff = False
		user.is_superuser = False


def _normalized_key(key):
	return (key or "").strip().lower().replace(" ", "_")


def _parse_group_names(value) -> list[str]:
	if value is None:
		return []
	s = str(value).strip()
	if not s:
		return []
	# Allow: single group, or multiple separated by comma/semicolon/pipe/newline.
	seps = [",", ";", "|", "\n", "\r"]
	for sep in seps:
		s = s.replace(sep, ",")
	parts = [p.strip() for p in s.split(",")]
	seen: set[str] = set()
	out: list[str] = []
	for p in parts:
		if not p:
			continue
		if p in seen:
			continue
		seen.add(p)
		out.append(p)
	return out


def _set_user_groups(user, group_names: list[str]):
	groups = []
	for name in group_names:
		g, _ = Group.objects.get_or_create(name=name)
		groups.append(g)
	user.groups.set(groups)


def _parse_csv(uploaded_file):
	data = uploaded_file.read()
	try:
		text = data.decode("utf-8-sig")
	except UnicodeDecodeError:
		text = data.decode("cp874", errors="replace")
	reader = csv.DictReader(text.splitlines())
	for row in reader:
		yield {(_normalized_key(k)): v for k, v in (row or {}).items()}


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


@method_decorator(staff_required, name="dispatch")
class ManageUserViews(TemplateView):
	template_name = "core/manage_user.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			if not (request.user.is_superuser or request.user.has_perm("core.add_user")):
				messages.error(request, "คุณไม่มีสิทธิ์ดาวน์โหลดไฟล์นำเข้าผู้ใช้งาน")
				return super().get(request, *args, **kwargs)
			return download_user_import_template(request)
		return super().get(request, *args, **kwargs)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request
		q = (request.GET.get("q") or "").strip()
		role = (request.GET.get("role") or "").strip().lower()
		status = (request.GET.get("status") or "").strip().lower()
		per_page_raw = (request.GET.get("per_page") or "").strip()
		page = (request.GET.get("page") or "1").strip() or "1"

		User = get_user_model()
		qs = (
			User.objects.select_related("profile")
			.prefetch_related("groups")
			.all()
			.order_by("username")
		)

		if q:
			qs = qs.filter(
				Q(username__icontains=q)
				| Q(email__icontains=q)
				| Q(first_name__icontains=q)
				| Q(last_name__icontains=q)
			)

		if role == "admin":
			qs = qs.filter(is_superuser=True)
		elif role == "staff":
			qs = qs.filter(is_staff=True, is_superuser=False)
		elif role == "user":
			qs = qs.filter(is_staff=False, is_superuser=False)

		if status == "active":
			qs = qs.filter(is_active=True)
		elif status == "disabled":
			qs = qs.filter(is_active=False)

		total_count = qs.count()

		allowed_per_page = {100, 200, 500, 1000}
		try:
			per_page = int(per_page_raw or 100)
		except Exception:
			per_page = 100
		if per_page not in allowed_per_page:
			per_page = 100

		paginator = Paginator(qs, per_page)
		page_obj = paginator.get_page(page)

		def _page_items(num_pages: int, current: int) -> list[int | None]:
			if num_pages <= 0:
				return []
			if num_pages <= 10:
				return list(range(1, num_pages + 1))
			items: list[int | None] = [1]
			if current > 4:
				items.append(None)
			start = max(2, current - 1)
			end = min(num_pages - 1, current + 1)
			if current <= 4:
				start, end = 2, 4
			if current >= num_pages - 3:
				start, end = num_pages - 3, num_pages - 1
			for n in range(start, end + 1):
				if 1 < n < num_pages:
					items.append(n)
			if current < num_pages - 3:
				items.append(None)
			items.append(num_pages)
			compressed: list[int | None] = []
			for it in items:
				if compressed and compressed[-1] == it:
					continue
				if it is None and compressed and compressed[-1] is None:
					continue
				compressed.append(it)
			return compressed

		ctx["users"] = list(page_obj.object_list)
		ctx["q"] = q
		ctx["role"] = role
		ctx["status"] = status
		ctx["total_count"] = total_count
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
		ctx["group_options"] = list(Group.objects.order_by("name").values_list("name", flat=True))
		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		User = get_user_model()

		def _deny(msg: str):
			messages.error(request, msg)
			return redirect(request.get_full_path())

		if action == "import":
			# Import can create AND update users, so require both add+change.
			if not (
				request.user.is_superuser
				or (
					request.user.has_perm("core.add_user")
					and request.user.has_perm("core.change_user")
				)
			):
				return _deny("คุณไม่มีสิทธิ์นำเข้าข้อมูลผู้ใช้งาน")

			uploaded = request.FILES.get("excel_file")
			if not uploaded:
				messages.error(request, "กรุณาเลือกไฟล์ Excel/CSV ก่อนนำเข้า")
				return redirect(request.get_full_path())

			filename = (uploaded.name or "").lower()
			if filename.endswith(".csv"):
				rows = _parse_csv(uploaded)
			elif filename.endswith(".xlsx"):
				try:
					rows = _parse_xlsx(uploaded)
				except RuntimeError:
					messages.error(request, "ยังไม่รองรับไฟล์ .xlsx ในสภาพแวดล้อมนี้ (ต้องติดตั้ง openpyxl)")
					return redirect(request.get_full_path())
			else:
				messages.error(request, "รองรับเฉพาะไฟล์ .xlsx หรือ .csv")
				return redirect(request.get_full_path())

			created = 0
			updated = 0
			skipped = 0

			with transaction.atomic():
				for row in rows:
					username = (row.get("username") or row.get("user") or "").strip()
					if not username:
						skipped += 1
						continue

					email = (row.get("email") or "").strip()
					full_name = (
						row.get("full_name") or row.get("name") or row.get("fullname") or ""
					)
					full_name = str(full_name).strip()
					role = (row.get("role") or "user").strip().lower()
					is_active = _bool_from_any(row.get("is_active"), default=True)
					shift = (row.get("shift") or "shift_day")
					shift = (str(shift).strip() if shift is not None else "shift_day") or "shift_day"
					# If header exists but blank => clear groups. If header missing => no change.
					group_raw = row.get("group")
					if group_raw is None:
						group_raw = row.get("groups")
					group_raw = ("" if group_raw is None else str(group_raw))
					password = row.get("password")
					password = (str(password).strip() if password is not None else "")
					provided_password = bool(password)
					if not password:
						password = secrets.token_urlsafe(10)

					first_name, last_name = _split_full_name(full_name)

					user = User.objects.filter(username=username).first()
					if user is None:
						user = User.objects.create_user(username=username, email=email, password=password)
						created += 1
					else:
						if email:
							user.email = email
						if provided_password:
							user.set_password(password)
						updated += 1

					if first_name:
						user.first_name = first_name
					if last_name:
						user.last_name = last_name
					user.is_active = is_active
					_apply_role(user, role)
					user.save()

					UserProfile.objects.update_or_create(user=user, defaults={"shift": shift})

					# Apply groups only when the column is present in import (including blank => clear).
					if (row.get("group") is not None) or (row.get("groups") is not None):
						_set_user_groups(user, _parse_group_names(group_raw))

			messages.success(
				request,
				f"นำเข้าผู้ใช้งานสำเร็จ: เพิ่มใหม่ {created} | อัปเดต {updated} | ข้าม {skipped}",
			)
			log_event(
				request,
				action="user:import",
				status="success",
				message="นำเข้าผู้ใช้งาน",
				metadata={
					"created": created,
					"updated": updated,
					"skipped": skipped,
					"filename": getattr(uploaded, "name", ""),
				},
			)
			return redirect(request.get_full_path())

		if action == "create":
			if not (request.user.is_superuser or request.user.has_perm("core.add_user")):
				return _deny("คุณไม่มีสิทธิ์เพิ่มผู้ใช้งาน")

			username = (request.POST.get("username") or "").strip()
			email = (request.POST.get("email") or "").strip()
			full_name = (request.POST.get("full_name") or "").strip()
			role = (request.POST.get("role") or "user").strip().lower()
			shift = (request.POST.get("shift") or "shift_day").strip() or "shift_day"
			group_name = (request.POST.get("group") or "").strip()
			is_active = (request.POST.get("is_active") or "") in {"1", "true", "on", "yes"}
			password = request.POST.get("password") or ""
			password_confirm = request.POST.get("password_confirm") or ""

			if not username:
				messages.error(request, "กรุณากรอก Username")
				return redirect(request.get_full_path())
			if not password:
				messages.error(request, "กรุณากรอกรหัสผ่าน")
				return redirect(request.get_full_path())
			if password != password_confirm:
				messages.error(request, "รหัสผ่านและยืนยันรหัสผ่านไม่ตรงกัน")
				return redirect(request.get_full_path())
			if User.objects.filter(username=username).exists():
				messages.error(request, "Username นี้มีอยู่แล้ว")
				return redirect(request.get_full_path())

			first_name, last_name = _split_full_name(full_name)
			user = User.objects.create_user(username=username, email=email, password=password)
			user.first_name = first_name
			user.last_name = last_name
			user.is_active = is_active
			_apply_role(user, role)
			user.save()
			UserProfile.objects.get_or_create(user=user, defaults={"shift": shift})
			if group_name:
				_set_user_groups(user, [group_name])
			else:
				user.groups.clear()
			messages.success(request, "เพิ่มผู้ใช้งานสำเร็จ")
			log_event(
				request,
				action="user:create",
				status="success",
				message="เพิ่มผู้ใช้งาน",
				metadata={"user_id": user.pk, "username": user.username},
			)
			return redirect(request.get_full_path())

		if action == "bulk_delete":
			if not (request.user.is_superuser or request.user.has_perm("core.delete_user")):
				return _deny("คุณไม่มีสิทธิ์ลบผู้ใช้งาน")

			raw_ids = request.POST.getlist("bulk_id")
			ids: list[int] = []
			for raw in raw_ids:
				raw = (raw or "").strip()
				if raw.isdigit():
					ids.append(int(raw))

			if not ids:
				messages.error(request, "กรุณาเลือกผู้ใช้งานที่ต้องการลบ")
				return redirect(request.get_full_path())

			# Safety: prevent deleting yourself
			if getattr(request, "user", None) is not None and request.user.is_authenticated:
				ids = [i for i in ids if i != request.user.pk]

			with transaction.atomic():
				deleted, _ = User.objects.filter(pk__in=ids).delete()
			messages.success(request, f"ลบสำเร็จ {deleted} ผู้ใช้งาน")
			log_event(
				request,
				action="user:bulk_delete",
				status="success",
				message="ลบผู้ใช้งานแบบ bulk",
				metadata={"deleted": deleted, "ids": ids},
			)
			return redirect(request.get_full_path())

		user_id = (request.POST.get("id") or "").strip()

		if not user_id.isdigit():
			messages.error(request, "ไม่พบรหัสผู้ใช้งาน")
			return redirect(request.get_full_path())

		try:
			target_user = User.objects.get(pk=int(user_id))
		except User.DoesNotExist:
			messages.error(request, "ไม่พบผู้ใช้งาน")
			return redirect(request.get_full_path())

		if action == "update":
			if not (request.user.is_superuser or request.user.has_perm("core.change_user")):
				return _deny("คุณไม่มีสิทธิ์แก้ไขผู้ใช้งาน")

			full_name = (request.POST.get("full_name") or "").strip()
			first_name = (request.POST.get("first_name") or "").strip()
			last_name = (request.POST.get("last_name") or "").strip()
			email = (request.POST.get("email") or "").strip()
			role = (request.POST.get("role") or "").strip().lower()
			shift = (request.POST.get("shift") or "shift_day").strip()
			group_name = (request.POST.get("group") or "").strip()
			is_active = (request.POST.get("is_active") or "") == "on"
			new_password = (request.POST.get("password") or "").strip()

			# Prefer full_name if provided (supports popup UI)
			if full_name:
				parts = [p for p in full_name.split() if p]
				if len(parts) == 1:
					first_name = parts[0]
					last_name = ""
				else:
					first_name = parts[0]
					last_name = " ".join(parts[1:])

			# Apply fields
			target_user.first_name = first_name
			target_user.last_name = last_name
			target_user.email = email
			target_user.is_active = is_active

			# Role mapping
			if role == "admin":
				target_user.is_superuser = True
				target_user.is_staff = True
			elif role == "staff":
				target_user.is_superuser = False
				target_user.is_staff = True
			else:
				target_user.is_superuser = False
				target_user.is_staff = False

			if new_password:
				target_user.set_password(new_password)

			target_user.save()

			# Update UserProfile shift
			profile, _ = UserProfile.objects.get_or_create(user=target_user)
			profile.shift = shift
			profile.save()

			# Update groups (single-select in UI). Blank => clear.
			if group_name:
				_set_user_groups(target_user, [group_name])
			else:
				target_user.groups.clear()

			messages.success(request, f"อัปเดตผู้ใช้งาน {target_user.username} สำเร็จ")
			log_event(
				request,
				action="user:update",
				status="success",
				message="อัปเดตผู้ใช้งาน",
				metadata={"user_id": target_user.pk, "username": target_user.username},
			)
			return redirect(request.get_full_path())

		if action == "delete":
			if not (request.user.is_superuser or request.user.has_perm("core.delete_user")):
				return _deny("คุณไม่มีสิทธิ์ลบผู้ใช้งาน")

			# Safety: prevent deleting yourself (if authenticated)
			if getattr(request, "user", None) is not None and request.user.is_authenticated:
				if request.user.pk == target_user.pk:
					messages.error(request, "ไม่สามารถลบบัญชีที่กำลังใช้งานอยู่")
					return redirect(request.get_full_path())

			target_username = target_user.username
			target_user.delete()
			messages.success(request, f"ลบผู้ใช้งาน {target_username} สำเร็จ")
			log_event(
				request,
				action="user:delete",
				status="success",
				message="ลบผู้ใช้งาน",
				metadata={"username": target_username},
			)
			return redirect(request.get_full_path())

		messages.error(request, "คำสั่งไม่ถูกต้อง")
		return redirect(request.get_full_path())

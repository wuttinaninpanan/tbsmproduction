from __future__ import annotations

from datetime import timedelta

from django.core.paginator import Paginator
from django.db.models import F, Q, Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView

from core.auth.decorators import user_required
from core.models.scrap_record import ScrapRecord


def _page_items(num_pages: int, current: int) -> list[int | None]:
    if num_pages <= 0:
        return []
    if num_pages <= 10:
        return list(range(1, num_pages + 1))
    items: list[int | None] = [1]
    if current > 4:
        items.append(None)
    start = max(2, current - 1)
    end = min(num_pages - 1, current + 1)
    if current <= 4:
        start, end = 2, 4
    if current >= num_pages - 3:
        start, end = num_pages - 3, num_pages - 1
    for n in range(start, end + 1):
        if 1 < n < num_pages:
            items.append(n)
    if current < num_pages - 3:
        items.append(None)
    items.append(num_pages)
    compressed: list[int | None] = []
    for it in items:
        if compressed and compressed[-1] == it:
            continue
        if it is None and compressed and compressed[-1] is None:
            continue
        compressed.append(it)
    return compressed


@method_decorator(user_required, name="dispatch")
class InspectionScrapDashboardView(TemplateView):
    """Read-only dashboard of scrap recorded by the Inspection M/C machines.

    Source = the legacy ``ScrapRecord`` table (what the inspection machines
    write to). Mirrors the old Manage Scrap table layout and supports an XLSX
    export of the filtered rows. No editing here — view + export only.
    """

    template_name = "inspection/inspection_scrap_dashboard.html"

    def get(self, request, *args, **kwargs):
        if (request.GET.get("action") or "").strip().lower() == "export_excel":
            return self._export_excel(request)
        return super().get(request, *args, **kwargs)

    # ------------------------------------------------------------------ filters
    def _filtered_qs(self, request):
        q = (request.GET.get("q") or "").strip()
        date_from_raw = (request.GET.get("date_from") or "").strip()
        date_to_raw = (request.GET.get("date_to") or "").strip()
        date_from = parse_date(date_from_raw) if date_from_raw else None
        date_to = parse_date(date_to_raw) if date_to_raw else None

        qs = ScrapRecord.objects.select_related(
            "production_line",
            "part_number",
            "defect_mode",
            "component_part",
            "created_by",
            "created_by__profile",
        ).all()

        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        if q:
            qs = qs.filter(
                Q(production_line__line_name__icontains=q)
                | Q(part_number__part_number__icontains=q)
                | Q(part_number__sd_code__icontains=q)
                | Q(part_number__sku__icontains=q)
                | Q(defect_mode__name_th__icontains=q)
                | Q(defect_mode__name_en__icontains=q)
                | Q(component_part__part_name__icontains=q)
                | Q(component_part__part_number__icontains=q)
                | Q(created_by__username__icontains=q)
                | Q(created_by__first_name__icontains=q)
                | Q(created_by__profile__shift__icontains=q)
            )
        return qs.order_by("-created_at")

    # --------------------------------------------------------------------- GET
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        request = self.request

        per_page_raw = (request.GET.get("per_page") or "").strip()
        page = (request.GET.get("page") or "1").strip() or "1"
        allowed_per_page = {100, 200, 500, 1000}
        try:
            per_page = int(per_page_raw or 100)
        except Exception:
            per_page = 100
        if per_page not in allowed_per_page:
            per_page = 100

        qs = self._filtered_qs(request)

        paginator = Paginator(qs, per_page)
        page_obj = paginator.get_page(page)

        ctx["records"] = list(page_obj.object_list)
        ctx["q"] = (request.GET.get("q") or "").strip()
        ctx["date_from"] = (request.GET.get("date_from") or "").strip()
        ctx["date_to"] = (request.GET.get("date_to") or "").strip()
        ctx["page_obj"] = page_obj
        ctx["paginator"] = paginator
        ctx["per_page"] = per_page
        ctx["rows_total"] = paginator.count
        ctx["total_count"] = paginator.count
        ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
        # KPI: total scrapped quantity across the filtered set.
        ctx["total_qty"] = qs.aggregate(s=Sum("quantity"))["s"] or 0
        ctx["charts"] = self._chart_data(qs, has_date=bool(ctx["date_from"] or ctx["date_to"]))
        return ctx

    # ------------------------------------------------------------------ charts
    def _chart_data(self, qs, has_date: bool) -> dict:
        """Aggregates for the dashboard charts, honouring the active filter.

        Daily trend uses the filtered date range, or the last 14 days when no
        date filter is set (so the line chart stays readable)."""
        daily_qs = qs
        if not has_date:
            today = timezone.localdate()
            daily_qs = daily_qs.filter(created_at__date__gte=today - timedelta(days=13))
        daily_rows = list(
            daily_qs.annotate(d=TruncDate("created_at"))
            .values("d")
            .annotate(s=Sum("quantity"))
            .order_by("d")
        )
        top_defect = list(
            qs.values(nm=F("defect_mode__name_th"))
            .annotate(s=Sum("quantity"))
            .order_by("-s", "nm")[:6]
        )
        top_lines = list(
            qs.values(nm=F("production_line__line_name"))
            .annotate(s=Sum("quantity"))
            .order_by("-s", "nm")[:8]
        )
        return {
            "daily": {
                "labels": [r["d"].strftime("%d %b") for r in daily_rows if r["d"]],
                "data": [int(r["s"] or 0) for r in daily_rows if r["d"]],
            },
            "top_defect": {
                "labels": [(r["nm"] or "-") for r in top_defect],
                "data": [int(r["s"] or 0) for r in top_defect],
            },
            "top_lines": {
                "labels": [(r["nm"] or "-") for r in top_lines],
                "data": [int(r["s"] or 0) for r in top_lines],
            },
        }

    # ------------------------------------------------------------------ export
    def _export_excel(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception:
            return HttpResponse(
                "XLSX export is not available (openpyxl is not installed).",
                status=400,
                content_type="text/plain; charset=utf-8",
            )

        qs = self._filtered_qs(request)

        wb = Workbook()
        ws = wb.active
        ws.title = "Scrap Records"
        headers = [
            "วันที่/เวลา",
            "ผู้ใช้งาน",
            "กะ",
            "Production line",
            "SD number",
            "Defect mode",
            "Comment",
            "Part name",
            "Quantity",
        ]
        ws.append(headers)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for r in qs:
            shift_display = "-"
            if r.created_by and getattr(r.created_by, "profile", None):
                shift_value = r.created_by.profile.shift
                if shift_value == "shift_a":
                    shift_display = "กะ A"
                elif shift_value == "shift_b":
                    shift_display = "กะ B"
                else:
                    shift_display = "กะ Day"
            created_at_local = timezone.localtime(r.created_at) if r.created_at else None
            ws.append([
                created_at_local.strftime("%d/%m/%Y %H:%M") if created_at_local else "-",
                r.created_by.get_short_name() if r.created_by else "-",
                shift_display,
                getattr(r.production_line, "code", "-"),
                getattr(r.part_number, "sd_code", "-") or "-",
                getattr(r.defect_mode, "name", "-"),
                r.comment or "-",
                getattr(r.component_part, "part_name", "-") or "-",
                r.quantity or 0,
            ])

        for col, width in zip("ABCDEFGHI", [18, 15, 10, 18, 18, 22, 22, 22, 12]):
            ws.column_dimensions[col].width = width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        ts = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        response["Content-Disposition"] = f'attachment; filename="InspectionScrapRecords_{ts}.xlsx"'
        wb.save(response)
        return response

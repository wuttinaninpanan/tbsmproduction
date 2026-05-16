from __future__ import annotations

import re
import uuid

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.shortcuts import redirect

from core.auth.decorators import staff_required
from core.models.businesspartner import BusinessPartner
from core.services.auditlog import log_event

try:
	import openpyxl  # type: ignore
except Exception:
	openpyxl = None


def _is_uuid(value: str) -> bool:
	try:
		uuid.UUID(str(value))
	except Exception:
		return False
	return True


def _normalized_key(key: str) -> str:
	key = (key or "").strip().lower()
	key = re.sub(r"[^0-9a-z]+", "_", key)
	key = re.sub(r"_+", "_", key).strip("_")
	return key


def _excel_to_str(value) -> str:
	if value is None:
		return ""
	if isinstance(value, bool):
		return "TRUE" if value else "FALSE"
	if isinstance(value, int):
		return str(value)
	if isinstance(value, float):
		if value.is_integer():
			return str(int(value))
		return str(value)
	return str(value).strip()


def _row_get_first(row: dict, *keys: str) -> str:
	for k in keys:
		if not k:
			continue
		v = row.get(k)
		s = _excel_to_str(v).strip()
		if s != "":
			return s
	return ""


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


def _download_businesspartner_template():
	headers = ["code", "name", "tax_id"]
	rows = [["BP001", "ชื่อบริษัท", "1234567890123"], ["BP002", "ตัวอย่าง", ""]]
	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "business_partners"
	ws.append(headers)
	for r in rows:
		ws.append(r)
	for col in range(1, len(headers) + 1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30
	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="business_partner_import_template.xlsx"'
	wb.save(response)
	return response


def _page_items(num_pages: int, current: int) -> list[int | None]:
	if num_pages <= 0:
		return []
	if num_pages <= 10:
		return list(range(1, num_pages + 1))
	items: list[int | None] = [1]
	if current > 4:
		items.append(None)
	start = max(2, current - 1)
	end = min(num_pages - 1, current + 1)
	if current <= 4:
		start, end = 2, 4
	if current >= num_pages - 3:
		start, end = num_pages - 3, num_pages - 1
	for n in range(start, end + 1):
		if 1 < n < num_pages:
			items.append(n)
	if current < num_pages - 3:
		items.append(None)
	items.append(num_pages)
	compressed: list[int | None] = []
	for it in items:
		if compressed and compressed[-1] == it:
			continue
		if it is None and compressed and compressed[-1] is None:
			continue
		compressed.append(it)
	return compressed


@method_decorator(staff_required, name="dispatch")
class ManageBusinessPartnerViews(TemplateView):
	template_name = "manage_businesspartner.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			return _download_businesspartner_template()
		return super().get(request, *args, **kwargs)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request

		q = (request.GET.get("q") or "").strip()
		per_page_raw = (request.GET.get("per_page") or "").strip()
		page = (request.GET.get("page") or "1").strip() or "1"

		allowed_per_page = {100, 200, 500, 1000}
		try:
			per_page = int(per_page_raw or 100)
		except Exception:
			per_page = 100
		if per_page not in allowed_per_page:
			per_page = 100

		qs = BusinessPartner.objects.all()
		if q:
			qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q) | Q(tax_id__icontains=q))
		qs = qs.order_by("code")
		paginator = Paginator(qs, per_page)
		page_obj = paginator.get_page(page)

		rows = []
		for obj in page_obj.object_list:
			rows.append({
				"id": str(obj.id),
				"code": obj.code,
				"name": obj.name,
				"tax_id": obj.tax_id or "",
			})

		ctx["rows"] = rows
		ctx["q"] = q
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
		ctx["total_count"] = paginator.count
		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		obj_id = (request.POST.get("id") or "").strip()
		code = (request.POST.get("code") or "").strip()
		name = (request.POST.get("name") or "").strip()
		tax_id = (request.POST.get("tax_id") or "").strip()

		if action == "import_master_data":
			if openpyxl is None:
				messages.error(request, "ไม่สามารถนำเข้า XLSX ได้: ยังไม่ได้ติดตั้ง openpyxl")
				return redirect(request.get_full_path())
			upload = request.FILES.get("excel_file")
			if upload is None:
				messages.error(request, "กรุณาเลือกไฟล์ Excel (.xlsx)")
				return redirect(request.get_full_path())
			filename = (getattr(upload, "name", "") or "").lower()
			if not filename.endswith(".xlsx"):
				messages.error(request, "รองรับเฉพาะไฟล์ .xlsx")
				return redirect(request.get_full_path())
			created = updated = skipped = 0
			try:
				with transaction.atomic():
					for row in _parse_xlsx(upload):
						bp_code = _row_get_first(row, "code", "bp_code", "partner_code")
						bp_name = _row_get_first(row, "name", "partner_name")
						bp_tax = _row_get_first(row, "tax_id", "tax", "taxid")
						if not bp_code or not bp_name:
							skipped += 1
							continue
						existing = BusinessPartner.objects.filter(code__iexact=bp_code).first()
						if existing is None:
							BusinessPartner.objects.create(code=bp_code, name=bp_name, tax_id=bp_tax)
							created += 1
						else:
							existing.name = bp_name
							existing.tax_id = bp_tax
							existing.save(update_fields=["name", "tax_id", "updated_at"])
							updated += 1
			except Exception as e:
				log_event(request, action="businesspartner:import", status="failure", message="นำเข้า Business Partner ไม่สำเร็จ", metadata={"error": str(e)})
				messages.error(request, f"เกิดข้อผิดพลาดระหว่างนำเข้า: {e}")
				return redirect(request.get_full_path())
			messages.success(request, f"นำเข้าสำเร็จ: +{created}, อัปเดต {updated}, ข้าม {skipped}")
			return redirect(request.get_full_path())
		if action == "bulk_delete":
			bulk_ids = request.POST.getlist("bulk_id")
			ids = [x for x in [b.strip() for b in bulk_ids] if _is_uuid(x)]
			if not ids:
				messages.error(request, "กรุณาเลือกรายการที่ต้องการลบ")
				return redirect(request.get_full_path())
			deleted = blocked = 0
			try:
				with transaction.atomic():
					for pk in ids:
						obj = BusinessPartner.objects.filter(pk=pk).first()
						if obj is None:
							continue
						try:
							obj.delete()
							deleted += 1
						except ProtectedError:
							blocked += 1
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return redirect(request.get_full_path())
			if blocked:
				messages.warning(request, f"ลบสำเร็จ {deleted} รายการ, ลบไม่ได้ {blocked} รายการ (มีข้อมูลอ้างอิง)")
			else:
				messages.success(request, f"ลบสำเร็จ {deleted} รายการ")
			return redirect(request.get_full_path())
		if action == "create":
			if not code or not name:
				messages.error(request, "กรุณากรอก Code และ Name")
				return redirect(request.get_full_path())
			try:
				with transaction.atomic():
					obj = BusinessPartner.objects.create(code=code, name=name, tax_id=tax_id)
					messages.success(request, "เพิ่ม Business Partner สำเร็จ")
					transaction.on_commit(lambda: log_event(request, action="businesspartner:create", message="เพิ่ม Business Partner", metadata={"id": str(obj.pk), "code": code, "name": name}))
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
			return redirect(request.get_full_path())
		if action == "update":
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return redirect(request.get_full_path())
			if not code or not name:
				messages.error(request, "กรุณากรอก Code และ Name")
				return redirect(request.get_full_path())
			try:
				with transaction.atomic():
					obj = BusinessPartner.objects.get(pk=obj_id)
					updated = []
					if obj.code != code:
						obj.code = code
						updated.append("code")
					if obj.name != name:
						obj.name = name
						updated.append("name")
					if (obj.tax_id or "") != tax_id:
						obj.tax_id = tax_id
						updated.append("tax_id")
					if updated:
						updated.append("updated_at")
						obj.save(update_fields=updated)
						messages.success(request, "บันทึกการแก้ไขสำเร็จ")
					else:
						messages.info(request, "ไม่มีการเปลี่ยนแปลง")
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
			return redirect(request.get_full_path())
		if action == "delete":
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return redirect(request.get_full_path())
			try:
				with transaction.atomic():
					obj = BusinessPartner.objects.get(pk=obj_id)
					obj.delete()
					messages.success(request, "ลบสำเร็จ")
			except ProtectedError:
				messages.error(request, "ลบไม่ได้ มีข้อมูลอ้างอิงอยู่")
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
			return redirect(request.get_full_path())
		messages.error(request, "ไม่รู้จัก action")
		return redirect(request.get_full_path())
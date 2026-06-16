import calendar
from datetime import datetime
from decimal import Decimal
from uuid import UUID

from django.db.models import DateField, Max, Q, Sum
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView

from core.auth.decorators import staff_required
from core.models.item_line import ItemLine
from core.models.item_list import Item_list
from core.models.line import Line
from core.models.process_defect import ProcessDefectScrap

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill
except Exception:  # pragma: no cover
    openpyxl = None


@method_decorator(staff_required, name="dispatch")
class ScrapWeightReportViews(TemplateView):
    template_name = "core/report_scrap_weight.html"

    def get(self, request, *args, **kwargs):
        export = (request.GET.get("export") or "").strip().lower()
        if export == "xlsx":
            return self._export_xlsx()
        return super().get(request, *args, **kwargs)

    def _parse_month_year(self):
        month_raw = (self.request.GET.get("month") or "").strip()
        now = timezone.localtime(timezone.now())
        year = now.year
        month = now.month
        if month_raw:
            try:
                parts = month_raw.split("-")
                if len(parts) == 2:
                    year = int(parts[0])
                    month = int(parts[1])
                    if not (1 <= month <= 12):
                        raise ValueError("month out of range")
            except Exception:
                year = now.year
                month = now.month
        return year, month

    def _get_month_range(self, year: int, month: int):
        tz = timezone.get_current_timezone()
        start_dt = timezone.make_aware(datetime(year, month, 1, 0, 0, 0), tz)
        if month == 12:
            end_dt = timezone.make_aware(datetime(year + 1, 1, 1, 0, 0, 0), tz)
        else:
            end_dt = timezone.make_aware(datetime(year, month + 1, 1, 0, 0, 0), tz)
        return start_dt, end_dt

    def _build_report(self, year: int, month: int, selected_line: str):
        start_dt, end_dt = self._get_month_range(year, month)

        days_in_month = calendar.monthrange(year, month)[1]
        days = list(range(1, days_in_month + 1))

        # New backbone: scrapped components live on ProcessDefectScrap, tied to
        # the produced part via process_defect → production_record.
        pr = "process_defect__production_record__"
        records_qs = ProcessDefectScrap.objects.filter(
            Q(**{f"{pr}production_date__gte": start_dt.date(), f"{pr}production_date__lt": end_dt.date()})
            | Q(**{f"{pr}production_date__isnull": True, "created_at__gte": start_dt, "created_at__lt": end_dt})
        )
        if selected_line:
            records_qs = records_qs.filter(
                process_defect__production_record__line__line_name__iexact=selected_line
            )

        day_field = Coalesce(f"{pr}production_date", TruncDate("created_at"), output_field=DateField())

        # Aggregate quantity by (produced line, produced SD number, day)
        line_key = f"{pr}line_id"
        sd_key = f"{pr}item__sd_code"
        agg = (
            records_qs.annotate(work_date=day_field)
            .values(line_key, sd_key, "work_date")
            .annotate(
                total_qty=Sum("quantity"),
                part_name=Max("process_defect__production_record__item__part_name"),
                weight_per_unit=Max("process_defect__production_record__item__weight"),
            )
        )

        key_day_qty: dict[tuple, int] = {}
        key_weight: dict[tuple, Decimal] = {}
        key_part_name: dict[tuple, str] = {}
        key_total_qty: dict[tuple, int] = {}
        for r in agg:
            line_id = r.get(line_key)
            sd_code = (r.get(sd_key) or "").strip()
            work_date = r.get("work_date")
            total_qty = int(r.get("total_qty") or 0)
            part_name = (r.get("part_name") or "").strip()
            weight_per_unit = Decimal(str(r.get("weight_per_unit") or 0))

            if not line_id or not sd_code or work_date is None:
                continue
            day = int(getattr(work_date, "day", 0) or 0)
            if 1 <= day <= days_in_month:
                key_day_qty[(line_id, sd_code, day)] = total_qty
                key_weight[(line_id, sd_code)] = weight_per_unit
                key_part_name[(line_id, sd_code)] = part_name
                key_total_qty[(line_id, sd_code)] = key_total_qty.get((line_id, sd_code), 0) + total_qty

        # Distinct defect comments per (line, SD) — e.g. the reason typed for an
        # "อื่นๆ" defect. Rolled up so the matrix stays one row per part.
        key_comments: dict[tuple, list[str]] = {}
        comment_rows = (
            records_qs.exclude(process_defect__comment__isnull=True)
            .exclude(process_defect__comment__exact="")
            .values(line_key, sd_key, "process_defect__comment")
            .distinct()
        )
        for c in comment_rows:
            lid = c.get(line_key)
            sd = (c.get(sd_key) or "").strip()
            cm = (c.get("process_defect__comment") or "").strip()
            if not lid or not sd or not cm:
                continue
            bucket = key_comments.setdefault((lid, sd), [])
            if cm not in bucket:
                bucket.append(cm)

        # Build rows from ItemLine master (shows all items even with 0)
        item_lines = ItemLine.objects.select_related("line", "item")
        if selected_line:
            item_lines = item_lines.filter(line__line_name__iexact=selected_line)
        item_lines = item_lines.order_by("line__line_name", "item__sd_code", "item__part_name")

        row_keys: list[tuple] = []
        lines: dict = {}
        row_part_name: dict[tuple, str] = {}
        row_weight: dict[tuple, Decimal] = {}
        for il in item_lines:
            if not il.line_id or not il.item_id:
                continue
            sd_code = (getattr(il.item, "sd_code", None) or "").strip()
            if not sd_code:
                continue
            key = (il.line_id, sd_code)
            if key not in row_part_name:
                row_part_name[key] = (getattr(il.item, "part_name", None) or "").strip()
                row_weight[key] = Decimal(str(getattr(il.item, "weight", 0) or 0))
                row_keys.append(key)
            if il.line_id not in lines:
                lines[il.line_id] = il.line

        # Fallback: build from aggregated data
        if not row_keys:
            row_keys = sorted(
                {(lid, sd) for (lid, sd, _d) in key_day_qty.keys()},
                key=lambda x: (str(x[0]), x[1]),
            )
            line_ids = {lid for (lid, _sd) in row_keys}
            lines = {l.id: l for l in Line.objects.filter(id__in=line_ids)}
            row_part_name = {k: v for k, v in key_part_name.items()}
            row_weight = {k: key_weight.get(k, Decimal("0")) for k in row_keys}

        grand_total_qty = 0

        column_totals = {d: Decimal("0") for d in days}
        grand_total = Decimal("0")
        rows = []

        for (line_id, sd_code) in row_keys:
            line_obj = lines.get(line_id)
            line_label = getattr(line_obj, "line_name", None) or getattr(line_obj, "code", None) or "-"
            part_name = (
                (row_part_name.get((line_id, sd_code)) or "").strip()
                or (key_part_name.get((line_id, sd_code)) or "").strip()
                or "-"
            )
            w = row_weight.get((line_id, sd_code)) or key_weight.get((line_id, sd_code)) or Decimal("0")

            values = []
            row_total = Decimal("0")
            row_qty = int(key_total_qty.get((line_id, sd_code), 0))
            for d in days:
                qty = int(key_day_qty.get((line_id, sd_code, d), 0))
                weight_val = (Decimal(qty) * w).quantize(Decimal("0.00"))
                values.append(weight_val)
                row_total += weight_val
                column_totals[d] += weight_val
            grand_total += row_total
            grand_total_qty += row_qty

            rows.append(
                {
                    "production_line": (str(line_label) or "-").strip() or "-",
                    "sd_number": (str(sd_code) or "-").strip() or "-",
                    "part_name": (str(part_name) or "-").strip() or "-",
                    "total_qty": row_qty,
                    "weight_per_unit": w,
                    "values": values,
                    "total": row_total.quantize(Decimal("0.00")),
                    "comment": " · ".join(key_comments.get((line_id, sd_code), [])),
                }
            )

        return (
            days,
            rows,
            [column_totals[d].quantize(Decimal("0.00")) for d in days],
            grand_total.quantize(Decimal("0.00")),
            grand_total_qty,
        )

    def _export_xlsx(self):
        if openpyxl is None:
            return HttpResponse(
                "XLSX export is not available (openpyxl is not installed).",
                status=400,
                content_type="text/plain; charset=utf-8",
            )

        selected_line = (self.request.GET.get("line") or "").strip()
        year, month = self._parse_month_year()
        days, rows, column_totals, grand_total, grand_total_qty = self._build_report(year, month, selected_line)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year:04d}-{month:02d}"

        header = ["Production Line", "SD number", "Part name", "จำนวนทั้งหมด", "Weight/unit (kg)"] + [str(d) for d in days] + ["Total (kg)"]
        ws.append(header)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for r in rows:
            ws.append([
                r["production_line"],
                r["sd_number"],
                r["part_name"],
                r["total_qty"],
                float(r["weight_per_unit"]),
                *[float(v) for v in r["values"]],
                float(r["total"]),
            ])

        ws.append(["", "", "Total (kg)", grand_total_qty, "", *[float(v) for v in column_totals], float(grand_total)])

        ws.freeze_panes = "F2"
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        suffix = f"_{selected_line}" if selected_line else ""
        response["Content-Disposition"] = (
            f'attachment; filename="scrap_weight_{year:04d}-{month:02d}{suffix}.xlsx"'
        )
        wb.save(response)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        selected_line = (self.request.GET.get("line") or "").strip()
        year, month = self._parse_month_year()
        days, rows, column_totals, grand_total, grand_total_qty = self._build_report(year, month, selected_line)

        ctx.update(
            {
                "year": year,
                "month": month,
                "month_value": f"{year:04d}-{month:02d}",
                "production_lines": list(Line.objects.order_by("line_name").values_list("line_name", flat=True)),
                "selected_line": selected_line,
                "days": days,
                "rows": rows,
                "column_totals": column_totals,
                "grand_total": grand_total,
                "grand_total_qty": grand_total_qty,
            }
        )
        return ctx

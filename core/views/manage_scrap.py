from __future__ import annotations

import uuid
from types import SimpleNamespace

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.shortcuts import redirect

from core.auth.decorators import staff_required
from core.models.bill_of_material import BillOfMaterial
from core.models.bill_of_material_item_master import BillOfMaterialItemMater
from core.models.defect_by_category import DefectByCategory
from core.models.defect_mode import DefectMode
from core.models.item_line import ItemLine
from core.models.item_list import Item_list
from core.models.line import Line
from core.models.process_defect import ProcessDefect, ProcessDefectScrap, ProductionRecord
from core.services.auditlog import log_event


def _is_uuid(value: str) -> bool:
    try:
        uuid.UUID(str(value))
    except Exception:
        return False
    return True


def _adapt_scrap(s: ProcessDefectScrap) -> SimpleNamespace:
    """Present a ProcessDefectScrap as a flat row using the legacy ScrapRecord
    field names the template/JS expect (production_line, part_number, ...)."""
    pr = s.process_defect.production_record
    return SimpleNamespace(
        id=s.id,
        created_at=s.created_at,
        production_date=pr.production_date,
        lot_number=pr.lot_number,
        shift=pr.shift,
        created_by=pr.created_by,
        production_line=pr.line,       # Line  → str = line_name, .code works
        part_number=pr.item,           # Item_list (produced part) → .sd_code, .id
        defect_mode=s.process_defect.defect_mode,
        component_part=s.component_part,
        quantity=s.quantity,
        comment=s.process_defect.comment,
        photo=None,                    # ProcessDefectScrap has no photo field
    )


def _page_items(num_pages: int, current: int) -> list[int | None]:
    if num_pages <= 0:
        return []
    if num_pages <= 10:
        return list(range(1, num_pages + 1))
    items: list[int | None] = [1]
    if current > 4:
        items.append(None)
    start = max(2, current - 1)
    end = min(num_pages - 1, current + 1)
    if current <= 4:
        start, end = 2, 4
    if current >= num_pages - 3:
        start, end = num_pages - 3, num_pages - 1
    for n in range(start, end + 1):
        if 1 < n < num_pages:
            items.append(n)
    if current < num_pages - 3:
        items.append(None)
    items.append(num_pages)
    compressed: list[int | None] = []
    for it in items:
        if compressed and compressed[-1] == it:
            continue
        if it is None and compressed and compressed[-1] is None:
            continue
        compressed.append(it)
    return compressed


@method_decorator(staff_required, name="dispatch")
class ManageScrapViews(TemplateView):
    template_name = "core/manage_scrap.html"

    def get(self, request, *args, **kwargs):
        export_action = (request.GET.get("action") or "").strip().lower()
        if export_action == "export_excel":
            return self._export_excel(request)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        request = self.request

        q = (request.GET.get("q") or "").strip()
        date_from_raw = (request.GET.get("date_from") or "").strip()
        date_to_raw = (request.GET.get("date_to") or "").strip()
        per_page_raw = (request.GET.get("per_page") or "").strip()
        page = (request.GET.get("page") or "1").strip() or "1"
        date_from = parse_date(date_from_raw) if date_from_raw else None
        date_to = parse_date(date_to_raw) if date_to_raw else None

        qs = ProcessDefectScrap.objects.select_related(
            "process_defect__production_record__line",
            "process_defect__production_record__item",
            "process_defect__production_record__shift",
            "process_defect__production_record__created_by",
            "process_defect__production_record__created_by__profile",
            "process_defect__defect_mode",
            "component_part",
        ).all()

        pr = "process_defect__production_record__"
        if date_from:
            qs = qs.filter(Q(**{f"{pr}production_date__gte": date_from}) | Q(**{f"{pr}production_date__isnull": True, "created_at__date__gte": date_from}))
        if date_to:
            qs = qs.filter(Q(**{f"{pr}production_date__lte": date_to}) | Q(**{f"{pr}production_date__isnull": True, "created_at__date__lte": date_to}))

        if q:
            qs = qs.filter(
                Q(**{f"{pr}line__line_name__icontains": q})
                | Q(**{f"{pr}lot_number__icontains": q})
                | Q(**{f"{pr}item__part_number__icontains": q})
                | Q(**{f"{pr}item__sd_code__icontains": q})
                | Q(**{f"{pr}item__sku__icontains": q})
                | Q(process_defect__defect_mode__name_th__icontains=q)
                | Q(process_defect__defect_mode__name_en__icontains=q)
                | Q(component_part__part_name__icontains=q)
                | Q(component_part__part_number__icontains=q)
                | Q(**{f"{pr}created_by__username__icontains": q})
                | Q(**{f"{pr}created_by__first_name__icontains": q})
                | Q(**{f"{pr}created_by__profile__shift__icontains": q})
            )

        allowed_per_page = {100, 200, 500, 1000}
        try:
            per_page = int(per_page_raw or 100)
        except Exception:
            per_page = 100
        if per_page not in allowed_per_page:
            per_page = 100

        qs = qs.order_by("-created_at")
        paginator = Paginator(qs, per_page)
        page_obj = paginator.get_page(page)

        ctx["component_part_records"] = [_adapt_scrap(s) for s in page_obj.object_list]
        ctx["can_edit"] = bool(getattr(request.user, "is_superuser", False))
        ctx["q"] = q
        ctx["date_from"] = date_from_raw
        ctx["date_to"] = date_to_raw
        ctx["page_obj"] = page_obj
        ctx["paginator"] = paginator
        ctx["per_page"] = per_page
        ctx["rows_total"] = paginator.count
        ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
        ctx["total_count"] = paginator.count
        ctx.setdefault("delete_action", "")

        ctx["record_data"] = self._build_record_data_payload()
        return ctx

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "").strip().lower()
        rec_id = (request.POST.get("id") or "").strip()

        if action == "bulk_delete":
            raw_ids = request.POST.getlist("bulk_id")
            ids = [rid for rid in raw_ids if _is_uuid((rid or "").strip())]
            if not ids:
                messages.error(request, "กรุณาเลือกรายการที่ต้องการลบ")
                return redirect(request.get_full_path())
            with transaction.atomic():
                deleted, _ = ProcessDefectScrap.objects.filter(pk__in=ids).delete()

            messages.success(request, f"ลบสำเร็จ {deleted} รายการ")
            transaction.on_commit(
                lambda: log_event(
                    request,
                    action="scrap:bulk_delete",
                    message="ลบ ScrapRecord แบบ bulk",
                    metadata={"selected": len(ids), "deleted": deleted, "ids": ids[:50]},
                )
            )
            return redirect(request.get_full_path())
        if action in {"delete", "update"}:
            if not _is_uuid(rec_id):
                messages.error(request, "ไม่พบรหัสรายการ")
                return redirect(request.get_full_path())
        if action == "delete":
            obj = (
                ProcessDefectScrap.objects.select_related(
                    "process_defect__production_record__line",
                    "process_defect__production_record__item",
                    "process_defect__defect_mode",
                    "component_part",
                )
                .filter(pk=rec_id)
                .first()
            )
            deleted, _ = ProcessDefectScrap.objects.filter(pk=rec_id).delete()
            if deleted:
                messages.success(request, "ลบรายการสำเร็จ")
                meta = {"record_id": rec_id}
                if obj is not None:
                    pr = obj.process_defect.production_record
                    meta.update(
                        {
                            "line_code": getattr(pr.line, "code", ""),
                            "part_number": getattr(pr.item, "sd_code", ""),
                            "defect": getattr(obj.process_defect.defect_mode, "name", ""),
                            "component_part": getattr(obj.component_part, "name", ""),
                            "quantity": obj.quantity,
                        }
                    )
                transaction.on_commit(
                    lambda: log_event(
                        request,
                        action="scrap:delete",
                        message="ลบ ProcessDefectScrap",
                        metadata=meta,
                    )
                )
            else:
                messages.error(request, "ไม่พบรายการ")
            return redirect(request.get_full_path())
        if action == "update":
            # Editing is restricted to Admin (superuser); others may view/delete only.
            if not getattr(request.user, "is_superuser", False):
                messages.error(request, "เฉพาะผู้ดูแลระบบ (Admin) เท่านั้นที่แก้ไขได้")
                return redirect(request.get_full_path())

            line_code = (request.POST.get("line_code") or "").strip()
            part_ref = (request.POST.get("part_number") or "").strip()
            defect_id = (request.POST.get("defect_id") or "").strip()
            component_part_id = (request.POST.get("component_part_id") or "").strip()
            qty_raw = (request.POST.get("quantity") or "").strip()
            comment = (request.POST.get("comment") or "").strip() or None

            try:
                quantity = int(qty_raw)
            except Exception:
                quantity = None

            if quantity is None or quantity < 1:
                messages.error(request, "กรุณาระบุ Quantity เป็นตัวเลข (>= 1)")
                return redirect(request.get_full_path())
            if not line_code or not part_ref or not _is_uuid(defect_id) or not _is_uuid(component_part_id):
                messages.error(request, "กรุณาเลือก Line / SD number / Defect / Part name ให้ครบ")
                return redirect(request.get_full_path())

            rec = (
                ProcessDefectScrap.objects.select_related("process_defect__production_record")
                .filter(pk=rec_id)
                .first()
            )
            if rec is None:
                messages.error(request, "ไม่พบรายการ")
                return redirect(request.get_full_path())
            pd = rec.process_defect
            pr = pd.production_record
            old_snapshot = {
                "line_id": str(pr.line_id),
                "part_id": str(pr.item_id),
                "defect_id": str(pd.defect_mode_id),
                "component_part_id": str(rec.component_part_id),
                "quantity": rec.quantity,
            }

            line = Line.objects.filter(line_name__iexact=line_code).first()
            if line is None:
                messages.error(request, "ไม่พบ Production line")
                return redirect(request.get_full_path())
            if _is_uuid(part_ref):
                part = Item_list.objects.filter(pk=part_ref).filter(item_lines__line=line).distinct().first()
            else:
                part = Item_list.objects.filter(part_number__iexact=part_ref).filter(item_lines__line=line).distinct().first()
            if part is None:
                messages.error(request, "ไม่พบ SD number ใน Production line ที่เลือก")
                return redirect(request.get_full_path())
            defect = DefectMode.objects.filter(pk=defect_id).first()
            if defect is None:
                messages.error(request, "ไม่พบ Defect mode")
                return redirect(request.get_full_path())
            component_part = Item_list.objects.filter(pk=component_part_id).first()
            if component_part is None:
                messages.error(request, "ไม่พบ Part name")
                return redirect(request.get_full_path())

            with transaction.atomic():
                changed: list[str] = []
                # Parent ProductionRecord (line / produced part) — shared by every
                # defect & scrap of the same production lot.
                pr_fields: list[str] = []
                if str(pr.line_id) != str(line.id):
                    pr.line = line
                    pr_fields.append("line")
                if str(pr.item_id) != str(part.id):
                    pr.item = part
                    pr_fields.append("item")
                if pr_fields:
                    pr.save(update_fields=pr_fields + ["updated_at"])
                    changed += pr_fields
                # Parent ProcessDefect (defect mode / comment) — shared by sibling scraps.
                pd_fields: list[str] = []
                if str(pd.defect_mode_id) != str(defect.id):
                    pd.defect_mode = defect
                    pd_fields.append("defect_mode")
                if pd.comment != comment:
                    pd.comment = comment
                    pd_fields.append("comment")
                if pd_fields:
                    pd.save(update_fields=pd_fields + ["updated_at"])
                    changed += pd_fields
                # The scrap row itself (component / quantity).
                rec_fields: list[str] = []
                if str(rec.component_part_id) != str(component_part.id):
                    rec.component_part = component_part
                    rec_fields.append("component_part")
                if rec.quantity != quantity:
                    rec.quantity = quantity
                    rec_fields.append("quantity")
                if rec_fields:
                    rec.save(update_fields=rec_fields + ["updated_at"])
                    changed += rec_fields

                if changed:
                    messages.success(request, "แก้ไขรายการสำเร็จ")
                    new_snapshot = {
                        "line_id": str(pr.line_id),
                        "part_id": str(pr.item_id),
                        "defect_id": str(pd.defect_mode_id),
                        "component_part_id": str(rec.component_part_id),
                        "quantity": rec.quantity,
                    }
                    transaction.on_commit(
                        lambda: log_event(
                            request,
                            action="scrap:update",
                            message="แก้ไข ProcessDefectScrap",
                            metadata={
                                "record_id": rec_id,
                                "changed_fields": changed,
                                "old": old_snapshot,
                                "new": new_snapshot,
                            },
                        )
                    )
                else:
                    messages.info(request, "ไม่มีการเปลี่ยนแปลง")

            return redirect(request.get_full_path())
        return redirect(request.get_full_path())
    def _build_record_data_payload(self) -> dict:
        lines = list(Line.objects.all().order_by("line_name"))
        line_names = [l.code for l in lines]

        item_lines = list(
            ItemLine.objects.select_related("item", "line")
            .filter(line__line_name__in=line_names)
            .order_by("line__line_name", "item__sd_code", "item__part_number")
        )

        items_by_line: dict[str, list[Item_list]] = {ln: [] for ln in line_names}
        item_ids: set[str] = set()
        for il in item_lines:
            items_by_line.setdefault(il.line.code, []).append(il.item)
            item_ids.add(str(il.item_id))

        parts = list(Item_list.objects.filter(pk__in=list(item_ids)))
        parts_by_id = {str(p.id): p for p in parts}

        all_defects = list(DefectMode.objects.all().order_by("name_th", "name_en"))

        category_ids = {str(p.category_id) for p in parts if getattr(p, "category_id", None)}
        cat_to_defects: dict[str, list[DefectMode]] = {}
        if category_ids:
            dbc_qs = (
                DefectByCategory.objects.filter(category_id__in=list(category_ids))
                .select_related("defect_mode")
                .order_by("defect_mode__name_th", "defect_mode__name_en")
            )
            grouped: dict[str, list[DefectMode]] = {}
            grouped_inlist: dict[str, list[DefectMode]] = {}
            for dbc in dbc_qs:
                cat_key = str(dbc.category_id)
                dm = dbc.defect_mode
                grouped.setdefault(cat_key, []).append(dm)
                if getattr(dbc, "is_inlist", False):
                    grouped_inlist.setdefault(cat_key, []).append(dm)
            for cat_key, defects_for_cat in grouped.items():
                preferred = grouped_inlist.get(cat_key) or defects_for_cat
                seen: set[str] = set()
                uniq: list[DefectMode] = []
                for dm in preferred:
                    dm_id = str(dm.id)
                    if dm_id in seen:
                        continue
                    seen.add(dm_id)
                    uniq.append(dm)
                cat_to_defects[cat_key] = uniq

        boms = list(
            BillOfMaterial.objects.filter(item__in=parts)
            .select_related("item")
            .prefetch_related(
                "items_master",
                "items_master__component",
            )
            .order_by("-updated_at")
        )
        components_by_item_id: dict[str, list[dict]] = {}
        for bom in boms:
            key = str(bom.item_id)
            if key in components_by_item_id:
                continue
            comps: list[dict] = []
            for it in getattr(bom, "items_master", []).all():
                comp = getattr(it, "component", None)
                if comp is None:
                    continue
                comps.append({"id": str(comp.id), "name": comp.part_name or comp.part_number or comp.sku})
            components_by_item_id[key] = comps

        production_lines_payload = []
        for line in lines:
            parts_payload = []
            for part_ref in items_by_line.get(line.code, []):
                part = parts_by_id.get(str(part_ref.id), part_ref)
                part_display_name = (getattr(part, "part_name", "") or getattr(part, "part_number", "") or getattr(part, "sku", "") or "").strip()
                # Editable Part name options = the product itself (FG) + its BOM
                # children — the same set a scrap row can reference.
                comps = [{"id": str(part.id), "name": f"FG · {part_display_name}".strip()}]
                comps.extend(components_by_item_id.get(str(part.id), []))

                defect_list = cat_to_defects.get(str(getattr(part, "category_id", ""))) or all_defects

                defects_payload = []
                for defect in defect_list:
                    defects_payload.append(
                        {
                            "id": str(defect.id),
                            "name": defect.name,
                            "component_parts": comps,
                        }
                    )

                parts_payload.append(
                    {
                        "id": str(part.id),
                        "sd_number": (getattr(part, "sd_code", "") or "").strip(),
                        "part_number": (getattr(part, "part_number", "") or "").strip(),
                        "part_name": (getattr(part, "part_name", "") or "").strip(),
                        "defects": defects_payload,
                        "component_parts": comps,
                    }
                )
            production_lines_payload.append({"code": line.code, "parts": parts_payload})

        return {"productionLines": production_lines_payload}

    def _export_excel(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception:
            messages.error(request, "ไม่สามารถ export Excel ได้เนื่องจากไม่มี openpyxl")
            return redirect(request.get_full_path())
        q = (request.GET.get("q") or "").strip()
        date_from_raw = (request.GET.get("date_from") or "").strip()
        date_to_raw = (request.GET.get("date_to") or "").strip()
        date_from = parse_date(date_from_raw) if date_from_raw else None
        date_to = parse_date(date_to_raw) if date_to_raw else None

        qs = ProcessDefectScrap.objects.select_related(
            "process_defect__production_record__line",
            "process_defect__production_record__item",
            "process_defect__production_record__shift",
            "process_defect__production_record__created_by",
            "process_defect__production_record__created_by__profile",
            "process_defect__defect_mode",
            "component_part",
        ).all()

        pr = "process_defect__production_record__"
        if date_from:
            qs = qs.filter(Q(**{f"{pr}production_date__gte": date_from}) | Q(**{f"{pr}production_date__isnull": True, "created_at__date__gte": date_from}))
        if date_to:
            qs = qs.filter(Q(**{f"{pr}production_date__lte": date_to}) | Q(**{f"{pr}production_date__isnull": True, "created_at__date__lte": date_to}))
        if q:
            qs = qs.filter(
                Q(**{f"{pr}line__line_name__icontains": q})
                | Q(**{f"{pr}lot_number__icontains": q})
                | Q(**{f"{pr}item__part_number__icontains": q})
                | Q(process_defect__defect_mode__name_th__icontains=q)
                | Q(process_defect__defect_mode__name_en__icontains=q)
                | Q(component_part__part_name__icontains=q)
                | Q(**{f"{pr}created_by__username__icontains": q})
                | Q(**{f"{pr}created_by__first_name__icontains": q})
                | Q(**{f"{pr}created_by__profile__shift__icontains": q})
            )

        qs = qs.order_by("-created_at")

        wb = Workbook()
        ws = wb.active
        ws.title = "Scrap Records"

        headers = [
            "วันที่/เวลา",
            "วันทำการ",
            "ผู้ใช้งาน",
            "กะ",
            "Production line",
            "Lot no.",
            "SD number",
            "Defect mode",
            "Comment",
            "Part name",
            "Quantity",
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for s in qs:
            r = _adapt_scrap(s)
            shift_display = "-"
            if r.shift:
                shift_display = r.shift.name
            elif r.created_by and hasattr(r.created_by, "profile") and r.created_by.profile:
                shift_value = r.created_by.profile.shift
                if shift_value == "shift_a":
                    shift_display = "กะ A"
                elif shift_value == "shift_b":
                    shift_display = "กะ B"
                else:
                    shift_display = "กะ Day"

            created_at_local = timezone.localtime(r.created_at) if r.created_at else None
            ws.append(
                [
                    created_at_local.strftime("%d/%m/%Y %H:%M") if created_at_local else "-",
                    r.production_date.strftime("%d/%m/%Y") if r.production_date else "-",
                    r.created_by.get_short_name() if r.created_by else "-",
                    shift_display,
                    getattr(r.production_line, "code", "-"),
                    r.lot_number or "-",
                    getattr(r.part_number, "sd_code", "-") or "-",
                    getattr(r.defect_mode, "name", "-"),
                    r.comment or "-",
                    getattr(r.component_part, "part_name", "-") or "-",
                    r.quantity or 0,
                ]
            )

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 24
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 22
        ws.column_dimensions["I"].width = 22
        ws.column_dimensions["J"].width = 22
        ws.column_dimensions["K"].width = 12

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename_ts = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        response["Content-Disposition"] = f'attachment; filename="ScrapRecords_{filename_ts}.xlsx"'
        wb.save(response)
        return response


# Backward compatible name (older code referenced this class name)
ManageComponentPartViews = ManageScrapViews

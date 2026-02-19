from __future__ import annotations

import re
import uuid
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView

from core.auth.decorators import staff_required
from core.models.bill_of_material import BillOfMaterial
from core.models.item_list import Item_list
from core.services.auditlog import log_event

try:
	import openpyxl  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None


def _is_uuid(value: str) -> bool:
	try:
		uuid.UUID(str(value))
	except Exception:
		return False
	return True


def _page_items(num_pages: int, current: int) -> list[int | None]:
	if num_pages <= 0:
		return []
	if num_pages <= 10:
		return list(range(1, num_pages + 1))
	items: list[int | None] = [1]
	if current > 4:
		items.append(None)
	start = max(2, current - 1)
	end = min(num_pages - 1, current + 1)
	if current <= 4:
		start, end = 2, 4
	if current >= num_pages - 3:
		start, end = num_pages - 3, num_pages - 1
	for n in range(start, end + 1):
		if 1 < n < num_pages:
			items.append(n)
	if current < num_pages - 3:
		items.append(None)
	items.append(num_pages)
	compressed: list[int | None] = []
	for it in items:
		if compressed and compressed[-1] == it:
			continue
		if it is None and compressed and compressed[-1] is None:
			continue
		compressed.append(it)
	return compressed


def _normalized_key(key: str) -> str:
	key = (key or "").strip().lower()
	key = re.sub(r"[^0-9a-z]+", "_", key)
	key = re.sub(r"_+", "_", key).strip("_")
	return key


def _excel_to_str(value) -> str:
	if value is None:
		return ""
	if isinstance(value, bool):
		return "TRUE" if value else "FALSE"
	if isinstance(value, int):
		return str(value)
	if isinstance(value, float):
		if value.is_integer():
			return str(int(value))
		return str(value)
	return str(value).strip()


def _row_get_first(row: dict, *keys: str) -> str:
	for k in keys:
		if not k:
			continue
		v = row.get(k)
		s = _excel_to_str(v).strip()
		if s != "":
			return s
	return ""


def _safe_decimal(value, default: Decimal = Decimal("0")) -> Decimal:
	if value is None:
		return default
	if isinstance(value, Decimal):
		return value
	if isinstance(value, (int, float)):
		try:
			return Decimal(str(value))
		except (InvalidOperation, ValueError):
			return default
	value = str(value).strip()
	if value == "":
		return default
	try:
		return Decimal(value)
	except (InvalidOperation, ValueError):
		return default


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


def download_manage_bill_of_material_import_template(request):
	"""Download a template for importing BOM headers."""
	sample_sku = "SKU-0001"
	item = Item_list.objects.order_by("sku").first()
	if item is not None:
		sample_sku = item.sku

	headers = ["sku", "revision", "lasted_eci", "scrap_percent"]
	rows = [[sample_sku, "A", "ECI-0001", 0]]
	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "bom_header"
	ws.append(headers)
	for r in rows:
		ws.append(r)
	for col in range(1, len(headers) + 1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="manage_bill_of_material_import_template.xlsx"'
	wb.save(response)
	return response


@method_decorator(staff_required, name="dispatch")
class ManageBillOfMaterialViews(TemplateView):
	template_name = "manage_bill_of_material.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			return download_manage_bill_of_material_import_template(request)
		return super().get(request, *args, **kwargs)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request

		q = (request.GET.get("q") or "").strip()
		per_page_raw = (request.GET.get("per_page") or "").strip()
		page = (request.GET.get("page") or "1").strip() or "1"

		allowed_per_page = {20, 50, 100, 200}
		try:
			per_page = int(per_page_raw or 20)
		except Exception:
			per_page = 20
		if per_page not in allowed_per_page:
			per_page = 20

		qs = BillOfMaterial.objects.select_related("item").all()
		if q:
			qs = qs.filter(
				Q(item__sku__icontains=q)
				| Q(item__part_number__icontains=q)
				| Q(item__part_name__icontains=q)
				| Q(revision__icontains=q)
				| Q(lasted_eci__icontains=q)
			)
		qs = qs.order_by("item__sku", "revision")
		paginator = Paginator(qs, per_page)
		page_obj = paginator.get_page(page)

		rows = []
		for obj in page_obj.object_list:
			rows.append(
				{
					"id": str(obj.id),
					"item_id": str(obj.item_id) if obj.item_id else "",
					"sku": getattr(obj.item, "sku", "") if obj.item_id else "",
					"part_name": getattr(obj.item, "part_name", "") if obj.item_id else "",
					"revision": obj.revision,
					"lasted_eci": obj.lasted_eci,
					"scrap_percent": str(obj.scrap_percent),
				}
			)

		ctx["rows"] = rows
		ctx["q"] = q
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
		ctx["total_count"] = paginator.count
		ctx["items"] = list(Item_list.objects.order_by("sku").values("id", "sku", "part_name"))
		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		obj_id = (request.POST.get("id") or "").strip()

		if action == "import_master_data":
			if openpyxl is None:
				messages.error(request, "ไม่สามารถนำเข้า XLSX ได้: ยังไม่ได้ติดตั้ง openpyxl")
				return self.get(request, *args, **kwargs)
			upload = request.FILES.get("excel_file")
			if upload is None:
				messages.error(request, "กรุณาเลือกไฟล์ Excel (.xlsx)")
				return self.get(request, *args, **kwargs)
			name = (getattr(upload, "name", "") or "").lower()
			if not name.endswith(".xlsx"):
				messages.error(request, "รองรับเฉพาะไฟล์ .xlsx")
				return self.get(request, *args, **kwargs)

			created = 0
			updated = 0
			skipped = 0
			item_not_found = 0
			try:
				with transaction.atomic():
					for row in _parse_xlsx(upload):
						sku = _row_get_first(row, "sku", "item_sku")
						revision = _row_get_first(row, "revision", "rev") or "A"
						lasted_eci = _row_get_first(row, "lasted_eci", "eci")
						scrap_percent = _safe_decimal(row.get("scrap_percent"), default=Decimal("0"))

						if not sku or not lasted_eci:
							skipped += 1
							continue
						item = Item_list.objects.filter(sku__iexact=sku).first()
						if item is None:
							item_not_found += 1
							continue

						obj = BillOfMaterial.objects.filter(item=item, revision__iexact=revision).first()
						if obj is None:
							BillOfMaterial.objects.create(
								item=item,
								revision=revision,
								lasted_eci=lasted_eci,
								scrap_percent=scrap_percent,
								user=request.user,
							)
							created += 1
						else:
							obj.revision = revision
							obj.lasted_eci = lasted_eci
							obj.scrap_percent = scrap_percent
							obj.save(update_fields=["revision", "lasted_eci", "scrap_percent", "updated_at"])
							updated += 1
			except Exception as e:
				log_event(
					request,
					action="bom:import_master_data",
					status="failure",
					message="นำเข้า BOM Header ไม่สำเร็จ",
					metadata={"filename": getattr(upload, "name", ""), "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาดระหว่างนำเข้า: {e}")
				return self.get(request, *args, **kwargs)

			messages.success(
				request,
				f"นำเข้าสำเร็จ: +{created}, อัปเดต {updated}, ข้าม {skipped}, ไม่พบ Item {item_not_found}",
			)
			transaction.on_commit(
				lambda: log_event(
					request,
					action="bom:import_master_data",
					message="นำเข้า BOM Header สำเร็จ",
					metadata={
						"filename": getattr(upload, "name", ""),
						"created": created,
						"updated": updated,
						"skipped": skipped,
						"item_not_found": item_not_found,
					},
				),
			)
			return self.get(request, *args, **kwargs)

		if action == "bulk_delete_boms":
			bulk_ids = request.POST.getlist("bulk_id")
			ids = [pk for pk in bulk_ids if _is_uuid((pk or "").strip())]
			if not ids:
				messages.error(request, "กรุณาเลือกรายการที่ต้องการลบ")
				return self.get(request, *args, **kwargs)

			deleted = 0
			blocked = 0
			not_found = 0
			try:
				with transaction.atomic():
					for pk in ids:
						obj = BillOfMaterial.objects.filter(pk=pk).first()
						if obj is None:
							not_found += 1
							continue
						try:
							obj.delete()
							deleted += 1
						except ProtectedError:
							blocked += 1
			except Exception as e:
				log_event(
					request,
					action="bom:bulk_delete",
					status="failure",
					message="ลบ BOM แบบ bulk ไม่สำเร็จ",
					metadata={"selected": len(ids), "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

			transaction.on_commit(
				lambda: log_event(
					request,
					action="bom:bulk_delete",
					message="ลบ BOM แบบ bulk",
					metadata={"selected": len(ids), "deleted": deleted, "blocked": blocked, "not_found": not_found},
				),
			)
			if blocked:
				messages.warning(request, f"ลบสำเร็จ {deleted} รายการ (ลบไม่ได้ {blocked}, ไม่พบ {not_found})")
			else:
				messages.success(request, f"ลบสำเร็จ {deleted} รายการ")
			return self.get(request, *args, **kwargs)

		if action in {"update", "delete"}:
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return self.get(request, *args, **kwargs)

		if action == "delete":
			obj = BillOfMaterial.objects.filter(pk=obj_id).first()
			if obj is None:
				messages.error(request, "ไม่พบรายการ")
				return self.get(request, *args, **kwargs)
			try:
				obj.delete()
			except ProtectedError:
				messages.error(request, "ไม่สามารถลบได้: รายการนี้ถูกใช้งานอยู่")
				return self.get(request, *args, **kwargs)
			transaction.on_commit(
				lambda: log_event(
					request,
					action="bom:delete",
					message="ลบ BOM",
					metadata={"id": obj_id},
				),
			)
			messages.success(request, "ลบรายการสำเร็จ")
			return self.get(request, *args, **kwargs)

		item_id = (request.POST.get("item_id") or "").strip()
		revision = (request.POST.get("revision") or "A").strip() or "A"
		lasted_eci = (request.POST.get("lasted_eci") or "").strip()
		scrap_percent_raw = (request.POST.get("scrap_percent") or "").strip()

		if not _is_uuid(item_id):
			messages.error(request, "กรุณาเลือก Item")
			return self.get(request, *args, **kwargs)
		item = Item_list.objects.filter(pk=item_id).first()
		if item is None:
			messages.error(request, "ไม่พบ Item")
			return self.get(request, *args, **kwargs)
		if not lasted_eci:
			messages.error(request, "กรุณาระบุ Lasted ECI")
			return self.get(request, *args, **kwargs)
		try:
			scrap_percent = _safe_decimal(scrap_percent_raw, default=Decimal("0"))
		except Exception:
			scrap_percent = Decimal("0")

		if action == "update":
			obj = BillOfMaterial.objects.filter(pk=obj_id).first()
			if obj is None:
				messages.error(request, "ไม่พบรายการ")
				return self.get(request, *args, **kwargs)
			obj.item = item
			obj.revision = revision
			obj.lasted_eci = lasted_eci
			obj.scrap_percent = scrap_percent
			try:
				obj.save(update_fields=["item", "revision", "lasted_eci", "scrap_percent", "updated_at"])
			except IntegrityError as e:
				messages.error(request, f"บันทึกไม่สำเร็จ: {e}")
				return self.get(request, *args, **kwargs)
			transaction.on_commit(
				lambda: log_event(
					request,
					action="bom:update",
					message="แก้ไข BOM Header",
					metadata={"id": obj_id, "sku": item.sku, "revision": revision},
				),
			)
			messages.success(request, "แก้ไขรายการสำเร็จ")
			return self.get(request, *args, **kwargs)

		# Default: create
		try:
			obj = BillOfMaterial.objects.create(
				item=item,
				revision=revision,
				lasted_eci=lasted_eci,
				scrap_percent=scrap_percent,
				user=request.user,
			)
		except Exception as e:
			log_event(
				request,
				action="bom:create",
				status="failure",
				message="เพิ่ม BOM Header ไม่สำเร็จ",
				metadata={"error": str(e)},
			)
			messages.error(request, f"เพิ่มข้อมูลไม่สำเร็จ: {e}")
			return self.get(request, *args, **kwargs)

		transaction.on_commit(
			lambda: log_event(
				request,
				action="bom:create",
				message="เพิ่ม BOM Header",
				metadata={"id": str(obj.id), "sku": item.sku, "revision": revision},
			),
		)
		messages.success(request, "เพิ่มข้อมูลสำเร็จ")
		return self.get(request, *args, **kwargs)

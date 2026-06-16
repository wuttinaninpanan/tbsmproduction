from __future__ import annotations

import re
import uuid

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.shortcuts import redirect

from core.auth.decorators import staff_required
from core.models.businesspartner import Address, BusinessPartner
from core.services.auditlog import log_event

try:
	import openpyxl  # type: ignore
except Exception:
	openpyxl = None


def _is_uuid(value: str) -> bool:
	try:
		uuid.UUID(str(value))
	except Exception:
		return False
	return True


def _normalized_key(key: str) -> str:
	key = (key or "").strip().lower()
	key = re.sub(r"[^0-9a-z]+", "_", key)
	key = re.sub(r"_+", "_", key).strip("_")
	return key


def _excel_to_str(value) -> str:
	if value is None:
		return ""
	if isinstance(value, bool):
		return "TRUE" if value else "FALSE"
	if isinstance(value, int):
		return str(value)
	if isinstance(value, float):
		if value.is_integer():
			return str(int(value))
		return str(value)
	return str(value).strip()


def _row_get_first(row: dict, *keys: str) -> str:
	for key in keys:
		if not key:
			continue
		value = row.get(key)
		text = _excel_to_str(value).strip()
		if text != "":
			return text
	return ""


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(header) if header is not None else "") for header in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			key = keys[idx] if idx < len(keys) else ""
			if not key:
				continue
			row[key] = value
		yield row


def _normalize_address_type(value: str) -> str:
	raw = (value or "").strip().lower()
	aliases = {
		"billing": "billing",
		"bill": "billing",
		"shipping": "shipping",
		"ship": "shipping",
		"head_office": "head_office",
		"head office": "head_office",
		"headoffice": "head_office",
		"สำนักงานใหญ่": "head_office",
	}
	return aliases.get(raw, raw)


def _download_address_template():
	headers = [
		"partner_code",
		"partner_name",
		"address_type",
		"address_line1",
		"address_line2",
		"subdistrict",
		"district",
		"province",
		"postal_code",
		"country",
	]
	rows = [
		["BP001", "ชื่อบริษัท", "billing", "123 ถนนสุขุมวิท", "ชั้น 5", "คลองเตย", "คลองเตย", "กรุงเทพมหานคร", "10110", "Thailand"],
		["BP001", "ชื่อบริษัท", "shipping", "88/8 นิคมอุตสาหกรรม", "อาคาร A", "ทับกวาง", "แก่งคอย", "สระบุรี", "18260", "Thailand"],
	]
	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "address_partners"
	ws.append(headers)
	for row in rows:
		ws.append(row)
	for col in range(1, len(headers) + 1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="address_partner_import_template.xlsx"'
	wb.save(response)
	return response


def _page_items(num_pages: int, current: int) -> list[int | None]:
	if num_pages <= 0:
		return []
	if num_pages <= 10:
		return list(range(1, num_pages + 1))
	items: list[int | None] = [1]
	if current > 4:
		items.append(None)
	start = max(2, current - 1)
	end = min(num_pages - 1, current + 1)
	if current <= 4:
		start, end = 2, 4
	if current >= num_pages - 3:
		start, end = num_pages - 3, num_pages - 1
	for n in range(start, end + 1):
		if 1 < n < num_pages:
			items.append(n)
	if current < num_pages - 3:
		items.append(None)
	items.append(num_pages)
	compressed: list[int | None] = []
	for it in items:
		if compressed and compressed[-1] == it:
			continue
		if it is None and compressed and compressed[-1] is None:
			continue
		compressed.append(it)
	return compressed


ADDRESS_TYPE_CHOICES = [
	("billing", "Billing"),
	("shipping", "Shipping"),
	("head_office", "Head Office"),
]


@method_decorator(staff_required, name="dispatch")
class ManageAddressPartnerViews(TemplateView):
	template_name = "core/manage_address_partner.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			return _download_address_template()
		return super().get(request, *args, **kwargs)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request

		q = (request.GET.get("q") or "").strip()
		per_page_raw = (request.GET.get("per_page") or "").strip()
		page = (request.GET.get("page") or "1").strip() or "1"

		allowed_per_page = {100, 200, 500, 1000}
		try:
			per_page = int(per_page_raw or 100)
		except Exception:
			per_page = 100
		if per_page not in allowed_per_page:
			per_page = 100

		qs = Address.objects.select_related("partner").all()
		if q:
			qs = qs.filter(
				Q(partner__name__icontains=q) |
				Q(partner__code__icontains=q) |
				Q(address_line1__icontains=q) |
				Q(province__icontains=q)
			)
		qs = qs.order_by("partner__code", "address_type")
		paginator = Paginator(qs, per_page)
		page_obj = paginator.get_page(page)

		rows = []
		for obj in page_obj.object_list:
			rows.append({
				"id": str(obj.id),
				"partner_id": str(obj.partner_id),
				"partner_name": obj.partner.name,
				"address_type": obj.address_type,
				"address_type_display": obj.get_address_type_display(),
				"address_line1": obj.address_line1,
				"address_line2": obj.address_line2 or "",
				"subdistrict": obj.subdistrict or "",
				"district": obj.district or "",
				"province": obj.province,
				"postal_code": obj.postal_code,
				"country": obj.country,
			})

		ctx["rows"] = rows
		ctx["partners"] = list(BusinessPartner.objects.order_by("code").values("id", "code", "name"))
		ctx["address_types"] = ADDRESS_TYPE_CHOICES
		ctx["q"] = q
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
		ctx["total_count"] = paginator.count
		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		obj_id = (request.POST.get("id") or "").strip()
		partner_id = (request.POST.get("partner_id") or "").strip()
		address_type = (request.POST.get("address_type") or "").strip()
		address_line1 = (request.POST.get("address_line1") or "").strip()
		address_line2 = (request.POST.get("address_line2") or "").strip()
		subdistrict = (request.POST.get("subdistrict") or "").strip()
		district = (request.POST.get("district") or "").strip()
		province = (request.POST.get("province") or "").strip()
		postal_code = (request.POST.get("postal_code") or "").strip()
		country = (request.POST.get("country") or "Thailand").strip()

		if action == "import_master_data":
			if openpyxl is None:
				messages.error(request, "ไม่สามารถนำเข้า XLSX ได้: ยังไม่ได้ติดตั้ง openpyxl")
				return redirect(request.get_full_path())
			upload = request.FILES.get("excel_file")
			if upload is None:
				messages.error(request, "กรุณาเลือกไฟล์ Excel (.xlsx)")
				return redirect(request.get_full_path())
			filename = (getattr(upload, "name", "") or "").lower()
			if not filename.endswith(".xlsx"):
				messages.error(request, "รองรับเฉพาะไฟล์ .xlsx")
				return redirect(request.get_full_path())
			created = updated = skipped = 0
			try:
				with transaction.atomic():
					for row in _parse_xlsx(upload):
						row_id = _row_get_first(row, "id")
						partner_code = _row_get_first(row, "partner_code", "code", "bp_code")
						partner_name = _row_get_first(row, "partner_name", "name")
						address_type_value = _normalize_address_type(_row_get_first(row, "address_type", "type"))
						address_line1_value = _row_get_first(row, "address_line1", "address1", "line1")
						address_line2_value = _row_get_first(row, "address_line2", "address2", "line2")
						subdistrict_value = _row_get_first(row, "subdistrict")
						district_value = _row_get_first(row, "district")
						province_value = _row_get_first(row, "province")
						postal_code_value = _row_get_first(row, "postal_code", "zipcode", "zip_code")
						country_value = _row_get_first(row, "country") or "Thailand"

						if not partner_code and not partner_name and not address_line1_value and not province_value and not postal_code_value:
							skipped += 1
							continue

						if address_type_value not in {choice for choice, _label in ADDRESS_TYPE_CHOICES}:
							skipped += 1
							continue
						if not address_line1_value or not province_value or not postal_code_value:
							skipped += 1
							continue

						partner = None
						if partner_code:
							partner = BusinessPartner.objects.filter(code__iexact=partner_code).first()
						if partner is None and partner_name:
							partner = BusinessPartner.objects.filter(name__iexact=partner_name).first()
						if partner is None:
							skipped += 1
							continue

						existing = None
						if _is_uuid(row_id):
							existing = Address.objects.filter(pk=row_id).first()
						if existing is None:
							existing = Address.objects.filter(
								partner=partner,
								address_type=address_type_value,
								address_line1__iexact=address_line1_value,
							).first()

						if existing is None:
							Address.objects.create(
								partner=partner,
								address_type=address_type_value,
								address_line1=address_line1_value,
								address_line2=address_line2_value,
								subdistrict=subdistrict_value,
								district=district_value,
								province=province_value,
								postal_code=postal_code_value,
								country=country_value,
							)
							created += 1
						else:
							existing.partner = partner
							existing.address_type = address_type_value
							existing.address_line1 = address_line1_value
							existing.address_line2 = address_line2_value
							existing.subdistrict = subdistrict_value
							existing.district = district_value
							existing.province = province_value
							existing.postal_code = postal_code_value
							existing.country = country_value
							existing.save(
								update_fields=[
									"partner",
									"address_type",
									"address_line1",
									"address_line2",
									"subdistrict",
									"district",
									"province",
									"postal_code",
									"country",
									"updated_at",
								]
							)
							updated += 1
			except Exception as e:
				log_event(request, action="address:import", status="failure", message="นำเข้า Address Partner ไม่สำเร็จ", metadata={"error": str(e)})
				messages.error(request, f"เกิดข้อผิดพลาดระหว่างนำเข้า: {e}")
				return redirect(request.get_full_path())
			log_event(request, action="address:import", message="นำเข้า Address Partner", metadata={"created": created, "updated": updated, "skipped": skipped})
			messages.success(request, f"นำเข้าสำเร็จ: +{created}, อัปเดต {updated}, ข้าม {skipped}")
			return redirect(request.get_full_path())
		if action == "bulk_delete":
			bulk_ids = request.POST.getlist("bulk_id")
			ids = [x for x in [b.strip() for b in bulk_ids] if _is_uuid(x)]
			if not ids:
				messages.error(request, "กรุณาเลือกรายการที่ต้องการลบ")
				return redirect(request.get_full_path())
			deleted = blocked = 0
			try:
				with transaction.atomic():
					for pk in ids:
						obj = Address.objects.filter(pk=pk).first()
						if obj is None:
							continue
						try:
							obj.delete()
							deleted += 1
						except ProtectedError:
							blocked += 1
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return redirect(request.get_full_path())
			if blocked:
				messages.warning(request, f"ลบสำเร็จ {deleted} รายการ, ลบไม่ได้ {blocked} รายการ")
			else:
				messages.success(request, f"ลบสำเร็จ {deleted} รายการ")
			return redirect(request.get_full_path())
		if action == "create":
			if not _is_uuid(partner_id) or not address_line1 or not province or not postal_code:
				messages.error(request, "กรุณากรอกข้อมูลที่จำเป็นให้ครบ")
				return redirect(request.get_full_path())
			try:
				with transaction.atomic():
					partner = BusinessPartner.objects.get(pk=partner_id)
					obj = Address.objects.create(
						partner=partner,
						address_type=address_type,
						address_line1=address_line1,
						address_line2=address_line2,
						subdistrict=subdistrict,
						district=district,
						province=province,
						postal_code=postal_code,
						country=country,
					)
					messages.success(request, "เพิ่ม Address สำเร็จ")
					transaction.on_commit(lambda: log_event(request, action="address:create", message="เพิ่ม Address", metadata={"id": str(obj.pk)}))
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
			return redirect(request.get_full_path())
		if action == "update":
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return redirect(request.get_full_path())
			if not address_line1 or not province or not postal_code:
				messages.error(request, "กรุณากรอกข้อมูลที่จำเป็นให้ครบ")
				return redirect(request.get_full_path())
			try:
				with transaction.atomic():
					obj = Address.objects.get(pk=obj_id)
					fields_map = {
						"address_type": address_type,
						"address_line1": address_line1,
						"address_line2": address_line2,
						"subdistrict": subdistrict,
						"district": district,
						"province": province,
						"postal_code": postal_code,
						"country": country,
					}
					if _is_uuid(partner_id):
						fields_map["partner_id"] = partner_id
					updated = []
					for field, val in fields_map.items():
						if str(getattr(obj, field) or "") != val:
							setattr(obj, field, val)
							updated.append(field)
					if updated:
						updated.append("updated_at")
						obj.save(update_fields=updated)
						messages.success(request, "บันทึกการแก้ไขสำเร็จ")
					else:
						messages.info(request, "ไม่มีการเปลี่ยนแปลง")
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
			return redirect(request.get_full_path())
		if action == "delete":
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return redirect(request.get_full_path())
			try:
				with transaction.atomic():
					obj = Address.objects.get(pk=obj_id)
					obj.delete()
					messages.success(request, "ลบสำเร็จ")
			except ProtectedError:
				messages.error(request, "ลบไม่ได้ มีข้อมูลอ้างอิงอยู่")
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
			return redirect(request.get_full_path())
		messages.error(request, "ไม่รู้จัก action")
		return redirect(request.get_full_path())
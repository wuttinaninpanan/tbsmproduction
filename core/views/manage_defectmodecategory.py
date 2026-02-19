from __future__ import annotations

import uuid

from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.core.paginator import Paginator

from core.auth.decorators import staff_required
from core.services.auditlog import log_event
from core.models.defect_by_category import DefectByCategory
from core.models.defect_mode import DefectMode
from core.models.item_category import ItemCategory

try:
	import openpyxl  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None


def _normalized_key(key: str) -> str:
	return (key or "").strip().lower().replace(" ", "_")


def _excel_to_str(value) -> str:
	if value is None:
		return ""
	if isinstance(value, bool):
		return "1" if value else "0"
	return str(value).strip()


def _parse_bool(value) -> bool:
	v = (value or "").strip().lower()
	return v in {"1", "true", "t", "yes", "y", "on"}


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


def download_manage_defectmodecategory_import_template(request):
	headers = [
		"category_name",
		"defect_name",
		"is_inlist",
		"title",
		"description",
	]
	rows = [
		[
			"Seat track",
			"Crack",
			"1",
			"Seat track - Crack",
			"",
		],
	]

	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "defectmodecategory"
	ws.append(headers)
	for r in rows:
		ws.append(r)
	for col in range(1, len(headers) + 1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 26

	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = (
		'attachment; filename="manage_defectmodecategory_import_template.xlsx"'
	)
	wb.save(response)
	return response


@method_decorator(staff_required, name="dispatch")
class ManageDefectModeCategoryViews(TemplateView):
	template_name = "manage_defectmodecategory.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			return download_manage_defectmodecategory_import_template(request)
		return super().get(request, *args, **kwargs)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request

		q = (request.GET.get("q") or "").strip()
		per_page_raw = (request.GET.get("per_page") or "").strip()
		page = (request.GET.get("page") or "1").strip() or "1"

		qs = DefectByCategory.objects.select_related("category", "defect_mode").all()
		if q:
			qs = qs.filter(
				Q(title__icontains=q)
				| Q(description__icontains=q)
				| Q(category__name__icontains=q)
				| Q(defect_mode__name_th__icontains=q)
				| Q(defect_mode__name_en__icontains=q)
				| Q(defect_mode__name_jp__icontains=q)
			)

		allowed_per_page = {20, 50, 100, 200}
		try:
			per_page = int(per_page_raw or 20)
		except Exception:
			per_page = 20
		if per_page not in allowed_per_page:
			per_page = 20

		qs = qs.order_by("category__name", "defect_mode__name_en", "title")
		paginator = Paginator(qs, per_page)
		page_obj = paginator.get_page(page)

		rows = []
		for d in page_obj.object_list:
			rows.append(
				{
					"id": str(d.id),
					"category_id": str(d.category_id) if d.category_id else "",
					"category_name": getattr(d.category, "name", "") or "",
					"defect_mode_id": str(d.defect_mode_id) if d.defect_mode_id else "",
					"defect_name": getattr(d.defect_mode, "name", "") or "",
					"is_inlist": bool(d.is_inlist),
					"title": d.title or "",
					"description": d.description or "",
				}
			)

		def _page_items(num_pages: int, current: int) -> list[int | None]:
			if num_pages <= 0:
				return []
			if num_pages <= 10:
				return list(range(1, num_pages + 1))
			items: list[int | None] = [1]
			if current > 4:
				items.append(None)
			start = max(2, current - 1)
			end = min(num_pages - 1, current + 1)
			if current <= 4:
				start, end = 2, 4
			if current >= num_pages - 3:
				start, end = num_pages - 3, num_pages - 1
			for n in range(start, end + 1):
				if 1 < n < num_pages:
					items.append(n)
			if current < num_pages - 3:
				items.append(None)
			items.append(num_pages)
			compressed: list[int | None] = []
			for it in items:
				if compressed and compressed[-1] == it:
					continue
				if it is None and compressed and compressed[-1] is None:
					continue
				compressed.append(it)
			return compressed

		ctx["rows"] = rows
		ctx["q"] = q
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
		ctx["total_count"] = paginator.count

		ctx["categories"] = list(ItemCategory.objects.all().order_by("name"))
		ctx["defects"] = list(DefectMode.objects.all().order_by("name_en", "name_th"))

		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		obj_id = (request.POST.get("id") or "").strip()
		uploaded = request.FILES.get("excel_file")

		category_id = (request.POST.get("category_id") or "").strip()
		defect_mode_id = (request.POST.get("defect_mode_id") or "").strip()
		title = (request.POST.get("title") or "").strip()
		description = (request.POST.get("description") or "").strip()
		is_inlist_raw = (request.POST.get("is_inlist") or "").strip()

		def _is_uuid(value: str) -> bool:
			try:
				uuid.UUID(str(value))
			except Exception:
				return False
			return True

		if action == "bulk_delete_defect_by_category":
			bulk_ids = request.POST.getlist("bulk_id")
			ids = [raw.strip() for raw in bulk_ids if _is_uuid((raw or "").strip())]
			if not ids:
				messages.error(request, "ไม่พบรายการที่เลือก")
				return self.get(request, *args, **kwargs)
			with transaction.atomic():
				deleted = 0
				for rid in ids:
					obj = DefectByCategory.objects.filter(pk=rid).first()
					if not obj:
						continue
					obj.delete()
					deleted += 1
					log_event(
						request,
						action="defectmodecategory:bulk_delete",
						message="delete",
						metadata={"id": rid},
					)
			messages.success(request, f"ลบสำเร็จ {deleted} รายการ")
			return self.get(request, *args, **kwargs)

		if action == "import_master_data":
			if uploaded is None:
				messages.error(request, "กรุณาเลือกไฟล์ .xlsx")
				return self.get(request, *args, **kwargs)
			if openpyxl is None:
				messages.error(request, "ยังไม่ได้ติดตั้ง openpyxl (ไม่สามารถ import XLSX ได้)")
				return self.get(request, *args, **kwargs)
			if not str(getattr(uploaded, "name", "")).lower().endswith(".xlsx"):
				messages.error(request, "รองรับเฉพาะไฟล์ .xlsx")
				return self.get(request, *args, **kwargs)

			created = 0
			updated = 0
			skipped = 0
			not_found = 0

			with transaction.atomic():
				for row in _parse_xlsx(uploaded):
					cat_name = _excel_to_str(
						row.get("category_name")
						or row.get("category")
						or row.get("item_category")
					)
					defect_name = _excel_to_str(
						row.get("defect_name")
						or row.get("defect")
						or row.get("defect_mode")
						or row.get("name_en")
						or row.get("name_th")
					)
					raw_is_inlist = _excel_to_str(row.get("is_inlist") or row.get("in_list") or "")
					row_title = _excel_to_str(row.get("title") or "")
					row_desc = _excel_to_str(row.get("description") or "")

					if not cat_name or not defect_name:
						skipped += 1
						continue

					category = ItemCategory.objects.filter(name__iexact=cat_name).order_by("pk").first()
					defect = DefectMode.objects.filter(
						Q(name_en__iexact=defect_name)
						| Q(name_th__iexact=defect_name)
						| Q(name_jp__iexact=defect_name)
					).order_by("pk").first()
					if not category or not defect:
						not_found += 1
						continue

					is_inlist = _parse_bool(raw_is_inlist)
					if not row_title:
						row_title = f"{category.name} - {defect.name_en}".strip()

					existing = DefectByCategory.objects.filter(category=category, defect_mode=defect).first()
					if existing:
						existing.title = row_title
						existing.description = row_desc
						existing.is_inlist = is_inlist
						existing.save(update_fields=["title", "description", "is_inlist", "updated_at"])
						updated += 1
						log_event(
							request,
							action="defectmodecategory:import_update",
							message="update",
							metadata={
								"id": str(existing.id),
								"category": category.name,
								"defect": defect.name,
							},
						)
					else:
						obj = DefectByCategory.objects.create(
							category=category,
							defect_mode=defect,
							is_inlist=is_inlist,
							title=row_title,
							description=row_desc,
							user=request.user,
						)
						created += 1
						log_event(
							request,
							action="defectmodecategory:import_create",
							message="create",
							metadata={
								"id": str(obj.id),
								"category": category.name,
								"defect": defect.name,
							},
						)

			messages.success(
				request,
				f"Import สำเร็จ: เพิ่ม {created}, อัปเดต {updated}, ข้าม {skipped}, ไม่พบข้อมูลอ้างอิง {not_found}",
			)
			return self.get(request, *args, **kwargs)

		# CRUD actions
		if action in {"create", "update"}:
			if not _is_uuid(category_id) or not _is_uuid(defect_mode_id):
				messages.error(request, "กรุณาเลือก Category และ Defect mode")
				return self.get(request, *args, **kwargs)

			category = ItemCategory.objects.filter(pk=category_id).first()
			defect = DefectMode.objects.filter(pk=defect_mode_id).first()
			if not category or not defect:
				messages.error(request, "ไม่พบ Category หรือ Defect mode")
				return self.get(request, *args, **kwargs)

			is_inlist = _parse_bool(is_inlist_raw)
			if not title:
				title = f"{category.name} - {defect.name_en}".strip()
			if description is None:
				description = ""

			with transaction.atomic():
				if action == "create":
					existing = DefectByCategory.objects.filter(category=category, defect_mode=defect).first()
					if existing:
						existing.title = title
						existing.description = description or ""
						existing.is_inlist = is_inlist
						existing.save(update_fields=["title", "description", "is_inlist", "updated_at"])
						log_event(
							request,
							action="defectmodecategory:create_as_update",
							message="create->update",
							metadata={"id": str(existing.id)},
						)
						messages.success(request, "บันทึกสำเร็จ (อัปเดตรายการเดิม)")
						return self.get(request, *args, **kwargs)

					obj = DefectByCategory.objects.create(
						category=category,
						defect_mode=defect,
						is_inlist=is_inlist,
						title=title,
						description=description or "",
						user=request.user,
					)
					log_event(
						request,
						action="defectmodecategory:create",
						message="create",
						metadata={"id": str(obj.id)},
					)
					messages.success(request, "เพิ่มข้อมูลสำเร็จ")
					return self.get(request, *args, **kwargs)

				# update
				if not _is_uuid(obj_id):
					messages.error(request, "ไม่พบ ID")
					return self.get(request, *args, **kwargs)
				obj = DefectByCategory.objects.filter(pk=obj_id).first()
				if not obj:
					messages.error(request, "ไม่พบข้อมูล")
					return self.get(request, *args, **kwargs)
				obj.category = category
				obj.defect_mode = defect
				obj.title = title
				obj.description = description or ""
				obj.is_inlist = is_inlist
				obj.save(update_fields=["category", "defect_mode", "title", "description", "is_inlist", "updated_at"])
				log_event(
					request,
					action="defectmodecategory:update",
					message="update",
					metadata={"id": str(obj.id)},
				)
				messages.success(request, "แก้ไขข้อมูลสำเร็จ")
				return self.get(request, *args, **kwargs)

		if action == "delete":
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบ ID")
				return self.get(request, *args, **kwargs)
			obj = DefectByCategory.objects.filter(pk=obj_id).first()
			if not obj:
				messages.error(request, "ไม่พบข้อมูล")
				return self.get(request, *args, **kwargs)
			with transaction.atomic():
				obj.delete()
				log_event(
					request,
					action="defectmodecategory:delete",
					message="delete",
					metadata={"id": obj_id},
				)
			messages.success(request, "ลบข้อมูลสำเร็จ")
			return self.get(request, *args, **kwargs)

		messages.error(request, "ไม่รองรับ action นี้")
		return self.get(request, *args, **kwargs)

import csv
import re

from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.db.models import Q

from core.models import DefectMode, PartNumber, ProductionLine, ComponentPart
from core.auth.decorators import staff_required
from core.services.auditlog import log_event

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None


def download_production_import_template(request):
    """Download a template for importing production master data.

    Default format is CSV (Excel-compatible). If `?format=xlsx` is provided and
    `openpyxl` is installed, an .xlsx file will be returned.

    Notes:
    - Reference images are not supported via CSV/XLSX import.
    - `component_parts` can contain multiple values separated by `;`.
    """
    fmt = (request.GET.get("format") or "csv").strip().lower()

    headers = ["line_code", "part_number", "defect_name", "defect_code", "component_parts"]
    rows = [
        ["DAA1", "DAR-54", "รอยขีดข่วน", "DEF-001", "Component part; อื่นๆ"],
        ["DAA1", "DAR-54", "สีไม่สม่ำเสมอ", "", "Component part"],
        ["DAA2", "XYZ-01", "ชิ้นงานแตก", "DEF-100", "Component part"],
    ]

    if fmt == "xlsx":
        if openpyxl is None:
            return HttpResponse(
                "XLSX format is not available (openpyxl is not installed).",
                status=400,
                content_type="text/plain; charset=utf-8",
            )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "master_data"
        ws.append(headers)
        for r in rows:
            ws.append(r)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="production_import_template.xlsx"'
        wb.save(response)
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="production_import_template.csv"'
    # Add UTF-8 BOM so Excel on Windows opens Thai correctly.
    response.write("\ufeff")
    writer = csv.writer(response, lineterminator="\r\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def _normalized_key(key):
    return (key or "").strip().lower().replace(" ", "_")


def _parse_csv(uploaded_file):
    data = uploaded_file.read()
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("cp874", errors="replace")
    reader = csv.DictReader(text.splitlines())
    for row in reader:
        yield {(_normalized_key(k)): v for k, v in (row or {}).items()}


def _parse_xlsx(uploaded_file):
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return
    keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
    for values in rows:
        row = {}
        for idx, value in enumerate(values):
            k = keys[idx] if idx < len(keys) else ""
            if not k:
                continue
            row[k] = value
        yield row


def _split_component_parts(value):
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    parts = re.split(r"[;\|,\n\r]+", s)
    return [p.strip() for p in parts if p and p.strip()]


@method_decorator(staff_required, name='dispatch')
class SettingsViews(TemplateView):
    template_name = "add_production.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        lines = list(ProductionLine.objects.all().order_by("code"))
        ctx["production_lines"] = [l.code for l in lines]

        ctx["stats"] = {
            "total_lines": ProductionLine.objects.count(),
            "total_parts": PartNumber.objects.count(),
            "total_defects": DefectMode.objects.count(),
        }

        # Build master_data structure compatible with templates/add_production.html and templates/record.html
        production_lines_payload = []
        for line in lines:
            parts_payload = []
            parts = list(PartNumber.objects.filter(production_line=line).order_by("number"))
            for part in parts:
                defects_payload = []
                defects = list(
                    DefectMode.objects.filter(Q(part=part) | Q(part__isnull=True)).order_by("name")
                )
                for defect in defects:
                    component_parts = list(
                        ComponentPart.objects.filter(part_number=part)
                        .order_by("name")
                        .values_list("name", flat=True)
                    )

                    if not component_parts:
                        component_parts = ["Component part"]
                    defects_payload.append(
                        {
                            "id": str(defect.pk),
                            "name": defect.name,
                            "code": defect.code or "",
                            "image_url": defect.reference_image.url if defect.reference_image else "",
                            "component_parts": component_parts,
                        }
                    )
                parts_payload.append(
                    {
                        "id": part.number,
                        "group": "Common",
                        "defects": defects_payload,
                    }
                )
            production_lines_payload.append(
                {
                    "id": line.code,
                    "parts": parts_payload,
                }
            )

        ctx["master_data"] = {"productionLines": production_lines_payload}
        return ctx

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "").strip().lower()

        try:
            with transaction.atomic():
                if action == "import_master_data":
                    uploaded = request.FILES.get("excel_file")
                    if not uploaded:
                        messages.error(request, "กรุณาเลือกไฟล์ Excel/CSV ก่อนนำเข้า")
                        return self.get(request, *args, **kwargs)

                    filename = (uploaded.name or "").lower()
                    if filename.endswith(".csv"):
                        rows = _parse_csv(uploaded)
                    elif filename.endswith(".xlsx"):
                        try:
                            rows = _parse_xlsx(uploaded)
                        except RuntimeError:
                            messages.error(request, "ยังไม่รองรับไฟล์ .xlsx ในสภาพแวดล้อมนี้ (ต้องติดตั้ง openpyxl)")
                            return self.get(request, *args, **kwargs)
                    else:
                        messages.error(request, "รองรับเฉพาะไฟล์ .xlsx หรือ .csv")
                        return self.get(request, *args, **kwargs)

                    created_lines = 0
                    created_parts = 0
                    created_defects = 0
                    updated_defects = 0
                    created_component_parts = 0
                    skipped = 0

                    for row in rows:
                        line_code = (row.get("line_code") or row.get("line") or row.get("production_line") or "").strip().upper()
                        part_number = (row.get("part_number") or row.get("part") or row.get("pn") or "").strip()
                        defect_name = (row.get("defect_name") or row.get("defect") or row.get("defect_mode") or "").strip()
                        defect_code = (row.get("defect_code") or row.get("defectmode_code") or "")
                        defect_code = (str(defect_code).strip() if defect_code is not None else "")

                        if not line_code or not part_number or not defect_name:
                            skipped += 1
                            continue

                        line, line_created = ProductionLine.objects.get_or_create(code=line_code)
                        if line_created:
                            created_lines += 1
                        part, part_created = PartNumber.objects.get_or_create(production_line=line, number=part_number)
                        if part_created:
                            created_parts += 1
                        defect, defect_created = DefectMode.objects.get_or_create(part=part, name=defect_name)
                        if defect_created:
                            created_defects += 1

                        if defect_code and defect.code != defect_code:
                            defect.code = defect_code
                            defect.save(update_fields=["code", "updated_at"])
                            if not defect_created:
                                updated_defects += 1

                        # component parts: preferred field `component_parts` (semicolon separated)
                        component_parts = []
                        component_parts.extend(_split_component_parts(row.get("component_parts")))
                        component_parts.extend(_split_component_parts(row.get("component_part")))
                        component_parts.extend(_split_component_parts(row.get("component")))
                        component_parts.extend(_split_component_parts(row.get("components")))

                        # also accept multiple columns like component_part_1, component_part_2, etc.
                        for k, v in (row or {}).items():
                            kk = _normalized_key(k)
                            if kk.startswith("component") and kk not in {
                                "component",
                                "components",
                                "component_part",
                                "component_parts",
                            }:
                                component_parts.extend(_split_component_parts(v))

                        # de-dup while preserving order
                        seen = set()
                        unique_component_parts = []
                        for s in component_parts:
                            if s in seen:
                                continue
                            seen.add(s)
                            unique_component_parts.append(s)

                        if unique_component_parts:
                            for component_part_name in unique_component_parts:
                                _, created = ComponentPart.objects.get_or_create(
                                    part_number=part,
                                    name=component_part_name,
                                )
                                if created:
                                    created_component_parts += 1
                        else:
                            # Keep downstream pages usable even if no component part was provided
                            ComponentPart.objects.get_or_create(part_number=part, name="Component part")

                    messages.success(
                        request,
                        "นำเข้า Master Data สำเร็จ: "
                        f"Line +{created_lines}, Part +{created_parts}, Defect +{created_defects} (อัปเดต {updated_defects}), Component Part +{created_component_parts} | ข้าม {skipped}",
                    )
                    transaction.on_commit(
                        lambda: log_event(
                            request,
                            action="settings:import_master_data",
                            message="นำเข้า master data (settings) สำเร็จ",
                            metadata={
                                "filename": getattr(uploaded, "name", ""),
                                "created_lines": created_lines,
                                "created_parts": created_parts,
                                "created_defects": created_defects,
                                "updated_defects": updated_defects,
                                "created_component_parts": created_component_parts,
                                "skipped": skipped,
                            },
                        )
                    )
                    return self.get(request, *args, **kwargs)

                # New unified form (templates/add_production.html)
                if action == "save_master_data":
                    line_code = (request.POST.get("line_code") or "").strip().upper()
                    part_number = (request.POST.get("part_number") or "").strip()
                    defect_name = (request.POST.get("defect_name") or "").strip()

                    defect_code = (request.POST.get("defect_code") or "").strip()
                    defect_image = request.FILES.get("defect_image")

                    component_parts = request.POST.getlist("component_parts[]")
                    if not component_parts:
                        component_parts = request.POST.getlist("component_parts")
                    component_parts = [s.strip() for s in component_parts if (s or "").strip()]

                    if not line_code or not part_number or not defect_name:
                        messages.error(request, "กรุณากรอกข้อมูลที่จำเป็นให้ครบ (Production line / Part number / Defect mode)")
                        return self.get(request, *args, **kwargs)

                    line, line_created = ProductionLine.objects.get_or_create(code=line_code)
                    part, part_created = PartNumber.objects.get_or_create(production_line=line, number=part_number)
                    defect, defect_created = DefectMode.objects.get_or_create(part=part, name=defect_name)

                    old_defect_code = defect.code or ""
                    had_image_before = bool(defect.reference_image)

                    updated_fields = []
                    if defect_code and defect.code != defect_code:
                        defect.code = defect_code
                        updated_fields.append("code")
                    if defect_image:
                        defect.reference_image = defect_image
                        updated_fields.append("reference_image")
                    if updated_fields:
                        updated_fields.append("updated_at")
                        defect.save(update_fields=updated_fields)

                    created_component_parts = 0
                    if component_parts:
                        for component_part_name in component_parts:
                            _, created = ComponentPart.objects.get_or_create(
                                part_number=part,
                                name=component_part_name,
                            )
                            if created:
                                created_component_parts += 1
                    else:
                        # Keep downstream pages usable even if no component part was provided
                        ComponentPart.objects.get_or_create(part_number=part, name="Component part")

                    messages.success(
                        request,
                        f"บันทึก Master Data สำเร็จ: Line {line_code}, Part {part_number}, Defect {defect_name}"
                        + (f" (+{created_component_parts} component part)" if created_component_parts else ""),
                    )

                    new_defect_code = defect.code or ""
                    has_image_after = bool(defect.reference_image)
                    transaction.on_commit(
                        lambda: log_event(
                            request,
                            action="settings:save_master_data",
                            message="บันทึก master data (settings)",
                            metadata={
                                "line_code": line_code,
                                "part_number": part_number,
                                "defect_id": defect.pk,
                                "defect_name": defect_name,
                                "line_created": line_created,
                                "part_created": part_created,
                                "defect_created": defect_created,
                                "defect_code": {"from": old_defect_code, "to": new_defect_code}
                                if old_defect_code != new_defect_code
                                else None,
                                "defect_image_changed": (had_image_before != has_image_after) or bool(defect_image),
                                "component_parts_created": created_component_parts,
                                "component_parts_count": len(component_parts),
                            },
                        )
                    )

                    return self.get(request, *args, **kwargs)

                elif action == "add_line":
                    line_code = (request.POST.get("line_code") or "").strip().upper()
                    if not line_code:
                        messages.error(request, "กรุณากรอก Production line code")
                        return self.get(request, *args, **kwargs)
                    _, created = ProductionLine.objects.get_or_create(code=line_code)
                    messages.success(request, f"เพิ่ม Production line {line_code} สำเร็จ")
                    transaction.on_commit(
                        lambda: log_event(
                            request,
                            action="settings:add_line",
                            message="เพิ่ม Production line",
                            metadata={"line_code": line_code, "created": created},
                        )
                    )

                elif action == "add_part":
                    line_code = (request.POST.get("line_for_part") or "").strip().upper()
                    part_number = (request.POST.get("part_number") or "").strip()
                    if not line_code or not part_number:
                        messages.error(request, "กรุณาเลือก Production line และกรอก Part number")
                        return self.get(request, *args, **kwargs)
                    line, line_created = ProductionLine.objects.get_or_create(code=line_code)
                    _, part_created = PartNumber.objects.get_or_create(production_line=line, number=part_number)
                    messages.success(request, f"เพิ่ม Part number {part_number} ใน {line_code} สำเร็จ")
                    transaction.on_commit(
                        lambda: log_event(
                            request,
                            action="settings:add_part",
                            message="เพิ่ม Part number",
                            metadata={
                                "line_code": line_code,
                                "part_number": part_number,
                                "line_created": line_created,
                                "part_created": part_created,
                            },
                        )
                    )

                elif action == "add_defect":
                    line_code = (request.POST.get("line_for_defect") or "").strip().upper()
                    part_number = (request.POST.get("part_for_defect") or "").strip()
                    defect_name = (request.POST.get("defect_name") or "").strip()
                    if not line_code or not part_number or not defect_name:
                        messages.error(request, "กรุณาเลือก Production line/Part number และกรอก Defect mode")
                        return self.get(request, *args, **kwargs)
                    line, line_created = ProductionLine.objects.get_or_create(code=line_code)
                    part, part_created = PartNumber.objects.get_or_create(production_line=line, number=part_number)
                    defect, defect_created = DefectMode.objects.get_or_create(part=part, name=defect_name)

                    # Ensure at least one scrap option exists for this part
                    if not ComponentPart.objects.filter(part_number=part).exists():
                        ComponentPart.objects.get_or_create(part_number=part, name="Component part")
                    messages.success(request, f"เพิ่ม Defect mode {defect_name} สำเร็จ")
                    transaction.on_commit(
                        lambda: log_event(
                            request,
                            action="settings:add_defect",
                            message="เพิ่ม Defect mode",
                            metadata={
                                "line_code": line_code,
                                "part_number": part_number,
                                "defect_id": defect.pk,
                                "defect_name": defect_name,
                                "line_created": line_created,
                                "part_created": part_created,
                                "defect_created": defect_created,
                            },
                        )
                    )

                elif action == "add_component_part":
                    # Accept defect PK (preferred) or defect name
                    defect_key = (request.POST.get("defect_for_scrap") or "").strip()
                    component_part_name = (request.POST.get("component_part") or "").strip()
                    if not defect_key or not component_part_name:
                        messages.error(request, "กรุณาระบุ Defect mode และ Component Part")
                        return self.get(request, *args, **kwargs)
                    defect = None
                    if defect_key.isdigit():
                        defect = DefectMode.objects.filter(pk=int(defect_key)).first()
                    if defect is None:
                        defect = DefectMode.objects.filter(name__iexact=defect_key).order_by("pk").first()
                    if defect is None:
                        messages.error(request, "ไม่พบ Defect mode ที่ระบุ")
                        return self.get(request, *args, **kwargs)
                    obj, created = ComponentPart.objects.get_or_create(part_number=defect.part, name=component_part_name)
                    messages.success(request, f"เพิ่ม Component Part {component_part_name} สำเร็จ")
                    transaction.on_commit(
                        lambda: log_event(
                            request,
                            action="settings:add_component_part",
                            message="เพิ่ม Component Part",
                            metadata={
                                "component_part_id": obj.pk,
                                "component_part_name": component_part_name,
                                "defect_id": defect.pk,
                                "line_code": defect.part.production_line.code if defect.part_id else "",
                                "part_number": defect.part.number if defect.part_id else "",
                                "created": created,
                            },
                        )
                    )

                else:
                    messages.error(request, "คำสั่งไม่ถูกต้อง")
                    return self.get(request, *args, **kwargs)
        except Exception as e:
            messages.error(request, f"บันทึกไม่สำเร็จ: {e}")

        return self.get(request, *args, **kwargs)

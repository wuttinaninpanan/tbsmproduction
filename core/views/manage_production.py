from __future__ import annotations

import os
from datetime import datetime
from itertools import chain
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.core.files.base import ContentFile
from django.core.paginator import Paginator

from core.services.auditlog import log_event
from core.models import PartNumber, ProductionLine, ComponentPart
from core.auth.decorators import staff_required

try:
	import openpyxl  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None


def _normalized_key(key: str) -> str:
	return (key or "").strip().lower().replace(" ", "_")


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


def _cell_to_text(value) -> str:
	if value is None:
		return ""
	# Excel numeric cells often come through as float.
	if isinstance(value, float) and value.is_integer():
		return str(int(value)).strip()
	return str(value).strip()


LINE_KEYS = {
	"line_code",
	"line",
	"production_line",
	"linecode",
	"line_code*",
	"ไลน์",
	"ไลน์ผลิต",
	"สายการผลิต",
}

PART_KEYS = {
	"part_number",
	"part",
	"pn",
	"partno",
	"part_no",
	"partnumber",
	"พาร์ท",
	"หมายเลขชิ้นงาน",
}

COMPONENT_PART_KEYS = {
	"component_part_name",
	"component_part",
	"component_part_item",
}

COMPONENT_PART_IMAGE_URL_KEYS = {
	"component_part_image_url",
	"image_url",
	"component_part_image",
}


def _get_any(row: dict, keys: set[str]):
	for k in keys:
		if k in row:
			return row.get(k)
	return None


def download_manage_production_import_template(request):
	"""Download a template for importing production master data (Line/Part/Scrap).

	Columns:
	- line_code (required)
	- part_number (required)
	- component_part_name (optional)
	- component_part_image_url (optional, http/https URL; will be downloaded and saved as reference_image)
	"""
	headers = ["line_code", "part_number", "component_part_name", "component_part_image_url"]
	rows = [
		["DAA1", "DAR-54", "Burr", "https://example.com/component-part/burr.jpg"],
		["DAA1", "DAR-54", "Component part", ""],
		["DAA2", "XYZ-01", "", ""],
	]

	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "master_data"
	ws.append(headers)
	for r in rows:
		ws.append(r)
	for col in range(1, len(headers) + 1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 26

	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="manage_production_import_template.xlsx"'
	wb.save(response)
	return response


@method_decorator(staff_required, name='dispatch')
class ManageProductionViews(TemplateView):
	template_name = "manage_production.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			return download_manage_production_import_template(request)
		if action == "export_excel":
			return self._export_excel(request)
		return super().get(request, *args, **kwargs)

	def _export_excel(self, request):
		"""Export filtered master rows (Line/Part/Component Part) to an Excel (.xlsx) file."""
		if openpyxl is None:
			return HttpResponse(
				"XLSX export is not available (openpyxl is not installed).",
				status=400,
				content_type="text/plain; charset=utf-8",
			)

		q = (request.GET.get("q") or "").strip()
		selected_line = (request.GET.get("line") or "").strip().upper()
		selected_part = (request.GET.get("part") or "").strip()

		lines_qs = (
			ProductionLine.objects.annotate(
				part_count=Count("parts", distinct=True),
				record_count=Count("component_part_records", distinct=True),
			)
			.order_by("code")
		)
		parts_qs = (
			PartNumber.objects.select_related("production_line")
			.annotate(record_count=Count("component_part_records", distinct=True))
			.order_by("production_line__code", "number")
		)
		component_parts_qs = (
			ComponentPart.objects.select_related("part_number__production_line")
			.annotate(record_count=Count("component_part_records", distinct=True))
			.order_by(
				"part_number__production_line__code",
				"part_number__number",
				"name",
			)
		)

		if selected_line:
			parts_qs = parts_qs.filter(production_line__code=selected_line)
			component_parts_qs = component_parts_qs.filter(part_number__production_line__code=selected_line)

		if selected_part:
			parts_qs = parts_qs.filter(number=selected_part)
			component_parts_qs = component_parts_qs.filter(part_number__number=selected_part)

		if q:
			lines_qs = lines_qs.filter(code__icontains=q)
			parts_qs = parts_qs.filter(Q(number__icontains=q) | Q(production_line__code__icontains=q))
			component_parts_qs = component_parts_qs.filter(
				Q(name__icontains=q)
				| Q(part_number__number__icontains=q)
				| Q(part_number__production_line__code__icontains=q)
			)

		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "master_data"
		headers = ["line_code", "part_number", "component_part_name", "component_part_image_url", "record_count"]
		ws.append(headers)

		seen_line_ids: set[int] = set()
		seen_part_ids: set[int] = set()

		for s in component_parts_qs.iterator():
			part = s.part_number
			line = part.production_line
			seen_line_ids.add(line.id)
			seen_part_ids.add(part.id)
			image_url = s.reference_image.url if getattr(s, "reference_image", None) else ""
			ws.append(
				[
					line.code,
					part.number,
					s.name,
					image_url,
					getattr(s, "record_count", 0) or 0,
				]
			)

		for p in parts_qs.iterator():
			if p.id in seen_part_ids:
				continue
			seen_part_ids.add(p.id)
			line = p.production_line
			seen_line_ids.add(line.id)
			ws.append(
				[
					line.code,
					p.number,
					"",
					"",
					getattr(p, "record_count", 0) or 0,
				]
			)

		for l in lines_qs.iterator():
			if l.id in seen_line_ids:
				continue
			ws.append([l.code, "", "", "", getattr(l, "record_count", 0) or 0])

		for col in range(1, len(headers) + 1):
			ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24

		now = timezone.localtime(timezone.now())
		stamp = now.strftime("%Y%m%d_%H%M%S")
		response = HttpResponse(
			content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)
		response["Content-Disposition"] = f'attachment; filename="master_data_{stamp}.xlsx"'
		wb.save(response)
		return response

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request

		q = (request.GET.get("q") or "").strip()
		selected_line = (request.GET.get("line") or "").strip().upper()
		selected_part = (request.GET.get("part") or "").strip()

		lines_qs = (
			ProductionLine.objects.annotate(
				part_count=Count("parts", distinct=True),
				record_count=Count("component_part_records", distinct=True),
			)
			.order_by("code")
		)
		parts_qs = (
			PartNumber.objects.select_related("production_line")
			.annotate(
				record_count=Count("component_part_records", distinct=True),
			)
			.order_by("production_line__code", "number")
		)
		component_parts_qs = (
			ComponentPart.objects.select_related("part_number__production_line")
			.annotate(record_count=Count("component_part_records", distinct=True))
			.order_by(
				"part_number__production_line__code",
				"part_number__number",
				"name",
			)
		)

		if selected_line:
			parts_qs = parts_qs.filter(production_line__code=selected_line)
			component_parts_qs = component_parts_qs.filter(part_number__production_line__code=selected_line)

		if selected_part:
			parts_qs = parts_qs.filter(number=selected_part)
			component_parts_qs = component_parts_qs.filter(part_number__number=selected_part)

		if q:
			lines_qs = lines_qs.filter(code__icontains=q)
			parts_qs = parts_qs.filter(Q(number__icontains=q) | Q(production_line__code__icontains=q))
			component_parts_qs = component_parts_qs.filter(
				Q(name__icontains=q)
				| Q(part_number__number__icontains=q)
				| Q(part_number__production_line__code__icontains=q)
			)

		# Dropdown options
		ctx["production_lines"] = list(ProductionLine.objects.order_by("code").values_list("code", flat=True))
		ctx["part_numbers"] = list(
			PartNumber.objects.filter(production_line__code=selected_line).order_by("number").values_list("number", flat=True)
			if selected_line
			else PartNumber.objects.order_by("number").values_list("number", flat=True)
		)

		component_parts = list(component_parts_qs[:5000])
		parts = list(parts_qs[:2000])
		lines = list(lines_qs[:500])

		rows = []
		seen_line_ids = set()
		seen_part_ids = set()
		for s in component_parts:
			part = s.part_number
			line = part.production_line
			seen_line_ids.add(line.id)
			seen_part_ids.add(part.id)
			rows.append(
				{
					"line_id": line.id,
					"line_code": line.code,
					"part_id": part.id,
					"part_number": part.number,
					"component_part_id": s.id,
					"component_part_name": s.name,
					"component_part_image_url": s.reference_image.url if getattr(s, "reference_image", None) else "",
					"record_count": getattr(s, "record_count", 0),
				}
			)

		for p in parts:
			if p.id in seen_part_ids:
				continue
			seen_part_ids.add(p.id)
			line = p.production_line
			seen_line_ids.add(line.id)
			rows.append(
				{
					"line_id": line.id,
					"line_code": line.code,
					"part_id": p.id,
					"part_number": p.number,
					"component_part_id": "",
					"component_part_name": "",
					"component_part_image_url": "",
					"record_count": getattr(p, "record_count", 0),
				}
			)

		# Add placeholder rows (lines without parts)
		for l in lines:
			if l.id in seen_line_ids:
				continue
			rows.append(
				{
					"line_id": l.id,
					"line_code": l.code,
					"part_id": "",
					"part_number": "",
					"component_part_id": "",
					"component_part_name": "",
					"component_part_image_url": "",
					"record_count": getattr(l, "record_count", 0),
				}
			)

		rows.sort(
			key=lambda r: (
				r.get("line_code") or "",
				r.get("part_number") or "",
				r.get("component_part_name") or "",
			)
		)

		allowed_per_page = {20, 50, 100, 200}
		try:
			per_page = int((request.GET.get("per_page") or "").strip() or 20)
		except Exception:
			per_page = 20
		if per_page not in allowed_per_page:
			per_page = 20
		page = (request.GET.get("page") or "1").strip() or "1"
		paginator = Paginator(rows, per_page)
		page_obj = paginator.get_page(page)

		def _page_items(num_pages: int, current: int) -> list[int | None]:
			if num_pages <= 0:
				return []
			if num_pages <= 10:
				return list(range(1, num_pages + 1))
			items: list[int | None] = [1]
			if current > 4:
				items.append(None)
			start = max(2, current - 1)
			end = min(num_pages - 1, current + 1)
			if current <= 4:
				start, end = 2, 4
			if current >= num_pages - 3:
				start, end = num_pages - 3, num_pages - 1
			for n in range(start, end + 1):
				if 1 < n < num_pages:
					items.append(n)
			if current < num_pages - 3:
				items.append(None)
			items.append(num_pages)
			# De-dupe consecutive duplicates/ellipses
			compressed: list[int | None] = []
			for it in items:
				if compressed and compressed[-1] == it:
					continue
				if it is None and compressed and compressed[-1] is None:
					continue
				compressed.append(it)
			return compressed

		ctx["rows"] = page_obj.object_list
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)

		ctx["q"] = q
		ctx["selected_line"] = selected_line
		ctx["selected_part"] = selected_part

		ctx["counts"] = {
			"lines": lines_qs.count(),
			"parts": parts_qs.count(),
			"component_parts": component_parts_qs.count(),
		}

		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		obj_id = (request.POST.get("id") or "").strip()
		value = (request.POST.get("value") or "").strip()

		line_id = (request.POST.get("line_id") or "").strip()
		part_id = (request.POST.get("part_id") or "").strip()
		component_part_id = (request.POST.get("component_part_id") or "").strip()

		line_code = (request.POST.get("line_code") or "").strip().upper()
		part_number = (request.POST.get("part_number") or "").strip()
		component_part_name = (request.POST.get("component_part_name") or "").strip()
		component_part_names = [
			(v or "").strip() for v in request.POST.getlist("component_part_names") if (v or "").strip()
		]
		component_part_image = request.FILES.get("component_part_image")
		uploaded = request.FILES.get("excel_file")

		def _download_image_to_component_part(component_part: ComponentPart, image_url: str) -> None:
			url = (image_url or "").strip()
			if not url:
				return
			parsed = urlparse(url)
			if parsed.scheme not in {"http", "https"}:
				raise ValueError("รองรับเฉพาะลิงก์ http/https")
			if (parsed.hostname or "").lower() in {"localhost", "127.0.0.1"}:
				raise ValueError("ไม่อนุญาตให้ดึงรูปจาก localhost")

			filename = os.path.basename(parsed.path or "") or "image.jpg"
			# Ensure we have an extension to help storage/content-type.
			if "." not in filename:
				filename = f"{filename}.jpg"
			req = Request(url, headers={"User-Agent": "tbsmproduction/1.0"})
			with urlopen(req, timeout=15) as resp:
				data = resp.read(10 * 1024 * 1024 + 1)  # 10MB max
				if len(data) > 10 * 1024 * 1024:
					raise ValueError("ไฟล์รูปภาพใหญ่เกินไป (เกิน 10MB)")
				component_part.reference_image.save(filename, ContentFile(data), save=True)

		if action == "import_master_data":
			if not uploaded:
				messages.error(request, "กรุณาเลือกไฟล์ Excel/CSV ก่อนนำเข้า")
				return self.get(request, *args, **kwargs)

			filename = (uploaded.name or "").lower()
			try:
				if filename.endswith(".xlsx"):
					rows_iter = _parse_xlsx(uploaded)
				else:
					messages.error(request, "รองรับเฉพาะไฟล์ Excel (.xlsx) เท่านั้น")
					return self.get(request, *args, **kwargs)
			except RuntimeError:
				messages.error(request, "ยังไม่รองรับไฟล์ .xlsx ในสภาพแวดล้อมนี้ (ต้องติดตั้ง openpyxl)")
				return self.get(request, *args, **kwargs)

			created_lines = 0
			created_parts = 0
			created_component_parts = 0
			updated_component_part_images = 0
			skipped_empty = 0
			skipped_missing_required = 0
			errors = 0

			rows_iter = iter(rows_iter)
			first_row = next(rows_iter, None)
			if first_row is None:
				messages.error(request, "ไฟล์ไม่มีข้อมูลสำหรับนำเข้า")
				return self.get(request, *args, **kwargs)

			detected_cols = sorted([k for k in (first_row or {}).keys() if k])
			# If required columns are missing, fail fast with a helpful message.
			if not (set(detected_cols) & LINE_KEYS) or not (set(detected_cols) & PART_KEYS):
				messages.error(
					request,
					"ไม่พบคอลัมน์ที่จำเป็นสำหรับนำเข้า (ต้องมี line_code และ part_number)"
					+ (f" | พบคอลัมน์: {', '.join(detected_cols[:30])}" if detected_cols else ""),
				)
				return self.get(request, *args, **kwargs)

			rows_iter = chain([first_row], rows_iter)

			try:
				with transaction.atomic():
					for row in rows_iter:
						line_code_row = _cell_to_text(_get_any(row, LINE_KEYS)).upper()
						part_number_row = _cell_to_text(_get_any(row, PART_KEYS))
						component_part_name_row = _cell_to_text(_get_any(row, COMPONENT_PART_KEYS))
						image_url_row = _cell_to_text(_get_any(row, COMPONENT_PART_IMAGE_URL_KEYS))

						# Skip fully empty rows
						if not (line_code_row or part_number_row or component_part_name_row or image_url_row):
							skipped_empty += 1
							continue

						if not line_code_row or not part_number_row:
							skipped_missing_required += 1
							continue

						line, line_created = ProductionLine.objects.get_or_create(code=line_code_row)
						if line_created:
							created_lines += 1
						part, part_created = PartNumber.objects.get_or_create(production_line=line, number=part_number_row)
						if part_created:
							created_parts += 1

						if component_part_name_row:
							component_part, component_part_created = ComponentPart.objects.get_or_create(
								part_number=part,
								name=component_part_name_row,
							)
							if component_part_created:
								created_component_parts += 1
							if image_url_row:
								try:
									_download_image_to_component_part(component_part, image_url_row)
									updated_component_part_images += 1
								except Exception:
									errors += 1
			except IntegrityError as e:
				log_event(
					request,
					action="production:import_master_data",
					status="failure",
					message="นำเข้า master data ไม่สำเร็จ (IntegrityError)",
					metadata={
						"filename": getattr(uploaded, "name", ""),
						"error": str(e),
					},
				)
				messages.error(request, f"นำเข้าไม่สำเร็จ (ข้อมูลซ้ำหรือผิดเงื่อนไข): {e}")
				return self.get(request, *args, **kwargs)
			except Exception as e:
				log_event(
					request,
					action="production:import_master_data",
					status="failure",
					message="นำเข้า master data ไม่สำเร็จ",
					metadata={
						"filename": getattr(uploaded, "name", ""),
						"error": str(e),
					},
				)
				messages.error(request, f"เกิดข้อผิดพลาดระหว่างนำเข้า: {e}")
				return self.get(request, *args, **kwargs)

			skipped_total = skipped_empty + skipped_missing_required
			if (
				created_lines == 0
				and created_parts == 0
				and created_component_parts == 0
				and updated_component_part_images == 0
			):
				messages.warning(
					request,
					"นำเข้าเสร็จแล้ว แต่ไม่มีรายการถูกเพิ่ม/อัปเดต"
					+ f" | ข้าม {skipped_total} (ว่าง {skipped_empty}, ขาด line/part {skipped_missing_required})"
					+ (f" | หัวคอลัมน์ที่พบ: {', '.join(detected_cols[:30])}" if detected_cols else ""),
				)
			else:
				messages.success(
					request,
					f"นำเข้าสำเร็จ: Line +{created_lines}, Part +{created_parts}, Component Part +{created_component_parts}, อัปเดตรูป +{updated_component_part_images}, ข้าม {skipped_total} (ว่าง {skipped_empty}, ขาด line/part {skipped_missing_required}), error รูป {errors}",
				)
				transaction.on_commit(
					lambda: log_event(
						request,
						action="production:import_master_data",
						message="นำเข้า master data สำเร็จ",
						metadata={
							"filename": getattr(uploaded, "name", ""),
							"created_lines": created_lines,
							"created_parts": created_parts,
							"created_component_parts": created_component_parts,
							"updated_component_part_images": updated_component_part_images,
							"skipped_empty": skipped_empty,
							"skipped_missing_required": skipped_missing_required,
							"errors": errors,
							"detected_cols": detected_cols[:50],
						},
					)
				)
			return self.get(request, *args, **kwargs)

		def _int_or_none(v: str):
			return int(v) if v and v.isdigit() else None

		line_pk = _int_or_none(line_id)
		part_pk = _int_or_none(part_id)
		component_part_pk = _int_or_none(component_part_id)

		if action == "bulk_delete_master_rows":
			bulk_rows = request.POST.getlist("bulk_row")
			if not bulk_rows:
				messages.error(request, "กรุณาเลือกรายการที่ต้องการลบ")
				return self.get(request, *args, **kwargs)

			component_part_ids: set[int] = set()
			part_ids: set[int] = set()
			line_ids: set[int] = set()
			for raw in bulk_rows:
				try:
					line_s, part_s, component_part_s = (raw or "").split("|", 2)
				except ValueError:
					continue
				line_id_i = _int_or_none(line_s)
				part_id_i = _int_or_none(part_s)
				component_part_id_i = _int_or_none(component_part_s)
				# Delete deepest object per selected row.
				if component_part_id_i is not None:
					component_part_ids.add(component_part_id_i)
				elif part_id_i is not None:
					part_ids.add(part_id_i)
				elif line_id_i is not None:
					line_ids.add(line_id_i)

			if not (component_part_ids or part_ids or line_ids):
				messages.error(request, "ไม่พบรายการที่ต้องการลบ")
				return self.get(request, *args, **kwargs)

			try:
				with transaction.atomic():
					deleted_component_parts, _ = (
						ComponentPart.objects.filter(pk__in=component_part_ids).delete()
						if component_part_ids
						else (0, {})
					)
					deleted_parts, _ = (PartNumber.objects.filter(pk__in=part_ids).delete() if part_ids else (0, {}))
					deleted_lines, _ = (ProductionLine.objects.filter(pk__in=line_ids).delete() if line_ids else (0, {}))
				messages.success(
					request,
					f"ลบสำเร็จ: Component Part {deleted_component_parts}, Part {deleted_parts}, Line {deleted_lines}",
				)
				transaction.on_commit(
					lambda: log_event(
						request,
						action="production:bulk_delete_master_rows",
						message="ลบ master rows แบบ bulk",
						metadata={
							"selected_rows": len(bulk_rows),
							"component_part_ids": len(component_part_ids),
							"part_ids": len(part_ids),
							"line_ids": len(line_ids),
							"deleted_component_parts": deleted_component_parts,
							"deleted_parts": deleted_parts,
							"deleted_lines": deleted_lines,
						},
					)
				)
				return self.get(request, *args, **kwargs)
			except ProtectedError:
				log_event(
					request,
					action="production:bulk_delete_master_rows",
					status="failure",
					message="ลบ master rows แบบ bulk ไม่สำเร็จ (ProtectedError)",
					metadata={
						"selected_rows": len(bulk_rows),
						"component_part_ids": len(component_part_ids),
						"part_ids": len(part_ids),
						"line_ids": len(line_ids),
					},
				)
				messages.error(
					request,
					"ลบไม่ได้: มีข้อมูล Component Part Record อ้างอิงอยู่ (โปรดลบ/ย้ายรายการในหน้า manage_component_part หรือหน้า record ก่อน)",
				)
				return self.get(request, *args, **kwargs)
			except Exception as e:
				log_event(
					request,
					action="production:bulk_delete_master_rows",
					status="failure",
					message="ลบ master rows แบบ bulk ไม่สำเร็จ",
					metadata={
						"selected_rows": len(bulk_rows),
						"component_part_ids": len(component_part_ids),
						"part_ids": len(part_ids),
						"line_ids": len(line_ids),
						"error": str(e),
					},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

		if action == "create_master_data":
			if not line_code or not part_number:
				messages.error(request, "กรุณากรอกข้อมูลที่จำเป็นให้ครบ (Line / Part)")
				return self.get(request, *args, **kwargs)

			# Prefer list input from the add modal; fall back to legacy single field.
			if not component_part_names and component_part_name:
				component_part_names = [component_part_name]

			if component_part_image is not None and not component_part_names:
				messages.error(request, "กรุณากรอกชื่อ Component Part ก่อนอัปโหลดรูปภาพ")
				return self.get(request, *args, **kwargs)
			if component_part_image is not None and len(component_part_names) != 1:
				messages.error(request, "อัปโหลดรูปภาพได้เฉพาะเมื่อเพิ่ม Component Part 1 รายการเท่านั้น")
				return self.get(request, *args, **kwargs)

			try:
				with transaction.atomic():
					line, _ = ProductionLine.objects.get_or_create(code=line_code)
					part, _ = PartNumber.objects.get_or_create(production_line=line, number=part_number)

					created_component_parts = 0
					unique_names: list[str] = []
					seen: set[str] = set()
					for raw_name in component_part_names:
						name = (raw_name or "").strip()
						if not name:
							continue
						key = name.lower()
						if key in seen:
							continue
						seen.add(key)
						unique_names.append(name)

					for name in unique_names:
						component_part, created = ComponentPart.objects.get_or_create(
							part_number=part,
							name=name,
						)
						if created:
							created_component_parts += 1
						# Only allowed when there is exactly one component part
						if component_part_image is not None:
							component_part.reference_image = component_part_image
							component_part.save(update_fields=["reference_image", "updated_at"])

					component_part_label = ""
					if unique_names:
						if len(unique_names) == 1:
							component_part_label = f" / {unique_names[0]}"
						else:
							component_part_label = f" / Component Part {len(unique_names)} รายการ"

					messages.success(request, f"เพิ่ม Master Data สำเร็จ: {line_code} / {part_number}{component_part_label}")
					transaction.on_commit(
						lambda: log_event(
							request,
							action="production:create_master_data",
							message="เพิ่ม master data",
							metadata={
								"line_code": line_code,
								"part_number": part_number,
								"component_parts_count": len(unique_names),
								"component_parts": unique_names[:50],
								"image_uploaded": component_part_image is not None,
							},
						)
					)
					return self.get(request, *args, **kwargs)
			except IntegrityError as e:
				log_event(
					request,
					action="production:create_master_data",
					status="failure",
					message="เพิ่ม master data ไม่สำเร็จ (IntegrityError)",
					metadata={
						"line_code": line_code,
						"part_number": part_number,
						"component_parts_count": len(component_part_names),
						"error": str(e),
					},
				)
				messages.error(request, f"บันทึกไม่สำเร็จ (ข้อมูลซ้ำหรือผิดเงื่อนไข): {e}")
				return self.get(request, *args, **kwargs)
			except Exception as e:
				log_event(
					request,
					action="production:create_master_data",
					status="failure",
					message="เพิ่ม master data ไม่สำเร็จ",
					metadata={
						"line_code": line_code,
						"part_number": part_number,
						"component_parts_count": len(component_part_names),
						"error": str(e),
					},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

		# Unified row actions (preferred by UI)
		if action in {"update_master_row", "delete_master_row"}:
			try:
				with transaction.atomic():
					if action == "update_master_row":
						audit_meta = {
							"line_id": line_pk,
							"part_id": part_pk,
							"component_part_id": component_part_pk,
							"input_line_code": line_code,
							"input_part_number": part_number,
							"input_component_part_name": component_part_name,
							"image_uploaded": component_part_image is not None,
							"changes": {},
						}
						# For Component Part rows: changing Line/Part should NOT rename shared objects.
						# Instead, move the ComponentPart to the target Line/Part (create if missing).
						if component_part_pk is not None and (line_code or part_number):
							component_part = ComponentPart.objects.select_related(
								"part_number__production_line",
							).get(pk=component_part_pk)
							current_part = component_part.part_number
							current_line = current_part.production_line
							audit_meta["changes"].update(
								{
									"from_line_code": current_line.code,
									"from_part_number": current_part.number,
								}
							)

							target_line_code = (line_code or current_line.code).strip().upper()
							target_part_number = (part_number or current_part.number).strip()
							if not target_line_code or not target_part_number:
								messages.error(request, "กรุณากรอก Line และ Part ให้ครบ")
								return self.get(request, *args, **kwargs)

							target_line, _ = ProductionLine.objects.get_or_create(code=target_line_code)
							target_part, _ = PartNumber.objects.get_or_create(
								production_line=target_line,
								number=target_part_number,
							)

							if component_part.part_number_id != target_part.id:
								component_part.part_number = target_part
								component_part.save(update_fields=["part_number", "updated_at"])
								audit_meta["changes"].update(
									{
										"to_line_code": target_line.code,
										"to_part_number": target_part.number,
										"moved": True,
									}
								)

						# For Part/Line rows (no component part): rename the shared objects (expected to affect all rows).
						if component_part_pk is None:
							# Rename-only updates for each existing object in the row.
							if line_pk is not None and line_code:
								line = ProductionLine.objects.get(pk=line_pk)
								if line.code != line_code:
									audit_meta["changes"].update({"line_code": {"from": line.code, "to": line_code}})
									line.code = line_code
									line.save(update_fields=["code", "updated_at"])

							if part_pk is not None and part_number:
								part = PartNumber.objects.get(pk=part_pk)
								if part.number != part_number:
									audit_meta["changes"].update({"part_number": {"from": part.number, "to": part_number}})
									part.number = part_number
									part.save(update_fields=["number", "updated_at"])

						if component_part_pk is not None and (component_part_name or component_part_image is not None):
							component_part = ComponentPart.objects.get(pk=component_part_pk)
							updated_fields = []
							if component_part_name and component_part.name != component_part_name:
								audit_meta["changes"].update(
									{"component_part_name": {"from": component_part.name, "to": component_part_name}}
								)
								component_part.name = component_part_name
								updated_fields.append("name")
							if component_part_image is not None:
								audit_meta["changes"].update({"component_part_image": "updated"})
								component_part.reference_image = component_part_image
								updated_fields.append("reference_image")
							if updated_fields:
								updated_fields.append("updated_at")
								component_part.save(update_fields=updated_fields)

						messages.success(request, "บันทึกการแก้ไขสำเร็จ")
						transaction.on_commit(
							lambda: log_event(
								request,
								action="production:update_master_row",
								message="แก้ไข master row",
								metadata=audit_meta,
							)
						)
						return self.get(request, *args, **kwargs)

					# delete_master_row: delete deepest object in the row
					if component_part_pk is not None:
						obj = ComponentPart.objects.select_related("part_number__production_line").get(pk=component_part_pk)
						meta = {
							"component_part_id": obj.pk,
							"component_part_name": obj.name,
							"part_number": obj.part_number.number,
							"line_code": obj.part_number.production_line.code,
						}
						obj.delete()
						messages.success(request, "ลบ Component Part สำเร็จ")
						transaction.on_commit(
							lambda: log_event(
								request,
								action="production:delete_component_part",
								message="ลบ Component Part",
								metadata=meta,
							)
						)
						return self.get(request, *args, **kwargs)
					if part_pk is not None:
						obj = PartNumber.objects.select_related("production_line").get(pk=part_pk)
						meta = {
							"part_id": obj.pk,
							"part_number": obj.number,
							"line_code": obj.production_line.code,
						}
						obj.delete()
						messages.success(request, "ลบ Part สำเร็จ")
						transaction.on_commit(
							lambda: log_event(
								request,
								action="production:delete_part",
								message="ลบ Part",
								metadata=meta,
							)
						)
						return self.get(request, *args, **kwargs)
					if line_pk is not None:
						obj = ProductionLine.objects.get(pk=line_pk)
						meta = {"line_id": obj.pk, "line_code": obj.code}
						obj.delete()
						messages.success(request, "ลบ Line สำเร็จ")
						transaction.on_commit(
							lambda: log_event(
								request,
								action="production:delete_line",
								message="ลบ Line",
								metadata=meta,
							)
						)
						return self.get(request, *args, **kwargs)

					messages.error(request, "ไม่พบข้อมูลให้ลบ")
					return self.get(request, *args, **kwargs)

			except ProtectedError:
				log_event(
					request,
					action=f"production:{action}",
					status="failure",
					message="ดำเนินการไม่สำเร็จ (ProtectedError)",
					metadata={
						"line_id": line_pk,
						"part_id": part_pk,
						"component_part_id": component_part_pk,
					},
				)
				messages.error(
					request,
					"ลบไม่ได้: มีข้อมูล Component Part Record อ้างอิงอยู่ (โปรดลบ/ย้ายรายการในหน้า manage_component_part หรือหน้า record ก่อน)",
				)
			except IntegrityError as e:
				log_event(
					request,
					action=f"production:{action}",
					status="failure",
					message="บันทึกไม่สำเร็จ (IntegrityError)",
					metadata={
						"line_id": line_pk,
						"part_id": part_pk,
						"component_part_id": component_part_pk,
						"error": str(e),
					},
				)
				messages.error(request, f"บันทึกไม่สำเร็จ (ข้อมูลซ้ำหรือผิดเงื่อนไข): {e}")
			except Exception as e:
				log_event(
					request,
					action=f"production:{action}",
					status="failure",
					message="เกิดข้อผิดพลาด",
					metadata={
						"line_id": line_pk,
						"part_id": part_pk,
						"component_part_id": component_part_pk,
						"error": str(e),
					},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")

			return self.get(request, *args, **kwargs)

		# Legacy per-level actions (kept for backwards compatibility)
		if not obj_id.isdigit():
			messages.error(request, "ไม่พบรหัสรายการ")
			return self.get(request, *args, **kwargs)

		pk = int(obj_id)

		try:
			with transaction.atomic():
				if action == "update_line":
					new_code = value.upper()
					if not new_code:
						messages.error(request, "กรุณากรอก Production line code")
						return self.get(request, *args, **kwargs)
					line = ProductionLine.objects.get(pk=pk)
					old = line.code
					line.code = new_code
					line.save(update_fields=["code", "updated_at"])
					messages.success(request, f"อัปเดต Production line {old} → {new_code} สำเร็จ")
					transaction.on_commit(
						lambda: log_event(
							request,
							action="production:update_line",
							message="อัปเดต Production line",
							metadata={"line_id": pk, "from": old, "to": new_code},
						)
					)

				elif action == "delete_line":
					line = ProductionLine.objects.get(pk=pk)
					code = line.code
					line.delete()
					messages.success(request, f"ลบ Production line {code} สำเร็จ")
					transaction.on_commit(
						lambda: log_event(
							request,
							action="production:delete_line",
							message="ลบ Production line",
							metadata={"line_id": pk, "line_code": code},
						)
					)

				elif action == "update_part":
					new_number = value
					if not new_number:
						messages.error(request, "กรุณากรอก Part number")
						return self.get(request, *args, **kwargs)
					part = PartNumber.objects.select_related("production_line").get(pk=pk)
					old = part.number
					part.number = new_number
					part.save(update_fields=["number", "updated_at"])
					messages.success(request, f"อัปเดต Part {part.production_line.code}: {old} → {new_number} สำเร็จ")
					transaction.on_commit(
						lambda: log_event(
							request,
							action="production:update_part",
							message="อัปเดต Part number",
							metadata={
								"part_id": pk,
								"line_code": part.production_line.code,
								"from": old,
								"to": new_number,
							},
						)
					)

				elif action == "delete_part":
					part = PartNumber.objects.select_related("production_line").get(pk=pk)
					label = f"{part.production_line.code} {part.number}"
					part.delete()
					messages.success(request, f"ลบ Part {label} สำเร็จ")
					transaction.on_commit(
						lambda: log_event(
							request,
							action="production:delete_part",
							message="ลบ Part",
							metadata={"part_id": pk, "label": label},
						)
					)

				else:
					messages.error(request, "คำสั่งไม่ถูกต้อง")

		except ProductionLine.DoesNotExist:
			messages.error(request, "ไม่พบ Production line")
		except PartNumber.DoesNotExist:
			messages.error(request, "ไม่พบ Part number")
		except ProtectedError:
			messages.error(
				request,
				"ลบไม่ได้: มีข้อมูล Component Part Record อ้างอิงอยู่ (โปรดลบ/ย้ายรายการในหน้า manage_component_part หรือหน้า record ก่อน)",
			)
		except IntegrityError as e:
			messages.error(request, f"บันทึกไม่สำเร็จ (ข้อมูลซ้ำหรือผิดเงื่อนไข): {e}")
		except Exception as e:
			messages.error(request, f"เกิดข้อผิดพลาด: {e}")

		return self.get(request, *args, **kwargs)

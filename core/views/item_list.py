from __future__ import annotations

import uuid
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import ProtectedError
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView

from core.auth.decorators import staff_required
from core.models.bill_of_material_item_master import BillOfMaterialItemMater
from core.models.item_category import ItemCategory
from core.models.item_line import ItemLine
from core.models.item_list import Item_list, is_spreadsheet_error
from core.services.item_import import SimilarSdIndex, fill_only_update
from core.models.item_stage import ItemStage

# Tabs available on the Item List page. Each maps to a queryset filter.
ALLOWED_TABS = {"all", "nosd", "no_partno", "no_partname", "no_bom", "add"}


def _is_uuid(value: str) -> bool:
    try:
        uuid.UUID(str(value))
    except Exception:
        return False
    return True


def _safe_decimal(value, default: Decimal = Decimal("0")) -> Decimal:
    if value is None:
        return default
    if isinstance(value, Decimal):
        return value
    value = str(value).strip()
    if value == "":
        return default
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return default


def _page_items(num_pages: int, current: int) -> list[int | None]:
    if num_pages <= 0:
        return []
    if num_pages <= 10:
        return list(range(1, num_pages + 1))
    items: list[int | None] = [1]
    if current > 4:
        items.append(None)
    start = max(2, current - 1)
    end = min(num_pages - 1, current + 1)
    if current <= 4:
        start, end = 2, 4
    if current >= num_pages - 3:
        start, end = num_pages - 3, num_pages - 1
    for n in range(start, end + 1):
        if 1 < n < num_pages:
            items.append(n)
    if current < num_pages - 3:
        items.append(None)
    items.append(num_pages)
    compressed: list[int | None] = []
    for it in items:
        if compressed and compressed[-1] == it:
            continue
        if it is None and compressed and compressed[-1] is None:
            continue
        compressed.append(it)
    return compressed


def _tab_filter(qs, tab: str):
    """Narrow a queryset to the rows belonging to the given tab."""
    if tab == "nosd":
        return qs.filter(sd_code__icontains="NOSD")
    if tab == "no_partno":
        return qs.filter(Q(part_number__isnull=True) | Q(part_number__exact=""))
    if tab == "no_partname":
        return qs.filter(Q(part_name__isnull=True) | Q(part_name__exact=""))
    if tab == "no_bom":
        # Items that are not a child (component) of any BoM — i.e. never
        # referenced as a component in BillOfMaterialItemMater. Whether the
        # item is itself a BoM header does not matter.
        return qs.exclude(
            id__in=BillOfMaterialItemMater.objects.values("component_id")
        )
    # "all" and "add" show the unfiltered set.
    return qs


@method_decorator(staff_required, name="dispatch")
class ItemListView(TemplateView):
    """Listing of EVERY item in Item_list (FG, components, raw materials),
    regardless of Line assignment, organised into tabs. The "Add new" tab
    exposes a form for creating a new item."""

    template_name = "core/item_list.html"

    def get(self, request, *args, **kwargs):
        if (request.GET.get("action") or "").strip().lower() == "import_template":
            return self._import_template_response()
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        request = self.request

        tab = (request.GET.get("tab") or "all").strip().lower()
        if tab not in ALLOWED_TABS:
            tab = "all"

        q = (request.GET.get("q") or "").strip()
        per_page_raw = (request.GET.get("per_page") or "").strip()
        page = (request.GET.get("page") or "1").strip() or "1"

        allowed_per_page = {100, 200, 500, 1000}
        try:
            per_page = int(per_page_raw or 100)
        except Exception:
            per_page = 100
        if per_page not in allowed_per_page:
            per_page = 100

        base_qs = (
            Item_list.objects
            .select_related("category", "stage", "portion", "side", "inout", "way")
            .order_by("item_code", "sd_code")
        )

        # Per-tab totals (independent of the search box) for the tab badges.
        ctx["count_all"] = base_qs.count()
        ctx["count_nosd"] = _tab_filter(base_qs, "nosd").count()
        ctx["count_no_partno"] = _tab_filter(base_qs, "no_partno").count()
        ctx["count_no_partname"] = _tab_filter(base_qs, "no_partname").count()
        ctx["count_no_bom"] = _tab_filter(base_qs, "no_bom").count()

        ctx["tab"] = tab
        ctx["q"] = q
        ctx["per_page"] = per_page
        ctx["categories"] = list(ItemCategory.objects.order_by("name").values("id", "name"))
        ctx["stages"] = list(
            ItemStage.objects.order_by("display_name").values("id", "display_name", "code_prefix")
        )

        # The "Add new" tab is a form only — no listing to build.
        if tab == "add":
            return ctx

        qs = _tab_filter(base_qs, tab)
        if q:
            qs = qs.filter(
                Q(item_code__icontains=q)
                | Q(sd_code__icontains=q)
                | Q(part_number__icontains=q)
                | Q(part_name__icontains=q)
                | Q(sku__icontains=q)
                | Q(category__name__icontains=q)
                | Q(stage__display_name__icontains=q)
                | Q(portion__title__icontains=q)
                | Q(side__title__icontains=q)
                | Q(inout__title__icontains=q)
                | Q(way__title__icontains=q)
                | Q(comment__icontains=q)
            )

        paginator = Paginator(qs, per_page)
        page_obj = paginator.get_page(page)

        # Map item_id -> line names (for the items shown on this page only).
        page_item_ids = [obj.id for obj in page_obj.object_list]
        lines_by_item: dict = {}
        if page_item_ids:
            for il in (
                ItemLine.objects
                .filter(item_id__in=page_item_ids)
                .select_related("line")
            ):
                lines_by_item.setdefault(il.item_id, []).append(
                    getattr(il.line, "line_name", "") or ""
                )

        rows = []
        for item in page_obj.object_list:
            image_url = ""
            try:
                if getattr(item, "reference_image", None):
                    image_url = item.reference_image.url
            except Exception:
                image_url = ""
            line_names = lines_by_item.get(item.id, [])
            rows.append({
                "id": str(item.id),
                "item_code": item.item_code or "",
                "sd_code": item.sd_code or "",
                "part_number": item.part_number or "",
                "part_name": item.part_name or "",
                "reference_image_url": image_url,
                "sku": item.sku or "",
                "weight": str(item.weight),
                "cost": str(item.cost),
                "purchased_price": str(item.purchased_price),
                "category_name": getattr(item.category, "name", "") if item.category_id else "",
                "stage_name": getattr(item.stage, "display_name", "") if item.stage_id else "",
                "portion_name": getattr(item.portion, "title", "") if item.portion_id else "",
                "side_name": getattr(item.side, "title", "") if item.side_id else "",
                "inout_name": getattr(item.inout, "title", "") if item.inout_id else "",
                "way_name": getattr(item.way, "title", "") if item.way_id else "",
                "comment": item.comment or "",
                "line_names": ", ".join(sorted({l for l in line_names if l})),
            })

        ctx["rows"] = rows
        ctx["page_obj"] = page_obj
        ctx["paginator"] = paginator
        ctx["rows_total"] = paginator.count
        ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
        ctx["total_count"] = paginator.count
        return ctx

    def post(self, request, *args, **kwargs):
        action = (request.POST.get("action") or "").strip().lower()
        if action == "create_item":
            return self._handle_create_item(request)
        if action == "import_excel":
            return self._handle_import_excel(request)
        if action == "delete_item":
            return self._handle_delete_item(request)
        messages.error(request, "ไม่รองรับการทำงานนี้")
        return redirect(request.get_full_path())

    def _handle_delete_item(self, request):
        item_id = (request.POST.get("id") or "").strip()
        if not item_id or not _is_uuid(item_id):
            messages.error(request, "ไม่พบรายการที่ต้องการลบ")
            return redirect(request.get_full_path())

        item = Item_list.objects.filter(pk=item_id).first()
        if item is None:
            messages.error(request, "ไม่พบรายการที่ต้องการลบ (อาจถูกลบไปแล้ว)")
            return redirect(request.get_full_path())

        label = item.item_code or item.sd_code or item.part_name or str(item.id)
        try:
            with transaction.atomic():
                item.delete()
        except ProtectedError:
            # Item ถูกอ้างอิงแบบ PROTECT (เป็น component ใน BOM / มีข้อมูล Scrap /
            # Defect) — ลบไม่ได้จนกว่าจะเอาการอ้างอิงเหล่านั้นออกก่อน
            messages.error(
                request,
                f"ลบ \"{label}\" ไม่ได้ — Item นี้ถูกใช้งานอยู่ "
                "(เป็น component ใน BOM หรือมีข้อมูล Scrap/Defect อ้างอิงอยู่) "
                "กรุณาเอาการอ้างอิงออกก่อนจึงจะลบได้",
            )
            return redirect(request.get_full_path())
        except Exception as e:
            messages.error(request, f"ลบไม่สำเร็จ: {e}")
            return redirect(request.get_full_path())

        messages.success(request, f"ลบ Item \"{label}\" สำเร็จ")
        return redirect(request.get_full_path())

    def _resolve_fk(self, model, value):
        value = (value or "").strip()
        if not value or not _is_uuid(value):
            return None
        return model.objects.filter(pk=value).first()

    def _handle_create_item(self, request):
        sd_code = (request.POST.get("sd_code") or "").strip()
        part_number = (request.POST.get("part_number") or "").strip()
        part_name = (request.POST.get("part_name") or "").strip()
        sku = (request.POST.get("sku") or "").strip()
        weight = _safe_decimal(request.POST.get("weight") or "0")
        cost = _safe_decimal(request.POST.get("cost") or "0")
        purchased_price = _safe_decimal(request.POST.get("purchased_price") or "0")
        comment = (request.POST.get("comment") or "").strip()
        reference_image = request.FILES.get("reference_image")

        if not part_number or not part_name or not sku:
            messages.error(request, "กรุณากรอก Part Number / Part Name / SKU")
            return redirect(request.get_full_path())

        # กัน SD Code ซ้ำ: หากมีหลายแถวใช้ sd_code เดียวกัน เครื่องจะ map ไม่ได้
        if sd_code and Item_list.objects.filter(sd_code=sd_code).exists():
            messages.error(
                request,
                f"มี SD Code \"{sd_code}\" อยู่แล้ว — ห้ามเพิ่มซ้ำ "
                "(ถ้าต้องการแก้ไข ให้เข้าไปแก้ที่รายการเดิม)",
            )
            return redirect(request.get_full_path())

        category = self._resolve_fk(ItemCategory, request.POST.get("category_id"))
        stage = self._resolve_fk(ItemStage, request.POST.get("stage_id"))

        try:
            with transaction.atomic():
                Item_list(
                    sd_code=sd_code,
                    part_number=part_number,
                    part_name=part_name,
                    sku=sku,
                    weight=weight,
                    cost=cost,
                    purchased_price=purchased_price,
                    comment=comment,
                    reference_image=reference_image,
                    category=category,
                    stage=stage,
                    user=request.user,
                ).save()
        except IntegrityError as e:
            messages.error(request, f"เพิ่มไม่สำเร็จ (ข้อมูลซ้ำ): {e}")
            return redirect(request.get_full_path())
        except Exception as e:
            messages.error(request, f"เกิดข้อผิดพลาด: {e}")
            return redirect(request.get_full_path())

        messages.success(request, "เพิ่ม Item สำเร็จ")
        return redirect(request.get_full_path())

    # ---- Bulk import from Excel -------------------------------------------

    # Column header → model field. Several aliases map to the same field.
    IMPORT_COLUMNS = [
        "sd_code", "part_number", "part_name", "sku",
        "weight", "cost", "purchased_price", "comment", "category", "stage",
    ]
    IMPORT_ALIASES = {
        "sd_code": {"sd_code", "sd", "sdcode", "sd_no"},
        "part_number": {"part_number", "part_no", "partno", "partnumber"},
        "part_name": {"part_name", "partname"},
        "sku": {"sku"},
        "weight": {"weight", "weight_kg", "weight_(kg)"},
        "cost": {"cost"},
        "purchased_price": {"purchased_price", "purchased", "purchase_price"},
        "comment": {"comment", "remark", "note"},
        "category": {"category", "category_name"},
        "stage": {"stage", "stage_name"},
    }

    def _import_template_response(self):
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception:
            messages.error(self.request, "ไม่สามารถสร้างเทมเพลตได้ (ไม่มี openpyxl)")
            return redirect("/item-list/?tab=add")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Items"
        ws.append(self.IMPORT_COLUMNS)
        fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        ws.append(["SD-EXAMPLE", "PN-0001", "Example Part", "SKU-0001", 1.5, 0, 0, "", "", ""])
        for i, w in enumerate([16, 18, 28, 16, 10, 10, 16, 24, 18, 18], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="item_import_template.xlsx"'
        wb.save(resp)
        return resp

    def _handle_import_excel(self, request):
        try:
            import openpyxl  # type: ignore
        except Exception:
            messages.error(request, "ไม่สามารถอ่านไฟล์ Excel ได้ (ไม่มี openpyxl)")
            return redirect(request.get_full_path())

        f = request.FILES.get("excel_file")
        if not f:
            messages.error(request, "กรุณาเลือกไฟล์ Excel (.xlsx)")
            return redirect(request.get_full_path())
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        except Exception as e:
            messages.error(request, f"เปิดไฟล์ไม่สำเร็จ: {e}")
            return redirect(request.get_full_path())

        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        try:
            header = next(row_iter)
        except StopIteration:
            messages.error(request, "ไฟล์ว่าง — ไม่มีหัวคอลัมน์")
            return redirect(request.get_full_path())

        def _norm(s) -> str:
            return str(s or "").strip().lower().replace(" ", "_")

        rev = {alias: field for field, al in self.IMPORT_ALIASES.items() for alias in al}
        col: dict[str, int] = {}
        for idx, h in enumerate(header):
            field = rev.get(_norm(h))
            if field and field not in col:
                col[field] = idx

        missing = [c for c in ("part_number", "part_name", "sku") if c not in col]
        if missing:
            messages.error(
                request,
                "ไฟล์ขาดคอลัมน์ที่จำเป็น: " + ", ".join(missing)
                + " (หัวคอลัมน์ต้องมี part_number, part_name, sku)",
            )
            return redirect(request.get_full_path())

        cat_by_name = {c.name.strip().lower(): c for c in ItemCategory.objects.all() if c.name}
        stage_by_name: dict[str, ItemStage] = {}
        for s in ItemStage.objects.all():
            if s.display_name:
                stage_by_name[s.display_name.strip().lower()] = s
            if getattr(s, "code_prefix", ""):
                stage_by_name.setdefault(s.code_prefix.strip().lower(), s)

        def cell(row, field):
            i = col.get(field)
            if i is None or i >= len(row):
                return None
            return row[i]

        created = updated = unchanged = errors = 0
        bad_value_skipped = no_sd_skipped = similar_skipped = 0
        notes: list[str] = []

        # Seed similar-sd detection with everything already in the DB; new rows
        # are added as we go so two similar NEW rows in one file are caught too.
        sim_index = SimilarSdIndex()
        sim_index.seed(Item_list.objects.exclude(sd_code="").values_list("sd_code", flat=True))

        def _num_or_none(v):
            """Decimal for a present cell; None for a blank one (so fill-only skips)."""
            return None if (v is None or str(v).strip() == "") else _safe_decimal(v)

        for ri, row in enumerate(row_iter, start=2):
            if row is None or all((c is None or str(c).strip() == "") for c in row):
                continue  # blank line
            sd_code = str(cell(row, "sd_code") or "").strip()
            part_number = str(cell(row, "part_number") or "").strip()
            part_name = str(cell(row, "part_name") or "").strip()
            sku = str(cell(row, "sku") or "").strip()

            # Reject spreadsheet error literals (e.g. "#REF!") in any key field.
            if any(is_spreadsheet_error(v) for v in (cell(row, "sd_code"), part_number, part_name, sku)):
                bad_value_skipped += 1
                if len(notes) < 8:
                    notes.append(f"แถว {ri}: พบค่า error จากสเปรดชีต (#REF! ฯลฯ) — ข้าม")
                continue

            # Rule 2: blank sd_code -> never insert or update.
            if not sd_code:
                no_sd_skipped += 1
                if len(notes) < 8:
                    notes.append(f"แถว {ri}: SD Code ว่าง — ข้าม")
                continue

            category = cat_by_name.get(str(cell(row, "category") or "").strip().lower())
            stage = stage_by_name.get(str(cell(row, "stage") or "").strip().lower())
            incoming = {
                "part_number": part_number or None,
                "part_name": part_name or None,
                "sku": sku or None,
                "comment": str(cell(row, "comment") or "").strip() or None,
                "weight": _num_or_none(cell(row, "weight")),
                "cost": _num_or_none(cell(row, "cost")),
                "purchased_price": _num_or_none(cell(row, "purchased_price")),
                "category": category,
                "stage": stage,
            }

            existing = Item_list.objects.filter(sd_code__iexact=sd_code).first()
            if existing is not None:
                # Rules 3-6: fill empty fields only, never overwrite.
                if incoming["sku"] and Item_list.objects.filter(
                    sku__iexact=incoming["sku"]
                ).exclude(pk=existing.pk).exists():
                    incoming["sku"] = None  # can't fill with a sku owned by another item
                try:
                    with transaction.atomic():
                        changed = fill_only_update(existing, incoming)
                        if changed:
                            existing.save(update_fields=changed + ["updated_at"])
                            updated += 1
                        else:
                            unchanged += 1
                except Exception as e:
                    errors += 1
                    if len(notes) < 8:
                        notes.append(f"แถว {ri}: {e}")
                continue

            # Rule 1: brand-new sd_code -> flag near-duplicates before inserting.
            sim = sim_index.similar_to(sd_code)
            if sim is not None:
                similar_skipped += 1
                if len(notes) < 8:
                    notes.append(f"แถว {ri}: SD Code \"{sd_code}\" คล้ายกับ \"{sim}\" ที่มีอยู่ — ข้าม (ตรวจซ้ำ)")
                continue
            if not part_number or not part_name or not sku:
                errors += 1
                if len(notes) < 8:
                    notes.append(f"แถว {ri}: SD ใหม่แต่ขาด part_number/part_name/sku")
                continue
            try:
                # Each row in its own savepoint so one failure doesn't poison the
                # rest, and the auto item_code sees prior rows' committed codes.
                with transaction.atomic():
                    Item_list(
                        sd_code=sd_code,
                        part_number=part_number,
                        part_name=part_name,
                        sku=sku,
                        weight=_safe_decimal(cell(row, "weight")),
                        cost=_safe_decimal(cell(row, "cost")),
                        purchased_price=_safe_decimal(cell(row, "purchased_price")),
                        comment=str(cell(row, "comment") or "").strip(),
                        category=category,
                        stage=stage,
                        user=request.user,
                    ).save()
                created += 1
                sim_index.add(sd_code)
            except Exception as e:
                errors += 1
                if len(notes) < 8:
                    notes.append(f"แถว {ri}: {e}")

        summary = [f"เพิ่มใหม่ {created}", f"อัปเดต(เติมช่องว่าง) {updated}"]
        if unchanged:
            summary.append(f"ไม่เปลี่ยน {unchanged}")
        if similar_skipped:
            summary.append(f"ข้ามคล้ายซ้ำ {similar_skipped}")
        if no_sd_skipped:
            summary.append(f"ข้าม SD ว่าง {no_sd_skipped}")
        if bad_value_skipped:
            summary.append(f"ข้ามค่า error {bad_value_skipped}")
        if errors:
            summary.append(f"ผิดพลาด {errors}")
        msg = " · ".join(summary)
        if notes:
            msg += " — " + " ; ".join(notes)
        if created or updated:
            messages.success(request, msg)
        else:
            messages.error(request, "ไม่มีการเปลี่ยนแปลง — " + msg)
        return redirect(request.get_full_path())

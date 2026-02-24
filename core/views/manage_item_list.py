from __future__ import annotations

import re
import uuid
from decimal import Decimal, InvalidOperation

from django.utils.text import slugify

from django.http import HttpResponse

from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.core.paginator import Paginator

from core.auth.decorators import staff_required
from core.services.auditlog import log_event
from core.models.item_list import Item_list
from core.models.item_category import ItemCategory


try:
	import openpyxl  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None


def _normalized_key(key: str) -> str:
	# Make header matching tolerant to spaces/punctuation/casing.
	# Examples: "Part number" -> "part_number", "part-number" -> "part_number".
	key = (key or "").strip().lower()
	key = re.sub(r"[^0-9a-z]+", "_", key)
	key = re.sub(r"_+", "_", key).strip("_")
	return key


def _excel_to_str(value) -> str:
	if value is None:
		return ""
	if isinstance(value, bool):
		return "TRUE" if value else "FALSE"
	if isinstance(value, int):
		return str(value)
	if isinstance(value, float):
		# Excel often stores codes as numbers; avoid '1234.0'
		if value.is_integer():
			return str(int(value))
		return str(value)
	return str(value).strip()


def _row_get_first(row: dict, *keys: str) -> str:
	for k in keys:
		if not k:
			continue
		v = row.get(k)
		s = _excel_to_str(v).strip()
		if s != "":
			return s
	return ""


def _parse_xlsx(uploaded_file):
	if openpyxl is None:
		raise RuntimeError("openpyxl is not installed")
	wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
	ws = wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		headers = next(rows)
	except StopIteration:
		return
	keys = [_normalized_key(str(h) if h is not None else "") for h in headers]
	for values in rows:
		row = {}
		for idx, value in enumerate(values):
			k = keys[idx] if idx < len(keys) else ""
			if not k:
				continue
			row[k] = value
		yield row


def download_manage_item_list_import_template(request):
	"""Download a template for importing item list data."""
	headers = [
		"sd_code",
		"part_number",
		"part_name",
		"sku",
		"weight",
		"cost",
		"purchased_price",
		"category_name",
		"level",
		"comment",
	]
	rows = [
		[
			"SD-0001",
			"PN-0001",
			"Sample Part",
			"SKU-0001",
			1.23,
			10.50,
			9.75,
			"(ไม่ระบุ)",
			1,
			"",
		],
	]
	if openpyxl is None:
		return HttpResponse(
			"XLSX format is not available (openpyxl is not installed).",
			status=400,
			content_type="text/plain; charset=utf-8",
		)
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "item_list"
	ws.append(headers)
	for r in rows:
		ws.append(r)
	for col in range(1, len(headers) + 1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = 'attachment; filename="manage_item_list_import_template.xlsx"'
	wb.save(response)
	return response



def _safe_decimal(value, default: Decimal = Decimal("0")) -> Decimal:
	if value is None:
		return default
	if isinstance(value, Decimal):
		return value
	if isinstance(value, (int, float)):
		try:
			return Decimal(str(value))
		except (InvalidOperation, ValueError):
			return default
	value = str(value).strip()
	if value == "":
		return default
	try:
		return Decimal(value)
	except (InvalidOperation, ValueError):
		return default


def _is_uuid(value: str) -> bool:
	try:
		uuid.UUID(str(value))
	except Exception:
		return False
	return True


def _sanitize_sku_base(value: str) -> str:
	"""Build a safe SKU base (uppercase, dash-separated)."""
	value = (value or "").strip()
	if not value:
		return ""
	# slugify produces lowercase + '-' and strips unsafe chars.
	base = slugify(value).replace("-", "_")
	base = (base or "").strip("_")
	return base.upper()


def _generate_unique_sku(*, part_number: str, sd_code: str) -> str:
	base = _sanitize_sku_base(part_number) or _sanitize_sku_base(sd_code) or "ITEM"
	# Ensure the final SKU fits model max_length=100.
	for _ in range(20):
		suffix = uuid.uuid4().hex[:8].upper()
		max_base_len = 100 - 1 - len(suffix)
		trimmed = base[:max_base_len]
		candidate = f"{trimmed}-{suffix}" if trimmed else suffix
		if not Item_list.objects.filter(sku__iexact=candidate).exists():
			return candidate
	# Very unlikely fallback
	return uuid.uuid4().hex[:12].upper()


def _page_items(num_pages: int, current: int) -> list[int | None]:
	if num_pages <= 0:
		return []
	if num_pages <= 10:
		return list(range(1, num_pages + 1))
	items: list[int | None] = [1]
	if current > 4:
		items.append(None)
	start = max(2, current - 1)
	end = min(num_pages - 1, current + 1)
	if current <= 4:
		start, end = 2, 4
	if current >= num_pages - 3:
		start, end = num_pages - 3, num_pages - 1
	for n in range(start, end + 1):
		if 1 < n < num_pages:
			items.append(n)
	if current < num_pages - 3:
		items.append(None)
	items.append(num_pages)
	compressed: list[int | None] = []
	for it in items:
		if compressed and compressed[-1] == it:
			continue
		if it is None and compressed and compressed[-1] is None:
			continue
		compressed.append(it)
	return compressed


@method_decorator(staff_required, name="dispatch")
class ManageItemListViews(TemplateView):
	template_name = "manage_item_list.html"

	def get(self, request, *args, **kwargs):
		action = (request.GET.get("action") or "").strip().lower()
		if action == "download_template":
			return download_manage_item_list_import_template(request)
		return super().get(request, *args, **kwargs)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		request = self.request

		q = (request.GET.get("q") or "").strip()
		sort = (request.GET.get("sort") or "").strip()
		dir_raw = (request.GET.get("dir") or "").strip().lower()
		sort_dir = "desc" if dir_raw == "desc" else "asc"
		per_page_raw = (request.GET.get("per_page") or "").strip()
		page = (request.GET.get("page") or "1").strip() or "1"

		allowed_per_page = {20, 50, 100, 200}
		try:
			per_page = int(per_page_raw or 20)
		except Exception:
			per_page = 20
		if per_page not in allowed_per_page:
			per_page = 20

		qs = Item_list.objects.select_related("category").all()
		if q:
			qs = qs.filter(
				Q(sku__icontains=q)
				| Q(part_number__icontains=q)
				| Q(part_name__icontains=q)
				| Q(sd_code__icontains=q)
			)

		allowed_sorts = {
			"sd_code": "sd_code",
			"part_number": "part_number",
			"part_name": "part_name",
			"sku": "sku",
			"weight": "weight",
			"cost": "cost",
			"purchased_price": "purchased_price",
			"category": "category__name",
			"level": "level",
		}
		order_field = allowed_sorts.get(sort) or "sku"
		prefix = "-" if sort_dir == "desc" else ""
		qs = qs.order_by(f"{prefix}{order_field}")
		paginator = Paginator(qs, per_page)
		page_obj = paginator.get_page(page)

		rows = []
		for item in page_obj.object_list:
			image_url = ""
			try:
				if getattr(item, "reference_image", None):
					image_url = item.reference_image.url
			except Exception:
				image_url = ""
			rows.append(
				{
					"id": str(item.id),
					"sd_code": item.sd_code,
					"part_number": item.part_number,
					"part_name": item.part_name,
					"reference_image_url": image_url,
					"sku": item.sku,
					"weight": str(item.weight),
					"category_id": str(item.category_id) if item.category_id else "",
					"category_name": getattr(item.category, "name", "") if item.category_id else "",
					"purchased_price": str(item.purchased_price),
					"cost": str(item.cost),
					"level": "" if item.level is None else str(item.level),
					"comment": item.comment or "",
				}
			)

		ctx["rows"] = rows
		ctx["q"] = q
		ctx["sort"] = sort if sort in allowed_sorts else "sku"
		ctx["dir"] = sort_dir
		ctx["page_obj"] = page_obj
		ctx["paginator"] = paginator
		ctx["per_page"] = per_page
		ctx["rows_total"] = paginator.count
		ctx["page_items"] = _page_items(paginator.num_pages, page_obj.number)
		ctx["total_count"] = paginator.count

		ctx["categories"] = list(ItemCategory.objects.order_by("name").values("id", "name"))

		return ctx

	def post(self, request, *args, **kwargs):
		action = (request.POST.get("action") or "").strip().lower()
		obj_id = (request.POST.get("id") or "").strip()

		if action == "import_master_data":
			if openpyxl is None:
				messages.error(request, "ไม่สามารถนำเข้า XLSX ได้: ยังไม่ได้ติดตั้ง openpyxl")
				return self.get(request, *args, **kwargs)
			upload = request.FILES.get("excel_file")
			if upload is None:
				messages.error(request, "กรุณาเลือกไฟล์ Excel (.xlsx)")
				return self.get(request, *args, **kwargs)
			name = (getattr(upload, "name", "") or "").lower()
			if not name.endswith(".xlsx"):
				messages.error(request, "รองรับเฉพาะไฟล์ .xlsx")
				return self.get(request, *args, **kwargs)

			created = 0
			duplicated = 0
			skipped = 0
			category_not_found = 0
			try:
				with transaction.atomic():
					for row in _parse_xlsx(upload):
						sd_code = _row_get_first(row, "sd_code", "sdcode", "sd")
						part_number = _row_get_first(row, "part_number", "part_no", "partnumber", "pn")
						part_name = _row_get_first(row, "part_name", "partname", "name")
						sku = _row_get_first(row, "sku", "item_sku")

						# Only require core fields. SKU can be blank (auto-generate).
						if not sd_code or not part_number or not part_name:
							skipped += 1
							continue
						if not sku:
							sku = _generate_unique_sku(part_number=part_number, sd_code=sd_code)

						if Item_list.objects.filter(sku__iexact=sku).exists():
							duplicated += 1
							continue

						category = None
						category_name = _row_get_first(row, "category_name", "category", "item_category")
						if category_name and category_name != "(ไม่ระบุ)":
							category = ItemCategory.objects.filter(name__iexact=category_name).first()
							if category is None:
								category_not_found += 1

						weight = _safe_decimal(row.get("weight"), default=Decimal("0"))
						cost = _safe_decimal(row.get("cost"), default=Decimal("0"))
						purchased_price = _safe_decimal(row.get("purchased_price"), default=Decimal("0"))
						level = row.get("level")
						level_int = None
						if level is not None and str(level).strip() != "":
							try:
								level_int = int(str(level).strip())
							except Exception:
								level_int = None

						comment = _excel_to_str(row.get("comment")).strip()

						Item_list.objects.create(
							sd_code=sd_code,
							part_number=part_number,
							part_name=part_name,
							sku=sku,
							weight=weight,
							category=category,
							purchased_price=purchased_price,
							cost=cost,
							level=level_int,
							comment=comment,
							user=request.user,
						)
						created += 1
			except Exception as e:
				log_event(
					request,
					action="item_list:import_master_data",
					status="failure",
					message="นำเข้า Item list ไม่สำเร็จ",
					metadata={"error": str(e)},
				)
				messages.error(request, f"นำเข้าไม่สำเร็จ: {e}")
				return self.get(request, *args, **kwargs)

			transaction.on_commit(
				lambda: log_event(
					request,
					action="item_list:import_master_data",
					message="นำเข้า Item list",
					metadata={
						"created": created,
						"duplicated": duplicated,
						"skipped": skipped,
						"category_not_found": category_not_found,
					},
				)
			)

			messages.success(
				request,
				f"นำเข้าสำเร็จ: เพิ่ม {created} รายการ, ซ้ำ {duplicated}, ข้าม {skipped}"
				+ (f", ไม่พบหมวดหมู่ {category_not_found}" if category_not_found else ""),
			)
			return self.get(request, *args, **kwargs)

		sd_code = (request.POST.get("sd_code") or "").strip()
		part_number = (request.POST.get("part_number") or "").strip()
		part_name = (request.POST.get("part_name") or "").strip()
		sku = (request.POST.get("sku") or "").strip()
		weight = _safe_decimal(request.POST.get("weight") or "0")
		category_id = (request.POST.get("category_id") or "").strip()
		purchased_price = _safe_decimal(request.POST.get("purchased_price") or "0")
		cost = _safe_decimal(request.POST.get("cost") or "0")
		level_raw = (request.POST.get("level") or "").strip()
		comment = (request.POST.get("comment") or "").strip()

		level = None
		if level_raw != "":
			try:
				level = int(level_raw)
			except Exception:
				messages.error(request, "Level ต้องเป็นตัวเลข")
				return self.get(request, *args, **kwargs)

		category = None
		if category_id:
			if not _is_uuid(category_id):
				messages.error(request, "หมวดหมู่ไม่ถูกต้อง")
				return self.get(request, *args, **kwargs)
			category = ItemCategory.objects.filter(pk=category_id).first()
			if category is None:
				messages.error(request, "ไม่พบหมวดหมู่")
				return self.get(request, *args, **kwargs)

		if action == "bulk_delete_items":
			bulk_ids = request.POST.getlist("bulk_id")
			ids: list[str] = []
			for raw in bulk_ids:
				raw = (raw or "").strip()
				if _is_uuid(raw):
					ids.append(raw)
			if not ids:
				messages.error(request, "กรุณาเลือกรายการที่ต้องการลบ")
				return self.get(request, *args, **kwargs)

			deleted = 0
			blocked = 0
			not_found = 0
			try:
				with transaction.atomic():
					for pk in ids:
						obj = Item_list.objects.filter(pk=pk).first()
						if obj is None:
							not_found += 1
							continue
						try:
							obj.delete()
							deleted += 1
						except ProtectedError:
							blocked += 1
			except Exception as e:
				log_event(
					request,
					action="item_list:bulk_delete_items",
					status="failure",
					message="ลบ Item list แบบ bulk ไม่สำเร็จ",
					metadata={"selected": len(ids), "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

			transaction.on_commit(
				lambda: log_event(
					request,
					action="item_list:bulk_delete_items",
					message="ลบ Item list แบบ bulk",
					metadata={
						"selected": len(ids),
						"deleted": deleted,
						"blocked": blocked,
						"not_found": not_found,
						"ids": ids[:50],
					},
				)
			)

			if blocked:
				messages.warning(request, f"ลบสำเร็จ {deleted} รายการ, ลบไม่ได้ {blocked} รายการ (มีข้อมูลอ้างอิง), ไม่พบ {not_found}")
			else:
				messages.success(request, f"ลบสำเร็จ {deleted} รายการ" + (f" (ไม่พบ {not_found})" if not_found else ""))
			return self.get(request, *args, **kwargs)

		if action == "create_item":
			if not sd_code or not part_number or not part_name:
				messages.error(request, "กรุณากรอก SD Code / Part number / Part name")
				return self.get(request, *args, **kwargs)
			if not sku:
				sku = _generate_unique_sku(part_number=part_number, sd_code=sd_code)
			reference_image = request.FILES.get("reference_image")
			try:
				with transaction.atomic():
					if Item_list.objects.filter(sku__iexact=sku).exists():
						raise IntegrityError("SKU ซ้ำ")
					obj = Item_list.objects.create(
						sd_code=sd_code,
						part_number=part_number,
						part_name=part_name,
						sku=sku,
						weight=weight,
						category=category,
						purchased_price=purchased_price,
						cost=cost,
						level=level,
						comment=comment,
						reference_image=reference_image,
						user=request.user,
					)
					messages.success(request, "เพิ่ม Item list สำเร็จ")
					transaction.on_commit(
						lambda: log_event(
							request,
							action="item_list:create_item",
							message="เพิ่ม Item list",
							metadata={"item_id": str(obj.pk), "sku": sku, "part_number": part_number},
						)
					)
					return self.get(request, *args, **kwargs)
			except IntegrityError as e:
				log_event(
					request,
					action="item_list:create_item",
					status="failure",
					message="เพิ่ม Item list ไม่สำเร็จ (IntegrityError)",
					metadata={"sku": sku, "error": str(e)},
				)
				messages.error(request, f"บันทึกไม่สำเร็จ (ข้อมูลซ้ำหรือผิดเงื่อนไข): {e}")
				return self.get(request, *args, **kwargs)
			except Exception as e:
				log_event(
					request,
					action="item_list:create_item",
					status="failure",
					message="เพิ่ม Item list ไม่สำเร็จ",
					metadata={"sku": sku, "error": str(e)},
				)
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

		if action == "update_item":
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return self.get(request, *args, **kwargs)
			if not sd_code or not part_number or not part_name:
				messages.error(request, "กรุณากรอก SD Code / Part number / Part name")
				return self.get(request, *args, **kwargs)
			reference_image = request.FILES.get("reference_image")
			try:
				with transaction.atomic():
					item = Item_list.objects.get(pk=obj_id)
					# SKU is optional on edit: if blank, keep existing value.
					if not sku:
						sku = item.sku
					if item.sku.lower() != sku.lower() and Item_list.objects.filter(sku__iexact=sku).exclude(pk=item.pk).exists():
						raise IntegrityError("SKU ซ้ำ")

					updated_fields = []
					old = {
						"sd_code": item.sd_code,
						"part_number": item.part_number,
						"part_name": item.part_name,
						"sku": item.sku,
						"weight": str(item.weight),
						"reference_image": bool(getattr(item, "reference_image", None)),
						"category_id": str(item.category_id) if item.category_id else "",
						"purchased_price": str(item.purchased_price),
						"cost": str(item.cost),
						"level": item.level,
						"comment": item.comment or "",
					}

					if item.sd_code != sd_code:
						item.sd_code = sd_code
						updated_fields.append("sd_code")
					if item.part_number != part_number:
						item.part_number = part_number
						updated_fields.append("part_number")
					if item.part_name != part_name:
						item.part_name = part_name
						updated_fields.append("part_name")
					if item.sku != sku:
						item.sku = sku
						updated_fields.append("sku")
					if item.weight != weight:
						item.weight = weight
						updated_fields.append("weight")
					if item.category_id != (category.id if category else None):
						item.category = category
						updated_fields.append("category")
					if item.purchased_price != purchased_price:
						item.purchased_price = purchased_price
						updated_fields.append("purchased_price")
					if item.cost != cost:
						item.cost = cost
						updated_fields.append("cost")
					if item.level != level:
						item.level = level
						updated_fields.append("level")
					if (item.comment or "") != comment:
						item.comment = comment
						updated_fields.append("comment")

					if reference_image is not None:
						item.reference_image = reference_image
						updated_fields.append("reference_image")

					if updated_fields:
						updated_fields.append("updated_at")
						item.save(update_fields=updated_fields)
						messages.success(request, "บันทึกการแก้ไขสำเร็จ")
						transaction.on_commit(
							lambda: log_event(
								request,
								action="item_list:update_item",
								message="แก้ไข Item list",
								metadata={
									"item_id": str(item.pk),
									"changed_fields": [f for f in updated_fields if f != "updated_at"],
									"from": old,
									"to": {
										"sd_code": item.sd_code,
										"part_number": item.part_number,
										"part_name": item.part_name,
										"sku": item.sku,
										"weight": str(item.weight),
										"reference_image": bool(getattr(item, "reference_image", None)),
										"category_id": str(item.category_id) if item.category_id else "",
										"purchased_price": str(item.purchased_price),
										"cost": str(item.cost),
										"level": item.level,
										"comment": item.comment or "",
									},
								},
							)
						)
					else:
						messages.info(request, "ไม่มีการเปลี่ยนแปลง")
					return self.get(request, *args, **kwargs)
			except ProtectedError:
				messages.error(request, "บันทึกไม่สำเร็จ: มีข้อมูลอ้างอิงอยู่")
				return self.get(request, *args, **kwargs)
			except IntegrityError as e:
				messages.error(request, f"บันทึกไม่สำเร็จ (ข้อมูลซ้ำหรือผิดเงื่อนไข): {e}")
				return self.get(request, *args, **kwargs)
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

		if action == "delete_item":
			if not _is_uuid(obj_id):
				messages.error(request, "ไม่พบรหัสรายการ")
				return self.get(request, *args, **kwargs)
			try:
				with transaction.atomic():
					item = Item_list.objects.get(pk=obj_id)
					meta = {"item_id": str(item.pk), "sku": item.sku, "part_number": item.part_number}
					item.delete()
					messages.success(request, "ลบ Item list สำเร็จ")
					transaction.on_commit(
						lambda: log_event(
							request,
							action="item_list:delete_item",
							message="ลบ Item list",
							metadata=meta,
						)
					)
					return self.get(request, *args, **kwargs)
			except ProtectedError:
				messages.error(request, "ลบไม่ได้: มีข้อมูลอ้างอิงอยู่")
				return self.get(request, *args, **kwargs)
			except Exception as e:
				messages.error(request, f"เกิดข้อผิดพลาด: {e}")
				return self.get(request, *args, **kwargs)

		messages.error(request, "ไม่รองรับการทำงานนี้")
		return self.get(request, *args, **kwargs)

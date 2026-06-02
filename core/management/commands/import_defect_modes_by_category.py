from __future__ import annotations

from pathlib import Path

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models.defect_by_category import DefectByCategory
from core.models.defect_mode import DefectMode
from core.models.item_category import ItemCategory


DEFAULT_INPUT = Path("core/management/ปัญหางาน_NG_แยกประเภท.xlsx")
DEFAULT_SHEET = "DefectModeByCategory"


def clean_cell(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def key(value: str) -> str:
    return clean_cell(value).casefold()


class Command(BaseCommand):
    help = (
        "Import DefectMode and DefectByCategory rows from the "
        "DefectModeByCategory sheet in the NG category workbook."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--input",
            default=str(DEFAULT_INPUT),
            help=f"Workbook path (default: {DEFAULT_INPUT}).",
        )
        parser.add_argument(
            "--sheet",
            default=DEFAULT_SHEET,
            help=f"Worksheet name (default: {DEFAULT_SHEET}).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Read and validate the workbook without writing to the database.",
        )
        parser.add_argument(
            "--no-create-categories",
            action="store_true",
            help="Fail when a category from the workbook does not already exist.",
        )

    def handle(self, *args, **options):
        input_path = Path(options["input"])
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]
        create_categories = not options["no_create_categories"]

        if not input_path.exists():
            raise CommandError(f"Workbook not found: {input_path}")

        rows = self._read_rows(input_path, sheet_name)
        if not rows:
            self.stdout.write(self.style.WARNING("No importable rows found."))
            return

        unique_defects = {key(row["name_en"]): row for row in rows}
        unique_categories = sorted({row["category"] for row in rows}, key=str.casefold)

        self.stdout.write(
            f"Read {len(rows)} row(s): {len(unique_defects)} unique defect mode(s), "
            f"{len(unique_categories)} category/categories."
        )

        if dry_run:
            self.stdout.write("Dry run only. No database changes were made.")
            return

        creator = self._get_creator()
        if creator is None:
            raise CommandError("No user found. Create at least one user before importing.")

        with transaction.atomic():
            stats = self._import_rows(rows, creator, create_categories)

        self.stdout.write(
            self.style.SUCCESS(
                "Import complete: "
                f"{stats['defects_created']} defect mode(s) created, "
                f"{stats['defects_updated']} updated, "
                f"{stats['categories_created']} category/categories created, "
                f"{stats['links_created']} link(s) created, "
                f"{stats['links_updated']} link(s) updated."
            )
        )

    def _read_rows(self, input_path: Path, sheet_name: str) -> list[dict[str, str]]:
        try:
            workbook = openpyxl.load_workbook(input_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"Could not open workbook {input_path}: {exc}") from exc

        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Sheet {sheet_name!r} not found. Available sheets: "
                f"{', '.join(workbook.sheetnames)}"
            )

        sheet = workbook[sheet_name]
        header = [clean_cell(cell.value) for cell in sheet[1]]
        header_map = {name: idx for idx, name in enumerate(header)}
        required = ["DefectMode(TH)", "DefectMode(EN)", "DefectMode(JP)", "Category"]
        missing = [name for name in required if name not in header_map]
        if missing:
            raise CommandError(f"Missing required column(s): {', '.join(missing)}")

        rows: list[dict[str, str]] = []
        errors: list[str] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            name_th = clean_cell(row[header_map["DefectMode(TH)"]])
            name_en = clean_cell(row[header_map["DefectMode(EN)"]])
            name_jp = clean_cell(row[header_map["DefectMode(JP)"]])
            category = clean_cell(row[header_map["Category"]])

            if not any([name_th, name_en, name_jp, category]):
                continue
            if not all([name_th, name_en, name_jp, category]):
                errors.append(f"row {row_number}: TH/EN/JP/Category must all be filled")
                continue

            rows.append(
                {
                    "name_th": name_th,
                    "name_en": name_en,
                    "name_jp": name_jp,
                    "category": category,
                }
            )

        if errors:
            preview = "\n".join(errors[:10])
            more = "" if len(errors) <= 10 else f"\n... and {len(errors) - 10} more"
            raise CommandError(f"Workbook validation failed:\n{preview}{more}")

        return rows

    def _get_creator(self):
        User = get_user_model()
        return (
            User.objects.filter(is_superuser=True).order_by("pk").first()
            or User.objects.order_by("pk").first()
        )

    def _import_rows(self, rows: list[dict[str, str]], creator, create_categories: bool):
        stats = {
            "defects_created": 0,
            "defects_updated": 0,
            "categories_created": 0,
            "links_created": 0,
            "links_updated": 0,
        }

        categories = {key(obj.name): obj for obj in ItemCategory.objects.all()}
        defects = {key(obj.name_en): obj for obj in DefectMode.objects.all() if obj.name_en}

        for row in rows:
            category = categories.get(key(row["category"]))
            if category is None:
                if not create_categories:
                    raise CommandError(f"Category not found: {row['category']}")
                category = ItemCategory.objects.create(
                    name=row["category"],
                    description="",
                    user=creator,
                )
                categories[key(category.name)] = category
                stats["categories_created"] += 1

            defect = defects.get(key(row["name_en"]))
            if defect is None:
                defect = DefectMode.objects.create(
                    name_th=row["name_th"],
                    name_en=row["name_en"],
                    name_jp=row["name_jp"],
                    defect_type=DefectMode.DefectType.PROCESS_NG,
                    user=creator,
                )
                defects[key(defect.name_en)] = defect
                stats["defects_created"] += 1
            else:
                fields = []
                for field in ("name_th", "name_en", "name_jp"):
                    if (getattr(defect, field) or "") != row[field]:
                        setattr(defect, field, row[field])
                        fields.append(field)
                if defect.defect_type != DefectMode.DefectType.PROCESS_NG:
                    defect.defect_type = DefectMode.DefectType.PROCESS_NG
                    fields.append("defect_type")
                if fields:
                    fields.append("updated_at")
                    defect.save(update_fields=fields)
                    stats["defects_updated"] += 1

            link = DefectByCategory.objects.filter(
                category=category,
                defect_mode=defect,
            ).first()
            title = f"{category.name} - {defect.name_en}".strip()
            if link is None:
                DefectByCategory.objects.create(
                    title=title,
                    category=category,
                    defect_mode=defect,
                    is_inlist=True,
                    description="",
                    user=creator,
                )
                stats["links_created"] += 1
            else:
                fields = []
                if link.title != title:
                    link.title = title
                    fields.append("title")
                if not link.is_inlist:
                    link.is_inlist = True
                    fields.append("is_inlist")
                if fields:
                    fields.append("updated_at")
                    link.save(update_fields=fields)
                    stats["links_updated"] += 1

        return stats

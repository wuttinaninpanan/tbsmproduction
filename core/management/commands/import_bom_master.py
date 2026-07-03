"""Rebuild the BOM structure from BoM_Master.xlsx.

Policy (agreed with the data owner):
  * KEEP everything tied to assemblies whose Item_list.sd_code matches DAR-<n> / DAL-<n>:
    their BillOfMaterial headers, their BillOfMaterialItemMater lines, the InspectionItem
    rows that hang off those lines, and every Item_list row used as a component in them.
  * DELETE all other BillOfMaterial / BillOfMaterialItemMater, then all other Item_list
    rows (this cascades ItemLine / ItemPrice / Routing / InspectionError / InspectionResult
    for the deleted parts — ItemLine is backed up to JSON first so it can be re-mapped).
  * RELOAD the non-DAR/DAL structure from BoM_Master.xlsx, ALL revisions:
    - new Item_list rows get sd_code = "Temp-<n>" (owner updates real SD codes later),
      taking only part_number + part_name from the sheet;
    - one BillOfMaterial header per assembly;
    - one BillOfMaterialItemMater line per component row, quantity from the sheet.

Defaults to a DRY RUN. Pass --commit to actually apply (inside one transaction).
"""

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl  # type: ignore
from django.core import serializers
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models.bill_of_material import BillOfMaterial
from core.models.bill_of_material_item_master import BillOfMaterialItemMater
from core.models.item_line import ItemLine
from core.models.item_list import Item_list

DARDAL_REGEX = r"^(DAR|DAL)-?[0-9]+"
_dardal_re = re.compile(DARDAL_REGEX, re.IGNORECASE)


def norm(v) -> str:
    return "" if v is None else str(v).strip()


def to_qty(v) -> Decimal:
    """Sheet quantity -> Decimal(12,4); the model forces <=0 to 1, so we mirror that."""
    try:
        d = Decimal(str(v))
    except (InvalidOperation, ValueError, TypeError):
        d = Decimal("1")
    if d <= 0:
        d = Decimal("1")
    return d.quantize(Decimal("0.0001"))


class Command(BaseCommand):
    help = "Wipe non-DAR/DAL BOM + Item_list and reload structure from BoM_Master.xlsx."

    def add_arguments(self, parser):
        default_xlsx = Path(__file__).resolve().parents[1] / "BoM_Master.xlsx"
        parser.add_argument("--xlsx", default=str(default_xlsx))
        parser.add_argument("--sheet", default="BoM")
        parser.add_argument("--user-id", default="1",
                            help="core_user.id to own all created rows (default 1).")
        parser.add_argument("--backup-dir", default="backups")
        parser.add_argument("--commit", action="store_true",
                            help="Apply changes. Without it the command only reports (dry run).")

    # ---- sheet parsing -------------------------------------------------
    def read_sheet(self, path, sheet):
        if not Path(path).exists():
            raise CommandError(f"xlsx not found: {path}")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            raise CommandError(f"sheet {sheet!r} not in {wb.sheetnames}")
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        next(it)  # header
        names = {}            # part_number -> part_name (first non-empty wins)
        component_rows = []   # (parent_pn, child_pn, qty)  for rows with a parent
        self_rows = 0         # rows where parent == child (a part listed as its own component)
        for row in it:
            parent, parent_name = norm(row[1]), norm(row[2])
            child, child_name = norm(row[3]), norm(row[4])
            qty = row[6]
            if child and child not in names and child_name:
                names[child] = child_name[:255]
            if parent and parent not in names and parent_name:
                names[parent] = parent_name[:255]
            if parent and child:
                # A part can never be a component of itself. The source sheet has
                # ~359 such rows; importing them makes the item edit page show the
                # assembly listed under its own BOM. Drop them here.
                if parent == child:
                    self_rows += 1
                    continue
                component_rows.append((parent, child, qty))
        if self_rows:
            self.stdout.write(self.style.WARNING(
                f"skipped {self_rows} self-referencing rows (parent == child)"))
        return names, component_rows

    def handle(self, *args, **opts):
        commit = opts["commit"]
        user_id = opts["user_id"]
        names, component_rows = self.read_sheet(opts["xlsx"], opts["sheet"])

        # ---- DAR/DAL keep-set from the live DB --------------------------
        dardal_items = Item_list.objects.filter(sd_code__regex=DARDAL_REGEX)
        dardal_item_ids = set(dardal_items.values_list("id", flat=True))
        dardal_pns = {norm(pn) for pn in dardal_items.values_list("part_number", flat=True)}

        keep_headers = list(
            BillOfMaterial.objects.filter(item_id__in=dardal_item_ids).values_list("id", flat=True)
        )
        comp_ids = set(
            BillOfMaterialItemMater.objects.filter(bom_id__in=keep_headers)
            .values_list("component_id", flat=True)
        )
        keep_item_ids = dardal_item_ids | comp_ids

        if not dardal_item_ids:
            raise CommandError(
                "No Item_list rows match DAR/DAL — refusing to run (would wipe everything). "
                "Check you are pointed at the right database."
            )

        # ---- safety: nothing outside keep-set may be PROTECT-referenced -
        # Walk every reverse FK that PROTECTs Item_list and confirm none of
        # them point at a row we're about to delete. Done up front so a
        # changed live state aborts cleanly instead of failing mid-delete.
        from django.db.models.deletion import PROTECT

        # BOM lines also PROTECT-reference Item_list, but we delete the
        # non-kept ones ourselves *before* touching Item_list, and the kept
        # (DAR/DAL) ones only reference keep-set components — so they never
        # block. Exclude them; check only the external operational tables.
        protect_hits = []
        for rel in Item_list._meta.related_objects:
            if getattr(rel, "on_delete", None) is not PROTECT:
                continue
            if rel.related_model is BillOfMaterialItemMater:
                continue
            col = f"{rel.field.name}_id"
            model = rel.related_model
            try:
                bad = (model.objects
                       .filter(**{f"{col}__isnull": False})
                       .exclude(**{f"{col}__in": keep_item_ids})
                       .count())
            except Exception as exc:  # table may not exist on an older schema
                self.stdout.write(self.style.WARNING(
                    f"  (skipped PROTECT check on {model.__name__}.{rel.field.name}: {exc})"))
                continue
            if bad:
                protect_hits.append(f"{model.__name__}.{rel.field.name}={bad}")
        if protect_hits:
            raise CommandError(
                "PROTECT references to Item_list rows outside the keep-set would block "
                "deletion — aborting (no changes made). Offenders: " + ", ".join(protect_hits)
            )

        total_items = Item_list.objects.count()
        del_items = total_items - len(keep_item_ids)
        del_headers = BillOfMaterial.objects.exclude(id__in=keep_headers).count()
        del_lines = BillOfMaterialItemMater.objects.exclude(bom_id__in=keep_headers).count()

        # ---- reload plan ------------------------------------------------
        nondardal_rows = [r for r in component_rows if r[0] not in dardal_pns]
        needed_pns = set()
        for parent, child, _q in nondardal_rows:
            needed_pns.add(parent)
            needed_pns.add(child)
        assemblies = {p for p, _c, _q in nondardal_rows}

        self.stdout.write(self.style.MIGRATE_HEADING("=== PLAN ==="))
        self.stdout.write(f"DAR/DAL kept: {len(dardal_item_ids)} assemblies, "
                          f"{len(keep_headers)} BOM headers, components {len(comp_ids)} "
                          f"(Item_list keep-set {len(keep_item_ids)})")
        self.stdout.write(f"DELETE: Item_list {del_items}, BOM headers {del_headers}, "
                          f"BOM lines {del_lines}")
        self.stdout.write(f"RELOAD: assemblies {len(assemblies)}, component lines "
                          f"{len(nondardal_rows)}, distinct part_numbers {len(needed_pns)}")

        if not commit:
            self.stdout.write(self.style.WARNING("\nDRY RUN — no changes. Re-run with --commit to apply."))
            return

        backup_dir = Path(opts["backup_dir"])
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = timezone.now().strftime("%Y%m%d_%H%M%S")

        with transaction.atomic():
            # 1) backup ItemLine that will cascade away (the ones whose item is deleted)
            doomed_lines = ItemLine.objects.exclude(item_id__in=keep_item_ids)
            backup_path = backup_dir / f"itemline_backup_{stamp}.json"
            with open(backup_path, "w", encoding="utf-8") as fh:
                fh.write(serializers.serialize("json", doomed_lines.select_related(None)))
            self.stdout.write(f"backed up {doomed_lines.count()} ItemLine -> {backup_path}")

            # 2) delete non-DAR/DAL BOM (lines before headers), then Item_list
            BillOfMaterialItemMater.objects.exclude(bom_id__in=keep_headers).delete()
            BillOfMaterial.objects.exclude(id__in=keep_headers).delete()
            Item_list.objects.exclude(id__in=keep_item_ids).delete()

            # 3) recreate Item_list (Temp-) for needed part_numbers not surviving
            survivors = {norm(i.part_number): i for i in Item_list.objects.all()}
            new_items = []
            temp_n = 0
            for pn in sorted(needed_pns):
                if pn in survivors:
                    continue
                temp_n += 1
                new_items.append(Item_list(
                    sd_code=f"Temp-{temp_n}",
                    part_number=pn[:255],
                    part_name=names.get(pn, "")[:255],
                    sku="",
                    user_id=user_id,
                ))
            Item_list.objects.bulk_create(new_items, batch_size=1000)
            for i in new_items:
                survivors[norm(i.part_number)] = i
            self.stdout.write(f"created {len(new_items)} Temp Item_list rows")

            # 4) BillOfMaterial header per assembly (skip those that already exist)
            existing_hdr_items = set(
                BillOfMaterial.objects.values_list("item_id", flat=True)
            )
            new_headers = []
            for pn in sorted(assemblies):
                item = survivors[pn]
                if item.id in existing_hdr_items:
                    continue
                new_headers.append(BillOfMaterial(
                    item=item, revision="A", latest_eci="", scrap_percent=0,
                    user_id=user_id,
                ))
            BillOfMaterial.objects.bulk_create(new_headers, batch_size=1000)
            header_by_item = {
                h.item_id: h for h in BillOfMaterial.objects.all()
            }
            self.stdout.write(f"created {len(new_headers)} BillOfMaterial headers")

            # 5) BillOfMaterialItemMater lines (all revisions)
            seq = {}
            new_lines = []
            for parent, child, qty in nondardal_rows:
                hdr = header_by_item[survivors[parent].id]
                seq[hdr.id] = seq.get(hdr.id, 0) + 1
                new_lines.append(BillOfMaterialItemMater(
                    bom=hdr,
                    component=survivors[child],
                    quantity=to_qty(qty),
                    unit="",
                    sequence=seq[hdr.id],
                    user_id=user_id,
                ))
            BillOfMaterialItemMater.objects.bulk_create(new_lines, batch_size=2000)
            self.stdout.write(self.style.SUCCESS(
                f"created {len(new_lines)} BillOfMaterialItemMater lines"))

        self.stdout.write(self.style.SUCCESS("DONE (committed)."))

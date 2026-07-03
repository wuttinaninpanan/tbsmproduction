"""Sync Item_list from the 'item list' sheet of itemlist New update.xlsx.

The sheet is the owner's latest authoritative item list (SD Code, Part number,
Part Name, Model). Policy, agreed with the data owner:

  * NEVER touch rows whose Item_list.sd_code matches DAR-<n> / DAL-<n> (the
    curated reclining-adjuster set). File rows carrying a DAR/DAL code are
    skipped too — that set is managed by hand.
  * UPDATE rows matched by part_number (exact, trimmed): overwrite sd_code and
    part_name with the sheet's values. This is how the Temp-<n> placeholders
    left by import_bom_master get their real SD codes.
  * CREATE an Item_list row for sheet part_numbers with no DB match
    (sku blank, no stage — same shape import_bom_master created).
  * DELETE non-DAR/DAL rows whose part_number is NOT in the sheet — but only
    when nothing references them. A row referenced by anything (BOM lines,
    inspection, scrap, ...) is KEPT and reported instead: deleting BOM
    components would tear out the structure rebuilt on 2026-07-01.

Before committing, the whole Item_list table is serialized to
backups/item_list_backup_<stamp>.json so the sync is reversible.

Defaults to a DRY RUN. Pass --commit to actually apply (inside one transaction).
"""

import re
from pathlib import Path

import openpyxl  # type: ignore
from django.core import serializers
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models.item_list import Item_list, is_spreadsheet_error

DARDAL_REGEX = r"^(DAR|DAL)-?[0-9]+"
_dardal_re = re.compile(DARDAL_REGEX, re.IGNORECASE)

# Real header is on the THIRD row of the sheet; the two rows above are blank.
HEADER_ROW_INDEX = 2
SD_COL, PN_COL, NAME_COL = 0, 1, 2


def norm(v) -> str:
    return "" if v is None else str(v).strip()


class Command(BaseCommand):
    help = "Replace non-DAR/DAL Item_list content with the 'item list' sheet (update/create/delete-unreferenced)."

    def add_arguments(self, parser):
        default_xlsx = Path(__file__).resolve().parents[1] / "itemlist New update.xlsx"
        parser.add_argument("--xlsx", default=str(default_xlsx))
        parser.add_argument("--sheet", default="item list")
        parser.add_argument("--user-id", default="1",
                            help="core_user.id to own created rows (default 1).")
        parser.add_argument("--backup-dir", default="backups")
        parser.add_argument("--show", type=int, default=20)
        parser.add_argument("--commit", action="store_true",
                            help="Apply changes. Without it the command only reports (dry run).")

    # ---- sheet parsing -------------------------------------------------
    def read_sheet(self, path, sheet):
        if not Path(path).exists():
            raise CommandError(f"xlsx not found: {path}")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            raise CommandError(f"sheet {sheet!r} not in {wb.sheetnames}")
        ws = wb[sheet]

        pn_map = {}          # part_number -> (sd_code, part_name)
        conflicts = {}
        skipped = {"blank_sd": 0, "sheet_error": 0, "dardal_in_file": 0}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= HEADER_ROW_INDEX or row is None:
                continue
            sd = norm(row[SD_COL]) if len(row) > SD_COL else ""
            pn = norm(row[PN_COL]) if len(row) > PN_COL else ""
            name = norm(row[NAME_COL]) if len(row) > NAME_COL else ""
            if not pn:
                continue
            if not sd:
                skipped["blank_sd"] += 1
                continue
            if is_spreadsheet_error(sd) or is_spreadsheet_error(pn) or is_spreadsheet_error(name):
                skipped["sheet_error"] += 1
                continue
            if _dardal_re.match(sd):
                skipped["dardal_in_file"] += 1
                continue
            if pn in pn_map and pn_map[pn][0] != sd:
                conflicts.setdefault(pn, {pn_map[pn][0]}).add(sd)
            pn_map[pn] = (sd[:32], name[:255])
        wb.close()

        if conflicts:
            raise CommandError(
                "The sheet maps the same part_number to different sd_codes; resolve first: "
                + "; ".join(f"{pn}={sorted(v)}" for pn, v in list(conflicts.items())[:20])
            )
        return pn_map, skipped

    # ---- reverse-relation census ---------------------------------------
    def referenced_ids(self, candidate_ids):
        """IDs among ``candidate_ids`` referenced by ANY reverse FK.

        Django-side on_delete is a mix of PROTECT (raises) and CASCADE
        (silently wipes operational data), so we refuse to delete anything
        referenced at all, whatever the rule.
        """
        referenced = set()
        for rel in Item_list._meta.related_objects:
            model = rel.related_model
            col = rel.field.name
            try:
                hit = (model.objects
                       .filter(**{f"{col}__in": candidate_ids})
                       .values_list(f"{col}_id", flat=True).distinct())
                referenced |= set(hit)
            except Exception as exc:  # table may not exist on an older schema
                self.stdout.write(self.style.WARNING(
                    f"  (skipped reverse check on {model.__name__}.{col}: {exc})"))
        return referenced

    def handle(self, *args, **opts):
        commit = opts["commit"]
        show = opts["show"]
        pn_map, skipped = self.read_sheet(opts["xlsx"], opts["sheet"])

        dardal_ids = set(
            Item_list.objects.filter(sd_code__regex=DARDAL_REGEX).values_list("id", flat=True)
        )
        if not dardal_ids:
            raise CommandError(
                "No Item_list rows match DAR/DAL — refusing to run. "
                "Check you are pointed at the right database."
            )

        db_by_pn = {}
        for pk, pn, sd, name in Item_list.objects.values_list(
                "id", "part_number", "sd_code", "part_name"):
            db_by_pn.setdefault(norm(pn), []).append((pk, norm(sd), norm(name)))

        to_update = []      # (id, pn, old_sd, new_sd, old_name, new_name)
        unchanged = 0
        to_create = []      # (pn, sd, name)
        for pn, (sd, name) in pn_map.items():
            rows = [r for r in db_by_pn.get(pn, []) if r[0] not in dardal_ids]
            if not rows:
                if not db_by_pn.get(pn):    # no match at all (not even DAR/DAL)
                    to_create.append((pn, sd, name))
                continue
            for pk, osd, oname in rows:
                if osd == sd and oname == name:
                    unchanged += 1
                else:
                    to_update.append((pk, pn, osd, sd, oname, name))

        stale = [
            (pk, sd, pn)
            for pn, rows in db_by_pn.items() if pn not in pn_map
            for pk, sd, _name in rows if pk not in dardal_ids
        ]
        referenced = self.referenced_ids([pk for pk, _sd, _pn in stale])
        to_delete = [x for x in stale if x[0] not in referenced]
        kept_referenced = [x for x in stale if x[0] in referenced]

        # ---- report -----------------------------------------------------
        h = self.style.MIGRATE_HEADING
        self.stdout.write(h("=== SHEET ==="))
        self.stdout.write(f"usable part_numbers: {len(pn_map)}  skipped: {skipped}")
        self.stdout.write(h("=== PLAN ==="))
        self.stdout.write(
            f"DAR/DAL rows untouched            : {len(dardal_ids)}\n"
            f"UPDATE sd_code/part_name          : {len(to_update)}\n"
            f"already identical                 : {unchanged}\n"
            f"CREATE new items                  : {len(to_create)}\n"
            f"DELETE (stale, unreferenced)      : {len(to_delete)}\n"
            f"KEEP (stale but referenced by BOM/etc.): {len(kept_referenced)}"
        )
        for title, rows_ in (
            ("updates", [f"{pn}: {osd!r}->{sd!r} name {oname!r}->{name!r}"
                         for _pk, pn, osd, sd, oname, name in to_update]),
            ("creates", [f"{pn} (sd {sd})" for pn, sd, _ in to_create]),
            ("deletes", [f"{pn} (sd {sd})" for _pk, sd, pn in to_delete]),
            ("kept-referenced", [f"{pn} (sd {sd})" for _pk, sd, pn in kept_referenced]),
        ):
            if rows_:
                self.stdout.write(h(f"--- {title} (first {show} of {len(rows_)}) ---"))
                for line in rows_[:show]:
                    self.stdout.write(f"    {line}")

        if not commit:
            self.stdout.write(self.style.WARNING(
                "\nDRY RUN — no changes. Re-run with --commit to apply."))
            return

        backup_dir = Path(opts["backup_dir"])
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"item_list_backup_{stamp}.json"
        with open(backup_path, "w", encoding="utf-8") as fh:
            fh.write(serializers.serialize("json", Item_list.objects.all()))
        self.stdout.write(f"backed up {Item_list.objects.count()} Item_list -> {backup_path}")

        with transaction.atomic():
            objs = [Item_list(id=pk, sd_code=sd, part_name=name[:255])
                    for pk, _pn, _osd, sd, _on, name in to_update]
            Item_list.objects.bulk_update(objs, ["sd_code", "part_name"], batch_size=1000)

            Item_list.objects.bulk_create(
                [Item_list(sd_code=sd, part_number=pn[:255], part_name=name[:255],
                           sku="", user_id=opts["user_id"])
                 for pn, sd, name in to_create],
                batch_size=1000,
            )

            if to_delete:
                Item_list.objects.filter(id__in=[pk for pk, _s, _p in to_delete]).delete()

        self.stdout.write(self.style.SUCCESS(
            f"DONE (committed) — updated {len(to_update)}, created {len(to_create)}, "
            f"deleted {len(to_delete)}, kept {len(kept_referenced)} stale-but-referenced."))

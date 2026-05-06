"""
import_bom_staging
──────────────────
Import BOM Update sheet จาก Excel ไฟล์เดียวเข้า BomStaging table

เงื่อนไข:
  - นำเข้าทีละไฟล์ (ระบุด้วย --file)
  - รักษาลำดับ row ตามต้นฉบับ Excel (เก็บใน row_index)
  - ไม่ dedupe — เก็บทุก row ที่พบ

Usage:
  python manage.py import_bom_staging --file="Document/Format_BOM_XXX.xlsx"
  python manage.py import_bom_staging --file="..." --clear-file   # ล้าง row ของไฟล์นี้ก่อน
  python manage.py import_bom_staging --file="..." --dry-run      # ทดสอบ
  python manage.py import_bom_staging --file="..." --sheet="BOM Update"
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models.bom_staging import BomStaging

User = get_user_model()

BOM_SHEET_CANDIDATES = ["MAS_BOM Update", "BOM Update", "BOM "]


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def safe_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.upper() in ("#N/A", "N/A", "NONE", "-") else s


def parse_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    try:
        s = str(val).strip()
        if not s:
            return None
        return Decimal(s)
    except (InvalidOperation, ValueError, TypeError):
        return None


# ─────────────────────────────────────────────
# Layout detection (รองรับ 3 แบบ: SD_col 6/7/8)
# ─────────────────────────────────────────────

def detect_layout(ws) -> Optional[dict]:
    """หา column positions ของ BOM Update sheet แบบยืดหยุ่น"""
    layout = {
        "header_row": None,
        "data_start_row": None,
        "level_col_start": None,
        "level_col_count": None,
        "sd_col": None,
        "pn_col": None,
        "name_col": None,
        "qty_col": None,
        "process_col": None,
        "line_col": None,
        "supplier_col": None,
        "model_col": None,
        "coil_press_col": None,
        "weight_col": None,
        "sd_fg_col": None,
        "sd_comp_col": None,
        "supplier2_col": None,
        "line_no2_col": None,
    }

    # สแกน 5 rows แรกหา header
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        for col_idx, val in enumerate(row):
            if not isinstance(val, str):
                continue
            v = val.strip().upper()

            if "SD_FG" in v and layout["sd_fg_col"] is None:
                layout["sd_fg_col"] = col_idx
            elif "SD" in v and "CODE" in v and "COMP" in v and layout["sd_comp_col"] is None:
                layout["sd_comp_col"] = col_idx
            elif "SD" in v and "CODE" in v and layout["sd_col"] is None:
                layout["sd_col"] = col_idx
                layout["header_row"] = row_idx + 1
            elif "PART" in v and ("NO" in v or "NUM" in v) and layout["pn_col"] is None:
                layout["pn_col"] = col_idx
            elif "PART" in v and "NAME" in v and layout["name_col"] is None:
                layout["name_col"] = col_idx
            elif "USAGE" in v and layout["qty_col"] is None:
                layout["qty_col"] = col_idx
            elif v == "PROCESS" and layout["process_col"] is None:
                layout["process_col"] = col_idx
            elif "LINE" in v and "NO" in v:
                if layout["line_col"] is None:
                    layout["line_col"] = col_idx
                elif layout["line_no2_col"] is None:
                    layout["line_no2_col"] = col_idx
            elif "SUPPLIER" in v:
                if layout["supplier_col"] is None:
                    layout["supplier_col"] = col_idx
                elif layout["supplier2_col"] is None:
                    layout["supplier2_col"] = col_idx
            elif v == "MODEL" and layout["model_col"] is None:
                layout["model_col"] = col_idx
            elif "COIL" in v and "PRESS" in v and layout["coil_press_col"] is None:
                layout["coil_press_col"] = col_idx
            elif v == "WEIGHT" and layout["weight_col"] is None:
                layout["weight_col"] = col_idx

    if layout["sd_col"] is None:
        return None

    # หา level_col_start (column ที่มีคำว่า "Level")
    level_start = None
    for row_idx in range(layout["header_row"]):
        row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))[0]
        for col_idx, val in enumerate(row):
            if isinstance(val, str) and "LEVEL" in val.strip().upper():
                level_start = col_idx
                break
        if level_start is not None:
            break

    layout["level_col_start"] = level_start if level_start is not None else 0
    layout["level_col_count"] = layout["sd_col"] - layout["level_col_start"]

    # data_start_row: ตรวจว่า row ถัดจาก header เป็น legend หรือข้อมูลจริง
    next_row_idx = layout["header_row"] + 1
    rows = list(ws.iter_rows(min_row=next_row_idx, max_row=next_row_idx, values_only=True))
    if not rows:
        layout["data_start_row"] = layout["header_row"] + 1
        return layout

    next_row = rows[0]
    sd_val = next_row[layout["sd_col"]] if layout["sd_col"] < len(next_row) else None
    lcs = layout["level_col_start"]
    lcc = layout["level_col_count"]
    level_vals = [next_row[lcs + i] for i in range(lcc) if (lcs + i) < len(next_row)]
    is_legend = (sd_val is None) and all(
        isinstance(v, (int, float)) or v is None for v in level_vals
    )
    layout["data_start_row"] = (layout["header_row"] + 2) if is_legend else (layout["header_row"] + 1)
    return layout


def extract_level(row, layout) -> tuple[int | None, str]:
    """
    คืน: (level, lv_star)
      - level   = column index ที่มี value (0=FG, 1..N=sub-levels)
      - lv_star = เลข level เป็น string (เช่น "0", "1", "5", "10")
    """
    lcs = layout["level_col_start"]
    lcc = layout["level_col_count"]

    level = None
    for i in range(lcc):
        idx = lcs + i
        if idx < len(row) and row[idx] is not None:
            level = i
            break

    lv_star = str(level) if level is not None else ""
    return level, lv_star


# ─────────────────────────────────────────────
# Command
# ─────────────────────────────────────────────

class Command(BaseCommand):
    help = "Import BOM Update sheet จาก Excel ไฟล์เดียว เข้า BomStaging (ไม่ dedupe)"

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to Excel file")
        parser.add_argument("--sheet", default=None, help="Sheet name override")
        parser.add_argument(
            "--clear-file", action="store_true",
            help="ลบ row เดิมของไฟล์นี้ก่อน import",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="ไม่บันทึกจริง แค่ parse ดู",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"ไม่พบไฟล์: {file_path}")

        sheet_override = options["sheet"]
        clear_file = options["clear_file"]
        dry_run = options["dry_run"]

        # ─── Open Excel ───
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as e:
            raise CommandError(f"เปิดไฟล์ไม่ได้: {e}")

        # หา BOM sheet
        if sheet_override:
            if sheet_override not in wb.sheetnames:
                wb.close()
                raise CommandError(
                    f"ไม่พบ sheet '{sheet_override}' — มี: {wb.sheetnames}"
                )
            sheet_name = sheet_override
        else:
            sheet_name = None
            for c in BOM_SHEET_CANDIDATES:
                if c in wb.sheetnames:
                    sheet_name = c
                    break
            if sheet_name is None:
                for s in wb.sheetnames:
                    if "BOM" in s.upper():
                        sheet_name = s
                        break
            if sheet_name is None:
                wb.close()
                raise CommandError(
                    f"ไม่พบ BOM sheet — sheets: {wb.sheetnames}"
                )

        self.stdout.write(f"📄 File : {file_path.name}")
        self.stdout.write(f"📑 Sheet: {sheet_name}")

        ws = wb[sheet_name]
        layout = detect_layout(ws)
        if layout is None:
            wb.close()
            raise CommandError("ตรวจจับ layout ของ sheet ไม่ได้ (ไม่พบ 'SD Code')")

        self.stdout.write(
            f"   SD_col={layout['sd_col']} | "
            f"level_cols={layout['level_col_start']}~{layout['sd_col']-1} | "
            f"data_start={layout['data_start_row']}"
        )

        # system user (optional)
        system_user = User.objects.filter(is_superuser=True).order_by("id").first()

        # ─── Parse rows ───
        def col(row, key) -> str:
            idx = layout.get(key)
            if idx is None or idx >= len(row):
                return ""
            return safe_str(row[idx])

        def col_decimal(row, key) -> Optional[Decimal]:
            idx = layout.get(key)
            if idx is None or idx >= len(row):
                return None
            return parse_decimal(row[idx])

        rows_to_create = []
        current_fg_sd = ""  # propagate SD ของ FG ปัจจุบัน

        row_iter = ws.iter_rows(
            min_row=layout["data_start_row"], values_only=True
        )

        for row_num, row in enumerate(row_iter, start=layout["data_start_row"]):
            # ข้าม row ที่ว่างเปล่า
            if not any(v is not None for v in row):
                continue

            level, lv_star = extract_level(row, layout)

            sd_code = col(row, "sd_col")
            part_no = col(row, "pn_col")
            part_name = col(row, "name_col")
            usage_rm = col_decimal(row, "qty_col")
            process = col(row, "process_col")
            line_no = col(row, "line_col")
            supplier_name = col(row, "supplier_col")
            model = col(row, "model_col")

            coil_press = col(row, "coil_press_col")
            weight = col_decimal(row, "weight_col")
            sd_fg_raw = col(row, "sd_fg_col")
            sd_code_component = col(row, "sd_comp_col")
            supplier2 = col(row, "supplier2_col")
            line_no_final = col(row, "line_no2_col")

            # ตั้ง FG context (สำหรับ rows ที่ SD_FG ว่างใน excel)
            if level == 0 and sd_code:
                current_fg_sd = sd_code

            sd_fg = sd_fg_raw or current_fg_sd

            rows_to_create.append(BomStaging(
                lv_star=lv_star,
                level=level,
                sd_code=sd_code,
                part_no=part_no,
                part_name=part_name,
                usage_rm=usage_rm,
                process=process,
                line_no=line_no,
                supplier_name=supplier_name,
                model=model,
                coil_press=coil_press,
                weight=weight,
                sd_fg=sd_fg,
                sd_code_component=sd_code_component,
                supplier=supplier2,
                line_no_final=line_no_final,
                source_file=file_path.name,
                source_sheet=sheet_name,
                row_index=row_num,
                user=system_user,
            ))

        wb.close()

        self.stdout.write(f"\n📊 พบ {len(rows_to_create)} data rows")

        # ─── Preview ตัวอย่าง 3 rows แรก ───
        if rows_to_create:
            self.stdout.write("\n🔍 ตัวอย่าง 3 rows แรก:")
            for s in rows_to_create[:3]:
                self.stdout.write(
                    f"   row{s.row_index:>4} L{s.level} | "
                    f"SD={s.sd_code!r:>12} | PN={s.part_no!r:>20} | "
                    f"Name={(s.part_name[:30] if s.part_name else '')!r}"
                )

        if dry_run:
            self.stdout.write(self.style.WARNING(
                "\n⚠ DRY RUN — ไม่บันทึกจริง (เพิ่ม --dry-run ออก เพื่อ import)"
            ))
            return

        # ─── Commit ───
        with transaction.atomic():
            if clear_file:
                deleted, _ = BomStaging.objects.filter(
                    source_file=file_path.name
                ).delete()
                if deleted:
                    self.stdout.write(f"🗑 ลบ row เดิมของไฟล์นี้: {deleted} rows")

            BomStaging.objects.bulk_create(rows_to_create, batch_size=500)

        self.stdout.write(self.style.SUCCESS(
            f"\n✅ Import สำเร็จ: +{len(rows_to_create)} rows จาก {file_path.name}"
        ))

        # ─── Summary ───
        total = BomStaging.objects.count()
        files_count = BomStaging.objects.values("source_file").distinct().count()
        self.stdout.write(
            f"   รวม BomStaging: {total:,} rows จาก {files_count} ไฟล์"
        )

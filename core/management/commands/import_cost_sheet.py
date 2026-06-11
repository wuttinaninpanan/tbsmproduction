from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None

from core.models.bill_of_material import BillOfMaterial
from core.models.bill_of_material_item_master import BillOfMaterialItemMater
from core.models.item_category import ItemCategory
from core.models.item_list import Item_list
from core.models.item_stage import ItemStage
from core.views.bom_template import _generate_unique_sku, _reclassify_bom_stages_and_codes


@dataclass
class Stats:
    items_created: int = 0
    items_updated: int = 0
    items_existing: int = 0
    bom_headers_created: int = 0
    bom_headers_existing: int = 0
    bom_links_created: int = 0
    bom_links_updated: int = 0
    bom_links_existing: int = 0
    skipped_items: int = 0
    skipped_bom_rows: int = 0


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def dec(value, default: Decimal = Decimal("0")) -> Decimal:
    s = clean(value)
    if not s:
        return default
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return default


def money(value) -> Decimal:
    return dec(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def quantity_decimal(value) -> Decimal:
    return dec(value, Decimal("1")).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


def row_has_values(row) -> bool:
    return any(clean(v) for v in row)


class Command(BaseCommand):
    help = "Import Item_list and BOM data from core/management/Cost Sheet.xlsx."

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="core/management/Cost Sheet.xlsx",
            help="Path to Cost Sheet workbook.",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Write changes. Without this flag the command rolls back as a dry-run.",
        )
        parser.add_argument(
            "--user",
            default="admin",
            help="Username to use for required user fields.",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl is not installed")

        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        User = get_user_model()
        user = User.objects.filter(username=options["user"]).first() or User.objects.order_by("id").first()
        if user is None:
            raise CommandError("No user found for required user fields")

        apply = bool(options["apply"])
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        stats = Stats()

        stages = {
            "fg": ItemStage.objects.filter(name="fg").first(),
            "raw_mat": ItemStage.objects.filter(name="raw_mat").first(),
            "wip": ItemStage.objects.filter(name="wip").first(),
            "wip_press": ItemStage.objects.filter(name="wip_press").first(),
        }
        categories = {(c.name or "").strip().lower(): c for c in ItemCategory.objects.all()}

        item_cache: dict[tuple[str, str], Item_list] = {}

        def find_item(sd_code: str, part_number: str) -> Item_list | None:
            key = (sd_code.lower(), part_number.lower())
            if key not in item_cache:
                item_cache[key] = Item_list.objects.filter(
                    sd_code__iexact=sd_code,
                    part_number__iexact=part_number,
                ).first()
            return item_cache.get(key)

        def upsert_item(
            sd_code,
            part_number,
            part_name="",
            *,
            stage_name="",
            category_name="",
            cost=None,
            purchased_price=None,
        ) -> Item_list | None:
            sd_code = clean(sd_code)
            part_number = clean(part_number)
            part_name = clean(part_name)
            if not sd_code or not part_number:
                stats.skipped_items += 1
                return None

            item = find_item(sd_code, part_number)
            if item is None:
                item = Item_list.objects.create(
                    sd_code=sd_code,
                    part_number=part_number,
                    part_name=part_name,
                    sku=_generate_unique_sku(part_number=part_number, sd_code=sd_code),
                    stage=stages.get(stage_name) if stage_name else None,
                    cost=money(cost) if cost is not None else Decimal("0"),
                    purchased_price=money(purchased_price) if purchased_price is not None else Decimal("0"),
                    user=user,
                )
                item_cache[(sd_code.lower(), part_number.lower())] = item
                stats.items_created += 1
                return item

            changed: list[str] = []
            if part_name and item.part_name != part_name:
                item.part_name = part_name
                changed.append("part_name")
            category = categories.get(clean(category_name).lower()) if category_name else None
            if category is not None and item.category_id != category.id:
                item.category = category
                changed.append("category")
            if cost is not None:
                cost_value = money(cost)
                if item.cost != cost_value:
                    item.cost = cost_value
                    changed.append("cost")
            if purchased_price is not None:
                price_value = money(purchased_price)
                if item.purchased_price != price_value:
                    item.purchased_price = price_value
                    changed.append("purchased_price")
            if changed:
                item.save(update_fields=changed + ["updated_at"])
                stats.items_updated += 1
            else:
                stats.items_existing += 1
            return item

        def ensure_bom(item: Item_list) -> BillOfMaterial:
            bom, created = BillOfMaterial.objects.get_or_create(
                item=item,
                defaults={"revision": "A", "latest_eci": "", "user": user},
            )
            if created:
                stats.bom_headers_created += 1
            else:
                stats.bom_headers_existing += 1
            return bom

        def upsert_bom_link(parent: Item_list | None, component: Item_list | None, qty, sequence: int):
            if parent is None or component is None:
                stats.skipped_bom_rows += 1
                return
            bom = ensure_bom(parent)
            quantity = quantity_decimal(qty) or Decimal("1.0000")
            obj = BillOfMaterialItemMater.objects.filter(bom=bom, component=component).first()
            if obj is None:
                BillOfMaterialItemMater.objects.create(
                    bom=bom,
                    component=component,
                    quantity=quantity,
                    unit="PCS",
                    sequence=sequence,
                    user=user,
                )
                stats.bom_links_created += 1
                return
            changed = []
            if obj.quantity != quantity:
                obj.quantity = quantity
                changed.append("quantity")
            if obj.unit != "PCS":
                obj.unit = "PCS"
                changed.append("unit")
            if obj.sequence != sequence:
                obj.sequence = sequence
                changed.append("sequence")
            if changed:
                obj.save(update_fields=changed + ["updated_at"])
                stats.bom_links_updated += 1
            else:
                stats.bom_links_existing += 1

        def import_master_items():
            masters: dict[tuple[str, str], dict] = {}

            def remember_item(sd_code, part_number, part_name="", *, stage_name="", category_name="", cost=None, purchased_price=None):
                sd_code = clean(sd_code)
                part_number = clean(part_number)
                if not sd_code or not part_number:
                    stats.skipped_items += 1
                    return
                key = (sd_code.lower(), part_number.lower())
                data = masters.setdefault(
                    key,
                    {
                        "sd_code": sd_code,
                        "part_number": part_number,
                        "part_name": "",
                        "stage_name": "",
                        "category_name": "",
                        "cost": None,
                        "purchased_price": None,
                    },
                )
                if clean(part_name):
                    data["part_name"] = clean(part_name)
                if stage_name:
                    data["stage_name"] = stage_name
                if clean(category_name):
                    data["category_name"] = clean(category_name)
                if cost is not None and clean(cost) != "":
                    data["cost"] = cost
                if purchased_price is not None and clean(purchased_price) != "":
                    data["purchased_price"] = purchased_price

            if "RM" in wb.sheetnames:
                ws = wb["RM"]
                for row in ws.iter_rows(min_row=4, values_only=True):
                    if not row_has_values(row):
                        continue
                    remember_item(row[0], row[1], row[2], stage_name="raw_mat", cost=row[5], purchased_price=row[5])
            if "FG" in wb.sheetnames:
                ws = wb["FG"]
                for row in ws.iter_rows(min_row=5, values_only=True):
                    if not row_has_values(row):
                        continue
                    remember_item(row[0], row[1], row[2], stage_name="fg", cost=row[7])
            if "WIP" in wb.sheetnames:
                ws = wb["WIP"]
                for row in ws.iter_rows(min_row=8, values_only=True):
                    if not row_has_values(row):
                        continue
                    remember_item(row[0], row[1], row[2], stage_name="wip", category_name=row[3], cost=row[6])
            if "Press" in wb.sheetnames:
                ws = wb["Press"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row_has_values(row):
                        continue
                    remember_item(row[0], row[1], "", stage_name="wip_press")

            for data in masters.values():
                upsert_item(**data)

        def import_boms():
            seq_by_parent: dict[tuple[str, str], int] = {}
            links: dict[tuple[str, str, str, str], tuple[object, object, object, object, object, int, str]] = {}

            def remember_link(parent_sd, parent_part, comp_sd, comp_part, qty, source):
                parent_sd = clean(parent_sd)
                parent_part = clean(parent_part)
                comp_sd = clean(comp_sd)
                comp_part = clean(comp_part)
                if not parent_sd or not parent_part or not comp_sd or not comp_part:
                    stats.skipped_bom_rows += 1
                    return
                parent_key = (parent_sd.lower(), parent_part.lower())
                seq_by_parent[parent_key] = seq_by_parent.get(parent_key, 0) + 1
                link_key = (parent_sd.lower(), parent_part.lower(), comp_sd.lower(), comp_part.lower())
                links[link_key] = (parent_sd, parent_part, comp_sd, comp_part, qty, seq_by_parent[parent_key], source)

            if "BOM_FG" in wb.sheetnames:
                ws = wb["BOM_FG"]
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if not row_has_values(row):
                        continue
                    remember_link(row[0], row[1], row[2], row[3], row[6] if len(row) > 6 else row[4], "fg")
            if "BOM_Sub Assy" in wb.sheetnames:
                ws = wb["BOM_Sub Assy"]
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if not row_has_values(row):
                        continue
                    remember_link(row[0], row[1], row[2], row[3], row[6] if len(row) > 6 else row[4], "wip")

            for parent_sd, parent_part, comp_sd, comp_part, qty, sequence, source in links.values():
                parent_stage = "fg" if source == "fg" else "wip"
                parent = upsert_item(parent_sd, parent_part, "", stage_name=parent_stage)
                component = upsert_item(comp_sd, comp_part, "", stage_name="raw_mat")
                upsert_bom_link(parent, component, qty, sequence)

        try:
            with transaction.atomic():
                import_master_items()
                import_boms()
                _reclassify_bom_stages_and_codes()
                if not apply:
                    transaction.set_rollback(True)
        finally:
            wb.close()

        mode = "APPLIED" if apply else "DRY-RUN"
        self.stdout.write(f"{mode} Cost Sheet import")
        for field in stats.__dataclass_fields__:
            self.stdout.write(f"{field}: {getattr(stats, field)}")

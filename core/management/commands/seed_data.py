"""Unified seed_data command — replaces seed_groups / seed_users /
create_initial_users / seed_bom_master / seed_master_data / seed_scrap_records.

Run with `python manage.py seed_data` (default = all stages, in dependency order).
Use `--stages=<list>` to run a subset, e.g.:
    python manage.py seed_data --stages=master,scrap

Stages (in execution order):
    groups              -> auth Groups (R&D, Production, ...)
    users               -> admin / staff / user (with --password)
    bom                 -> Item_list, BillOfMaterial, BillOfMaterialItemMater
                            from 1.BoM_Master_Summary.xlsx
    line_process        -> LineProcess from Bom master template.xlsx
    item_category       -> ItemCategory (ensures CHC / Cylinder Head Cover)
    line                -> Line from Bom template + NG file + extras
    defect_mode         -> DefectMode from ปัญหางาน NG แยกประเภท.xlsx
    defect_by_category  -> DefectByCategory linking defect <-> category
    item_line           -> ItemLine from BoM template item_and_line sheet
    item_line_extra     -> ItemLine from line.xlsx + variant copy
    scrap               -> ScrapRecord from line.xlsx Append2

Conveniences:
    --stages=all (default)    run every stage above
    --stages=master           groups,users,bom,line_process,item_category,line,
                              defect_mode,defect_by_category,item_line,item_line_extra
    --stages=defects_only     defect_mode,defect_by_category
"""

from __future__ import annotations

from collections import OrderedDict
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models.bill_of_material import BillOfMaterial
from core.models.bill_of_material_item_master import BillOfMaterialItemMater
from core.models.defect_by_category import DefectByCategory
from core.models.defect_mode import DefectMode
from core.models.inout import InOut
from core.models.item_category import ItemCategory
from core.models.item_line import ItemLine
from core.models.item_list import Item_list
from core.models.item_stage import ItemStage
from core.models.line import Line
from core.models.line_process import LineProcess
from core.models.plant import Plant
from core.models.portion import Portion
from core.models.process import Process
from core.models.scrap_record import ScrapRecord
from core.models.side import Side
from core.models.user_profile import UserProfile
from core.models.way import Way


# -----------------------------------------------------------------------------
# Paths to bundled Excel sources
# -----------------------------------------------------------------------------
MGMT_DIR = Path(__file__).resolve().parent.parent
BOM_MASTER_XLSX = MGMT_DIR / "1.BoM_Master_Summary.xlsx"
BOM_TEMPLATE_XLSX = MGMT_DIR / "Bom master template.xlsx"
NG_XLSX = MGMT_DIR / "ปัญหางาน NG แยกประเภท.xlsx"
LINE_XLSX = MGMT_DIR / "line.xlsx"


# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------
DEFAULT_GROUPS = ["R&D", "Production", "Accounting", "Production Control"]

DEFAULT_USERS = [
    # (username, is_staff, is_superuser)
    ("admin", True, True),
    ("staff", True, False),
    ("user", False, False),
]

# BoM master sheet column indices (0-based, row 1 = header).
BOM_SHEET = "BoMItemMaster"
COL_SDPN = 13
COL_M2M = 14
COL_SD = 15
COL_PN = 16
COL_PNAME = 17
COL_QTY = 18

DEFAULT_LINE_PROCESS = "assy"
ITEM_AND_LINE_DEFAULT_STAGE = "fg"  # match the canonical lookup-table key

# -----------------------------------------------------------------------------
# Lookup-table seeds (was migrations 0030_seed_item_stages + 0039_seed_base_lookups)
# -----------------------------------------------------------------------------
LOOKUP_CATEGORIES = [
    "Seat tract", "Lower arm", "Side frame",
    "Hinge", "Round recliner", "Loop handle",
    "CHC",
]
LOOKUP_PORTIONS = ["Front", "Rear"]
LOOKUP_SIDES = ["RH", "LH"]
LOOKUP_INOUTS = ["inner", "outer"]
LOOKUP_WAYS = ["M4", "M6"]
LOOKUP_PLANTS = ["In-house", "Out source"]
LOOKUP_PROCESSES = [
    "Receiving or Incomming",
    "Part supply",
    "Press",
    "Sub-assembly",
    "ED paint",
    "Assembly",
    "Final inspection",
    "Anti rust",
    "Hardening",
]

# (name, display_name, code_prefix) — base 5 + 7 WIP variants.
LOOKUP_ITEM_STAGES = [
    ("raw_mat", "Raw material", "R"),
    ("wip", "Work in process", "W"),
    ("semi_fg", "Semi finished goods", "S"),
    ("fg", "Finished goods", "G"),
    ("delivery", "Delivery", "G"),
    ("wip_press", "WIP(Press)", "W"),
    ("wip_washing", "WIP(Washing)", "W"),
    ("wip_sub_line", "WIP(Sub line)", "W"),
    ("wip_hardening", "WIP(Hardening)", "W"),
    ("wip_anti_rust_oil", "WIP(Anti Rust oil)", "W"),
    ("wip_edp", "WIP(EDP)", "W"),
    ("component_part", "COPONENT PART", "R"),
]

# Friendlier description override for select few categories — empty string keeps
# the column non-null without overriding existing user-entered descriptions.
CATEGORY_DESCRIPTIONS = {"CHC": "Cylinder Head Cover"}

ITEM_CODE_PADDING = 6

# NG sheet name -> ItemCategory.name
NG_SHEET_TO_CATEGORY = {
    "Frame": "Side frame",
    "Arm": "Lower arm",
    "ST": "Seat tract",
    "CHC": "CHC",
    "RR": "Round recliner",
    "Loop": "Loop handle",
}

# Lines explicitly requested by stakeholders that may not appear in any Excel.
EXTRA_LINES = [
    ("CHC1", "CHC"),
    ("CHC2", "CHC"),
]

# line.xlsx (Append2) line names that map to multiple DB lines.
LINE_NAME_ALIAS: dict[str, list[str]] = {
    "570": ["570-1", "570-2"],
    "572": ["572-2"],
    "CFR,CFL": ["CFR", "CFL"],
    "CTU-S,CTA-S": ["CTA-S"],
    "ETI6,ITO4-2": ["ITO4-2"],
    "HFRI,HFRO": ["HFRI"],
    "TAI,TAO": ["TAI", "TAO"],
    "DAA-1": ["DAA"],
    "IF1": ["IF1L", "IF1R"],
    "IF2": ["IF2L", "IF2R"],
    "IFB": ["IFBL", "IFBR"],
    "IFJ1": ["IFJL1", "IFJR1"],
    "IFJ2": ["IFJL2"],
    "IFJ3": ["IFJL3", "IFJR3"],
    "ITA": ["ITA-I", "ITA-O"],
    "ITSL": ["ITSL-1", "ITSL-2"],
    "ITSL-03": ["ITSL-3I", "ITSL-3O"],
    "ITSLB-INN": ["ITSLBI"],
    "ITSLB-OUT": ["ITSLBO"],
    "PRESS 150 T": ["Press"],
    "TFR": ["TFR1"],
    "TFL": ["TFL1"],
    "TTLS": ["TTLS", "HTLS"],
}

# Variant lines that share items with a base line (target -> base).
LINE_VARIANT_BASE: dict[str, str] = {
    "DTA-1": "DTA",
    "TAI-S": "TAI",
    "TAO-S": "TAO",
    "TT1-S": "TT1",
    "DL-S": "DLR-S",
    "DUR-S2.S1": "DUR-S",
    "EL-S": "ELR-S",
    "ETUS": "ITUS",
    "OT-S": "OTA",
    "TG-S": "TGA-01",
    "TH-S": "TGH-01",
    "TGO-01": "TGA-01",
    "TGO-02": "TGA-02",
}

# Stage names accepted on --stages, in canonical execution order.
ALL_STAGES = [
    "groups",
    "users",
    "lookups",
    "bom",
    "bom_classify",
    "line_process",
    "line",
    "defect_mode",
    "defect_by_category",
    "item_line",
    "item_line_extra",
    "scrap",
]
# Convenience aliases.
STAGE_ALIASES = {
    "all": ALL_STAGES,
    "master": [
        "groups", "users", "lookups", "bom", "bom_classify",
        "line_process", "line", "defect_mode", "defect_by_category",
        "item_line", "item_line_extra",
    ],
    "defects_only": ["defect_mode", "defect_by_category"],
    "lines_only": ["line_process", "line", "item_line", "item_line_extra"],
    "lookups_only": ["lookups"],
    # Backward-compat alias for the old separate stage name.
    "item_category": ["lookups"],
}


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _to_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except (ValueError, TypeError):
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0


# -----------------------------------------------------------------------------
# Command
# -----------------------------------------------------------------------------
class Command(BaseCommand):
    help = "Unified seed: Groups, Users, BoM, Lines, Defects, ItemLines, ScrapRecords."

    # ------------------------------------------------------------ arguments

    def add_arguments(self, parser):
        parser.add_argument(
            "--stages", default="all",
            help=(
                "Comma-separated stages to run, or one of: "
                f"{', '.join(STAGE_ALIASES)}. Individual stages: "
                f"{', '.join(ALL_STAGES)}."
            ),
        )
        parser.add_argument("--user", default="admin",
                            help="Username used as the FK owner on seeded rows.")
        parser.add_argument("--password", default="1234",
                            help="Password for users stage (default '1234').")
        parser.add_argument("--bom-master", default=str(BOM_MASTER_XLSX))
        parser.add_argument("--bom-template", default=str(BOM_TEMPLATE_XLSX))
        parser.add_argument("--ng-file", default=str(NG_XLSX))
        parser.add_argument("--line-file", default=str(LINE_XLSX))
        parser.add_argument("--scrap-defect", default="Scrap",
                            help="DefectMode.name_th to attach to scrap stage rows.")
        parser.add_argument("--include-zero-scrap", action="store_true",
                            help="Insert scrap rows whose Total is 0 (default skips).")
        parser.add_argument("--reset-bom-items", action="store_true",
                            help="Delete existing BillOfMaterialItemMater rows before bom stage.")
        parser.add_argument("--dry-run", action="store_true")

    # ------------------------------------------------------------ entrypoint

    def handle(self, *args, **options):
        stages = self._resolve_stages(options["stages"])
        dry_run = bool(options["dry_run"])
        username = options["user"]
        password = options["password"]

        self.stdout.write(self.style.MIGRATE_HEADING(
            f"seed_data: stages={stages} dry_run={dry_run}"
        ))

        # Stage `users` is responsible for ensuring `admin` exists; later stages
        # need the user object. We resolve it lazily just before the first
        # stage that actually needs it.
        owner_user = None

        def get_owner():
            nonlocal owner_user
            if owner_user is None:
                User = get_user_model()
                try:
                    owner_user = User.objects.get(username=username)
                except User.DoesNotExist as exc:
                    raise CommandError(
                        f"User '{username}' does not exist. Run --stages=users first "
                        f"or pass --user=<existing>."
                    ) from exc
            return owner_user

        # Validate Excel paths only for the stages that will use them.
        bom_master_path = Path(options["bom_master"])
        bom_template_path = Path(options["bom_template"])
        ng_path = Path(options["ng_file"])
        line_path = Path(options["line_file"])

        if "bom" in stages and not bom_master_path.exists():
            raise CommandError(f"BoM master file not found: {bom_master_path}")
        needs_template = bool({"line_process", "line", "item_line"} & set(stages))
        if needs_template and not bom_template_path.exists():
            raise CommandError(f"BoM template not found: {bom_template_path}")
        needs_ng = bool({"line", "defect_mode", "defect_by_category"} & set(stages))
        if needs_ng and not ng_path.exists():
            raise CommandError(f"NG file not found: {ng_path}")
        needs_line_xlsx = bool({"item_line_extra", "scrap"} & set(stages))
        if needs_line_xlsx and not line_path.exists():
            raise CommandError(f"line.xlsx not found: {line_path}")

        # Run inside a single transaction so partial seeds don't leave the DB
        # in a half-baked state. (`dry_run` short-circuits writes inside each
        # stage method.)
        with transaction.atomic():
            # ---- 1. groups
            if "groups" in stages:
                self._stage_groups(dry_run)

            # ---- 2. users
            if "users" in stages:
                self._stage_users(password, dry_run)

            # ---- 3. lookups (categories, portions, sides, inouts, ways, plants,
            #              processes, item_stages — replaces old migrations 0030 & 0039)
            if "lookups" in stages:
                self._stage_lookups(get_owner(), dry_run)

            # ---- 4. bom
            if "bom" in stages:
                self._stage_bom(
                    bom_master_path, get_owner(), dry_run,
                    reset_items=bool(options["reset_bom_items"]),
                )

            # ---- 5. bom_classify (replaces old migrations 0032/0033/0034)
            if "bom_classify" in stages:
                self._stage_bom_classify(dry_run)

            # ---- 6-11. master data (Excel-driven)
            lp_map: dict[str, LineProcess] = {}
            cat_map: dict[str, ItemCategory] = {}
            line_map: dict[str, Line] = {}
            dm_map: dict[str, DefectMode] = {}
            ng_data: dict[str, dict] = {}

            need_ng_parse = bool({"line", "defect_mode", "defect_by_category"} & set(stages))
            if need_ng_parse:
                ng_data = self._read_ng_file(ng_path)

            if "line_process" in stages:
                lp_map = self._stage_line_process(bom_template_path, get_owner(), dry_run)
            else:
                lp_map = {p.name: p for p in LineProcess.objects.all()}

            cat_map = {c.name.strip().lower(): c for c in ItemCategory.objects.all()}

            if "line" in stages:
                line_map = self._stage_line(
                    bom_template_path, ng_data, lp_map, get_owner(), dry_run
                )
            else:
                line_map = {l.line_name: l for l in Line.objects.select_related("line_process").all()}

            if "defect_mode" in stages:
                dm_map = self._stage_defect_mode(ng_data, get_owner(), dry_run)
            else:
                dm_map = {d.name_th: d for d in DefectMode.objects.all()}

            if "defect_by_category" in stages:
                self._stage_defect_by_category(ng_data, cat_map, dm_map, get_owner(), dry_run)

            if "item_line" in stages:
                self._stage_item_line(bom_template_path, line_map, get_owner(), dry_run)

            if "item_line_extra" in stages:
                self._stage_item_line_extra(line_path, line_map, get_owner(), dry_run)
                self._stage_propagate_variants(line_map, get_owner(), dry_run)

            # ---- 11. scrap
            if "scrap" in stages:
                self._stage_scrap(
                    line_path, options["scrap_defect"],
                    bool(options["include_zero_scrap"]),
                    get_owner(), dry_run,
                )

        if dry_run:
            self.stdout.write(self.style.WARNING("Dry-run complete (no DB writes)."))
        else:
            self.stdout.write(self.style.SUCCESS("seed_data done."))

    # ------------------------------------------------------------ stages

    def _resolve_stages(self, raw: str) -> list[str]:
        if raw in STAGE_ALIASES:
            return STAGE_ALIASES[raw]
        wanted = [s.strip() for s in raw.split(",") if s.strip()]
        unknown = [s for s in wanted if s not in ALL_STAGES and s not in STAGE_ALIASES]
        if unknown:
            raise CommandError(f"Unknown stages: {unknown}. Valid: {ALL_STAGES + list(STAGE_ALIASES)}")
        # Expand aliases that may appear in the list.
        out: list[str] = []
        for s in wanted:
            if s in STAGE_ALIASES:
                out.extend(STAGE_ALIASES[s])
            else:
                out.append(s)
        # Preserve canonical order, deduped.
        seen: set[str] = set()
        ordered: list[str] = []
        for s in ALL_STAGES:
            if s in out and s not in seen:
                ordered.append(s)
                seen.add(s)
        return ordered

    # -- 1. groups
    def _stage_groups(self, dry_run: bool) -> None:
        self.stdout.write(self.style.MIGRATE_LABEL("[groups]"))
        if dry_run:
            self.stdout.write(f"  would ensure groups: {DEFAULT_GROUPS}")
            return
        created = 0
        for name in DEFAULT_GROUPS:
            _, was_created = Group.objects.get_or_create(name=name)
            if was_created:
                created += 1
        self.stdout.write(self.style.SUCCESS(f"  groups ensured (created={created})"))

    # -- 2. users
    def _stage_users(self, password: str, dry_run: bool) -> None:
        self.stdout.write(self.style.MIGRATE_LABEL("[users]"))
        if dry_run:
            self.stdout.write(f"  would ensure users: {[u for u, *_ in DEFAULT_USERS]}")
            return
        User = get_user_model()
        for username, is_staff, is_superuser in DEFAULT_USERS:
            user, created = User.objects.get_or_create(
                username=username,
                defaults={"is_active": True},
            )
            user.set_password(password)
            user.is_staff = is_staff
            user.is_superuser = is_superuser
            user.is_active = True
            user.save()
            UserProfile.objects.get_or_create(
                user=user,
                defaults={"display_name": username},
            )
            tag = "created" if created else "updated"
            self.stdout.write(f"  {username}: {tag} (staff={is_staff} super={is_superuser})")

    # -- 3. bom
    def _stage_bom(self, path: Path, user, dry_run: bool, reset_items: bool) -> None:
        self.stdout.write(self.style.MIGRATE_LABEL("[bom]"))
        items, edges, stats = self._parse_bom_master(path)
        self.stdout.write(
            f"  parsed: rows={stats['rows']} items={len(items)} edges={len(edges)} "
            f"parents={stats['parents']} skipped_unmatched_m2m={stats['skipped_m2m']}"
        )
        if dry_run:
            return

        sdpn_to_item: dict[str, Item_list] = {}
        for sdpn, data in items.items():
            obj, _ = Item_list.objects.update_or_create(
                sd_code=sdpn,
                defaults={
                    "part_number": data["part_number"],
                    "part_name": data["part_name"],
                    "sku": data["sku"],
                    "user": user,
                },
            )
            sdpn_to_item[sdpn] = obj

        parent_sdpns = {p for p, _, _ in edges}
        sdpn_to_bom: dict[str, BillOfMaterial] = {}
        for sdpn in parent_sdpns:
            item = sdpn_to_item[sdpn]
            bom, _ = BillOfMaterial.objects.update_or_create(
                item=item,
                defaults={
                    "revision": "A",
                    "latest_eci": "",
                    "scrap_percent": Decimal("0"),
                    "user": user,
                },
            )
            sdpn_to_bom[sdpn] = bom

        if reset_items and sdpn_to_bom:
            deleted, _ = BillOfMaterialItemMater.objects.filter(
                bom__in=sdpn_to_bom.values()
            ).delete()
            self.stdout.write(f"  reset BillOfMaterialItemMater rows: {deleted}")

        existing_pairs: set[tuple] = set()
        if not reset_items and sdpn_to_bom:
            existing_pairs = set(
                BillOfMaterialItemMater.objects.filter(
                    bom__in=sdpn_to_bom.values()
                ).values_list("bom_id", "component_id")
            )

        seq_by_parent: dict[str, int] = {}
        new_rows: list[BillOfMaterialItemMater] = []
        for parent_sdpn, child_sdpn, qty in edges:
            bom = sdpn_to_bom[parent_sdpn]
            comp = sdpn_to_item[child_sdpn]
            if (bom.id, comp.id) in existing_pairs:
                continue
            seq_by_parent[parent_sdpn] = seq_by_parent.get(parent_sdpn, 0) + 1
            new_rows.append(BillOfMaterialItemMater(
                bom=bom,
                component=comp,
                quantity=qty,
                unit="PC",
                sequence=seq_by_parent[parent_sdpn],
                user=user,
            ))
        BillOfMaterialItemMater.objects.bulk_create(new_rows, batch_size=500)
        self.stdout.write(self.style.SUCCESS(
            f"  items={len(sdpn_to_item)} boms={len(sdpn_to_bom)} bom_items_inserted={len(new_rows)}"
        ))

    # -- 4. line_process
    def _stage_line_process(self, path: Path, user, dry_run: bool) -> dict[str, LineProcess]:
        self.stdout.write(self.style.MIGRATE_LABEL("[line_process]"))
        rows = self._read_line_processes(path)
        self.stdout.write(f"  parsed: {len(rows)}")
        out: dict[str, LineProcess] = {}
        if dry_run:
            return {p.name: p for p in LineProcess.objects.all()}
        for name, display in rows:
            obj, _ = LineProcess.objects.update_or_create(
                name=name,
                defaults={"display_name": display, "user": user},
            )
            out[name] = obj
        return out

    # -- 3. lookups (formerly migrations 0030 + 0039)
    def _stage_lookups(self, user, dry_run: bool) -> None:
        self.stdout.write(self.style.MIGRATE_LABEL("[lookups]"))
        plan = {
            "ItemCategory": len(LOOKUP_CATEGORIES),
            "Portion": len(LOOKUP_PORTIONS),
            "Side": len(LOOKUP_SIDES),
            "InOut": len(LOOKUP_INOUTS),
            "Way": len(LOOKUP_WAYS),
            "Plant": len(LOOKUP_PLANTS),
            "Process": len(LOOKUP_PROCESSES),
            "ItemStage": len(LOOKUP_ITEM_STAGES),
        }
        self.stdout.write(f"  ensuring lookup rows: {plan}")
        if dry_run:
            return

        for name in LOOKUP_CATEGORIES:
            desc = CATEGORY_DESCRIPTIONS.get(name, "")
            ItemCategory.objects.update_or_create(
                name=name,
                defaults={"description": desc, "user": user},
            )
        for title in LOOKUP_PORTIONS:
            Portion.objects.update_or_create(title=title, defaults={"user": user})
        for title in LOOKUP_SIDES:
            Side.objects.update_or_create(title=title, defaults={"user": user})
        for title in LOOKUP_INOUTS:
            InOut.objects.update_or_create(title=title, defaults={"user": user})
        for title in LOOKUP_WAYS:
            Way.objects.update_or_create(title=title, defaults={"user": user})
        for title in LOOKUP_PLANTS:
            Plant.objects.update_or_create(title=title, defaults={"user": user})
        for title in LOOKUP_PROCESSES:
            Process.objects.update_or_create(title=title, defaults={"user": user})
        for name, display, prefix in LOOKUP_ITEM_STAGES:
            ItemStage.objects.update_or_create(
                name=name,
                defaults={
                    "display_name": display,
                    "code_prefix": prefix,
                    "user": user,
                },
            )

        # One-time: collapse any legacy uppercase "FG" stage created by an
        # older seed run into the canonical lowercase "fg".
        legacy = ItemStage.objects.filter(name="FG").first()
        canonical = ItemStage.objects.filter(name="fg").first()
        if legacy and canonical and legacy.id != canonical.id:
            from django.db.models import F  # noqa: WPS433
            Item_list.objects.filter(stage=legacy).update(stage=canonical)
            ItemLine.objects.filter(item_stage=legacy).update(item_stage=canonical)
            legacy.delete()
            self.stdout.write(self.style.SUCCESS(
                "  collapsed legacy ItemStage 'FG' -> 'fg'"
            ))

    # -- 5. bom_classify (formerly migrations 0032/0033/0034)
    def _stage_bom_classify(self, dry_run: bool) -> None:
        """Reclassify Item_list.stage by BoM tree depth and renumber item_codes
        with a global counter. Items not present in any BoM are left alone.
        """
        self.stdout.write(self.style.MIGRATE_LABEL("[bom_classify]"))
        if dry_run:
            return

        pairs = list(
            BillOfMaterialItemMater.objects.values_list("component_id", "bom__item_id")
        )
        item_to_parents: dict = {}
        parent_ids: set = set()
        component_ids: set = set()
        for component_id, parent_item_id in pairs:
            if component_id and parent_item_id:
                item_to_parents.setdefault(component_id, set()).add(parent_item_id)
            if parent_item_id:
                parent_ids.add(parent_item_id)
            if component_id:
                component_ids.add(component_id)

        all_in_boms = parent_ids | component_ids
        if not all_in_boms:
            self.stdout.write("  no BoM rows; skipping classify")
            return

        # depth from any FG (top of tree). FG = parent-only items.
        levels: dict = {}
        visiting: set = set()

        def compute_level(item_id):
            if item_id in levels:
                return levels[item_id]
            parents = item_to_parents.get(item_id)
            if not parents:
                levels[item_id] = 0
                return 0
            if item_id in visiting:
                return 0
            visiting.add(item_id)
            try:
                parent_levels = [compute_level(pid) for pid in parents]
            finally:
                visiting.discard(item_id)
            lvl = max(parent_levels) + 1 if parent_levels else 0
            levels[item_id] = lvl
            return lvl

        for item_id in all_in_boms:
            compute_level(item_id)

        max_level = max(levels.values()) if levels else 0
        stages = {s.name: s for s in ItemStage.objects.filter(name__in=["fg", "wip", "raw_mat"])}
        fg_stage = stages.get("fg")
        wip_stage = stages.get("wip")
        raw_stage = stages.get("raw_mat")
        if not (fg_stage and wip_stage and raw_stage):
            self.stdout.write(self.style.WARNING(
                "  base ItemStages missing; run --stages=lookups first"
            ))
            return

        fg_items = [iid for iid, lvl in levels.items() if lvl == 0]
        if max_level > 0:
            raw_items = [iid for iid, lvl in levels.items() if lvl == max_level]
            wip_items = [iid for iid, lvl in levels.items() if 0 < lvl < max_level]
        else:
            raw_items = []
            wip_items = []

        # Wipe stage + item_code on items being reclassified, then assign fresh.
        Item_list.objects.filter(id__in=all_in_boms).update(stage=None, item_code=None)

        # Two-phase global renumber to avoid unique-constraint clashes.
        ordered = list(
            Item_list.objects.exclude(item_code__isnull=True).exclude(item_code="")
            .order_by("created_at", "sd_code", "id")
        )
        for idx, item in enumerate(ordered, start=1):
            item.item_code = f"__pending_{idx}__"
            item.save(update_fields=["item_code", "updated_at"])

        def _format_code(prefix, num):
            return f"{prefix}{num:0{ITEM_CODE_PADDING}d}"

        def assign_group(item_ids, stage):
            if not stage or not item_ids:
                return
            prefix = (stage.code_prefix or "").strip()
            items = list(
                Item_list.objects.filter(id__in=item_ids)
                .order_by("created_at", "sd_code", "id")
            )
            for item in items:
                item.stage = stage
                item.save(update_fields=["stage", "updated_at"])
                if prefix:
                    # Use a global counter based on the existing max (any prefix).
                    pass  # second pass below assigns code

        assign_group(fg_items, fg_stage)
        assign_group(wip_items, wip_stage)
        assign_group(raw_items, raw_stage)

        # Final pass: every item with a stage gets a fresh, globally unique
        # numeric portion based on a single counter.
        all_with_stage = list(
            Item_list.objects.exclude(stage__isnull=True)
            .select_related("stage")
            .order_by("created_at", "sd_code", "id")
        )
        for idx, item in enumerate(all_with_stage, start=1):
            prefix = (item.stage.code_prefix or "").strip()
            new_code = _format_code(prefix, idx) if prefix else None
            if item.item_code != new_code:
                item.item_code = new_code
                item.save(update_fields=["item_code", "updated_at"])

        self.stdout.write(self.style.SUCCESS(
            f"  classified fg={len(fg_items)} wip={len(wip_items)} raw={len(raw_items)} "
            f"renumbered={len(all_with_stage)}"
        ))

    # -- 6. line
    def _stage_line(
        self,
        bom_template: Path,
        ng_data: dict[str, dict],
        lp_map: dict[str, LineProcess],
        user,
        dry_run: bool,
    ) -> dict[str, Line]:
        self.stdout.write(self.style.MIGRATE_LABEL("[line]"))
        bom_lines = self._read_bom_lines(bom_template)

        default_lp = lp_map.get(DEFAULT_LINE_PROCESS)
        if not default_lp and not dry_run:
            default_lp, _ = LineProcess.objects.get_or_create(
                name=DEFAULT_LINE_PROCESS,
                defaults={"display_name": "Assembly", "user": user},
            )
            lp_map[DEFAULT_LINE_PROCESS] = default_lp

        name_to_lp: "OrderedDict[str, str]" = OrderedDict()
        for name, lp in bom_lines:
            name_to_lp.setdefault(name, lp)
        for sheet, data in ng_data.items():
            for line_name in data["lines"]:
                name_to_lp.setdefault(line_name, DEFAULT_LINE_PROCESS)
        for name, _hint in EXTRA_LINES:
            name_to_lp.setdefault(name, DEFAULT_LINE_PROCESS)

        self.stdout.write(f"  upserting {len(name_to_lp)} Line rows")
        out: dict[str, Line] = {}
        if dry_run:
            return {l.line_name: l for l in Line.objects.all()}
        for name, lp_name in name_to_lp.items():
            lp = lp_map.get(lp_name) or default_lp
            obj, _ = Line.objects.update_or_create(
                line_name=name,
                defaults={"line_process": lp, "user": user},
            )
            out[name] = obj
        return out

    # -- 7. defect_mode
    def _stage_defect_mode(self, ng_data: dict[str, dict], user, dry_run: bool) -> dict[str, DefectMode]:
        self.stdout.write(self.style.MIGRATE_LABEL("[defect_mode]"))
        all_defects: "OrderedDict[str, None]" = OrderedDict()
        for sheet, data in ng_data.items():
            for d in data["defects"]:
                all_defects.setdefault(d, None)
        self.stdout.write(f"  unique defects: {len(all_defects)}")
        if dry_run:
            return {d.name_th: d for d in DefectMode.objects.all()}
        out: dict[str, DefectMode] = {}
        for text in all_defects:
            obj, _ = DefectMode.objects.update_or_create(
                name_th=text,
                defaults={
                    "name_en": text,
                    "name_jp": "",
                    "defect_type": DefectMode.DefectType.PROCESS_NG,
                    "user": user,
                },
            )
            out[text] = obj
        return out

    # -- 8. defect_by_category
    def _stage_defect_by_category(
        self,
        ng_data: dict[str, dict],
        cat_map: dict[str, ItemCategory],
        dm_map: dict[str, DefectMode],
        user,
        dry_run: bool,
    ) -> None:
        self.stdout.write(self.style.MIGRATE_LABEL("[defect_by_category]"))
        if dry_run:
            for sheet, data in ng_data.items():
                self.stdout.write(
                    f"  would link sheet={sheet} -> cat={NG_SHEET_TO_CATEGORY[sheet]} "
                    f"defects={len(data['defects'])}"
                )
            return
        created = skipped_no_cat = skipped_no_defect = 0
        for sheet, data in ng_data.items():
            cat = cat_map.get(NG_SHEET_TO_CATEGORY[sheet].lower())
            if not cat:
                skipped_no_cat += len(data["defects"])
                continue
            for defect_text in data["defects"]:
                dm = dm_map.get(defect_text)
                if not dm:
                    skipped_no_defect += 1
                    continue
                _, was_created = DefectByCategory.objects.update_or_create(
                    category=cat,
                    defect_mode=dm,
                    defaults={
                        "title": defect_text[:255],
                        "is_inlist": True,
                        "description": defect_text[:100],
                        "user": user,
                    },
                )
                if was_created:
                    created += 1
        self.stdout.write(self.style.SUCCESS(
            f"  created={created} skipped_no_cat={skipped_no_cat} skipped_no_defect={skipped_no_defect}"
        ))

    # -- 9. item_line
    def _stage_item_line(
        self, bom_template: Path, line_map: dict[str, Line], user, dry_run: bool,
    ) -> None:
        self.stdout.write(self.style.MIGRATE_LABEL("[item_line]"))
        rows = self._read_item_and_line(bom_template)
        self.stdout.write(f"  parsed (forward-filled): {len(rows)}")
        if dry_run:
            return
        stage = self._ensure_stage(user)
        sd_to_item, pn_to_item = self._build_item_lookups()
        existing_pairs = set(ItemLine.objects.values_list("item_id", "line_id"))
        new_rows: list[ItemLine] = []
        not_found_lines: set[str] = set()
        not_found_items = 0
        for line_name, sd, pn, _pname in rows:
            line = line_map.get(line_name)
            if not line:
                not_found_lines.add(line_name)
                continue
            item = sd_to_item.get(sd) if sd else None
            if not item and pn:
                item = pn_to_item.get(pn)
            if not item:
                not_found_items += 1
                continue
            if (item.id, line.id) in existing_pairs:
                continue
            existing_pairs.add((item.id, line.id))
            new_rows.append(ItemLine(item=item, line=line, item_stage=stage, user=user))
        ItemLine.objects.bulk_create(new_rows, batch_size=500, ignore_conflicts=True)
        self.stdout.write(self.style.SUCCESS(
            f"  inserted={len(new_rows)} missing_lines={len(not_found_lines)} "
            f"missing_items={not_found_items}"
        ))

    # -- 10. item_line_extra (line.xlsx)
    def _stage_item_line_extra(
        self, line_path: Path, line_map: dict[str, Line], user, dry_run: bool,
    ) -> None:
        self.stdout.write(self.style.MIGRATE_LABEL("[item_line_extra]"))
        rows = self._read_line_xlsx(line_path)
        self.stdout.write(f"  parsed: {len(rows)}")
        if dry_run or not rows:
            return
        stage = self._ensure_stage(user)
        sd_to_item, pn_to_item = self._build_item_lookups()
        existing_pairs = set(ItemLine.objects.values_list("item_id", "line_id"))

        default_lp = LineProcess.objects.filter(name=DEFAULT_LINE_PROCESS).first()
        if not default_lp:
            default_lp, _ = LineProcess.objects.get_or_create(
                name=DEFAULT_LINE_PROCESS,
                defaults={"display_name": "Assembly", "user": user},
            )

        seen_pair: set[tuple[str, str]] = set()
        new_rows: list[ItemLine] = []
        not_found_lines: set[str] = set()
        not_found_items = 0
        autocreated = 0

        for line_name, sd, pn in rows:
            target_names = LINE_NAME_ALIAS.get(line_name, [line_name])
            target_lines: list[Line] = []
            for tname in target_names:
                tl = line_map.get(tname)
                if not tl:
                    tl, was_created = Line.objects.get_or_create(
                        line_name=tname,
                        defaults={"line_process": default_lp, "user": user},
                    )
                    line_map[tname] = tl
                    if was_created:
                        autocreated += 1
                target_lines.append(tl)
            if not target_lines:
                not_found_lines.add(line_name)
                continue
            item = sd_to_item.get(sd) if sd else None
            if not item and pn:
                item = pn_to_item.get(pn)
            if not item:
                not_found_items += 1
                continue
            for tl in target_lines:
                key = (str(item.id), str(tl.id))
                if key in seen_pair or (item.id, tl.id) in existing_pairs:
                    continue
                seen_pair.add(key)
                new_rows.append(ItemLine(item=item, line=tl, item_stage=stage, user=user))

        ItemLine.objects.bulk_create(new_rows, batch_size=500, ignore_conflicts=True)
        self.stdout.write(self.style.SUCCESS(
            f"  inserted={len(new_rows)} new_lines_autocreated={autocreated} "
            f"missing_lines={len(not_found_lines)} missing_items={not_found_items}"
        ))

    def _stage_propagate_variants(self, line_map: dict[str, Line], user, dry_run: bool) -> None:
        if dry_run:
            return
        try:
            stage = ItemStage.objects.get(name=ITEM_AND_LINE_DEFAULT_STAGE)
        except ItemStage.DoesNotExist:
            return
        existing_pairs = set(ItemLine.objects.values_list("item_id", "line_id"))
        new_rows: list[ItemLine] = []
        copied: dict[str, int] = {}
        for variant_name, base_name in LINE_VARIANT_BASE.items():
            variant = line_map.get(variant_name)
            base = line_map.get(base_name)
            if not variant or not base:
                continue
            base_items = ItemLine.objects.filter(line=base).values_list("item_id", flat=True)
            count = 0
            for item_id in base_items:
                if (item_id, variant.id) in existing_pairs:
                    continue
                existing_pairs.add((item_id, variant.id))
                new_rows.append(ItemLine(
                    item_id=item_id, line=variant, item_stage=stage, user=user,
                ))
                count += 1
            if count:
                copied[variant_name] = count
        ItemLine.objects.bulk_create(new_rows, batch_size=500, ignore_conflicts=True)
        if new_rows:
            self.stdout.write(self.style.SUCCESS(
                f"  variant copy inserted={len(new_rows)} variants={copied}"
            ))

    # -- 11. scrap
    def _stage_scrap(
        self, line_path: Path, defect_name: str, include_zero: bool,
        user, dry_run: bool,
    ) -> None:
        self.stdout.write(self.style.MIGRATE_LABEL("[scrap]"))
        rows = self._read_scrap_rows(line_path, include_zero)
        self.stdout.write(f"  candidate rows: {len(rows)}")
        if dry_run:
            return
        defect, _ = DefectMode.objects.get_or_create(
            name_th=defect_name,
            defaults={
                "name_en": defect_name, "name_jp": "",
                "defect_type": DefectMode.DefectType.PROCESS_NG, "user": user,
            },
        )
        line_map = {l.line_name.strip(): l for l in Line.objects.all()}
        sd_to_item, pn_to_item = self._build_item_lookups()
        new_records: list[ScrapRecord] = []
        skipped_no_line = skipped_no_item = 0
        for line_name, sd, pn, total in rows:
            line = line_map.get(line_name)
            if not line:
                skipped_no_line += 1
                continue
            item = sd_to_item.get(sd) if sd else None
            if not item and pn:
                item = pn_to_item.get(pn)
            if not item:
                skipped_no_item += 1
                continue
            new_records.append(ScrapRecord(
                production_line=line,
                part_number=item,
                defect_mode=defect,
                component_part=item,
                quantity=max(total, 0) or 1,
                comment=f"Imported from {line_path.name} (Total={total})",
                created_by=user,
            ))
        ScrapRecord.objects.bulk_create(new_records, batch_size=500)
        self.stdout.write(self.style.SUCCESS(
            f"  inserted={len(new_records)} skipped_no_line={skipped_no_line} "
            f"skipped_no_item={skipped_no_item}"
        ))

    # ------------------------------------------------------------ readers

    def _parse_bom_master(self, path: Path):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if BOM_SHEET not in wb.sheetnames:
            raise CommandError(f"Sheet '{BOM_SHEET}' not found in {path.name}")
        ws = wb[BOM_SHEET]
        items: "OrderedDict[str, dict]" = OrderedDict()
        raw_edges: list[tuple[str, str, Decimal]] = []
        rows_read = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) <= COL_PNAME:
                continue
            # COL_SDPN (col N) is the composite "SD-PartNumber" identifier used as
            # an in-memory join key for parent/child edges. The actual sd_code
            # written to Item_list comes from COL_SD (col P), which holds the
            # bare SD code (e.g., "DAR-56") without the part-number suffix.
            sdpn = _clean(row[COL_SDPN])
            if not sdpn:
                continue
            rows_read += 1
            if sdpn not in items:
                sd_only = _clean(row[COL_SD]) if len(row) > COL_SD else ""
                part_no = _clean(row[COL_PN])
                items[sdpn] = {
                    "sd_code": sd_only or sdpn,
                    "part_number": part_no,
                    "part_name": _clean(row[COL_PNAME]),
                    "sku": part_no or sd_only or sdpn,
                }
            m2m = _clean(row[COL_M2M])
            if m2m:
                raw_edges.append((m2m, sdpn, _to_decimal(row[COL_QTY])))
        wb.close()
        seen: set[tuple[str, str]] = set()
        edges: list[tuple[str, str, Decimal]] = []
        skipped_m2m = 0
        for parent, child, qty in raw_edges:
            if parent not in items:
                skipped_m2m += 1
                continue
            key = (parent, child)
            if key in seen:
                continue
            seen.add(key)
            edges.append((parent, child, qty))
        stats = {
            "rows": rows_read,
            "parents": len({p for p, _, _ in edges}),
            "skipped_m2m": skipped_m2m,
        }
        return items, edges, stats

    def _read_line_processes(self, path: Path) -> list[tuple[str, str]]:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb["lineProcess"]
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = _clean(row[1] if len(row) > 1 else None)
            display = _clean(row[2] if len(row) > 2 else None) or name
            if name:
                out.append((name, display))
        wb.close()
        return out

    def _read_bom_lines(self, path: Path) -> list[tuple[str, str]]:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb["Line"]
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = _clean(row[1] if len(row) > 1 else None)
            lp = _clean(row[2] if len(row) > 2 else None) or DEFAULT_LINE_PROCESS
            if name:
                out.append((name, lp))
        wb.close()
        return out

    def _read_ng_file(self, path: Path) -> dict[str, dict]:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        out: dict[str, dict] = {}
        for sheet in NG_SHEET_TO_CATEGORY:
            if sheet not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"NG sheet '{sheet}' missing; skipping."))
                continue
            ws = wb[sheet]
            defects: list[str] = []
            lines: list[str] = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                problem = _clean(row[2] if len(row) > 2 else None)
                if problem and not problem.lower().startswith("test"):
                    defects.append(problem)
                line_val = row[4] if len(row) > 4 else None
                line_name = _clean(line_val)
                if line_name and line_name.upper() != "LINE  NO":
                    lines.append(line_name)
            out[sheet] = {
                "defects": list(OrderedDict.fromkeys(defects)),
                "lines": list(OrderedDict.fromkeys(lines)),
            }
        wb.close()
        return out

    def _read_item_and_line(self, path: Path) -> list[tuple[str, str, str, str]]:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb["item_and_line"]
        out: list[tuple[str, str, str, str]] = []
        current_line = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            line_cell = _clean(row[0] if len(row) > 0 else None)
            if line_cell:
                current_line = line_cell
            sd = _clean(row[1] if len(row) > 1 else None)
            pn = _clean(row[2] if len(row) > 2 else None)
            pname = _clean(row[3] if len(row) > 3 else None)
            if not (sd or pn) or not current_line:
                continue
            out.append((current_line, sd, pn, pname))
        wb.close()
        return out

    def _read_line_xlsx(self, path: Path) -> list[tuple[str, str, str]]:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet = "Append2" if "Append2" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        out: list[tuple[str, str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) < 3:
                continue
            line_name = _clean(row[0])
            sd = _clean(row[1])
            pn = _clean(row[2])
            if not line_name or not (sd or pn):
                continue
            out.append((line_name, sd, pn))
        wb.close()
        return out

    def _read_scrap_rows(self, path: Path, include_zero: bool) -> list[tuple[str, str, str, int]]:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet = "Append2" if "Append2" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]
        out: list[tuple[str, str, str, int]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) < 5:
                continue
            line_name = _clean(row[0])
            sd = _clean(row[1])
            pn = _clean(row[2])
            total = _to_int(row[35] if len(row) > 35 else None)
            if not line_name or not (sd or pn):
                continue
            if not include_zero and total <= 0:
                continue
            out.append((line_name, sd, pn, total))
        wb.close()
        return out

    # ------------------------------------------------------------ misc helpers

    def _ensure_stage(self, user) -> ItemStage:
        try:
            return ItemStage.objects.get(name=ITEM_AND_LINE_DEFAULT_STAGE)
        except ItemStage.DoesNotExist:
            return ItemStage.objects.create(
                name=ITEM_AND_LINE_DEFAULT_STAGE,
                display_name="Finished goods",
                user=user,
            )

    def _build_item_lookups(self) -> tuple[dict[str, Item_list], dict[str, Item_list]]:
        sd_to_item: dict[str, Item_list] = {}
        for it in Item_list.objects.exclude(sd_code="").only("id", "sd_code"):
            sd_to_item.setdefault(it.sd_code.strip(), it)
        pn_to_item: dict[str, Item_list] = {}
        for it in Item_list.objects.exclude(part_number="").only("id", "part_number"):
            pn_to_item.setdefault(it.part_number.strip(), it)
        return sd_to_item, pn_to_item

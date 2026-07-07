"""Rebuild Item_list / BOM / ItemLine from BoM_Update_Master.xlsx.

Policy (agreed with the data owner, 2026-07-06):
  * KEEP untouched: every Item_list whose sd_code matches DAR-<n>/DAL-<n> (15 rows,
    14 sd codes — DAR-46 exists twice and both are PROTECT-referenced), their
    BillOfMaterial headers, their BOM lines and the InspectionItem rows on them,
    plus the component Item_list rows those lines PROTECT-reference.
  * DELETE everything else: non-kept BOM lines, then headers, then Item_list rows.
    Existing ItemLine rows are KEPT (owner's choice) — the sheet only adds missing ones.
  * RELOAD from core/management/BoM_Update_Master.xlsx:
      - 'item list' sheet  -> Item_list (sd/pn/name, Model -> comment) + one
        BillOfMaterial header per row;
      - 'Billofmaterialitemmaster' sheet -> hierarchy from the Level/LevN
        columns -> BillOfMaterialItemMater lines, quantity from "Usage RM";
      - 'ItemLine' sheet   -> ItemLine (item x Line, stage by display_name),
        skipping pairs that already exist;
      - 'Sheet1'           -> Item_list.weight (kg), matched by sd_code.

Data-quality rules (from pre-analysis of the file):
  * A BoM row is valid only when exactly the Lev<Level> column is filled
    (a Level with no filled LevN column is a broken runaway block -> skipped).
  * A row whose parent level is missing re-attaches to the nearest filled
    ancestor (source files shifted some blocks one level down).
  * parent == child (same part number) rows are dropped.
  * Duplicate (parent, child) pairs: the LAST occurrence in the file wins
    (later blocks come from the per-product Format_BOM ... (OK).xlsx files).
  * Spreadsheet error literals (#N/A, #REF!, ...) in Model -> stored blank;
    in sd/pn/name -> row skipped.
  * Existing non-DAR/DAL items matched by part_number are updated in place
    (sd_code / part_name / comment) instead of duplicated.

Defaults to a DRY RUN. Pass --commit to apply everything in one transaction.
"""

import re
from collections import OrderedDict
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl  # type: ignore
from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction
from django.utils import timezone

from core.models.bill_of_material import BillOfMaterial
from core.models.bill_of_material_item_master import BillOfMaterialItemMater
from core.models.item_line import ItemLine
from core.models.item_list import Item_list, is_spreadsheet_error
from core.models.item_stage import ItemStage
from core.models.line import Line

DARDAL_REGEX = r"^(DAR|DAL)-?[0-9]+"


def norm(v) -> str:
    return "" if v is None else str(v).strip()


def pad(row, width):
    """read_only worksheets trim trailing empty cells — pad back to width."""
    return tuple(row) + (None,) * max(0, width - len(row))


# Field limits: quantity numeric(12,4) -> 8 integer digits; weight numeric(10,2) -> 8.
_QTY_MAX = Decimal("99999999.9999")
_WEIGHT_MAX = Decimal("99999999.99")


def to_qty(v) -> Decimal:
    """Sheet quantity -> Decimal(12,4); the model forces <=0 to 1, so mirror that.

    NaN/Infinity and out-of-range magnitudes fall back to 1 so one bad cell can
    never raise inside the import transaction.
    """
    try:
        d = Decimal(str(v))
        if not d.is_finite() or d <= 0 or d > _QTY_MAX:
            d = Decimal("1")
        return d.quantize(Decimal("0.0001"))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("1")


def to_weight(v):
    """Sheet weight -> Decimal(10,2) kg, or None if not a positive in-range number."""
    try:
        d = Decimal(str(v))
        if not d.is_finite() or d <= 0 or d > _WEIGHT_MAX:
            return None
        return d.quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return None


class Command(BaseCommand):
    help = "Wipe non-DAR/DAL Item_list + BOM and reload from BoM_Update_Master.xlsx."

    def add_arguments(self, parser):
        default_xlsx = Path(__file__).resolve().parents[1] / "BoM_Update_Master.xlsx"
        parser.add_argument("--xlsx", default=str(default_xlsx))
        parser.add_argument("--user-id", default="1",
                            help="core_user.id to own all created rows (default 1).")
        parser.add_argument("--report-dir", default="backups")
        parser.add_argument("--commit", action="store_true",
                            help="Apply changes. Without it the command only reports (dry run).")

    # ------------------------------------------------------------------ #
    # sheet parsing                                                      #
    # ------------------------------------------------------------------ #
    def load_workbook(self, path):
        if not Path(path).exists():
            raise CommandError(f"xlsx not found: {path}")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sheet in ("item list", "Billofmaterialitemmaster", "ItemLine", "Sheet1"):
            if sheet not in wb.sheetnames:
                raise CommandError(f"sheet {sheet!r} not in {wb.sheetnames}")
        return wb

    def read_item_list(self, wb, notes):
        """'item list' sheet -> {pn: {sd, name, model}} (pn unique in the sheet)."""
        ws = wb["item list"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = next((i for i, r in enumerate(rows) if norm(r[0]) == "SD Code"), None)
        if hdr is None:
            raise CommandError("'item list' sheet: header row with 'SD Code' not found")
        items = OrderedDict()
        model_errs = skipped = dup_pn = 0
        for row in rows[hdr + 1:]:
            sd, pn, name, model = (norm(c) for c in pad(row, 4)[:4])
            if not sd and not pn and not name:
                continue
            if is_spreadsheet_error(model):
                model_errs += 1
                model = ""
            if any(is_spreadsheet_error(x) for x in (sd, pn, name)) or not pn:
                skipped += 1
                notes.append(f"item list: skipped row sd={sd!r} pn={pn!r} name={name!r}")
                continue
            if pn in items:
                dup_pn += 1
                notes.append(f"item list: duplicate part_number {pn!r} — first row kept")
                continue
            items[pn] = {"sd": sd, "name": name, "model": model}
        self.stdout.write(f"item list sheet: {len(items)} items "
                          f"(Model errors blanked: {model_errs}, skipped rows: {skipped}, "
                          f"dup pn: {dup_pn})")
        return items

    def read_bom(self, wb, notes):
        """'BoM Update Master' sheet -> per-parent ordered children + pn meta.

        Returns (children, meta):
          children: {parent_pn: OrderedDict{child_pn: usage_raw}}   (last usage wins)
          meta:     {pn: {sd, name}}  first non-empty per pn
          roots:    [pn] level-0 assemblies in sheet order (deduped)
        """
        ws = wb["Billofmaterialitemmaster"]
        rows = list(ws.iter_rows(values_only=True))
        meta = {}
        children = OrderedDict()
        roots = []
        stats = {"garbage": 0, "reattached": 0, "selfref": 0, "err": 0,
                 "dup_pair": 0, "usage_conflict": 0}
        stack = {}  # level -> pn
        for rn, row in enumerate(rows[1:], start=2):
            row = pad(row, 16)
            lev_cell = row[11]
            sd, pn, name = norm(row[12]), norm(row[13]), norm(row[14])
            usage = row[15]
            if not sd and not pn and lev_cell is None:
                continue
            if any(is_spreadsheet_error(x) for x in (sd, pn, name, norm(usage))) or not pn:
                stats["err"] += 1
                notes.append(f"BoM row {rn}: skipped (error value / empty pn) sd={sd!r} pn={pn!r}")
                continue
            filled = [j for j in range(11) if norm(row[j]) != ""]
            try:
                lev = int(str(lev_cell).strip())
            except (ValueError, TypeError):
                lev = None
            if lev is None and len(filled) == 1:
                lev = filled[0]
            if lev is None or filled != [lev]:
                stats["garbage"] += 1
                notes.append(f"BoM row {rn}: skipped garbage (Level={lev_cell!r}, "
                             f"filled LevN={filled}) sd={sd!r} pn={pn!r}")
                continue
            if pn not in meta:
                meta[pn] = {"sd": sd, "name": name}
            elif not meta[pn]["name"] and name:
                meta[pn]["name"] = name
            if lev == 0:
                if pn not in children:
                    roots.append(pn)
                stack = {0: pn}
                children.setdefault(pn, OrderedDict())
                continue
            parent = stack.get(lev - 1)
            if parent is None:
                anc = [k for k in stack if k < lev]
                if not anc:
                    stats["garbage"] += 1
                    notes.append(f"BoM row {rn}: no ancestor at all — skipped pn={pn!r}")
                    continue
                parent = stack[max(anc)]
                stats["reattached"] += 1
            if parent == pn:
                stats["selfref"] += 1
            else:
                kids = children.setdefault(parent, OrderedDict())
                if pn in kids:
                    stats["dup_pair"] += 1
                    if norm(kids[pn]) != norm(usage):
                        stats["usage_conflict"] += 1
                        notes.append(f"BoM: usage conflict {parent} -> {pn}: "
                                     f"{kids[pn]!r} vs {usage!r} (kept later value)")
                    kids[pn] = usage  # last occurrence wins
                else:
                    kids[pn] = usage
            stack[lev] = pn
            for k in [k for k in stack if k > lev]:
                del stack[k]
        n_pairs = sum(len(v) for v in children.values())
        self.stdout.write(
            f"BoM sheet: {len(roots)} root assemblies, {len(children)} parents, "
            f"{n_pairs} unique (parent,child) lines "
            f"(garbage rows: {stats['garbage']}, reattached orphans: {stats['reattached']}, "
            f"self-ref dropped: {stats['selfref']}, dup pairs merged: {stats['dup_pair']}, "
            f"usage conflicts (last wins): {stats['usage_conflict']}, errors: {stats['err']})")
        return children, meta, roots

    def read_main_part(self, wb, notes):
        """'ItemLine' sheet -> [(pn, sd, name, line_name, stage_display)]."""
        ws = wb["ItemLine"]
        rows = list(ws.iter_rows(values_only=True))
        out = []
        skipped = 0
        for rn, row in enumerate(rows[1:], start=2):
            line, sd, pn, name, stage = (norm(c) for c in pad(row, 5)[:5])
            if not any((line, sd, pn, name, stage)):
                continue
            if any(is_spreadsheet_error(x) for x in (line, sd, pn, name, stage)) \
                    or not pn or not line or not stage:
                skipped += 1
                notes.append(f"ItemLine row {rn}: skipped sd={sd!r} pn={pn!r} "
                             f"line={line!r} stage={stage!r}")
                continue
            out.append((pn, sd, name, line, stage))
        self.stdout.write(f"ItemLine sheet: {len(out)} rows (skipped: {skipped})")
        return out

    def read_weights(self, wb, notes):
        """'Sheet1' -> {sd_code: Decimal weight (kg)} (first positive value per sd)."""
        ws = wb["Sheet1"]
        rows = list(ws.iter_rows(values_only=True))
        weights = {}
        conflicts = 0
        for row in rows[1:]:
            sd, _sup, wt = (norm(c) for c in pad(row, 3)[:3])
            if not sd:
                continue
            w = to_weight(wt)
            if w is None:
                continue
            if sd in weights:
                if weights[sd] != w:
                    conflicts += 1
                    notes.append(f"Sheet1: weight conflict for sd={sd!r}: "
                                 f"{weights[sd]} vs {w} (kept first)")
                continue
            weights[sd] = w
        self.stdout.write(f"Sheet1: weights for {len(weights)} sd_codes "
                          f"(conflicts: {conflicts})")
        return weights

    # ------------------------------------------------------------------ #
    def handle(self, *args, **opts):
        commit = opts["commit"]
        user_id = opts["user_id"]
        notes = []  # detailed anomalies -> report file

        wb = self.load_workbook(opts["xlsx"])
        il_items = self.read_item_list(wb, notes)
        children, bom_meta, roots = self.read_bom(wb, notes)
        main_part = self.read_main_part(wb, notes)
        weights = self.read_weights(wb, notes)

        # ---- Line / ItemStage lookups must fully resolve up front ------
        lines_by_name = {norm(n): pk for pk, n in
                         Line.objects.values_list("id", "line_name")}
        stages_by_display = {norm(d): pk for pk, d in
                             ItemStage.objects.values_list("id", "display_name")}
        missing_lines = {l for _pn, _sd, _n, l, _s in main_part} - set(lines_by_name)
        missing_stages = {s for _pn, _sd, _n, _l, s in main_part} - set(stages_by_display)
        if missing_lines or missing_stages:
            raise CommandError(f"unresolved main-part lookups — lines: {sorted(missing_lines)}, "
                               f"stages: {sorted(missing_stages)}")

        # ---- keep-set from the live DB ---------------------------------
        dardal_items = Item_list.objects.filter(sd_code__regex=DARDAL_REGEX)
        dardal_item_ids = set(dardal_items.values_list("id", flat=True))
        dardal_pns = {norm(pn) for pn in dardal_items.values_list("part_number", flat=True)}
        if not dardal_item_ids:
            raise CommandError("No Item_list rows match DAR/DAL — refusing to run "
                               "(wrong database?).")

        keep_headers = set(
            BillOfMaterial.objects.filter(item_id__in=dardal_item_ids)
            .values_list("id", flat=True))
        comp_ids = set(
            BillOfMaterialItemMater.objects.filter(bom_id__in=keep_headers)
            .values_list("component_id", flat=True))
        keep_item_ids = dardal_item_ids | comp_ids

        # ---- safety: nothing outside keep-set may be PROTECT-referenced
        from django.db.models.deletion import PROTECT
        protect_hits = []
        for rel in Item_list._meta.related_objects:
            if getattr(rel, "on_delete", None) is not PROTECT:
                continue
            if rel.related_model is BillOfMaterialItemMater:
                continue  # we delete non-kept lines ourselves before Item_list
            col = f"{rel.field.name}_id"
            model = rel.related_model
            try:
                bad = (model.objects
                       .filter(**{f"{col}__isnull": False})
                       .exclude(**{f"{col}__in": keep_item_ids})
                       .count())
            except Exception as exc:
                self.stdout.write(self.style.WARNING(
                    f"  (skipped PROTECT check on {model.__name__}.{rel.field.name}: {exc})"))
                continue
            if bad:
                protect_hits.append(f"{model.__name__}.{rel.field.name}={bad}")
        if protect_hits:
            raise CommandError(
                "PROTECT references to Item_list rows outside the keep-set — aborting "
                "(no changes made). Offenders: " + ", ".join(protect_hits))

        # ---- build the reload plan --------------------------------------
        # item universe: item-list sheet first, then BoM-sheet-only pns,
        # then main-part-only pns. sd/name priority follows that order.
        universe = OrderedDict()
        for pn, d in il_items.items():
            universe[pn] = {"sd": d["sd"], "name": d["name"], "model": d["model"]}
        for pn, m in bom_meta.items():
            if pn not in universe:
                universe[pn] = {"sd": m["sd"], "name": m["name"], "model": ""}
        for pn, sd, name, _l, _s in main_part:
            if pn not in universe:
                universe[pn] = {"sd": sd, "name": name, "model": ""}
        # weight (kg) matched by sd_code from Sheet1
        for pn, d in universe.items():
            d["weight"] = weights.get(d["sd"])

        # parents keyed off the BoM sheet must exist even if not in 'item list'
        header_pns = [pn for pn in il_items] + \
                     [p for p in children if p not in il_items]
        # never touch DAR/DAL assemblies: their headers/lines stay as-is
        header_pns = [p for p in header_pns if p not in dardal_pns]
        skipped_dardal_parents = [p for p in children if p in dardal_pns]
        line_plan = {p: kids for p, kids in children.items() if p not in dardal_pns}
        n_lines_planned = sum(len(k) for k in line_plan.values())

        total_items = Item_list.objects.count()
        del_items = total_items - len(keep_item_ids)
        del_headers = BillOfMaterial.objects.exclude(id__in=keep_headers).count()
        del_lines = BillOfMaterialItemMater.objects.exclude(bom_id__in=keep_headers).count()

        self.stdout.write(self.style.MIGRATE_HEADING("=== PLAN ==="))
        self.stdout.write(
            f"KEEP: {len(dardal_item_ids)} DAR/DAL items, {len(keep_headers)} headers "
            f"(+{BillOfMaterialItemMater.objects.filter(bom_id__in=keep_headers).count()} lines), "
            f"{len(comp_ids)} components; ItemLine rows kept: {ItemLine.objects.count()}")
        self.stdout.write(f"DELETE: Item_list {del_items}, BOM headers {del_headers}, "
                          f"BOM lines {del_lines}")
        self.stdout.write(f"RELOAD: item universe {len(universe)} pns, headers {len(header_pns)}, "
                          f"BOM lines {n_lines_planned} "
                          f"(DAR/DAL parents skipped: {len(skipped_dardal_parents)}), "
                          f"main-part rows {len(main_part)}")

        report_dir = Path(opts["report_dir"])
        report_dir.mkdir(parents=True, exist_ok=True)
        stamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        report_path = report_dir / f"import_bom_update_report_{stamp}.txt"

        if not commit:
            report_path.write_text("\n".join(notes), encoding="utf-8")
            self.stdout.write(f"anomaly details -> {report_path}")
            self.stdout.write(self.style.WARNING(
                "\nDRY RUN — no changes. Re-run with --commit to apply."))
            return

        with transaction.atomic():
            # 1) delete non-kept BOM (lines before headers), then Item_list
            BillOfMaterialItemMater.objects.exclude(bom_id__in=keep_headers).delete()
            BillOfMaterial.objects.exclude(id__in=keep_headers).delete()
            Item_list.objects.exclude(id__in=keep_item_ids).delete()

            # 2) update surviving non-DAR/DAL components in place, create the rest
            survivors = {norm(i.part_number): i for i in Item_list.objects.all()}
            updated = 0
            to_update = []
            for pn, d in universe.items():
                item = survivors.get(pn)
                if item is None or item.id in dardal_item_ids:
                    continue
                new_sd = d["sd"][:32]
                new_name = d["name"][:255]
                new_comment = d["model"][:255]
                new_weight = d["weight"]
                changed = False
                if item.sd_code != new_sd:
                    item.sd_code = new_sd; changed = True
                if item.part_name != new_name:
                    item.part_name = new_name; changed = True
                if new_comment and item.comment != new_comment:
                    item.comment = new_comment; changed = True
                if new_weight is not None and item.weight != new_weight:
                    item.weight = new_weight; changed = True
                if changed:
                    to_update.append(item)
                    updated += 1
            Item_list.objects.bulk_update(
                to_update, ["sd_code", "part_name", "comment", "weight"], batch_size=500)
            self.stdout.write(f"updated {updated} surviving Item_list rows in place")

            new_items = []
            for pn, d in universe.items():
                if pn in survivors:
                    continue
                new_items.append(Item_list(
                    sd_code=d["sd"][:32],
                    part_number=pn[:255],
                    part_name=d["name"][:255],
                    sku="",
                    comment=d["model"][:255],
                    weight=d["weight"] or 0,
                    user_id=user_id,
                ))
            Item_list.objects.bulk_create(new_items, batch_size=1000)
            for i in new_items:
                survivors[norm(i.part_number)] = i
            self.stdout.write(f"created {len(new_items)} Item_list rows")

            # 3) headers: one per item-list row + per non-DAR/DAL BoM parent
            existing_hdr_items = set(BillOfMaterial.objects.values_list("item_id", flat=True))
            new_headers = []
            for pn in header_pns:
                item = survivors[pn]
                if item.id in existing_hdr_items:
                    continue
                existing_hdr_items.add(item.id)
                new_headers.append(BillOfMaterial(
                    item=item, revision="A", latest_eci="", scrap_percent=0,
                    user_id=user_id,
                ))
            BillOfMaterial.objects.bulk_create(new_headers, batch_size=1000)
            header_by_item = {h.item_id: h for h in BillOfMaterial.objects.all()}
            self.stdout.write(f"created {len(new_headers)} BillOfMaterial headers")

            # 4) BOM lines from the deduped hierarchy
            new_lines = []
            for parent_pn, kids in line_plan.items():
                hdr = header_by_item[survivors[parent_pn].id]
                for seq, (child_pn, usage) in enumerate(kids.items(), start=1):
                    new_lines.append(BillOfMaterialItemMater(
                        bom=hdr,
                        component=survivors[child_pn],
                        quantity=to_qty(usage),
                        unit="",
                        sequence=seq,
                        user_id=user_id,
                    ))
            BillOfMaterialItemMater.objects.bulk_create(new_lines, batch_size=2000)
            self.stdout.write(f"created {len(new_lines)} BillOfMaterialItemMater lines")

            # 5) ItemLine: add missing (item, line) pairs, keep existing rows
            existing_pairs = set(ItemLine.objects.values_list("item_id", "line_id"))
            new_ils, il_skipped = [], 0
            for pn, _sd, _name, line_name, stage_disp in main_part:
                item = survivors[pn]
                key = (item.id, lines_by_name[line_name])
                if key in existing_pairs:
                    il_skipped += 1
                    continue
                existing_pairs.add(key)
                new_ils.append(ItemLine(
                    item=item,
                    line_id=lines_by_name[line_name],
                    item_stage_id=stages_by_display[stage_disp],
                    user_id=user_id,
                ))
            ItemLine.objects.bulk_create(new_ils, batch_size=1000)
            self.stdout.write(f"created {len(new_ils)} ItemLine rows "
                              f"(already present / duplicate in sheet: {il_skipped})")

        report_path.write_text("\n".join(notes), encoding="utf-8")
        self.stdout.write(f"anomaly details -> {report_path}")
        self.stdout.write(self.style.SUCCESS("DONE (committed)."))

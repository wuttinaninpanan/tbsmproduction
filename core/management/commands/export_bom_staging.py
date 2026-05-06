"""
export_bom_staging
──────────────────
Export ทุก row ใน BomStaging → Excel ไฟล์เดียว
Layout ตรงกับ BOM Update template:
  Level (1/2/3/4/*) | SD Code | Part No. | Part Name | Usage RM | Process |
  Line No. | Supplier Name | Model | Coil Press | Weight |
  SD_FG | SD Code_Component | Supplier | Line No. |
  (Source File) | (Row)

Usage:
  python manage.py export_bom_staging
  python manage.py export_bom_staging --output=merged.xlsx
  python manage.py export_bom_staging --source-file="Format_BOM_XXX.xlsx"   # filter
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from django.core.management.base import BaseCommand

from core.models.bom_staging import BomStaging


# ─── Column layout (ตาม template ของผู้ใช้) ───
# (header_key, sub_header, width)
COLUMNS = [
    # Level
    ("lv_star", "", 6),
    # Main
    ("sd_code", "", 14),
    ("part_no", "", 18),
    ("part_name", "", 40),
    ("usage_rm", "", 10),
    ("process", "", 12),
    ("line_no", "", 10),
    ("supplier_name", "", 18),
    ("model", "", 10),
    # Derived (blue)
    ("coil_press", "", 10),
    ("weight", "", 10),
    # Derived (yellow/cyan)
    ("sd_fg", "", 12),
    ("sd_code_component", "", 14),
    ("supplier", "", 18),
    ("line_no_final", "", 10),
    # Metadata
    ("source_file", "", 35),
    ("row_index", "", 7),
]

HEADER_GROUPS = [
    # (label, start_col_index, end_col_index, fill_color)
    ("Level", 1, 1, "FCE4D6"),           # orange
    ("SD Code", 2, 2, "A9D08E"),         # green
    ("Part No.", 3, 3, "A9D08E"),
    ("Part Name", 4, 4, "A9D08E"),
    ("Usage RM", 5, 5, "A9D08E"),
    ("Process", 6, 6, "A9D08E"),
    ("Line No.", 7, 7, "A9D08E"),
    ("Supplier Name", 8, 8, "A9D08E"),
    ("Model", 9, 9, "A9D08E"),
    ("Coil Press", 10, 10, "BDD7EE"),    # blue
    ("Weight", 11, 11, "BDD7EE"),
    ("SD_FG", 12, 12, "FFE699"),          # yellow
    ("SD Code_Component", 13, 13, "BDD7EE"),  # blue
    ("Supplier", 14, 14, "FFE699"),       # yellow
    ("Line No.", 15, 15, "BDD7EE"),       # blue
    ("Source File", 16, 16, "D9D9D9"),    # grey
    ("Row", 17, 17, "D9D9D9"),
]


def _border():
    side = Side(border_style="thin", color="808080")
    return Border(left=side, right=side, top=side, bottom=side)


class Command(BaseCommand):
    help = "Export BomStaging ทั้งหมด → Excel ไฟล์เดียว (Merged BOM)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output", default="bom_staging_merged.xlsx",
            help="Output file path (default: bom_staging_merged.xlsx)",
        )
        parser.add_argument(
            "--source-file", default=None,
            help="กรองเฉพาะ row ของ source file ที่ระบุ (icontains)",
        )

    def handle(self, *args, **options):
        output_path = Path(options["output"])
        source_filter = options["source_file"]

        qs = BomStaging.objects.all()
        if source_filter:
            qs = qs.filter(source_file__icontains=source_filter)
        qs = qs.order_by("source_file", "row_index")

        total = qs.count()
        if total == 0:
            self.stdout.write(self.style.WARNING("⚠ ไม่มีข้อมูลใน BomStaging"))
            return

        self.stdout.write(f"📊 จะ export {total:,} rows → {output_path}")

        # ─── สร้าง workbook ───
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM Merged"

        bold = Font(bold=True, size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = _border()

        # ─── Row 1: Group headers ───
        for label, start_col, end_col, color in HEADER_GROUPS:
            cell = ws.cell(row=1, column=start_col, value=label)
            cell.font = bold
            cell.alignment = center
            cell.fill = PatternFill("solid", fgColor=color)
            cell.border = border
            if end_col > start_col:
                ws.merge_cells(
                    start_row=1, start_column=start_col,
                    end_row=1, end_column=end_col,
                )
                # Apply style to merged area
                for c in range(start_col + 1, end_col + 1):
                    cc = ws.cell(row=1, column=c)
                    cc.fill = PatternFill("solid", fgColor=color)
                    cc.border = border

        # ─── Row 2: Sub-headers (Level 1/2/3/4/*) ───
        for col_idx, (field, sub, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx)
            if sub:
                cell.value = sub
            cell.font = bold
            cell.alignment = center
            cell.border = border
            # Level column uses orange fill
            if field == "lv_star":
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header rows + first few cols
        ws.freeze_panes = "F3"

        # ─── Data rows ───
        row_pointer = 3
        for obj in qs.iterator(chunk_size=500):
            for col_idx, (field, _sub, _w) in enumerate(COLUMNS, start=1):
                val = getattr(obj, field, "")
                if val is None:
                    val = ""
                # Convert Decimal to float for Excel
                if hasattr(val, "quantize"):
                    val = float(val)
                cell = ws.cell(row=row_pointer, column=col_idx, value=val)
                cell.border = border
                if field in ("part_name",):
                    cell.alignment = left_align
                else:
                    cell.alignment = center
            row_pointer += 1

        # ─── Save ───
        wb.save(output_path)

        self.stdout.write(self.style.SUCCESS(
            f"\n✅ Export สำเร็จ: {total:,} rows → {output_path.resolve()}"
        ))

        # ─── Summary by source file ───
        from django.db.models import Count
        self.stdout.write("\n📋 สรุปตามไฟล์ต้นทาง:")
        summary = (
            BomStaging.objects.values("source_file")
            .annotate(n=Count("id"))
            .order_by("source_file")
        )
        for s in summary:
            self.stdout.write(f"   {s['source_file']:<50} {s['n']:>6,} rows")

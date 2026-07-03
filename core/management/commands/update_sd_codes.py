"""Update Item_list.sd_code from the 'itemlist' sheet of an Excel workbook.

Background: import_bom_master.py rebuilt the non-DAR/DAL Item_list rows with
placeholder sd_codes ("Temp-<n>"), leaving the owner to fill in the real SD
codes later. This command does that fill-in: it reads the real codes from the
spreadsheet and writes them onto the matching Item_list rows.

Matching is by part_number == the sheet's 'Child Parts No.' (exact, trimmed).
Only sd_code is touched; nothing else on the row changes, and no rows are
created or deleted.

The 'itemlist' sheet layout (real header is on the THIRD row, rows 0-1 blank):
    col 0  sd code           <- the value we write
    col 1  Child Parts No.   <- matched against Item_list.part_number
    col 2  Child Parts Name
    ...

Guards:
  * Spreadsheet error literals (#REF! etc.) are never written (see
    core.models.item_list.is_spreadsheet_error).
  * Blank sd codes are skipped (never blanks out an existing code).
  * A sd_code claimed by more than one part_number is reported (sd_code is not
    unique in the schema, so this is allowed but usually worth a look).

Defaults to a DRY RUN. Pass --commit to actually apply (inside one transaction).
"""

from pathlib import Path

import openpyxl  # type: ignore
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models.item_list import Item_list, is_spreadsheet_error

# The real header sits on the third row of the sheet; the two rows above it are
# blank spacers in the source workbook.
HEADER_ROW_INDEX = 2
SD_COL = 0
PART_NUMBER_COL = 1


def norm(v) -> str:
    return "" if v is None else str(v).strip()


class Command(BaseCommand):
    help = "Update Item_list.sd_code from the 'itemlist' sheet of an Excel workbook."

    def add_arguments(self, parser):
        default_xlsx = (
            Path(__file__).resolve().parents[1] / "Item list from BoM Master.xlsx"
        )
        parser.add_argument("--xlsx", default=str(default_xlsx))
        parser.add_argument("--sheet", default="itemlist")
        parser.add_argument(
            "--commit",
            action="store_true",
            help="Apply changes. Without it the command only reports (dry run).",
        )
        parser.add_argument(
            "--show",
            type=int,
            default=30,
            help="How many rows of each detail list to print (default 30).",
        )

    # ---- sheet parsing -------------------------------------------------
    def read_mapping(self, path, sheet):
        if not Path(path).exists():
            raise CommandError(f"xlsx not found: {path}")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            raise CommandError(f"sheet {sheet!r} not in {wb.sheetnames}")
        ws = wb[sheet]

        pn_to_sd = {}          # part_number -> sd_code (last non-blank wins)
        skipped_error = 0      # rows whose sd_code was a spreadsheet error literal
        skipped_blank = 0      # rows with a part_number but no sd_code
        conflicts = {}         # part_number -> {differing sd_codes} within the sheet

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= HEADER_ROW_INDEX:
                continue
            if row is None:
                continue
            sd = norm(row[SD_COL]) if len(row) > SD_COL else ""
            pn = norm(row[PART_NUMBER_COL]) if len(row) > PART_NUMBER_COL else ""
            if not pn:
                continue
            if not sd:
                skipped_blank += 1
                continue
            if is_spreadsheet_error(sd):
                skipped_error += 1
                continue
            if pn in pn_to_sd and pn_to_sd[pn] != sd:
                conflicts.setdefault(pn, {pn_to_sd[pn]}).add(sd)
            pn_to_sd[pn] = sd

        return pn_to_sd, skipped_error, skipped_blank, conflicts

    def handle(self, *args, **opts):
        commit = opts["commit"]
        show = opts["show"]
        pn_to_sd, skipped_error, skipped_blank, conflicts = self.read_mapping(
            opts["xlsx"], opts["sheet"]
        )

        if conflicts:
            # Same part_number given two different codes in the sheet -> the data
            # owner must resolve it; we refuse to silently pick one.
            raise CommandError(
                "The sheet maps the same part_number to different sd_codes; "
                "resolve these before running: "
                + "; ".join(f"{pn}={sorted(v)}" for pn, v in list(conflicts.items())[:20])
            )

        # sd_codes claimed by more than one part_number (allowed, but reported).
        sd_owners = {}
        for pn, sd in pn_to_sd.items():
            sd_owners.setdefault(sd, []).append(pn)
        shared_sd = {sd: pns for sd, pns in sd_owners.items() if len(pns) > 1}

        # ---- diff against the live DB ----------------------------------
        db_items = list(
            Item_list.objects.all().values_list("id", "part_number", "sd_code")
        )
        db_by_pn = {}
        for pk, pn, sd in db_items:
            db_by_pn.setdefault(norm(pn), []).append((pk, norm(sd)))

        to_update = []         # (id, part_number, old_sd, new_sd)
        unchanged = 0          # part matched, sd already correct
        unmatched_in_sheet = []  # sheet part_numbers with no Item_list row
        matched_pns = set()

        for pn, new_sd in pn_to_sd.items():
            rows = db_by_pn.get(pn)
            if not rows:
                unmatched_in_sheet.append((pn, new_sd))
                continue
            matched_pns.add(pn)
            for pk, old_sd in rows:
                if old_sd == new_sd:
                    unchanged += 1
                else:
                    to_update.append((pk, pn, old_sd, new_sd))

        db_pns_not_in_sheet = sorted(
            pn for pn in db_by_pn if pn and pn not in pn_to_sd
        )

        # ---- report -----------------------------------------------------
        h = self.style.MIGRATE_HEADING
        self.stdout.write(h("=== SHEET ==="))
        self.stdout.write(
            f"part_numbers with a real sd_code : {len(pn_to_sd)}\n"
            f"rows skipped (blank sd_code)      : {skipped_blank}\n"
            f"rows skipped (spreadsheet error)  : {skipped_error}\n"
            f"sd_codes shared by >1 part_number : {len(shared_sd)}"
        )
        for sd, pns in list(shared_sd.items())[:show]:
            self.stdout.write(f"    shared sd {sd!r}: {pns}")

        self.stdout.write(h("\n=== DB DIFF ==="))
        self.stdout.write(
            f"Item_list rows total              : {len(db_items)}\n"
            f"sd_code WILL CHANGE               : {len(to_update)}\n"
            f"already correct (no change)       : {unchanged}\n"
            f"sheet part_numbers not in DB      : {len(unmatched_in_sheet)}\n"
            f"DB part_numbers not in sheet      : {len(db_pns_not_in_sheet)}"
        )

        if to_update:
            self.stdout.write(h(f"\n--- changes (first {show}) ---"))
            for pk, pn, old_sd, new_sd in to_update[:show]:
                self.stdout.write(f"    #{pk} {pn}: {old_sd!r} -> {new_sd!r}")
            if len(to_update) > show:
                self.stdout.write(f"    ... and {len(to_update) - show} more")

        if unmatched_in_sheet:
            self.stdout.write(h(f"\n--- sheet part_numbers not in DB (first {show}) ---"))
            for pn, new_sd in unmatched_in_sheet[:show]:
                self.stdout.write(f"    {pn} (sd {new_sd})")
            if len(unmatched_in_sheet) > show:
                self.stdout.write(f"    ... and {len(unmatched_in_sheet) - show} more")

        if not commit:
            self.stdout.write(
                self.style.WARNING(
                    "\nDRY RUN — no changes. Re-run with --commit to apply."
                )
            )
            return

        if not to_update:
            self.stdout.write(self.style.SUCCESS("\nNothing to update — DB already matches."))
            return

        with transaction.atomic():
            objs = []
            for pk, _pn, _old, new_sd in to_update:
                objs.append(Item_list(id=pk, sd_code=new_sd))
            Item_list.objects.bulk_update(objs, ["sd_code"], batch_size=1000)

        self.stdout.write(
            self.style.SUCCESS(f"\nDONE (committed) — updated {len(to_update)} sd_codes.")
        )

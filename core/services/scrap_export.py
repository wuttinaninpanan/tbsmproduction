"""Shared Excel workbook builder for scrap / production exports.

Both the Manage Production page and the Inspection Scrap Dashboard export to the
*same* 4-sheet template so the files look and chart identically:

- ``Production Record``  : one row per produced part / production record.
- ``Scrap``              : one row per scrapped sub-part.
- ``สรุปตาม Line``        : defect qty per Production line + a bar chart.
- ``สรุปตาม Defect mode`` : defect qty per Defect mode + a pie chart.

Callers assemble plain rows/dicts (no model coupling) and pass them in.
"""

from __future__ import annotations

RECORD_HEADERS = [
    "วันที่",
    "วันทำการ",
    "ผู้ใช้งาน",
    "กะ",
    "Production line",
    "Lot no.",
    "SD number",
    "part name",
    "จำนวนผลิต",
    "จำนวนของเสีย",
    "อัตราของเสีย",
]
SCRAP_HEADERS = [
    "วันที่",
    "วันทำการ",
    "ผู้ใช้งาน",
    "กะ",
    "Production line",
    "Lot no.",
    "SD number (ชิ้นส่วนที่ทิ้ง)",
    "part number (ชิ้นส่วนที่ทิ้ง)",
    "part name (ชิ้นส่วนที่ทิ้ง)",
    "จำนวนที่ทิ้ง",
]
INSPECTION_HEADERS = [
    "วันที่",
    "SD number",
    "Production line",
    "QR (เลขงาน)",
    "ผลตรวจ",
]


def build_scrap_workbook(record_rows, scrap_rows, line_totals, defect_totals):
    """Build the 4-sheet workbook. Raises ImportError if openpyxl is missing.

    Args:
        record_rows: iterable of 11-tuples matching ``RECORD_HEADERS``. The last
            value (อัตราของเสีย) should be a fraction (e.g. 0.04) to render as a
            percentage, or ``None`` to print ``-``. ``จำนวนผลิต`` may be ``None``
            (prints ``-``) when the source has no production-quantity concept.
            ``วันทำการ`` (2nd value) may be ``"-"`` for sources without a
            working date (e.g. legacy inspection scrap).
        scrap_rows: iterable of 10-tuples matching ``SCRAP_HEADERS``.
        line_totals: mapping ``{production_line: defect_qty}``.
        defect_totals: mapping ``{defect_mode: defect_qty}``.
    """
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    def set_widths(ws, widths):
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width

    wb = Workbook()

    # ----- Sheet 1: Production Record -----
    ws1 = wb.active
    ws1.title = "Production Record"
    ws1.append(RECORD_HEADERS)
    style_header(ws1)
    row_idx = 1
    for date_str, prod_date_str, user, shift, line, lot, sd, part_name, products_qty, total_defect, rate in record_rows:
        row_idx += 1
        ws1.append(
            [
                date_str,
                prod_date_str,
                user,
                shift,
                line,
                lot,
                sd,
                part_name,
                products_qty if products_qty is not None else "-",
                total_defect,
                rate if rate is not None else "-",
            ]
        )
        if rate is not None:
            ws1.cell(row=row_idx, column=11).number_format = "0.0%"
    set_widths(ws1, [18, 14, 15, 10, 18, 24, 16, 26, 11, 13, 12])

    # ----- Sheet 2: Scrap -----
    ws2 = wb.create_sheet(title="Scrap")
    ws2.append(SCRAP_HEADERS)
    style_header(ws2)
    for row in scrap_rows:
        ws2.append(list(row))
    set_widths(ws2, [18, 14, 15, 10, 18, 24, 24, 24, 26, 12])

    # ----- Sheet 3: Summary + bar chart (defect qty per Production line) -----
    ws3 = wb.create_sheet(title="สรุปตาม Line")
    ws3.append(["Production line", "จำนวนของเสีย"])
    style_header(ws3)
    for name in sorted(line_totals):
        ws3.append([name, line_totals[name]])
    set_widths(ws3, [22, 14])
    n = len(line_totals)
    if n:
        chart = BarChart()
        chart.type = "col"
        chart.title = "จำนวนของเสียตาม Production line"
        chart.y_axis.title = "จำนวนของเสีย"
        chart.x_axis.title = "Production line"
        chart.legend = None
        data = Reference(ws3, min_col=2, min_row=1, max_row=n + 1)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=n + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        ws3.add_chart(chart, "D2")

    # ----- Sheet 4: Summary + pie chart (defect proportion per Defect mode) -----
    ws4 = wb.create_sheet(title="สรุปตาม Defect mode")
    ws4.append(["Defect mode", "จำนวนของเสีย"])
    style_header(ws4)
    for name in sorted(defect_totals):
        ws4.append([name, defect_totals[name]])
    set_widths(ws4, [26, 14])
    m = len(defect_totals)
    if m:
        pie = PieChart()
        pie.title = "สัดส่วนของเสียตาม Defect mode"
        data = Reference(ws4, min_col=2, min_row=1, max_row=m + 1)
        cats = Reference(ws4, min_col=1, min_row=2, max_row=m + 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.height = 10
        pie.width = 20
        ws4.add_chart(pie, "D2")

    return wb


def build_inspection_workbook(result_rows, result_totals):
    """Build a 2-sheet workbook for Inspection Machine results.

    - ``Inspection Result`` : one row per inspection result.
    - ``สรุปผลตรวจ``         : count per result value + a pie chart.

    Args:
        result_rows: iterable of 5-tuples matching ``INSPECTION_HEADERS``
            (date_str, sd_number, production_line, qr_work, result).
        result_totals: mapping ``{result_value: count}``.

    Raises ImportError if openpyxl is missing.
    """
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    def set_widths(ws, widths):
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width

    wb = Workbook()

    # ----- Sheet 1: Inspection Result -----
    ws1 = wb.active
    ws1.title = "Inspection Result"
    ws1.append(INSPECTION_HEADERS)
    style_header(ws1)
    for row in result_rows:
        ws1.append(list(row))
    set_widths(ws1, [20, 18, 18, 22, 14])

    # ----- Sheet 2: Summary + pie chart (count per result value) -----
    ws2 = wb.create_sheet(title="สรุปผลตรวจ")
    ws2.append(["ผลตรวจ", "จำนวน"])
    style_header(ws2)
    for name in sorted(result_totals):
        ws2.append([name or "(ไม่ระบุ)", result_totals[name]])
    set_widths(ws2, [22, 12])
    m = len(result_totals)
    if m:
        pie = PieChart()
        pie.title = "สัดส่วนผลตรวจ"
        data = Reference(ws2, min_col=2, min_row=1, max_row=m + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=m + 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.height = 10
        pie.width = 20
        ws2.add_chart(pie, "D2")

    return wb

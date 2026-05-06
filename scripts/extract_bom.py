"""
Extract and consolidate BOM data from all Excel files in Document/ folder.

Outputs (scripts/bom_output/):
  items.json         : All unique items (→ Item_list)
  relationships.json : All unique parent-child BOM links (→ BillOfMaterial + BOMItemMaster)
  bom_tree.csv       : Human-readable BOM tree for review
  summary.json       : Stats and data quality warnings
"""

import openpyxl
import os
import json
import csv
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import Optional


DOCUMENT_DIR = Path("Document")
OUTPUT_DIR = Path("scripts/bom_output")

# ชื่อ Sheet ที่จะค้นหา (เรียงตาม priority)
BOM_SHEET_CANDIDATES = ["MAS_BOM Update", "BOM Update", "BOM "]


# ─────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────

@dataclass
class BomItem:
    unique_key: str      # sd_code หรือ "PN:<part_no>" กรณี DUM
    sd_code: str
    part_no: str
    part_name: str
    process: str = ""
    line_no: str = ""
    supplier: str = ""
    model: str = ""
    min_level: int = 99  # ระดับต่ำสุด (0=FG) ที่พบ item นี้
    source_files: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


@dataclass
class BomRelationship:
    parent_key: str
    child_key: str
    quantity: float
    process: str = ""
    line_no: str = ""
    source_file: str = ""


# ─────────────────────────────────────────────
# Helper functions
# ─────────────────────────────────────────────

def safe_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in ("#N/A", "N/A", "None", "-") else s


def parse_qty(val) -> float:
    if val is None:
        return 1.0
    if isinstance(val, (int, float)):
        return round(float(val), 6)
    try:
        return round(float(str(val).strip()), 6)
    except (ValueError, TypeError):
        return 1.0


def get_unique_key(sd_code: str, part_no: str) -> Optional[str]:
    """
    SD Code เปรียบเหมือน IP address — ใช้เป็น primary key
    กรณีพิเศษ: SD Code = 'DUM' (placeholder) → ใช้ Part No. แทน
    """
    sd = safe_str(sd_code)
    pn = safe_str(part_no)

    if not sd:
        return f"PN:{pn}" if pn else None

    if sd.upper() == "DUM":
        return f"PN:{pn}" if pn else None

    return sd


# ─────────────────────────────────────────────
# Sheet layout detection
# ─────────────────────────────────────────────

def detect_layout(ws) -> Optional[dict]:
    """
    หา column positions โดยอ่าน header row
    รองรับ 3 แบบ: SD_col=6 / SD_col=7 / SD_col=8 (มี sequence col)
    """
    layout = {
        "header_row": None,
        "data_start_row": None,
        "level_col_start": None,
        "level_col_count": None,
        "sd_col": None,
        "pn_col": None,
        "name_col": None,
        "qty_col": None,
        "process_col": None,
        "line_col": None,
        "supplier_col": None,
        "model_col": None,
    }

    # สแกน 5 row แรกหา header
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        found_sd = False
        for col_idx, val in enumerate(row):
            if not isinstance(val, str):
                continue
            v = val.strip().upper()
            if ("SD" in v and "CODE" in v) and layout["sd_col"] is None:
                layout["sd_col"] = col_idx
                layout["header_row"] = row_idx + 1
                found_sd = True
            elif "PART NO" in v and layout["pn_col"] is None:
                layout["pn_col"] = col_idx
            elif "PART NAME" in v and layout["name_col"] is None:
                layout["name_col"] = col_idx
            elif "USAGE" in v and layout["qty_col"] is None:
                layout["qty_col"] = col_idx
            elif v == "PROCESS" and layout["process_col"] is None:
                layout["process_col"] = col_idx
            elif "LINE NO" in v and layout["line_col"] is None:
                layout["line_col"] = col_idx
            elif "SUPPLIER" in v and layout["supplier_col"] is None:
                layout["supplier_col"] = col_idx
            elif v == "MODEL" and layout["model_col"] is None:
                layout["model_col"] = col_idx
        if found_sd:
            break

    if layout["sd_col"] is None:
        return None

    # หา level_col_start (column ที่มีคำว่า "Level")
    level_start = None
    for row_idx in range(layout["header_row"]):
        row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))[0]
        for col_idx, val in enumerate(row):
            if isinstance(val, str) and "LEVEL" in val.strip().upper():
                level_start = col_idx
                break
        if level_start is not None:
            break

    layout["level_col_start"] = level_start if level_start is not None else 0
    layout["level_col_count"] = layout["sd_col"] - layout["level_col_start"]

    # ตรวจสอบว่า row ถัดจาก header เป็น legend row (0,1,2,3...) หรือ data จริง
    next_row_idx = layout["header_row"] + 1
    rows = list(ws.iter_rows(min_row=next_row_idx, max_row=next_row_idx, values_only=True))
    if not rows:
        layout["data_start_row"] = layout["header_row"] + 1
        return layout

    next_row = rows[0]
    sd_val = next_row[layout["sd_col"]] if layout["sd_col"] < len(next_row) else None
    lcs = layout["level_col_start"]
    lcc = layout["level_col_count"]
    level_vals = [next_row[lcs + i] for i in range(lcc) if (lcs + i) < len(next_row)]
    is_legend = (sd_val is None) and all(
        isinstance(v, (int, float)) or v is None for v in level_vals
    )

    layout["data_start_row"] = (layout["header_row"] + 2) if is_legend else (layout["header_row"] + 1)
    return layout


# ─────────────────────────────────────────────
# Core parser
# ─────────────────────────────────────────────

def get_level(row, layout) -> Optional[int]:
    """หา level จาก column position ที่ไม่เป็น None"""
    lcs = layout["level_col_start"]
    lcc = layout["level_col_count"]
    for i in range(lcc):
        idx = lcs + i
        if idx < len(row) and row[idx] is not None:
            return i
    return None


def col_val(row, layout, col_key) -> str:
    idx = layout.get(col_key)
    if idx is None or idx >= len(row):
        return ""
    return safe_str(row[idx])


def parse_sheet(ws, layout, filename) -> tuple[dict, list]:
    """
    อ่าน BOM sheet แล้ว return:
      items         : dict {unique_key -> BomItem}
      relationships : list[BomRelationship]
    """
    items: dict[str, BomItem] = {}
    relationships: list[BomRelationship] = []
    parent_stack: dict[int, str] = {}  # level -> unique_key

    for row in ws.iter_rows(min_row=layout["data_start_row"], values_only=True):
        if not any(v is not None for v in row):
            continue

        level = get_level(row, layout)
        if level is None:
            continue

        sd_code = col_val(row, layout, "sd_col")
        part_no = col_val(row, layout, "pn_col")
        part_name = col_val(row, layout, "name_col")
        process = col_val(row, layout, "process_col")
        line_no = col_val(row, layout, "line_col")
        supplier = col_val(row, layout, "supplier_col")
        model = col_val(row, layout, "model_col")

        qty_raw = row[layout["qty_col"]] if layout.get("qty_col") and layout["qty_col"] < len(row) else None
        qty = parse_qty(qty_raw)

        key = get_unique_key(sd_code, part_no)
        if not key:
            continue

        # ─── Register / update item ───
        if key not in items:
            items[key] = BomItem(
                unique_key=key,
                sd_code=sd_code,
                part_no=part_no,
                part_name=part_name,
                process=process,
                line_no=line_no,
                supplier=supplier,
                model=model,
                min_level=level,
                source_files=[filename],
            )
        else:
            item = items[key]
            # ตรวจสอบความขัดแย้งของ Part No. / Part Name
            existing_pn = safe_str(item.part_no)
            existing_name = safe_str(item.part_name)
            if existing_pn and part_no and existing_pn != part_no:
                item.warnings.append(
                    f"Part No. conflict: '{existing_pn}' vs '{part_no}' (file: {filename})"
                )
            if existing_name and part_name and existing_name != part_name:
                item.warnings.append(
                    f"Part Name conflict: '{existing_name}' vs '{part_name}' (file: {filename})"
                )
            if filename not in item.source_files:
                item.source_files.append(filename)
            if level < item.min_level:
                item.min_level = level

        # ─── Register relationship ───
        if level > 0:
            parent_key = parent_stack.get(level - 1)
            if parent_key:
                relationships.append(BomRelationship(
                    parent_key=parent_key,
                    child_key=key,
                    quantity=qty,
                    process=process,
                    line_no=line_no,
                    source_file=filename,
                ))

        # ─── Update parent stack ───
        parent_stack[level] = key
        for deeper in [k for k in parent_stack if k > level]:
            del parent_stack[deeper]

    return items, relationships


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    all_items: dict[str, BomItem] = {}
    all_relationships: list[BomRelationship] = []

    files = sorted(f for f in DOCUMENT_DIR.iterdir()
                   if f.suffix in (".xlsx", ".xls") and f.is_file())

    skipped_files = []
    processed_files = []

    for fpath in files:
        fname = fpath.name
        print(f"\n📄 {fname}")

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except Exception as e:
            print(f"   ⚠ Cannot open: {e}")
            skipped_files.append({"file": fname, "reason": str(e)})
            continue

        # หา BOM sheet
        sheet_name = None
        for candidate in BOM_SHEET_CANDIDATES:
            if candidate in wb.sheetnames:
                sheet_name = candidate
                break
        # fallback: หา sheet ที่มี "BOM" ใน name
        if sheet_name is None:
            for s in wb.sheetnames:
                if "BOM" in s.upper():
                    sheet_name = s
                    break

        if sheet_name is None:
            print(f"   ⚠ No BOM sheet found (sheets: {wb.sheetnames})")
            skipped_files.append({"file": fname, "reason": "No BOM sheet"})
            wb.close()
            continue

        ws = wb[sheet_name]
        layout = detect_layout(ws)

        if layout is None:
            print(f"   ⚠ Cannot detect layout")
            skipped_files.append({"file": fname, "reason": "Layout detection failed"})
            wb.close()
            continue

        print(f"   sheet={sheet_name} | SD_col={layout['sd_col']} | "
              f"level_cols={layout['level_col_start']}~{layout['sd_col']-1} | "
              f"data_start={layout['data_start_row']}")

        items, relationships = parse_sheet(ws, layout, fname)
        wb.close()

        # Merge items
        for key, item in items.items():
            if key not in all_items:
                all_items[key] = item
            else:
                existing = all_items[key]
                existing.warnings.extend(item.warnings)
                for f in item.source_files:
                    if f not in existing.source_files:
                        existing.source_files.append(f)
                if item.min_level < existing.min_level:
                    existing.min_level = item.min_level

        all_relationships.extend(relationships)
        processed_files.append({"file": fname, "sheet": sheet_name,
                                 "items": len(items), "relationships": len(relationships)})
        print(f"   ✓ {len(items)} items, {len(relationships)} relationships")

    # ─── Deduplicate relationships ───
    # key = (parent, child) — เก็บ qty ล่าสุด (หรือ max)
    rel_map: dict[tuple, BomRelationship] = {}
    for rel in all_relationships:
        rkey = (rel.parent_key, rel.child_key)
        if rkey not in rel_map:
            rel_map[rkey] = rel
        else:
            # ถ้า qty ต่างกัน → เก็บ max (conservative)
            if rel.quantity > rel_map[rkey].quantity:
                rel_map[rkey] = rel
    unique_relationships = list(rel_map.values())

    # ─── Output: items.json ───
    items_out = []
    for item in sorted(all_items.values(), key=lambda x: (x.min_level, x.unique_key)):
        d = asdict(item)
        items_out.append(d)

    with open(OUTPUT_DIR / "items.json", "w", encoding="utf-8") as f:
        json.dump(items_out, f, ensure_ascii=False, indent=2)

    # ─── Output: relationships.json ───
    rels_out = [asdict(r) for r in unique_relationships]
    with open(OUTPUT_DIR / "relationships.json", "w", encoding="utf-8") as f:
        json.dump(rels_out, f, ensure_ascii=False, indent=2)

    # ─── Output: bom_tree.csv ───
    # สร้าง lookup สำหรับ item name
    item_lookup = {k: v for k, v in all_items.items()}

    # สร้าง children map
    children_map: dict[str, list] = {}
    for rel in unique_relationships:
        children_map.setdefault(rel.parent_key, []).append(rel)

    # หา FG items (level 0)
    fg_items = [item for item in all_items.values() if item.min_level == 0]

    def build_tree_rows(key, indent=0, visited=None):
        if visited is None:
            visited = set()
        if key in visited:
            return [{"indent": indent, "key": key, "note": "(circular ref)"}]
        visited = visited | {key}

        item = item_lookup.get(key)
        rows = [{
            "level_indent": "  " * indent + str(indent),
            "unique_key": key,
            "sd_code": item.sd_code if item else "",
            "part_no": item.part_no if item else "",
            "part_name": item.part_name if item else "",
            "process": item.process if item else "",
            "line_no": item.line_no if item else "",
            "supplier": item.supplier if item else "",
            "quantity": "",
        }]
        for rel in children_map.get(key, []):
            child_rows = build_tree_rows(rel.child_key, indent + 1, visited)
            if child_rows:
                child_rows[0]["quantity"] = rel.quantity
            rows.extend(child_rows)
        return rows

    csv_rows = []
    for fg in sorted(fg_items, key=lambda x: x.unique_key):
        csv_rows.extend(build_tree_rows(fg.unique_key))
        csv_rows.append({})  # blank line between FGs

    with open(OUTPUT_DIR / "bom_tree.csv", "w", newline="", encoding="utf-8-sig") as f:
        fieldnames = ["level_indent", "unique_key", "sd_code", "part_no",
                      "part_name", "process", "line_no", "supplier", "quantity"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in csv_rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})

    # ─── Output: summary.json ───
    items_with_warnings = [k for k, v in all_items.items() if v.warnings]
    orphan_children = [r.child_key for r in unique_relationships
                       if r.child_key not in all_items]
    missing_parents = [r.parent_key for r in unique_relationships
                       if r.parent_key not in all_items]

    summary = {
        "processed_files": len(processed_files),
        "skipped_files": len(skipped_files),
        "total_unique_items": len(all_items),
        "total_unique_relationships": len(unique_relationships),
        "fg_count": len(fg_items),
        "files_detail": processed_files,
        "skipped_detail": skipped_files,
        "data_quality": {
            "items_with_conflicts": len(items_with_warnings),
            "conflict_keys": items_with_warnings[:20],
            "orphan_children": list(set(orphan_children))[:20],
            "missing_parents": list(set(missing_parents))[:20],
        },
    }
    with open(OUTPUT_DIR / "summary.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    # ─── Print summary ───
    print("\n" + "=" * 60)
    print("✅ EXTRACTION COMPLETE")
    print(f"   Files processed : {len(processed_files)}")
    print(f"   Files skipped   : {len(skipped_files)}")
    print(f"   Unique items    : {len(all_items)}")
    print(f"   Unique FG       : {len(fg_items)}")
    print(f"   Unique BOM links: {len(unique_relationships)}")
    print(f"   Items w/conflicts: {len(items_with_warnings)}")
    print(f"\n   Output → {OUTPUT_DIR.resolve()}/")
    print("   - items.json")
    print("   - relationships.json")
    print("   - bom_tree.csv")
    print("   - summary.json")


if __name__ == "__main__":
    main()
